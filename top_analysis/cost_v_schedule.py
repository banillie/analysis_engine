'''
Creates costs v schedule graph.

Code still in development but working.

Follow instructions at end.
'''

from analysis.data import list_of_masters_all, milestone_bl_index, \
    costs_bl_index, root_path, abbreviations, a14, hs2_1, gwrm, south_west_route_capacity, \
    a358, east_coast_mainline, ewr_western, hs2_2a, thameslink, crossrail, hexagon, a417
from analysis.engine_functions import all_milestone_data_bulk
from openpyxl import Workbook
from openpyxl.chart import Series, Reference, BubbleChart
from collections import Counter
import datetime


def bubble_chart(ws, rag_count):

    chart = BubbleChart()
    chart.style = 18  # use a preset style

    # add the first series of data
    amber_stop = 2 + rag_count['Amber']
    xvalues = Reference(ws, min_col=3, min_row=3, max_row= amber_stop)
    yvalues = Reference(ws, min_col=4, min_row=3, max_row= amber_stop)
    size = Reference(ws, min_col=5, min_row=3, max_row= amber_stop)
    series = Series(values=yvalues, xvalues=xvalues, zvalues=size, title="Amber")
    chart.series.append(series)
    series.graphicalProperties.solidFill = "fce553"

    # add the second
    amber_g_stop = amber_stop + rag_count['Amber/Green']
    xvalues = Reference(ws, min_col=3, min_row= amber_stop + 1, max_row= amber_g_stop)
    yvalues = Reference(ws, min_col=4, min_row= amber_stop + 1, max_row= amber_g_stop)
    size = Reference(ws, min_col=5, min_row= amber_stop + 1, max_row= amber_g_stop)
    series = Series(values=yvalues, xvalues=xvalues, zvalues=size, title="Amber/Green")
    chart.series.append(series)
    series.graphicalProperties.solidFill = "a5b700"

    amber_r_stop = amber_g_stop + rag_count['Amber/Red']
    xvalues = Reference(ws, min_col=3, min_row=amber_g_stop + 1, max_row=amber_r_stop)
    yvalues = Reference(ws, min_col=4, min_row=amber_g_stop + 1, max_row=amber_r_stop)
    size = Reference(ws, min_col=5, min_row=amber_g_stop + 1, max_row=amber_r_stop)
    series = Series(values=yvalues, xvalues=xvalues, zvalues=size, title="Amber/Red")
    chart.series.append(series)
    series.graphicalProperties.solidFill = "f97b31"

    green_stop = amber_r_stop + rag_count['Green']
    xvalues = Reference(ws, min_col=3, min_row=amber_r_stop + 1, max_row=green_stop)
    yvalues = Reference(ws, min_col=4, min_row=amber_r_stop + 1, max_row=green_stop)
    size = Reference(ws, min_col=5, min_row=amber_r_stop + 1, max_row=green_stop)
    series = Series(values=yvalues, xvalues=xvalues, zvalues=size, title="Green")
    chart.series.append(series)
    series.graphicalProperties.solidFill = "17960c"

    red_stop = green_stop + rag_count['Red']
    xvalues = Reference(ws, min_col=3, min_row=green_stop + 1, max_row=red_stop)
    yvalues = Reference(ws, min_col=4, min_row=green_stop + 1, max_row=red_stop)
    size = Reference(ws, min_col=5, min_row=green_stop + 1, max_row=red_stop)
    series = Series(values=yvalues, xvalues=xvalues, zvalues=size, title="Red")
    chart.series.append(series)
    series.graphicalProperties.solidFill = "cb1f00"

    ws.add_chart(chart, "E1")

    return ws


def calculate_schedule_change(project_name):

    not_foc_projects = ['Crossrail Programme', 'Thameslink Programme',
                        'Hexagon', 'HS2 Phase2a', 'A14 Cambridge to Huntingdon Improvement Scheme',
                        'M4 Junctions 3 to 12 Smart Motorway', 'A303 Amesbury to Berwick Down',
                        'A358 Taunton to Southfields Dualling', 'East Coast Digital Programme',
                        'East Coast Mainline Programme', 'A428 Black Cat to Caxton Gibbet',
                        'Intercity Express Programme', 'Midland Main Line Programme',
                        'Great Western Route Modernisation (GWRM) including electrification',
                        'South West Route Capacity', 'A417 Air Balloon']

    '''full operation current date'''
    proj_milestones = all_milestone_data_bulk([project_name], list_of_masters_all[0])

    for i in range(1):
        if project_name not in not_foc_projects:
            foc_one = tuple(proj_milestones[project_name]['Full Operations'])[0]
        else:
            if project_name == 'Crossrail Programme':
                foc_one = tuple(proj_milestones[project_name]['Gateway Review 5'])[0]
            if project_name == 'Thameslink Programme':
                foc_one = tuple(proj_milestones[project_name]['Thameslink 24tph Peak Timetable'])[0]
            if project_name == 'Hexagon':
                foc_one = tuple(proj_milestones[project_name]['Full Operations'])[0]
            if project_name == 'HS2 Phase2a':
                foc_one = tuple(proj_milestones[project_name]['Start of Operation'])[0]
            if project_name == 'A14 Cambridge to Huntingdon Improvement Scheme':
                foc_one = tuple(proj_milestones[project_name]['Start of Operation'])[0]
            if project_name == 'M4 Junctions 3 to 12 Smart Motorway':
                foc_one = tuple(proj_milestones[project_name]['Project End Date'])[0]
            if project_name == 'A303 Amesbury to Berwick Down':
                foc_one = tuple(proj_milestones[project_name]['Start of Operation'])[0]
            if project_name == 'A358 Taunton to Southfields Dualling':
                foc_one = tuple(proj_milestones[project_name]['Start of Operation'])[0]
            if project_name == 'East Coast Digital Programme':
                foc_one = tuple(proj_milestones[project_name]['Project End Date'])[0]
            if project_name == 'East Coast Mainline Programme':
                foc_one = tuple(proj_milestones[project_name]['Project End Date'])[0]
            if project_name == 'A428 Black Cat to Caxton Gibbet':
                foc_one = tuple(proj_milestones[project_name]['Start of Operation'])[0]
            if project_name == 'Intercity Express Programme':
                foc_one = tuple(proj_milestones[project_name]['IEP rolling stock fully into service on GWML and ECML'])[0]
            if project_name == 'Midland Main Line Programme':
                foc_one = tuple(proj_milestones[project_name]['Start of Operation'])[0]
            if project_name == 'Great Western Route Modernisation (GWRM) including electrification':
                foc_one = tuple(proj_milestones[project_name]['Project End Date'])[0]
            if project_name == 'South West Route Capacity':
                foc_one = tuple(proj_milestones[project_name]['Start of Operation'])[0]
            if project_name == 'A417 Air Balloon':
                foc_one = tuple(proj_milestones[project_name]['Start of Operation'])[0]

    '''full operation baseline date'''
    proj_milestones_bl = all_milestone_data_bulk([project_name], list_of_masters_all[milestone_bl_index[project_name][2]])

    for i in range(1):
        try:
            sop = tuple(proj_milestones_bl[project_name]['Start of Project'])[0]
            if sop is None:
                sop = tuple(proj_milestones[project_name]['Start of Project'])[0]
        except KeyError:
            if project_name == 'Hexagon':
                sop = datetime.date(2016, 2, 1)

        if project_name not in not_foc_projects:
            foc_two = tuple(proj_milestones_bl[project_name]['Full Operations'])[0]
        else:
            if project_name == 'Crossrail Programme':
                foc_two = tuple(proj_milestones_bl[project_name]['Gateway Review 5'])[0]
            if project_name == 'Thameslink Programme':
                foc_two = tuple(proj_milestones_bl[project_name]['Thameslink 24tph Peak Timetable'])[0]
            if project_name == 'Hexagon':
                foc_two = datetime.date(2020, 5, 31)
            if project_name == 'HS2 Phase2a':
                foc_two = tuple(proj_milestones_bl[project_name]['Start of Operation'])[0]
            if project_name == 'A14 Cambridge to Huntingdon Improvement Scheme':
                foc_two = tuple(proj_milestones_bl[project_name]['Start of Operation'])[0]
            if project_name == 'M4 Junctions 3 to 12 Smart Motorway':
                foc_two = tuple(proj_milestones_bl[project_name]['Project End Date'])[0]
            if project_name == 'A303 Amesbury to Berwick Down':
                foc_two = tuple(proj_milestones_bl[project_name]['Start of Operation'])[0]
            if project_name == 'A358 Taunton to Southfields Dualling':
                foc_two = tuple(proj_milestones_bl[project_name]['Start of Operation'])[0]
            if project_name == 'East Coast Digital Programme':
                foc_two = tuple(proj_milestones_bl[project_name]['Project End Date'])[0]
            if project_name == 'East Coast Mainline Programme':
                foc_two = tuple(proj_milestones_bl[project_name]['Project End Date'])[0]
            if project_name == 'A428 Black Cat to Caxton Gibbet':
                foc_two = tuple(proj_milestones_bl[project_name]['Start of Operation'])[0]
            if project_name == 'Intercity Express Programme':
                foc_two = tuple(proj_milestones_bl[project_name]['IEP rolling stock fully into service on GWML and ECML'])[0]
            if project_name == 'Midland Main Line Programme':
                foc_two = tuple(proj_milestones_bl[project_name]['Start of Operation'])[0]
            if project_name == 'Great Western Route Modernisation (GWRM) including electrification':
                foc_two = tuple(proj_milestones_bl[project_name]['Project End Date'])[0]
            if project_name == 'South West Route Capacity':
                foc_two = tuple(proj_milestones_bl[project_name]['Start of Operation'])[0]
            if project_name == 'A417 Air Balloon':
                foc_two = tuple(proj_milestones_bl[project_name]['Start of Operation'])[0]

    try:
        proj_length = (foc_two - sop).days  # project length
    except TypeError:
        proj_length = None
    try:
        foc_change = (foc_one - foc_two).days
    except TypeError:
        foc_change = None

    try:
        percent_change = int((foc_change / proj_length) * 100)
    except TypeError:
        percent_change = 'couldn\'t calculate'

    return percent_change

def calculate_schedule_change_full_check(project_name, ws, x):
    '''this function isn't to be used but contains the workings for reaching the change figure so keeping in case
    helpful in future'''

    ws.cell(row=2, column=3).value = 'project Full Operation. NOW'
    ws.cell(row=2, column=4).value = 'project Start of Operation'
    ws.cell(row=2, column=5).value = 'project Full Operation BL'
    ws.cell(row=2, column=6).value = 'project length'
    ws.cell(row=2, column=7).value = 'length change'
    ws.cell(row=2, column=8).value = 'percentage change'

    '''full operation current date'''
    proj_milestones = all_milestone_data_bulk([project_name], list_of_masters_all[0])

    try:
        # foc_one = tuple(proj_milestones['Full Operating Capacity (FOC)'])[0]
        foc_one = tuple(proj_milestones[project_name]['Project End Date'])[0]

        if foc_one is None:
            try:
                foc_one = tuple(proj_milestones[project_name]['Full Operations'])[0]
                ws.cell(row=x + 3, column=3).value = foc_one
            except (KeyError, TypeError):
                foc_one = None
                ws.cell(row=x + 4, column=3).value = foc_one
        else:
            ws.cell(row=x + 3, column=3).value = foc_one

    except KeyError:
        foc_one = None
        ws.cell(row=x + 3, column=3).value = foc_one

    '''full operation baseline date'''
    proj_milestones_bl = all_milestone_data_bulk([project_name],
                                                 list_of_masters_all[milestone_bl_index[project_name][2]])

    try:
        sop = tuple(proj_milestones_bl[project_name]['Start of Project'])[0]
        ws.cell(row=x + 3, column=4).value = sop
    except KeyError:
        sop = None
        ws.cell(row=x + 3, column=4).value = sop

    try:
        # foc_two = tuple(proj_milestones_bl['Full Operating Capacity (FOC)'])[0]
        foc_two = tuple(proj_milestones_bl[project_name]['Project End Date'])[0]

        if foc_two is None:
            try:
                foc_two = tuple(proj_milestones_bl[project_name]['Full Operations'])[0]
                ws.cell(row=x + 3, column=5).value = foc_two
            except (KeyError, TypeError):
                foc_two = None
                ws.cell(row=x + 3, column=5).value = foc_two
        else:
            ws.cell(row=x + 3, column=5).value = foc_two
    except KeyError:
        foc_two = None
        ws.cell(row=x + 3, column=5).value = foc_two

    try:
        proj_length = (foc_two - sop).days  # project length
        ws.cell(row=x + 3, column=6).value = proj_length
    except TypeError:
        proj_length = None
        ws.cell(row=x + 3, column=6).value = proj_length
    try:
        foc_change = (foc_one - foc_two).days
        ws.cell(row=x + 3, column=7).value = foc_change
    except TypeError:
        foc_change = None
        ws.cell(row=x + 3, column=7).value = foc_change

    try:
        percent_change = int((foc_change / proj_length) * 100)
        ws.cell(row=x + 3, column=8).value = percent_change
    except TypeError:
        ws.cell(row=x + 3, column=8).value = 'couldn\'t calculate'

current_milestones_all = all_milestone_data_bulk(list_of_masters_all[0].projects,
                                                 list_of_masters_all[0])

filtered_project_list = [a14, hs2_1, gwrm, south_west_route_capacity,
                         a358, east_coast_mainline, ewr_western, hs2_2a, thameslink, crossrail, hexagon]
'''INSTRUCTIONS

Enter project list variable into function. Recommend firstly doing so for all projects (e.g. latest_quarter_project
_names) to identify projects of interest and then placing those projects into the filtered_project_list above '''

p_list = get_project_by_stage(list_of_masters_all[0], ['Full Business Case', 'Outline Business Case'])

run = cost_v_schedule_chart(p_list)
run.save(root_path/'output/cost_v_schedule_matrix_q2_2021.xlsx')