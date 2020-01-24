'''
programme that generate the information required for the dandelion chart. The output data is then sorted in excel and
placed into the dandelion graph wb.

Follow instructions below.

Note: all master data is taken from the data file. Make sure this is up to date and that all relevant data is in
the import statement.
'''

from openpyxl import Workbook
from analysis.data import list_of_masters_all, root_path

def dandelion_data():
    '''
    Simple function that returns data required for the dandelion graph. Sorting done via excel.

    :param master_data: quarter master data
    :return: excel wb
    '''

    wb = Workbook()
    ws = wb.active

    for i, project_name in enumerate(list_of_masters_all[0].projects):
        ws.cell(row=2 + i, column=1).value = list_of_masters_all[0].data[project_name]['DfT Group']
        ws.cell(row=2 + i, column=2).value = project_name
        ws.cell(row=2 + i, column=3).value = list_of_masters_all[0].data[project_name]['Total Forecast']

    ws.cell(row=1, column=1).value = 'Group'
    ws.cell(row=1, column=2).value = 'Project Name'
    ws.cell(row=1, column=3).value = 'WLC (forecast)'

    return wb

'''  RUNNING PROGRAMME '''

'''simply run the programme'''
output = dandelion_data()
output.save(root_path/'output/dandelion.xlsx')