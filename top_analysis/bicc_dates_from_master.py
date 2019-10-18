'''
THIS PROGRAMME IS NOT WORKING AND MAY NOW BE DEFUNCT AS WE NO LONGER COLLECT BICC DATES FROM PROJECTS.

Use this programme to extract the dates at bicc being reported by projects and check
whether they are correct. It produces a workbook with five columns A) project name
B) reported last at bicc date C) a column for manually editing last at bicc dates D) reported next
at bicc date E) column for manually editing next at bicc dates.

If the reported dates needs to be change they should be inserted into columns C and E. The workbook should
then be saved. If no manual edits are required then the relevant project cell should be left blank. i.e. only enter dates
into the manual edit columns if they need to be changed.

It's important that any manual edits are correctly entered into the workbook - otherwise this may cause issues when
transferring dates back into master and then eventually into the summary dashboard.

To run the programme the correct file paths need to be provided.
'''

from openpyxl import Workbook
from openpyxl.styles import Font
from analysis.data import q2_1920, q1_1920

def data_return(project_list, data_key_list, master_one, master_two):
    wb = Workbook()
    ws = wb.active

    '''lists project names in ws'''
    for x in range(0, len(project_list)):
        ws.cell(row=x + 2, column=1, value=project_list[x])

    for row_num in range(2, ws.max_row + 1):
        red_text = Font(color="00fc2525")
        project_name = ws.cell(row=row_num, column=1).value
        print(project_name)
        col_start = 2
        if project_name in master_one.projects:
            for item in data_key_list:
                ws.cell(row=row_num, column=col_start).value = master_one.data[project_name][item]
                if master_one.data[project_name][item] == None:
                    ws.cell(row=row_num, column=col_start).value = 'Not reported'
                try:
                    if master_one.data[project_name][item] != master_two.data[project_name][item]:
                        ws.cell(row=row_num, column=col_start).font = red_text
                except KeyError:
                    pass
                col_start += 2
        else:
            ws.cell(row=row_num, column=col_start).value = 'None'
            col_start += 2

    ws.cell(row=1, column=1, value='Project')
    ws.cell(row=1, column=2, value='Last @ BICC')
    ws.cell(row=1, column=3, value='Manual amend: Last @ BICC')
    ws.cell(row=1, column=4, value='Next @ BICC')
    ws.cell(row=1, column=5, value='Manual amend: Next @ BICC')

    return wb

'''  RUNNING THE PROGRAMME   '''
'''ONE. get list of projects'''
project_quarter_list = q2_1920.projects

'''TWO. Set the data of interest, which in fact won't change. 
Note this keys now need to be inserted into the master. do this at the bottom'''
data_interest = ['Last time at BICC', 'Next at BICC']

'''THREE. run the programme and specify the file path'''
run = data_return(project_quarter_list, data_interest, q2_1920, q1_1920)

'''Specify file path and name of document to be saved'''
run.save('C:\\Users\\Standalone\\general\\masters folder\\portfolio_dashboards\\q2_1920_bicc_dates_doc.xlsx')