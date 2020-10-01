'''

Programme for creating an new portfolio dashboards

input documents:
1) Dashboard master document - this is an excel file. This is the dashboard, but all data fields left blank.
Note. If project data does not get placed into the correct part of the master, check that the project name is
consistent with the name in master data, because names need to be exactly the same for information to be exported.

output document:
1) Dashboard with all project data placed into dashboard and formatted correctly.

Instructions:
1) provide path to dashboard master
3) provide path and specify file name for output document

Note: all master data is taken from the data file. Make sure this is up to date and that all relevant data is in
the import statement, and that bicc_date is correct.

TODO: refactor
'''

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from analysis.data import list_of_masters_all, root_path, ipdc_date, milestone_bl_index, \
    costs_bl_index, benefits_bl_index, bc_index
from analysis.engine_functions import all_milestone_data_bulk, convert_rag_text, convert_bc_stage_text, \
    project_time_difference, ap_p_milestone_data_bulk, concatenate_dates, highlight_close_dates_ipdc

def place_in_excel(wb):
    '''
    function that places all information into the master dashboard sheet
    :param wb:
    :return:
    '''

    financial_info(wb)

    schedule_info(wb)

    benefits_info(wb)

    overall_info(wb)

    return wb

def financial_info(wb):

    ws = wb.worksheets[0]

    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=3).value
        if project_name in list_of_masters_all[0].projects:
            '''BC Stage'''
            bc_stage = list_of_masters_all[0].data[project_name]['IPDC approval point']
            ws.cell(row=row_num, column=4).value = convert_bc_stage_text(bc_stage)
            try:
                bc_stage_lst_qrt = list_of_masters_all[0].data[project_name]['IPDC approval point']
                if bc_stage != bc_stage_lst_qrt:
                    ws.cell(row=row_num, column=4).font = Font(name='Arial', size=10, color='00fc2525')
            except KeyError:
                pass

            '''planning stage'''
            plan_stage = list_of_masters_all[0].data[project_name]['Project stage']
            ws.cell(row=row_num, column=5).value = plan_stage
            try:
                plan_stage_lst_qrt = list_of_masters_all[1].data[project_name]['Project stage']
                if plan_stage != plan_stage_lst_qrt:
                    ws.cell(row=row_num, column=5).font = Font(name='Arial', size=10, color='00fc2525')
            except KeyError:
                pass


            '''Total WLC'''
            wlc_now = list_of_masters_all[0].data[project_name]['Total Forecast']
            ws.cell(row=row_num, column=6).value = wlc_now
            '''WLC variance against lst quarter'''
            try:
                wlc_lst_quarter = list_of_masters_all[1].data[project_name]['Total Forecast']
                diff_lst_qrt = wlc_now - wlc_lst_quarter
                if float(diff_lst_qrt) > 0.49 or float(diff_lst_qrt) < -0.49:
                    ws.cell(row=row_num, column=7).value = diff_lst_qrt
                else:
                    ws.cell(row=row_num, column=7).value = '-'

                try:
                    percentage_change = ((wlc_now - wlc_lst_quarter) / wlc_now) * 100
                    if percentage_change > 5 or percentage_change < -5:
                        ws.cell(row=row_num, column=7).font = Font(name='Arial', size=10, color='00fc2525')
                except ZeroDivisionError:
                    pass

            except KeyError:
                ws.cell(row=row_num, column=7).value = '-'

            '''WLC variance against baseline quarter'''
            wlc_baseline = list_of_masters_all[costs_bl_index[project_name][2]].data[project_name]['Total Forecast']
            #print(str(project_name) + '' + str(wlc_baseline))
            try:
                diff_bl = wlc_now - wlc_baseline
                if float(diff_bl) > 0.49 or float(diff_bl) < -0.49:
                    ws.cell(row=row_num, column=8).value = diff_bl
                else:
                    ws.cell(row=row_num, column=8).value = '-'
            except TypeError: # exception is here as some projects e.g. Hs2 phase 2b have (real) written into historical totals
                pass

            try:
                percentage_change = ((wlc_now - wlc_baseline) / wlc_now) * 100
                if percentage_change > 5 or percentage_change < -5:
                    ws.cell(row=row_num, column=8).font = Font(name='Arial', size=10, color='00fc2525')
            except (ZeroDivisionError, TypeError): # zerodivision error obvious, type error handling as above
                pass

            '''Aggregate Spent'''
            '''Committed spend'''
            '''remaining'''
            '''P-Value'''

            '''Contigency'''
            ws.cell(row=row_num, column=13).value = \
                list_of_masters_all[0].data[project_name]['Overall contingency (£m)']

            '''OB'''
            ws.cell(row=row_num, column=14).value = \
                list_of_masters_all[0].data[project_name]['Overall figure for Optimism Bias (£m)']

            '''financial DCA rating - this quarter'''
            ws.cell(row=row_num, column=15).value = convert_rag_text(list_of_masters_all[0].data
                                                                     [project_name]['SRO Finance confidence'])
            '''financial DCA rating - last qrt'''
            try:
                ws.cell(row=row_num, column=16).value = convert_rag_text(list_of_masters_all[1].data
                                                                     [project_name]['SRO Finance confidence'])
            except KeyError:
                ws.cell(row=row_num, column=16).value = ''
            '''financial DCA rating - 2 qrts ago'''
            try:
                ws.cell(row=row_num, column=17).value = convert_rag_text(list_of_masters_all[2].data
                                                                     [project_name]['SRO Finance confidence'])
            except KeyError:
                ws.cell(row=row_num, column=17).value = ''
            '''financial DCA rating - 3 qrts ago'''
            try:
                ws.cell(row=row_num, column=18).value = convert_rag_text(list_of_masters_all[3].data
                                                                     [project_name]['SRO Finance confidence'])
            except KeyError:
                ws.cell(row=row_num, column=18).value = ''
            '''financial DCA rating - baseline'''
            ws.cell(row=row_num, column=19).value = \
                convert_rag_text(list_of_masters_all[costs_bl_index[project_name][2]].data[project_name]
                                 ['SRO Finance confidence'])

    '''list of columns with conditional formatting'''
    list_columns = ['o', 'p', 'q', 'r', 's']

    '''same loop but the text is black. In addition these two loops go through the list_columns list above'''
    for column in list_columns:
        for i, dca in enumerate(rag_txt_list):
            text = black_text
            fill = fill_colour_list[i]
            dxf = DifferentialStyle(font=text, fill=fill)
            rule = Rule(type="containsText", operator="containsText", text=dca, dxf=dxf)
            for_rule_formula = 'NOT(ISERROR(SEARCH("' + dca + '",' + column + '5)))'
            rule.formula = [for_rule_formula]
            ws.conditional_formatting.add('' + column + '5:' + column + '60', rule)

    for row_num in range(2, ws.max_row + 1):
        for col_num in range(5, ws.max_column+1):
            if ws.cell(row=row_num, column=col_num).value == 0:
                ws.cell(row=row_num, column=col_num).value = '-'


    return wb

def schedule_info(wb):

    ws = wb.worksheets[1]

    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=3).value
        if project_name in list_of_masters_all[0].projects:
            '''IPDC approval point'''
            bc_stage = list_of_masters_all[0].data[project_name]['IPDC approval point']
            ws.cell(row=row_num, column=4).value = convert_bc_stage_text(bc_stage)
            try:
                bc_stage_lst_qrt = list_of_masters_all[1].data[project_name]['IPDC approval point']
                if bc_stage != bc_stage_lst_qrt:
                    ws.cell(row=row_num, column=4).font = Font(name='Arial', size=10, color='00fc2525')
            except KeyError:
                pass
            '''stage'''
            plan_stage = list_of_masters_all[0].data[project_name]['Project stage']
            ws.cell(row=row_num, column=5).value = plan_stage
            try:
                plan_stage_lst_qrt = list_of_masters_all[1].data[project_name]['Project stage']
                if plan_stage != plan_stage_lst_qrt:
                    ws.cell(row=row_num, column=5).font = Font(name='Arial', size=10, color='00fc2525')
            except KeyError:
                pass

            '''Next milestone name and variance'''
            local_milestone_dates = []
            for x, key in enumerate(current_milestones_ap_p[project_name].keys()):
                date = tuple(current_milestones_ap_p[project_name][key])[0]
                if date is None:
                    pass
                elif date > ipdc_date:
                        local_milestone_dates.append((date, x))

            if len(local_milestone_dates) != 0: # checks if list is empty
                try:
                    next_milestone_name = list(current_milestones_ap_p[project_name].keys())[local_milestone_dates[0][1]]
                    ws.cell(row=row_num, column=6).value = next_milestone_name
                    next_milestone_date = local_milestone_dates[0][0]
                    ws.cell(row=row_num, column=7).value = next_milestone_date
                except (TypeError, KeyError):
                    ws.cell(row=row_num, column=6).value = 'None Scheduled'
                try:
                    lst_qrt_diff = first_diff_data[project_name][next_milestone_name]
                    ws.cell(row=row_num, column=8).value = lst_qrt_diff
                    if lst_qrt_diff > 25:
                        ws.cell(row=row_num, column=8).font = Font(name='Arial', size=10, color='00fc2525')
                except (TypeError, KeyError):
                    ws.cell(row=row_num, column=8).value = ''
                try:
                    bl_qrt_diff = second_diff_data[project_name][next_milestone_name]
                    ws.cell(row=row_num, column=9).value = bl_qrt_diff
                    if bl_qrt_diff > 46:
                        ws.cell(row=row_num, column=9).font = Font(name='Arial', size=10, color='00fc2525')
                except (TypeError, KeyError):
                    ws.cell(row=row_num, column=9).value = ''
            else:
                pass


            '''start of construction (soc) current date'''
            try:
                current_soc = tuple(current_milestones_all[project_name]['Start of Construction/build'])[0]
                ws.cell(row=row_num, column=10).value = current_soc
                if current_soc < ipdc_date:
                    ws.cell(row=row_num, column=10).value = 'Completed'
            except (KeyError, TypeError):
                ws.cell(row=row_num, column=10).value = ''
            '''soc variance against lst quarter'''
            try:
                soc_lst_qrt_diff = first_diff_data[project_name]['Start of Construction/build']
                ws.cell(row=row_num, column=11).value = soc_lst_qrt_diff
                if soc_lst_qrt_diff > 46:
                    ws.cell(row=row_num, column=11).font = Font(name='Arial', size=10, color='00fc2525')
            except (KeyError, TypeError):
                ws.cell(row=row_num, column=11).value = ''
            '''soc variance against baseline'''
            try:
                soc_bl_diff = second_diff_data[project_name]['Start of Construction/build']
                ws.cell(row=row_num, column=12).value = soc_bl_diff
                if soc_bl_diff > 85:
                    ws.cell(row=row_num, column=12).font = Font(name='Arial', size=10, color='00fc2525')
            except (KeyError, TypeError):
                ws.cell(row=row_num, column=12).value = ''

            '''start of operation (sop) current date'''
            try:
                current_sop = tuple(current_milestones_all[project_name]['Start of Operation'])[0]
                ws.cell(row=row_num, column=13).value = current_sop
                if current_sop < ipdc_date:
                    ws.cell(row=row_num, column=13).value = 'Completed'
            except (KeyError, TypeError):
                ws.cell(row=row_num, column=13).value = ''
            '''sop variance against lst quarter'''
            try:
                sop_lst_qrt_diff = first_diff_data[project_name]['Start of Operation']
                ws.cell(row=row_num, column=14).value = sop_lst_qrt_diff
                if sop_lst_qrt_diff > 46:
                    ws.cell(row=row_num, column=14).font = Font(name='Arial', size=10, color='00fc2525')
            except (KeyError, TypeError):
                ws.cell(row=row_num, column=14).value = ''
            '''sop variance against baseline'''
            try:
                sop_bl_diff = second_diff_data[project_name]['Start of Operation']
                ws.cell(row=row_num, column=15).value = sop_bl_diff
                if sop_bl_diff > 86:
                    ws.cell(row=row_num, column=15).font = Font(name='Arial', size=10, color='00fc2525')
            except (KeyError, TypeError):
                ws.cell(row=row_num, column=15).value = ''

            '''full operation current date'''
            try:
                foc = tuple(current_milestones_all[project_name]['Full Operations'])[0]
                ws.cell(row=row_num, column=16).value = foc
                if foc < ipdc_date:
                    ws.cell(row=row_num, column=16).value = 'Completed'
                else:
                    ws.cell(row=row_num, column=16).value = foc
            except (KeyError, TypeError):
                ws.cell(row=row_num, column=16).value = ''

            '''fop against lst quarter'''
            try:
                foc_lst_qrt_diff = first_diff_data[project_name]['Full Operations']
                ws.cell(row=row_num, column=17).value = foc_lst_qrt_diff
                if foc_lst_qrt_diff > 46:
                    ws.cell(row=row_num, column=17).font = Font(name='Arial', size=10, color='00fc2525')
            except (KeyError, TypeError):
                ws.cell(row=row_num, column=17).value = ''
            '''fop against baseline'''
            try:
                foc_bl_diff = second_diff_data[project_name]['Full Operations']
                ws.cell(row=row_num, column=18).value = foc_bl_diff
                if foc_bl_diff > 86:
                    ws.cell(row=row_num, column=18).font = Font(name='Arial', size=10, color='00fc2525')
            except (KeyError, TypeError):
                ws.cell(row=row_num, column=18).value = ''

            '''project end date'''
            try:
                ped = tuple(current_milestones_all[project_name]['Project End Date'])[0]
                ws.cell(row=row_num, column=19).value = ped
            except (KeyError, TypeError):
                ws.cell(row=row_num, column=19).value = ''
            '''ped variance against lst quarter'''
            try:
                ped_lst_qrt_diff = first_diff_data[project_name]['Project End Date']
                ws.cell(row=row_num, column=20).value = ped_lst_qrt_diff
                if ped_lst_qrt_diff > 46:
                    ws.cell(row=row_num, column=20).font = Font(name='Arial', size=10, color='00fc2525')
            except (KeyError, TypeError):
                ws.cell(row=row_num, column=20).value = ''
            '''ped variance against baseline'''
            try:
                ped_bl_diff = second_diff_data[project_name]['Project End Date']
                ws.cell(row=row_num, column=21).value = ped_bl_diff
                if ped_bl_diff > 86:
                    ws.cell(row=row_num, column=21).font = Font(name='Arial', size=10, color='00fc2525')
            except (KeyError, TypeError):
                ws.cell(row=row_num, column=21).value = ''

            '''schedule DCA rating - this quarter'''
            ws.cell(row=row_num, column=22).value = convert_rag_text(list_of_masters_all[0].data
                                                                     [project_name]['SRO Schedule Confidence'])
            '''schedule DCA rating - last qrt'''
            try:
                ws.cell(row=row_num, column=23).value = convert_rag_text(list_of_masters_all[1].data
                                                                         [project_name]['SRO Schedule Confidence'])
            except KeyError:
                ws.cell(row=row_num, column=23).value = ''
            '''schedule DCA rating - 2 qrts ago'''
            try:
                ws.cell(row=row_num, column=24).value = convert_rag_text(list_of_masters_all[2].data
                                                                         [project_name]['SRO Schedule Confidence'])
            except KeyError:
                ws.cell(row=row_num, column=24).value = ''
            '''schedule DCA rating - 3 qrts ago'''
            try:
                ws.cell(row=row_num, column=25).value = convert_rag_text(list_of_masters_all[3].data
                                                                         [project_name]['SRO Schedule Confidence'])
            except KeyError:
                ws.cell(row=row_num, column=25).value = ''
            '''schedule DCA rating - baseline'''
            try:
                ws.cell(row=row_num, column=26).value = \
                    convert_rag_text(list_of_masters_all[milestone_bl_index[project_name][2]].data[project_name]
                                     ['SRO Schedule Confidence'])
            except:
                ws.cell(row=row_num, column=26).value = ''

    '''list of columns with conditional formatting'''
    list_columns = ['v', 'w', 'x', 'y', 'z']

    '''same loop but the text is black. In addition these two loops go through the list_columns list above'''
    for column in list_columns:
        for i, dca in enumerate(rag_txt_list):
            text = black_text
            fill = fill_colour_list[i]
            dxf = DifferentialStyle(font=text, fill=fill)
            rule = Rule(type="containsText", operator="containsText", text=dca, dxf=dxf)
            for_rule_formula = 'NOT(ISERROR(SEARCH("' + dca + '",' + column + '5)))'
            rule.formula = [for_rule_formula]
            ws.conditional_formatting.add('' + column + '5:' + column + '60', rule)

    for row_num in range(2, ws.max_row + 1):
        for col_num in range(5, ws.max_column+1):
            if ws.cell(row=row_num, column=col_num).value == 0:
                ws.cell(row=row_num, column=col_num).value = '-'

    return wb

def benefits_info(wb):

    ws = wb.worksheets[2]

    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=3).value
        if project_name in list_of_masters_all[0].projects:

            '''BICC approval point'''
            bc_stage = list_of_masters_all[0].data[project_name]['IPDC approval point']
            ws.cell(row=row_num, column=4).value = convert_bc_stage_text(bc_stage)
            try:
                bc_stage_lst_qrt = list_of_masters_all[1].data[project_name]['IPDC approval point']
                if bc_stage != bc_stage_lst_qrt:
                    ws.cell(row=row_num, column=4).font = Font(name='Arial', size=10, color='00fc2525')
            except:
                pass
            '''Next stage'''
            proj_stage = list_of_masters_all[0].data[project_name]['Project stage']
            ws.cell(row=row_num, column=5).value = proj_stage
            try:
                proj_stage_lst_qrt = list_of_masters_all[1].data[project_name]['Project stage']
                if proj_stage != proj_stage_lst_qrt:
                    ws.cell(row=row_num, column=5).font = Font(name='Arial', size=10, color='00fc2525')
            except:
                pass

            '''initial bcr'''
            initial_bcr = list_of_masters_all[0].data[project_name]['Initial Benefits Cost Ratio (BCR)']
            ws.cell(row=row_num, column=6).value = initial_bcr
            '''initial bcr baseline'''
            try:
                baseline_initial_bcr = \
                    list_of_masters_all[benefits_bl_index[project_name][2]].data[project_name]['Initial Benefits Cost Ratio (BCR)']
                ws.cell(row=row_num, column=7).value = baseline_initial_bcr
                if initial_bcr != baseline_initial_bcr:
                    if baseline_initial_bcr is None:
                        pass
                    else:
                        ws.cell(row=row_num, column=6).font = Font(name='Arial', size=10, color='00fc2525')
                        ws.cell(row=row_num, column=7).font = Font(name='Arial', size=10, color='00fc2525')
            except TypeError:
                ws.cell(row=row_num, column=7).value = ''

            '''adjusted bcr'''
            adjusted_bcr = list_of_masters_all[0].data[project_name]['Adjusted Benefits Cost Ratio (BCR)']
            ws.cell(row=row_num, column=8).value = adjusted_bcr
            '''adjusted bcr baseline'''
            try:
                baseline_adjusted_bcr = \
                    list_of_masters_all[benefits_bl_index[project_name][2]].data[project_name]['Adjusted Benefits Cost Ratio (BCR)']
                ws.cell(row=row_num, column=9).value = baseline_adjusted_bcr
                if adjusted_bcr != baseline_adjusted_bcr:
                    if baseline_adjusted_bcr is None:
                        pass
                    else:
                        ws.cell(row=row_num, column=8).font = Font(name='Arial', size=10, color='00fc2525')
                        ws.cell(row=row_num, column=9).font = Font(name='Arial', size=10, color='00fc2525')
            except TypeError:
                ws.cell(row=row_num, column=9).value = ''

            '''vfm category now'''
            if list_of_masters_all[0].data[project_name]['VfM Category single entry'] is None:
                vfm_cat = str(list_of_masters_all[0].data[project_name]['VfM Category lower range']) + ' - ' + \
                          str(list_of_masters_all[0].data[project_name]['VfM Category upper range'])
                ws.cell(row=row_num, column=10).value = vfm_cat
            else:
                vfm_cat = list_of_masters_all[0].data[project_name]['VfM Category single entry']
                ws.cell(row=row_num, column=10).value = vfm_cat

            '''vfm category baseline'''
            try:
                if list_of_masters_all[benefits_bl_index[project_name][2]].data[project_name]['VfM Category single entry'] is None:
                    vfm_cat_baseline = str(list_of_masters_all[benefits_bl_index[project_name][2]].data[project_name][
                                               'VfM Category lower range']) + ' - ' + \
                                       str(list_of_masters_all[benefits_bl_index[project_name][2]].data[project_name][
                                               'VfM Category upper range'])
                    ws.cell(row=row_num, column=11).value = vfm_cat_baseline
                else:
                    vfm_cat_baseline = list_of_masters_all[benefits_bl_index[project_name[2]]].data[project_name][
                        'VfM Category single entry']
                    ws.cell(row=row_num, column=11).value = vfm_cat_baseline

            except KeyError:
                try:
                    vfm_cat_baseline = list_of_masters_all[benefits_bl_index[project_name][2]].data[project_name]['VfM Category single entry']
                    ws.cell(row=row_num, column=11).value = vfm_cat_baseline
                except KeyError:
                    vfm_cat_baseline = list_of_masters_all[benefits_bl_index[project_name][2]].data[project_name]['VfM Category']
                    ws.cell(row=row_num, column=11).value = vfm_cat_baseline

            if vfm_cat != vfm_cat_baseline:
                if vfm_cat_baseline is None:
                    pass
                else:
                    ws.cell(row=row_num, column=10).font = Font(name='Arial', size=10, color='00fc2525')
                    ws.cell(row=row_num, column=11).font = Font(name='Arial', size=10, color='00fc2525')

            '''total monetised benefits'''
            tmb = list_of_masters_all[0].data[project_name]['Total BEN Forecast - Total Monetised Benefits']
            ws.cell(row=row_num, column=12).value = tmb
            '''tmb variance'''
            baseline_tmb = list_of_masters_all[benefits_bl_index[project_name][2]].data[project_name]['Total BEN Forecast - Total Monetised Benefits']
            ws.cell(row=row_num, column=13).value = tmb - baseline_tmb
            try:
                percentage_change = ((tmb - baseline_tmb) / tmb) * 100
                if percentage_change > 5 or percentage_change < -5:
                    ws.cell(row=row_num, column=13).font = Font(name='Arial', size=10, color='00fc2525')
            except ZeroDivisionError:
                pass

            '''in year benefits'''

            '''benefits DCA rating - this quarter'''
            ws.cell(row=row_num, column=16).value = convert_rag_text(list_of_masters_all[0].data
                                                                     [project_name]['SRO Benefits RAG'])
            '''benefits DCA rating - last qrt'''
            try:
                ws.cell(row=row_num, column=17).value = convert_rag_text(list_of_masters_all[1].data
                                                                         [project_name]['SRO Benefits RAG'])
            except KeyError:
                ws.cell(row=row_num, column=17).value = ''
            '''benefits DCA rating - 2 qrts ago'''
            try:
                ws.cell(row=row_num, column=18).value = convert_rag_text(list_of_masters_all[2].data
                                                                         [project_name]['SRO Benefits RAG'])
            except KeyError:
                ws.cell(row=row_num, column=18).value = ''
            '''benefits DCA rating - 3 qrts ago'''
            try:
                ws.cell(row=row_num, column=19).value = convert_rag_text(list_of_masters_all[3].data
                                                                         [project_name]['SRO Benefits RAG'])
            except KeyError:
                ws.cell(row=row_num, column=19).value = ''
            '''benefits DCA rating - baseline'''
            try:
                ws.cell(row=row_num, column=20).value = \
                    convert_rag_text(list_of_masters_all[benefits_bl_index[project_name][2]].data[project_name]
                                     ['SRO Benefits RAG'])
            except:
                ws.cell(row=row_num, column=20).value = ''

        '''list of columns with conditional formatting'''
        list_columns = ['p', 'q', 'r', 's', 't']

        '''loops below place conditional formatting (cf) rules into the wb. There are two as the dashboard currently has 
        two distinct sections/headings, which do not require cf. Therefore, cf starts and ends at the stated rows. this
        is hard code that will need to be changed should the position of information in the dashboard change. It is an
        easy change however'''

        '''same loop but the text is black. In addition these two loops go through the list_columns list above'''
        for column in list_columns:
            for i, dca in enumerate(rag_txt_list):
                text = black_text
                fill = fill_colour_list[i]
                dxf = DifferentialStyle(font=text, fill=fill)
                rule = Rule(type="containsText", operator="containsText", text=dca, dxf=dxf)
                for_rule_formula = 'NOT(ISERROR(SEARCH("' + dca + '",' + column + '5)))'
                rule.formula = [for_rule_formula]
                ws.conditional_formatting.add('' + column + '5:' + column + '60', rule)

    for row_num in range(2, ws.max_row + 1):
        for col_num in range(5, ws.max_column+1):
            if ws.cell(row=row_num, column=col_num).value == 0:
                ws.cell(row=row_num, column=col_num).value = '-'

    return wb

def overall_info(wb):
    ws = wb.worksheets[3]

    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=2).value
        if project_name in list_of_masters_all[0].projects:
            '''BC Stage'''
            bc_stage = list_of_masters_all[0].data[project_name]['IPDC approval point']
            ws.cell(row=row_num, column=3).value = convert_bc_stage_text(bc_stage)
            try:
                bc_stage_lst_qrt = list_of_masters_all[1].data[project_name]['IPDC approval point']
                if bc_stage != bc_stage_lst_qrt:
                    ws.cell(row=row_num, column=3).font = Font(name='Arial', size=10, color='00fc2525')
            except KeyError:
                pass

            '''planning stage'''
            plan_stage = list_of_masters_all[0].data[project_name]['Project stage']
            ws.cell(row=row_num, column=4).value = plan_stage
            try:
                plan_stage_lst_qrt = list_of_masters_all[1].data[project_name]['Project stage']
                if plan_stage != plan_stage_lst_qrt:
                    ws.cell(row=row_num, column=4).font = Font(name='Arial', size=10, color='00fc2525')
            except KeyError:
                pass

            '''Total WLC'''
            wlc_now = list_of_masters_all[0].data[project_name]['Total Forecast']
            ws.cell(row=row_num, column=5).value = wlc_now
            '''WLC variance against lst quarter'''
            try:
                wlc_lst_quarter = list_of_masters_all[1].data[project_name]['Total Forecast']
                diff_lst_qrt = wlc_now - wlc_lst_quarter
                if float(diff_lst_qrt) > 0.49 or float(diff_lst_qrt) < -0.49:
                    ws.cell(row=row_num, column=6).value = diff_lst_qrt
                else:
                    ws.cell(row=row_num, column=6).value = '-'

                try:
                    percentage_change = ((wlc_now - wlc_lst_quarter) / wlc_now) * 100
                    if percentage_change > 5 or percentage_change < -5:
                        ws.cell(row=row_num, column=6).font = Font(name='Arial', size=10, color='00fc2525')
                except ZeroDivisionError:
                    pass

            except KeyError:
                ws.cell(row=row_num, column=6).value = '-'

            '''WLC variance against baseline quarter'''
            wlc_baseline = list_of_masters_all[costs_bl_index[project_name][2]].data[project_name][
                'Total Forecast']
            try:
                diff_bl = wlc_now - wlc_baseline
                if float(diff_bl) > 0.49 or float(diff_bl) < -0.49:
                    ws.cell(row=row_num, column=7).value = diff_bl
                else:
                    ws.cell(row=row_num, column=7).value = '-'
                #print(project_name)
                try:
                    percentage_change = ((wlc_now - wlc_baseline) / wlc_now) * 100
                    if percentage_change > 5 or percentage_change < -5:
                        ws.cell(row=row_num, column=7).font = Font(name='Arial', size=10, color='00fc2525')
                except ZeroDivisionError:
                    pass

            except TypeError:
                pass

            '''vfm category now'''
            if list_of_masters_all[0].data[project_name]['VfM Category single entry'] is None:
                vfm_cat = str(list_of_masters_all[0].data[project_name]['VfM Category lower range']) + ' - ' + \
                          str(list_of_masters_all[0].data[project_name]['VfM Category upper range'])
                ws.cell(row=row_num, column=8).value = vfm_cat
            else:
                vfm_cat = list_of_masters_all[0].data[project_name]['VfM Category single entry']
                ws.cell(row=row_num, column=8).value = vfm_cat

            '''vfm category last quarter'''
            try:
                if list_of_masters_all[1].data[project_name]['VfM Category single entry'] is None:
                    vfm_cat_last_quarter = str(list_of_masters_all[1].data[project_name][
                                               'VfM Category lower range']) + ' - ' + \
                                       str(list_of_masters_all[1].data[project_name][
                                               'VfM Category upper range'])
                else:
                    vfm_cat_last_quarter = list_of_masters_all[1].data[project_name][
                        'VfM Category single entry']

            except KeyError:
                # handles projects that have been reporting for latest quarter only
                try:
                    vfm_cat_last_quarter = list_of_masters_all[costs_bl_index[project_name][2]].data[project_name][
                        'VfM Category single entry']
                except KeyError:
                    # try:
                    vfm_cat_last_quarter = list_of_masters_all[costs_bl_index[project_name][2]].data[project_name]['VfM Category']
                    # except:
                    #     vfm_cat_last_quarter = None

            if vfm_cat != vfm_cat_last_quarter:
                # if vfm_cat_last_quarter is None:
                #     pass
                # else:
                    ws.cell(row=row_num, column=8).font = Font(name='Arial', size=10, color='00fc2525')

            '''full operation current date'''
            try:
                foc = tuple(current_milestones_all[project_name]['Full Operations'])[0]
                ws.cell(row=row_num, column=9).value = foc
                if foc < ipdc_date:
                    ws.cell(row=row_num, column=9).value = 'Completed'
                else:
                    ws.cell(row=row_num, column=9).value = foc
            except (KeyError, TypeError):
                ws.cell(row=row_num, column=9).value = ''

            '''fop against lst quarter'''
            try:
                foc_lst_qrt_diff = first_diff_data[project_name]['Full Operations']
                ws.cell(row=row_num, column=10).value = foc_lst_qrt_diff
                if foc_lst_qrt_diff > 46:
                    ws.cell(row=row_num, column=10).font = Font(name='Arial', size=10, color='00fc2525')
            except (KeyError, TypeError):
                ws.cell(row=row_num, column=10).value = ''
            '''fop against baseline'''
            try:
                foc_bl_diff = second_diff_data[project_name]['Full Operations']
                ws.cell(row=row_num, column=11).value = foc_bl_diff
                if foc_bl_diff > 86:
                    ws.cell(row=row_num, column=11).font = Font(name='Arial', size=10, color='00fc2525')
            except (KeyError, TypeError):
                ws.cell(row=row_num, column=11).value = ''


            try:
                ws.cell(row=row_num, column=12).value = highlight_close_dates_ipdc\
                    (concatenate_dates(list_of_masters_all[0].data[project_name]['Last time at BICC'], ipdc_date))
                ws.cell(row=row_num, column=13).value = highlight_close_dates_ipdc\
                    (concatenate_dates(list_of_masters_all[0].data[project_name]['Next at BICC'], ipdc_date))
            except (KeyError, TypeError):
                print(project_name + ' last at / next at ipdc data could not be calculated. Check data.')


            '''IPA DCA rating'''
            ipa_dca = convert_rag_text(list_of_masters_all[0].data[project_name]['GMPP - IPA DCA'])
            ws.cell(row=row_num, column=15).value = ipa_dca
            if ipa_dca == 'None':
                ws.cell(row=row_num, column=15).value = ''

            '''DCA rating - this quarter'''
            ws.cell(row=row_num, column=17).value = convert_rag_text(list_of_masters_all[0].data
                                                                     [project_name]['Departmental DCA'])
            '''DCA rating - last qrt'''
            try:
                ws.cell(row=row_num, column=19).value = convert_rag_text(list_of_masters_all[1].data
                                                                         [project_name]['Departmental DCA'])
            except KeyError:
                ws.cell(row=row_num, column=19).value = ''
            '''DCA rating - 2 qrts ago'''
            try:
                ws.cell(row=row_num, column=20).value = convert_rag_text(list_of_masters_all[2].data
                                                                         [project_name]['Departmental DCA'])
            except KeyError:
                ws.cell(row=row_num, column=20).value = ''
            '''DCA rating - 3 qrts ago'''
            try:
                ws.cell(row=row_num, column=21).value = convert_rag_text(list_of_masters_all[3].data
                                                                         [project_name]['Departmental DCA'])
            except KeyError:
                ws.cell(row=row_num, column=21).value = ''
            '''DCA rating - baseline'''
            try:
                ws.cell(row=row_num, column=23).value = \
                    convert_rag_text(list_of_masters_all[bc_index[project_name][2]].data[project_name]
                                     ['Departmental DCA'])
            except:
                ws.cell(row=row_num, column=23).value = ''

        '''list of columns with conditional formatting'''
        list_columns = ['o', 'q', 's', 't', 'u', 'w']

        '''same loop but the text is black. In addition these two loops go through the list_columns list above'''
        for column in list_columns:
            for i, dca in enumerate(rag_txt_list):
                text = black_text
                fill = fill_colour_list[i]
                dxf = DifferentialStyle(font=text, fill=fill)
                rule = Rule(type="containsText", operator="containsText", text=dca, dxf=dxf)
                for_rule_formula = 'NOT(ISERROR(SEARCH("' + dca + '",' + column + '5)))'
                rule.formula = [for_rule_formula]
                ws.conditional_formatting.add(column + '5:' + column + '60', rule)

        for row_num in range(2, ws.max_row + 1):
            for col_num in range(5, ws.max_column + 1):
                if ws.cell(row=row_num, column=col_num).value == 0:
                    ws.cell(row=row_num, column=col_num).value = '-'

    return wb


'''highlight cells that contain RAG text, with background and text the same colour'''

'''store of different colours'''
ag_text = Font(color="92a700") # text same colour as background
ag_fill = PatternFill(bgColor="92a700")
ar_text = Font(color="e77200") # text same colour as background
ar_fill = PatternFill(bgColor="e77200")
red_text = Font(color="c00000") # text same colour as background
red_fill = PatternFill(bgColor="c00000")
green_text = Font(color="007d00") # text same colour as background
green_fill = PatternFill(bgColor="007d00")
amber_text = Font(color="ffba00") # text same colour as background
amber_fill = PatternFill(bgColor="ffba00")

black_text = Font(color="00000000")

'''NOTE. these three lists need to have rag ratings in the same order'''
'''different colours are placed into a list'''
txt_colour_list = [ag_text, ar_text, red_text, green_text, amber_text]
fill_colour_list = [ag_fill, ar_fill, red_fill, green_fill, amber_fill]
'''list of how rag ratings are shown in document'''
rag_txt_list = ["A/G", "A/R", "R", "G", "A"]



'''python dictionary of all project milestone for the latest and last quarter are put into variables here, because these
are single source calculations. The baseline milestones are calcuated in functions and project baseline quarter are 
project specific'''
current_milestones_all = all_milestone_data_bulk(list_of_masters_all[0].projects, list_of_masters_all[0])
last_qrt_milestone_all = all_milestone_data_bulk(list_of_masters_all[0].projects, list_of_masters_all[1])
current_milestones_ap_p = ap_p_milestone_data_bulk(list_of_masters_all[0].projects, list_of_masters_all[0])

'''calcualting milestone time deltas'''
current_milestones_data = {}
last_milestones_data = {}
oldest_milestones_data = {}
for project_name in list_of_masters_all[0].projects:
    p_current_milestones_data = all_milestone_data_bulk([project_name], list_of_masters_all[0])
    current_milestones_data.update(p_current_milestones_data)
    p_last_milestones_data = all_milestone_data_bulk([project_name], list_of_masters_all[1])
    last_milestones_data.update(p_last_milestones_data)
    p_oldest_milestones_data = all_milestone_data_bulk([project_name], list_of_masters_all[milestone_bl_index[project_name][2]])
    oldest_milestones_data.update(p_oldest_milestones_data)

    '''calculate time current and last quarter'''
    first_diff_data = project_time_difference(current_milestones_data, last_milestones_data)
    second_diff_data = project_time_difference(current_milestones_data, oldest_milestones_data)

''' RUNNING THE PROGRAMME '''

'''ONE. Provide file path to dashboard master'''
dashboard_master = load_workbook(root_path/'input/dashboards_master_q1_20_21.xlsx')

'''THREE. place arguments into the place_in_excle function and provide file path for saving output wb'''
dashboard_completed = place_in_excel(dashboard_master)
dashboard_completed.save(root_path/'output/dashboards_q1_2021_populated.xlsx')