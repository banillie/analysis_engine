'''

Programme for creating an aggregate project dashboard

input documents:
1) Dashboard master document - this is an excel file. This is the dashboard, but all data fields left blank.
Note. If project data does not get placed into the correct part of the master, check that the project name is
consistent with the name in master data, because names need to be exactly the same for information to be exported.

output document:
1) Dashboard with all project data placed into dashboard and formatted correctly.

Instructions:
1) provide path to dashboard master
2) change bicc_date variable
3) provide path and specify file name for output document

Supplementary instructions:
These things need to be done to check and assure the data going into the dashboard. Use the other programmes available
for undertaking these tasks.
1) Check that project stage/last at BICC data is correct. This is done via the bc_stage_from_master and
bc_amended_to_master programmes.
2) insert into the master document last at / next at BICC project data. This is done via the bicc_dates_from_master and
bicc_dates_amended_to_master programmes.

Note that some manual adjustments need to be made to:
1) Project WLC totals e.g. Hs2 Phases
2) The last/next at BICC specification. e.g. Hs2 Prog should be changed to 'often'


NOTE - code should ideally be refactored in line with latest structure as sort of hacked together at mo to accommodate
the change in data keys (i.e. new template).

Note: all master data is taken from the data file. Make sure this is up to date and that all relevant data is in
the import statement.
'''

from openpyxl import load_workbook
import datetime
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, IconSet, FormatObject
from analysis.data import q2_1920, q1_1920
from analysis.engine_functions import all_milestone_data_bulk, convert_rag_text, concatenate_dates, up_or_down, \
    convert_bc_stage_text

def highlight_close_dates_bicc(concate_calculation):
    '''
    function that further converts dates that have been concatenated so it is clearer whether/when the project
    was last at BICC
    :param concate_calculation: a concatenated date. provided by concatenate_dates function
    :return: relevant text for dashboard
    '''

    if concate_calculation == '-2 weeks':
        return 'Last BICC'
    if concate_calculation == '2 weeks':
        return 'Next BICC'
    if concate_calculation == 'Today':
        return 'This BICC'
    if concate_calculation == '-2 weeks':
        return 'Last BICC'
    if concate_calculation == '2 weeks':
        return 'Next BICC'
    if concate_calculation == 'Today':
        return 'This BICC'
    else:
        return concate_calculation

def place_in_excel(master_one, master_two, wb):
    '''
    function that places all information into the master dashboard sheet
    :param master_one:
    :param master_two:
    :return:
    '''

    ws = wb.active

    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=3).value
        print(project_name)
        if project_name in master_one.projects:
            ws.cell(row=row_num, column=4).value = master_one.data[project_name]['Total Forecast']
            try:
                dca_now = master_one.data[project_name]['Departmental DCA']
                dca_past = master_two.data[project_name]['Departmental DCA']
                ws.cell(row=row_num, column=6).value = up_or_down(dca_now, dca_past)
            except KeyError:
                ws.cell(row=row_num, column=6).value = 'NEW'
            ws.cell(row=row_num, column=7).value = convert_rag_text(master_one.data[project_name]['Departmental DCA'])
            ws.cell(row=row_num, column=8).value = convert_rag_text(master_one.data[project_name]
                                                                    ['GMPP - IPA DCA last quarter'])
            ws.cell(row=row_num, column=9).value = convert_bc_stage_text(master_one.data[project_name]
                                                                         ['BICC approval point'])

            p_m_data = all_milestone_data_bulk([project_name], master_one)
            try:
                ws.cell(row=row_num, column=10).value = concatenate_dates\
                    (list(p_m_data[project_name]['Start of Operation'])[0])
            except KeyError:
                ws.cell(row=row_num, column=10).value = 'no data'
            try:
                ws.cell(row=row_num, column=11).value = concatenate_dates\
                    (list(p_m_data[project_name]['Project End Date'])[0])
            except KeyError:
                ws.cell(row=row_num, column=10).value = 'no data'

            ws.cell(row=row_num, column=12).value = convert_rag_text(master_one.data[project_name]
                                                                     ['SRO Finance confidence'])
            try:
                ws.cell(row=row_num, column=13).value = highlight_close_dates_bicc\
                    (concatenate_dates(master_one.data[project_name]['Last time at BICC']))
                ws.cell(row=row_num, column=14).value = highlight_close_dates_bicc\
                    (concatenate_dates(master_one.data[project_name]['Next at BICC']))
            except KeyError:
                print('programme has crashed due to last time at BICC and Next and BICC keys not being inserted into'
                      'master')

    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=3).value
        if project_name in master_two.projects:
            ws.cell(row=row_num, column=5).value = convert_rag_text(master_two.data[project_name]['Departmental DCA'])

    '''highlight cells that contain RAG text, with background and text the same colour'''

    '''list of how rag ratings are shown in document'''
    rag_txt_list = ["A/G", "A/R", "R", "G", "A"]

    '''store of different colours'''
    ag_text = Font(color="00a5b700")
    ag_fill = PatternFill(bgColor="00a5b700")
    ar_text = Font(color="00f97b31")
    ar_fill = PatternFill(bgColor="00f97b31")
    red_text = Font(color="00fc2525")
    red_fill = PatternFill(bgColor="00fc2525")
    green_text = Font(color="0017960c")
    green_fill = PatternFill(bgColor="0017960c")
    amber_text = Font(color="00fce553")
    amber_fill = PatternFill(bgColor="00fce553")

    '''placed into a list'''
    txt_colour_list = [ag_text, ar_text, red_text, green_text, amber_text]
    fill_colour_list = [ag_fill, ar_fill, red_fill, green_fill, amber_fill]

    '''list of columns with conditional formatting'''
    list_columns = ['e', 'g', 'h', 'l']

    '''loops below place conditional formatting (cf) rules into the wb. There are two as the dashboard currently has 
    two distinct sections/headings, which do not require cf. Therefore, cf starts and ends at the stated rows. this
    is hard code that will need to be changed should the position of information in the dashboard change. It is an
    easy change however'''
    for column in list_columns:
        for i, dca in enumerate(rag_txt_list):
            text = txt_colour_list[i]
            fill = fill_colour_list[i]
            dxf = DifferentialStyle(font=text, fill=fill)
            rule = Rule(type="containsText", operator="containsText", text=dca, dxf=dxf)
            for_rule_formula = 'NOT(ISERROR(SEARCH("' + dca + '",' + column + '9)))'
            rule.formula = [for_rule_formula]
            ws.conditional_formatting.add('' + column + '9:' + column + '29', rule)

    for column in list_columns:
        for i, dca in enumerate(rag_txt_list):
            text = txt_colour_list[i]
            fill = fill_colour_list[i]
            dxf = DifferentialStyle(font=text, fill=fill)
            rule = Rule(type="containsText", operator="containsText", text=dca, dxf=dxf)
            for_rule_formula = 'NOT(ISERROR(SEARCH("' + dca + '",' + column + '34)))'
            rule.formula = [for_rule_formula]
            ws.conditional_formatting.add('' + column + '34:' + column + '60', rule)

    '''this conditional formatting highlights new projects'''
    red_text = Font(color="00fc2525")
    white_fill = PatternFill(bgColor="000000")
    dxf = DifferentialStyle(font=red_text, fill=white_fill)
    rule = Rule(type="uniqueValues", operator="equal", text="NEW", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("NEW",F1)))']
    ws.conditional_formatting.add('F1:F100', rule)

    '''this conditional formatting assigns the icon set to a rule'''
    first = FormatObject(type='num', val=-1)
    second = FormatObject(type='num', val=0)
    third = FormatObject(type='num', val=1)
    iconset = IconSet(iconSet='3Arrows', cfvo=[first, second, third], percent=None, reverse=None)
    rule = Rule(type='iconSet', iconSet=iconset)
    ws.conditional_formatting.add('F1:F100', rule)

    # highlights specific text in bold
    ft = Font(bold=True)
    for row_num in range(2, ws.max_row + 1):
        lis = ['This week', 'Next week', 'Last week', 'Two weeks',
               'Two weeks ago', 'This mth', 'Last mth', 'Next mth',
               '2 mths', '3 mths', '-2 mths', '-3 mths', '-2 weeks',
               'Today', 'Last BICC', 'Next BICC', 'This BICC',
               'Later this mth']
        if ws.cell(row=row_num, column=10).value in lis:
            ws.cell(row=row_num, column=10).font = ft
        if ws.cell(row=row_num, column=11).value in lis:
            ws.cell(row=row_num, column=11).font = ft
        if ws.cell(row=row_num, column=13).value in lis:
            ws.cell(row=row_num, column=13).font = ft
        if ws.cell(row=row_num, column=14).value in lis:
            ws.cell(row=row_num, column=14).font = ft

    return wb


''' RUNNING THE PROGRAMME '''

'''ONE. Provide file path to dashboard master'''
dashboard_master = load_workbook('C:\\Users\\Standalone\\general\\masters folder\\portfolio_dashboards\\master.xlsx')

'''TWO. Provide list of projects on which to provide analysis'''
project_list = q2_1920.projects

'''THREE. place arguments into the place_in_excle function and provide file path for saving output wb'''
dashboard_completed = place_in_excel(q2_1920, q1_1920, dashboard_master)
dashboard_completed.save('C:\\Users\\Standalone\\general\\masters folder\\portfolio_dashboards\\test.xlsx')