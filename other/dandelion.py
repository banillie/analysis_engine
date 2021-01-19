"""
programme that generate the information required for the dandelion chart. The output data is then sorted in excel and
placed into the dandelion graph wb.

Follow instructions below.

Note: all master data is taken from the data file. Make sure this is up to date and that all relevant data is in
the import statement.
"""

from openpyxl import Workbook
from analysis_engine.data import list_of_masters_all, root_path, abbreviations

def dandelion_data():
    '''
    Simple function that returns data required for the dandelion graph. Sorting done via excel.

    :param master_data: quarter master data
    :return: excel wb
    '''

    wb = Workbook()
    ws = wb.active

    for i, project_name in enumerate(list_of_masters_all[0].projects):
        ws.cell(row=2 + i, column=1).value = list_of_masters_all[0].data[project_name]['DfT Group']
        total = int(list_of_masters_all[0].data[project_name]['Total Forecast'])
        total_len = len(str(total))
        try:
            if total_len <= 3:
                round_total = int(round(total, -1))
                string_append = str(round_total) + 'm'
            if total_len == 4:
                round_total = int(round(total, -2))
                string_append = str(round_total)[0] + ',' + str(round_total)[1] + 'bn'
            if total_len == 5:
                round_total = int(round(total, -2))
                string_append = str(round_total)[:2] + ',' + str(round_total)[2] + 'bn'
        except ValueError:
            string_append = str(total)
        ws.cell(row=2 + i, column=2).value = abbreviations[project_name] + ', £' + string_append
        ws.cell(row=2 + i, column=3).value = total
        ws.cell(row=2 + i, column=4).value = list_of_masters_all[0].data[project_name]['Departmental DCA']

    ws.cell(row=1, column=1).value = 'Group'
    ws.cell(row=1, column=2).value = 'Project details'
    ws.cell(row=1, column=3).value = 'WLC (forecast)'
    ws.cell(row=1, column=4).value = 'DCA'

    return wb

'''  RUNNING PROGRAMME '''

'''simply run the programme'''
output = dandelion_data()
output.save(root_path/'output/dandelion_q1_20_21.xlsx')