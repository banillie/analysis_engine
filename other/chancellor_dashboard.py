
from openpyxl import load_workbook
from analysis.data import list_of_masters_all, root_path
from analysis.engine_functions import all_milestone_data_bulk, convert_bc_stage_text

def info(wb):

    ws = wb.worksheets[0]

    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=2).value
        print(project_name)
        if project_name in list_of_masters_all[0].projects:
            #print(project_name)
            '''BC Stage'''
            bc_stage = list_of_masters_all[0].data[project_name]['IPDC approval point']
            ws.cell(row=row_num, column=3).value = convert_bc_stage_text(bc_stage)

            '''Total WLC'''
            wlc_now = list_of_masters_all[0].data[project_name]['Total Forecast']
            ws.cell(row=row_num, column=4).value = wlc_now

            # '''initial bcr'''
            # initial_bcr = list_of_masters_all[0].data[project_name]['Initial Benefits Cost Ratio (BCR)']
            # ws.cell(row=row_num, column=5).value = initial_bcr

            '''adjusted bcr'''
            adjusted_bcr = list_of_masters_all[0].data[project_name]['Adjusted Benefits Cost Ratio (BCR)']
            ws.cell(row=row_num, column=5).value = adjusted_bcr

            '''vfm category now'''
            if list_of_masters_all[0].data[project_name]['VfM Category single entry'] is None:
                vfm_cat = str(list_of_masters_all[0].data[project_name]['VfM Category lower range']) + ' - ' + \
                          str(list_of_masters_all[0].data[project_name]['VfM Category upper range'])
                ws.cell(row=row_num, column=6).value = vfm_cat
            else:
                vfm_cat = list_of_masters_all[0].data[project_name]['VfM Category single entry']
                ws.cell(row=row_num, column=6).value = vfm_cat

            try:
                current_soc = tuple(current_milestones_all[project_name]['Start of Construction/build'])[0]
                ws.cell(row=row_num, column=7).value = current_soc
                # if current_soc < ipdc_date:
                #     ws.cell(row=row_num, column=10).value = 'Completed'
            except (KeyError, TypeError):
                ws.cell(row=row_num, column=7).value = 'Not reported'

            try:
                current_sop = tuple(current_milestones_all[project_name]['Start of Operation'])[0]
                ws.cell(row=row_num, column=8).value = current_sop
                # if current_sop < ipdc_date:
                #     ws.cell(row=row_num, column=13).value = 'Completed'
            except (KeyError, TypeError):
                ws.cell(row=row_num, column=8).value = 'Not reported'

            try:
                foc = tuple(current_milestones_all[project_name]['Full Operations'])[0]
                ws.cell(row=row_num, column=9).value = foc
                # if foc < ipdc_date:
                #     ws.cell(row=row_num, column=16).value = 'Completed'
                # else:
                #     ws.cell(row=row_num, column=16).value = foc
            except (KeyError, TypeError):
                ws.cell(row=row_num, column=9).value = 'Not reported'


    return wb

current_milestones_all = all_milestone_data_bulk(list_of_masters_all[0].projects, list_of_masters_all[0])


'''ONE. Provide file path to dashboard master'''
dashboard_master = load_workbook(root_path/'input/no10_commission_master.xlsx')

'''THREE. place arguments into the place_in_excel function and provide file path for saving output wb'''
dashboard_completed = info(dashboard_master)
dashboard_completed.save(root_path/'output/test.xlsx')