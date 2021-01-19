"""
This programme creates a master spreadsheet to share with IPA for gmpp reporting and the populating of individual
gmpp reporting templates.

Documents required to run the programme are:
1) the latest dft quarter_master spreadsheet (i.e. the quarter that is being reported). Entered via file path at 94.
2) the gmpp datamap (make sure you have the latest version). Entered via file path at 98.

Once you have provide these documents run the programme and it will produce two documents in the output folder,
which are:
1) gmpp_dataset[quarter_name].xlsx. This contains all the output data for sharing with IPA.
2) no_match_dft_gmpp_dataset.xlsx. If there are any keys which the programme could not match these will be
recorded in this document. This is useful for checking if data has transferred properly.
"""

from datetime import date
from typing import Dict, Union

from openpyxl import load_workbook, Workbook, workbook
from analysis_engine.engine_functions import filter_gmpp

from analysis_engine.data import root_path
from datamaps.api import project_data_from_master


def create_master(
        gmpp_wb: workbook, master: Dict[str, Union[str, int, date, float]]
) -> None:
    ws = gmpp_wb.active

    keys_relating_to_costs = [
        "RDEL",
        "CDEL",
        "Non-Gov",
        "Income",
        "BEN",
    ]  # list of cost/ben types. used to remove none value entries

    gmpp_project_names = filter_gmpp(master)  # gets gmpp project names

    for i, project_name in enumerate(gmpp_project_names):
        print(project_name)  # to show progress of programme.
        ws.cell(
            row=1, column=7 + i
        ).value = project_name  # place project names in file.
        ws.cell(
            row=2, column=7 + i
        ).value = project_name  # place project names in file twice. this time against project/programme name

        keys_not_found = []
        for row_num in range(2, ws.max_row + 1):  # for loop for placing data into the worksheet
            key = ws.cell(row=row_num, column=1).value
            if key is not None:  # remove None types
                try:
                    if (
                            key != "Project/Programme Name"
                    ):  # this keys will not be excel document.
                        ws.cell(row=row_num, column=7 + i).value = master.data[
                            project_name
                        ][key]
                        for cost_type in keys_relating_to_costs:
                            if cost_type in key:
                                if master.data[project_name][key] is None:
                                    ws.cell(row=row_num, column=7 + i).value = 0
                except KeyError:
                    try:
                        key_altered = key.replace(
                            ",", ""
                        )  # This handles keys names in excel document have comma's which are not
                        # present in the python dictionary
                        ws.cell(row=row_num, column=7 + i).value = master.data[
                            project_name
                        ][key_altered]
                        for cost_type in keys_relating_to_costs:
                            if cost_type in key_altered:
                                if master.data[project_name][key_altered] is None:
                                    ws.cell(row=row_num, column=7 + i).value = 0
                    except KeyError:
                        keys_not_found.append(key)

    wb_keys_not_found = Workbook()
    ws = wb_keys_not_found.active
    for x, key in enumerate(keys_not_found):
        ws.cell(row=x + 2, column=1).value = key

    ws.cell(
        row=1, column=1
    ).value = "Keys with no match between DfT datamap and GMPP datamap."
    wb_keys_not_found.save(root_path / 'output/no_match_dft_gmpp_datamaps.xlsx')

    quarter = str(quarter_master.quarter)
    quarter = quarter.replace(
        "/", "_"
    )
    gmpp_wb.save(root_path / 'output/gmpp_dataset_{}.xlsx'.format(quarter))


# place file path to quarter master here:
quarter_master = project_data_from_master(
    root_path / "core_data/master_2_2020.xlsx", 2, 2020
)
# place file path to gmpp datamap here:
master_dm = load_workbook(root_path / "input/new_gmpp_oscar_II_datamap_master_v3.xlsx")

create_master(master_dm, quarter_master)
