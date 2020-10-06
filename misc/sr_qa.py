from datamaps.api import project_data_from_master
from openpyxl import Workbook
from data_mgmt.data import root_path


def get_project_names():
    return project_data_from_master(root_path / 'sr20/project_naming_t.xlsx', 4, 2099)

def get_sr_data():
    return project_data_from_master(root_path / 'sr20/SR20_bid_data_t.xlsx', 1, 2020)

def get_trail_data():
    return project_data_from_master(root_path / 'sr20/Trail_bid_data_t.xlsx', 1, 2020)

def integrate_data():
    wb = Workbook()
    ws = wb.active

    for i, p in enumerate(p_names.projects):
        ws.cell(row=i+3, column=1).value = p
        sr_p_name = p_names.data[p]['Corresponding CEN']
        ws.cell(row=i+3, column=2).value = sr_p_name

        for x, k in enumerate(sr_keys):
            try:
                ws.cell(row=i+3, column=4+x).value = sr_data.data[sr_p_name][k]
            except KeyError:
                #print("Could not match " + str(sr_p_name))
                ws.cell(row=i+3, column=4+x).value = 'no match'

        for x, k in enumerate(sr_keys):
            ws.cell(row=2, column=4+x).value = k

        for y, tk in enumerate(trail_keys):
            try:
                ws.cell(row=i+3, column=len(sr_keys)+5+y).value = trail_data[p][tk]
            except KeyError:
                ws.cell(row=i + 3, column=len(sr_keys) + 5 + y).value = 'no match'

        for y, tk in enumerate(trail_keys):
            ws.cell(row=2, column=len(sr_keys) + 5 + y).value = tk

    ws.cell(row=2, column=1).value = 'Trail Project Name'
    ws.cell(row=2, column=2).value = 'SR (CEN) Project Name'
    ws.cell(row=1, column=4).value = 'HMT CS20 unified template'
    ws.cell(row=1, column=len(sr_keys)+5).value = 'SR Check of Figures - TRAIL Download'

    return wb.save(root_path/'sr20/sr_integrated_data.xlsx')


sr_keys = ['Discounted public sector whole life cost (WLC)', 'Discounted social benefits', 'Discounted social costs', 'NPSV', 'NPSV/WLC', 'Climate Impacts (Y/N)', 'Productivity impacts (Y/N)', 'Additional regional impacts (Y/N)']
trail_keys = ['Total_PVC_Option_1', 'PVC_Local_Authority_Option_1', 'PVC_Revenue_Amount_Option_1', 'PVC_Efficiency_Amount_Option_1', 'PVC_Data_Robustness', 'Total_PVB_Option_1', 'Total_PVB_Level_1_and_2_Option_1', 'Total_PVB_Level_3_Option_1', 'Total_PVB_Other_Level_3_Option_1', 'PVB_Data_Robustness', 'NPV_Option_1', 'Initial_BCR_Option_1', 'Adjusted_BCR_Option_1', 'VFM_Category_1', 'VFM_Assessment_Data_Robustness']


p_names = get_project_names()
sr_data = get_sr_data()
trail_data = get_trail_data()
integrate_data()
