from openpyxl import load_workbook
from data_mgmt.data import root_path
from datamaps.api import project_data_from_master

# abbreviations. Used in analysis instead of full projects names
abbreviations = {'2nd Generation UK Search and Rescue Aviation': 'SARH2',
                 'A12 Chelmsford to A120 widening': 'A12',
                 'A14 Cambridge to Huntingdon Improvement Scheme': 'A14',
                 'A303 Amesbury to Berwick Down': 'A303',
                 'A358 Taunton to Southfields Dualling': 'A358',
                 'A417 Air Balloon': 'A417',
                 'A428 Black Cat to Caxton Gibbet': 'A428',
                 'A66 Northern Trans-Pennine': 'A66',
                 'Crossrail Programme': 'Crossrail',
                 'East Coast Digital Programme': 'ECDP',
                 'East Coast Mainline Programme': 'ECMP',
                 'East West Rail Programme (Central Section)': 'EWR (Central)',
                 'East West Rail Programme (Western Section)': 'EWR (Western)',
                 "East West Rail Configuration State 1": "EWR Config 1",
                 "East West Rail Configuration State 2": "EWR Config 2",
                 "East West Rail Configuration State 3": "EWR Config 3",
                 'Future Theory Test Service (FTTS)': 'FTTS',
                 'Great Western Route Modernisation (GWRM) including electrification': 'GWRM',
                 'Heathrow Expansion': 'HEP',
                 'Hexagon': 'Hexagon',
                 'High Speed Rail Programme (HS2)': 'HS2 Prog',
                 'HS2 Phase 2b': 'HS2 2b',
                 'HS2 Phase1': 'HS2 1',
                 'HS2 Phase2a': 'HS2 2a',
                 'Integrated and Smart Ticketing - creating an account based back office': 'IST',
                 'Intercity Express Programme': 'IEP',
                 'Lower Thames Crossing': 'LTC',
                 'M4 Junctions 3 to 12 Smart Motorway': 'M4',
                 'Manchester North West Quadrant': 'MNWQ',
                 'Midland Main Line Programme': 'MML Prog',
                 'Midlands Rail Hub': 'Mid Rail Hub',
                 'North Western Electrification': 'NWE',
                 'Northern Powerhouse Rail': 'NPR',
                 'Oxford-Cambridge Expressway': 'Ox-Cam Expressway',
                 'Rail Franchising Programme': 'Rail Franchising',
                 'South West Route Capacity': 'SWRC',
                 'Thameslink Programme': 'Thameslink',
                 'Transpennine Route Upgrade (TRU)': 'TRU',
                 'Western Rail Link to Heathrow': 'WRLtH'}

test_abbreviations = {'Sea of Tranquility': 'SoT',
            'Apollo 11': 'A11',
            'Apollo 13': 'A13',
            'Falcon 9': 'F9',
            'Columbia': 'Columbia',
            'Mars': 'Mars'}

path = root_path / "core_data/project_group_id_no.xlsx"
test_path = "/home/will/code/python/analysis_engine/tests/resources/test_project_group_id_no.xlsx"

def put_in_abbreviations(abbreviations):
    wb = load_workbook(test_path)
    ws = wb.active

    for i in range(1, ws.max_column+2):
        c = ws.cell(row=1, column=i).value
        if c in list(abbreviations.keys()):
            ws.cell(row=4, column=i).value = abbreviations[c]

    wb.save(test_path)

def put_in_status():
    master = project_data_from_master(root_path / "core_data/master_1_2020.xlsx", 1, 2099)
    wb = load_workbook(root_path / "core_data/project_info.xlsx")
    ws = wb.active

    for i in range(1, ws.max_column+2):
        c = ws.cell(row=1, column=i).value
        if c in master.projects:
            ws.cell(row=5, column=i).value = 'Live'

    wb.save(root_path / "core_data/project_info.xlsx")

put_in_abbreviations(test_abbreviations)
