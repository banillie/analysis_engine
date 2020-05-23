'''code to compile covid dashboard'''

from datamaps.api import project_data_from_master
from analysis.data import root_path
from analysis.engine_functions import convert_rag_text
from openpyxl.styles import Font
from openpyxl import load_workbook

def place_in_excel(wb):

    keys = ['Group', 'BC Stage', 'SRO', 'SRO Reallocated', 'PD Reallocated', 'Key Staff Reallocated']
    ws = wb.worksheets[0]

    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=2).value

        if project_name in covid_master_list[0].projects:
            for x, y in enumerate(keys):
                v = covid_master_list[0].data[project_name][y]
                ws.cell(row=row_num, column=x+3).value = v
                try:
                    v_lst_qrt = covid_master_list[1].data[project_name][y]
                    if v != v_lst_qrt:
                        ws.cell(row=row_num, column=x+3).font = Font(name='Arial', size=10, color='00fc2525')
                except KeyError:
                    pass

            '''DCA impact'''
            ws.cell(row=row_num, column=10).value = convert_rag_text(covid_master_list[0].data[project_name]
                                                                     ['Impact RAG Rating'])
            try:
                ws.cell(row=row_num, column=11).value = convert_rag_text(covid_master_list[1].data[project_name]
                                                                     ['Impact RAG Rating'])
            except KeyError:
                pass

            try:
                ws.cell(row=row_num, column=12).value = convert_rag_text(covid_master_list[2].data[project_name]
                                                                     ['Impact RAG Rating'])
            except KeyError:
                pass



    return wb

'''Ensure the latest master is being uploaded and placed in the masters list'''

master_29_05 = project_data_from_master(root_path/'core_data/covid_19/master_290520.xlsx', 1, 2020)
master_13_05 = project_data_from_master(root_path/'core_data/covid_19/master_130520.xlsx', 1, 2020)
master_01_05 = project_data_from_master(root_path/'core_data/covid_19/master_010520.xlsx', 1, 2020)

covid_master_list = [master_29_05,
                     master_13_05,
                     master_01_05]

'''file path to where the dashboard master is'''
c_dashboard_master = load_workbook(root_path/'input/covid_19/covid_dasboard_template.xlsx')

dashboard_completed = place_in_excel(c_dashboard_master)
dashboard_completed.save(root_path/'output/covid_19/covid_dashboard_completed.xlsx')