'''
Programme that creates a financial profiles for a group of projects i.e. can produce the portfolio profile or a chosen
set of projects profile.

Output document is an excel wb with four tabs containing the aggregate financial profile for 'latest quarter',
'last quarter', 'baseline 1' (i.e. nearest baseline of approvals) 'baseline 2' (second nearest baselines).

Just data handling for now. No graphics produced. Graph needs to be done manually.

See instructions below.

'''

from openpyxl import Workbook
from analysis.data import cost_list, income_list, year_list, latest_cost_profiles, last_cost_profiles, \
    baseline_1_cost_profiles, latest_income_profiles, list_of_masters_all, \
    last_income_profiles, baseline_1_income_profiles, dont_double_count, root_path, northern_powerhouse
from analysis.engine_functions import filter_project_group, calculate_group_project_total

dont_double_count = [northern_powerhouse]


def place_in_excel(project_name_list):
    wb = Workbook()

    financial_profile_list = ['Latest Profile', 'Last quarter profile', 'Baseline 1 Profile']
    cost_profile_data_list = [latest_cost_profiles, last_cost_profiles, baseline_1_cost_profiles]
    income_profile_data_list = [latest_income_profiles, last_income_profiles, baseline_1_income_profiles]

    for p, profile in enumerate(financial_profile_list):
        '''worksheet is created for each project'''
        ws = wb.create_sheet(profile, p)  # creating worksheets
        ws.title = profile  # title of worksheet

        '''place information in each sheet'''
        ws.cell(row=1, column=1).value = 'Project'
        for i, project_name in enumerate(project_name_list):
            '''lists project names in row one'''
            ws.cell(row=1, column=i + 2).value = project_name

            '''iterates through financial dictionary - placing financial data in ws'''
            row_number = 1
            for cost in cost_list:
                for year in year_list:
                    try:
                        ws.cell(row=row_number + 1, column=i+2).value = \
                            cost_profile_data_list[p][project_name][year + cost]
                    except KeyError:
                        pass
                    row_number += 1

        '''places totals in final column'''

        cost_totals = calculate_group_project_total(project_name_list, cost_profile_data_list[p], dont_double_count,
                                                    cost_list, year_list)

        ws.cell(row=1, column=len(project_name_list) + 2).value = 'Total'
        for i, value in enumerate(cost_totals.values()):
            ws.cell(row=i + 2, column=len(project_name_list)+2).value = value

        '''places keys into the chart in the first column'''
        for i, key in enumerate(cost_totals.keys()):
            ws.cell(row=i+2, column=1).value = key

        '''information on which projects are not included in totals'''
        ws.cell(row=1, column=len(project_name_list) + 4).value = 'Projects that have been removed to avoid double counting'
        for i, project in enumerate(dont_double_count):
            ws.cell(row=i + 2, column=len(project_name_list) + 4).value = project

        '''total cost data for output chart'''
        for z, year in enumerate(year_list):
            for x, type in enumerate(cost_list):
                ws.cell(row=z + 41, column=x + 2, value=cost_totals[year + type])

        '''labeling for total figure table '''
        ws.cell(row=40, column=1, value='Year')
        labeling_list_type = ['RDEL', 'CDEL', 'Non-Gov', 'Total']
        for i, label in enumerate(labeling_list_type):
            ws.cell(row=40, column=2 + i, value=label)

        '''labeling of years down the side'''
        for i, label in enumerate(year_list):
            ws.cell(row=41 + i, column=1, value=label)

        '''income total data'''
        income_totals = calculate_group_project_total(project_name_list, income_profile_data_list[p], dont_double_count,
                                                    income_list, year_list)

        for y, year in enumerate(year_list):
            for t, type in enumerate(income_list):
                ws.cell(row=y + 55, column=t + 2, value=income_totals[year + type])

        '''labeling for total figure table '''
        ws.cell(row=54, column=1, value='Year')
        ws.cell(row=54, column=2, value='Income Totals')

        '''labeling of years down the side'''
        for i, label in enumerate(year_list):
            ws.cell(row=55 + i, column=1, value=label)

    return wb


''' RUNNING PROGRAMME'''

'''To run the programme place list of project names into function. 
NOTE: Default option is list of all current projects in portfolio. In majority of cases user should just run the 
default programme'''
output = place_in_excel(list_of_masters_all[0].projects)
output.save(root_path/'output/portfolio_financial_profile_q1_2021.xlsx')