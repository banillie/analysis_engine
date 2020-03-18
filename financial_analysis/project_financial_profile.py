'''
This programme calculates and provides cost profiles for each project.

Output document:
1) one excel workbook with a cost profile for each project on different tabs. Three different cost profiles calculated:
i) latest,
ii) last,
iii) baseline (nearest).

See instructions below.

'''

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font
from analysis.data import cost_list, income_list, year_list, baseline_bc_stamp, latest_cost_profiles, last_cost_profiles, \
    baseline_1_cost_profiles, latest_income_profiles, last_income_profiles, baseline_1_income_profiles, root_path

#TODO update code so it includes older baseline profiles as is now the case for the portfolio profile.

def place_in_excel_one_wb():

    project_name_list = list(latest_cost_profiles.keys())

    wb = Workbook()

    cost_list.append(' total')

    for i, project_name in enumerate(project_name_list):
        '''worksheet is created for each project'''
        ws = wb.create_sheet(project_name, i)  # creating worksheets
        ws.title = project_name  # title of worksheet

        '''COST PROFILE CHART'''

        ''''places in raw/reported cost data'''
        for i, year in enumerate(year_list):
            for x, type in enumerate(cost_list):
                ws.cell(row=i+3, column=x+2, value=baseline_1_cost_profiles[project_name][year + type])
                try:
                    ws.cell(row=i+3, column=x+6, value=last_cost_profiles[project_name][year + type])
                except KeyError:
                    pass
                ws.cell(row=i+3, column=x+10, value=latest_cost_profiles[project_name][year + type])


        '''labeling data in table'''
        labeling_list_quarter = ['Baseline', 'Last Quarter', 'Latest quarter']
        ws.cell(row=1, column=2, value=labeling_list_quarter[0])
        ws.cell(row=1, column=6, value=labeling_list_quarter[1])
        ws.cell(row=1, column=10, value=labeling_list_quarter[2])

        '''top reference to which data being used'''
        labeling_list_type = ['RDEL', 'CDEL', 'Non-Gov', 'Total']
        repeat = 3
        c = 0
        while repeat > 0:
            for i, label in enumerate(labeling_list_type):
                ws.cell(row=2, column=2+i+c, value=label)
            c += 4
            repeat -= 1

        '''labeling of years down the side'''
        for i, label in enumerate(year_list):
            ws.cell(row=3+i, column=1, value=label)
        ws.cell(row=2, column=1, value='Year')

        '''record of which baseline is being used'''
        ws.cell(row=1, column=16).value = 'Baseline quarter'
        ws.cell(row=2, column=16).value = baseline_bc_stamp[project_name][0][1]

        '''Total cost profile. starting with data placement'''
        for i, year in enumerate(year_list):
            for x, type in enumerate([' total']):
                ws.cell(row=i+16, column=x+2, value=baseline_1_cost_profiles[project_name][year + type])
                ws.cell(row=i+16, column=x+3, value=last_cost_profiles[project_name][year + type])
                ws.cell(row=i+16, column=x+4, value=latest_cost_profiles[project_name][year + type])

        '''data for graph labeling'''
        for i, quarter in enumerate(labeling_list_quarter):
            ws.cell(row=15, column=i + 2, value=quarter)

        for i, label in enumerate(year_list):
            ws.cell(row=16+i, column=1, value=label)
        ws.cell(row=15, column=1, value='Year')

        '''Cost chart information'''
        chart = LineChart()
        chart.title = str(project_name) + ' Cost Profile'
        chart.style = 4
        chart.x_axis.title = 'Financial Year'
        chart.y_axis.title = 'Cost £m'

        '''styling chart'''
        # axis titles
        font = Font(typeface='Calibri')
        size = 1200  # 12 point size
        cp = CharacterProperties(latin=font, sz=size, b=True)  # Bold
        pp = ParagraphProperties(defRPr=cp)
        rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
        chart.x_axis.title.tx.rich.p[0].pPr = pp
        chart.y_axis.title.tx.rich.p[0].pPr = pp
        # chart.title.tx.rich.p[0].pPr = pp

        # title
        size_2 = 1400
        cp_2 = CharacterProperties(latin=font, sz=size_2, b=True)
        pp_2 = ParagraphProperties(defRPr=cp_2)
        rtp_2 = RichText(p=[Paragraph(pPr=pp_2, endParaRPr=cp_2)])
        chart.title.tx.rich.p[0].pPr = pp_2

        '''unprofiled costs not included in the chart'''
        data = Reference(ws, min_col=2, min_row=15, max_col=4, max_row=25)
        chart.add_data(data, titles_from_data=True)
        cats = Reference(ws, min_col=1, min_row=16, max_row=25)
        chart.set_categories(cats)

        s3 = chart.series[0]
        s3.graphicalProperties.line.solidFill = "cfcfea"  # light blue
        s8 = chart.series[1]
        s8.graphicalProperties.line.solidFill = "5097a4"  # medium blue
        s9 = chart.series[2]
        s9.graphicalProperties.line.solidFill = "0e2f44"  # dark blue'''

        ws.add_chart(chart, "H15")

        '''INCOME CHART'''

        '''Quick calculation to see if project is reporting income. If not then no graph is created'''
        total_income = sum(list(latest_income_profiles[project_name].values()))

        '''income data is inserted'''
        if total_income > 0:
            for i, year in enumerate(year_list):
                for type in income_list:
                    ws.cell(row=i + 32, column=2, value=baseline_1_income_profiles[project_name][year + type])
                    ws.cell(row=i + 32, column=3, value=last_income_profiles[project_name][year + type])
                    ws.cell(row=i + 32, column=4, value=latest_income_profiles[project_name][year + type])

            '''labeling of the data table'''
            for i, quarter in enumerate(labeling_list_quarter):
                ws.cell(row=32, column=i + 2, value=quarter)

            for i, label in enumerate(year_list):
                ws.cell(row=33 + i, column=1, value=label)
            ws.cell(row=32, column=1, value='Year')

            '''income graph created'''
            chart = LineChart()
            chart.title = str(project_name) + ' Income Profile'
            chart.style = 4
            chart.x_axis.title = 'Financial Year'
            chart.y_axis.title = 'Cost £m'

            font = Font(typeface='Calibri')
            size = 1200  # 12 point size
            cp = CharacterProperties(latin=font, sz=size, b=True)  # Bold
            pp = ParagraphProperties(defRPr=cp)
            rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
            chart.x_axis.title.tx.rich.p[0].pPr = pp
            chart.y_axis.title.tx.rich.p[0].pPr = pp
            # chart.title.tx.rich.p[0].pPr = pp

            # title
            size_2 = 1400
            cp_2 = CharacterProperties(latin=font, sz=size_2, b=True)
            pp_2 = ParagraphProperties(defRPr=cp_2)
            rtp_2 = RichText(p=[Paragraph(pPr=pp_2, endParaRPr=cp_2)])
            chart.title.tx.rich.p[0].pPr = pp_2

            #unprofiled costs not included in the chart
            data = Reference(ws, min_col=2, min_row=32, max_col=4, max_row=42)
            chart.add_data(data, titles_from_data=True)
            cats = Reference(ws, min_col=1, min_row=33, max_row=42)
            chart.set_categories(cats)

            s3 = chart.series[0]
            s3.graphicalProperties.line.solidFill = "e2f1bb"  # light green
            s8 = chart.series[1]
            s8.graphicalProperties.line.solidFill = "a0db8e"  # medium green
            s9 = chart.series[2]
            s9.graphicalProperties.line.solidFill = "29ab87"  # dark green

            ws.add_chart(chart, "H31")

        else:
            pass

    return wb


'''RUNNING PROGRAMME'''

'''
Only one part of programme is to be amended each quarter. place which ever quarter information being produced at 
end of output file name e.g. q4_1920. Note make sure file ends with .xlsx format

Note when code completes it may state:
UserWarning: Title is more than 31 characters. Some applications may not be able to read the file 
warnings.warn("Title is more than 31 characters. Some applications may not be able to read the file"). 
However, you can ignore and open the file as usual.
'''

output = place_in_excel_one_wb()
output.save(root_path/'output/ind_project_financial_profiles_q4_1920.xlsx')