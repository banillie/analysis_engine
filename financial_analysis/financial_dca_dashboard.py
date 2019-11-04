'''

Programme for creating an aggregate portfolio financial dashboard

input documents:
1) Dashboard master document - this is an excel file. It should have the dashboard design, with all projects structured
in the correct way (order), but all data fields left blank. Note if project data does not get placed into the correct
part of the master, check that the project name is consistent with the name in master data. The names need to be
exactly the same for information to be released.
2) Master data for two quarters - this will usually be latest and previous quarter. now handled via the analysis.data
import at tip.

output document:
3) Dashboard with all project data placed into dashboard. The aim of this programme is to get all relevant data into one
document. From this point on only projects of interest. i.e. those with red confidence ratings or that have changed in
financial confidence should remain on the dashboard. the others should be delete. The financial narrative provided for
each project should be checked.

Instructions:
1) provide path to dashboard master
32 provide path and specify file name for output document

Note some manual adjustments may need to be made to output document, this includes:
1) Project WLC totals e.g. Hs2 Phases

'''

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, IconSet, FormatObject
from analysis.data import q2_1920, q1_1920, all_project_names
from analysis.engine_functions import up_or_down, convert_rag_text

def place_in_excel(master_data_latest, master_data_last):
    '''
    function that places all information into the master dashboard sheet
    :param master_one:
    :param master_two:
    :return:
    '''

    ws = wb.active

    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=3).value
        print(project_name)
        if project_name in all_project_names:
            ws.cell(row=row_num, column=4).value = master_data_latest.data[project_name]['Total Forecast']
            try:
                dca_now = master_data_latest.data[project_name]['SRO Finance confidence']
                dca_past = master_data_last.data[project_name]['SRO Finance confidence']
                ws.cell(row=row_num, column=6).value = up_or_down(dca_now, dca_past)
            except KeyError:
                ws.cell(row=row_num, column=6).value = 'NEW'
            ws.cell(row=row_num, column=7).value = convert_rag_text(master_data_latest.data[project_name]
                                                                    ['SRO Finance confidence'])
            ws.cell(row=row_num, column=8).value = combine_narrtives(project_name, master_data_latest.data,
                                                                     narrative_keys)

    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=3).value
        if project_name in master_data_last.projects:
            ws.cell(row=row_num, column=5).value = convert_rag_text(master_data_last.data[project_name]
                                                                    ['SRO Finance confidence'])

    '''highlight cells that contain RAG text, with background and text the same colour'''

    '''store of different colours'''
    ag_text = Font(color="00a5b700") # text same colour as background
    ag_fill = PatternFill(bgColor="00a5b700")
    ar_text = Font(color="00f97b31") # text same colour as background
    ar_fill = PatternFill(bgColor="00f97b31")
    red_text = Font(color="00fc2525") # text same colour as background
    red_fill = PatternFill(bgColor="00fc2525")
    green_text = Font(color="0017960c") # text same colour as background
    green_fill = PatternFill(bgColor="0017960c")
    amber_text = Font(color="00fce553") # text same colour as background
    amber_fill = PatternFill(bgColor="00fce553")

    black_text = Font(color="00000000")

    '''NOTE. these three lists need to have rag ratings in the same order'''
    '''different colours are placed into a list'''
    txt_colour_list = [ag_text, ar_text, red_text, green_text, amber_text]
    fill_colour_list = [ag_fill, ar_fill, red_fill, green_fill, amber_fill]
    '''list of how rag ratings are shown in document'''
    rag_txt_list = ["A/G", "A/R", "R", "G", "A"]


    '''loops below place conditional formatting (cf) rules into the wb. There are two as the dashboard currently has 
    two distinct sections/headings, which do not require cf. Therefore, cf starts and ends at the stated rows. this
    is hard code that will need to be changed should the position of information in the dashboard change. It is an
    easy change however'''

    '''these two loops provide conditional formatting in column e with text and fill colours the same'''
    for i, dca in enumerate(rag_txt_list):
        text = txt_colour_list[i]
        fill = fill_colour_list[i]
        dxf = DifferentialStyle(font=text, fill=fill)
        rule = Rule(type="containsText", operator="containsText", text=dca, dxf=dxf)
        for_rule_formula = 'NOT(ISERROR(SEARCH("' + dca + '",e9)))'
        rule.formula = [for_rule_formula]
        ws.conditional_formatting.add('e9:e27', rule)

    for i, dca in enumerate(rag_txt_list):
        text = txt_colour_list[i]
        fill = fill_colour_list[i]
        dxf = DifferentialStyle(font=text, fill=fill)
        rule = Rule(type="containsText", operator="containsText", text=dca, dxf=dxf)
        for_rule_formula = 'NOT(ISERROR(SEARCH("' + dca + '",e32)))'
        rule.formula = [for_rule_formula]
        ws.conditional_formatting.add('e32:e60', rule)

    '''these two loops provide conditional formatting in column e with text and fill colours the same'''
    for i, dca in enumerate(rag_txt_list):
        text = txt_colour_list[i]
        fill = fill_colour_list[i]
        dxf = DifferentialStyle(font=text, fill=fill)
        rule = Rule(type="containsText", operator="containsText", text=dca, dxf=dxf)
        for_rule_formula = 'NOT(ISERROR(SEARCH("' + dca + '",g9)))'
        rule.formula = [for_rule_formula]
        ws.conditional_formatting.add('g9:g27', rule)

    for i, dca in enumerate(rag_txt_list):
        text = txt_colour_list[i]
        fill = fill_colour_list[i]
        dxf = DifferentialStyle(font=text, fill=fill)
        rule = Rule(type="containsText", operator="containsText", text=dca, dxf=dxf)
        for_rule_formula = 'NOT(ISERROR(SEARCH("' + dca + '",g32)))'
        rule.formula = [for_rule_formula]
        ws.conditional_formatting.add('g32:g60', rule)


    '''this conditional formatting highlights new projects'''
    red_text = Font(color="00fc2525")
    white_fill = PatternFill(bgColor="000000")
    dxf = DifferentialStyle(font=red_text, fill=white_fill)
    rule = Rule(type="uniqueValues", operator="equal", text="NEW", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("NEW",F1)))']
    ws.conditional_formatting.add('F1:F100', rule)

    '''this conditional formatting assigns the icon set to a rule'''
    first = FormatObject(type='num', val=-1)
    second = FormatObject(type='num', val=0)
    third = FormatObject(type='num', val=1)
    iconset = IconSet(iconSet='3Arrows', cfvo=[first, second, third], percent=None, reverse=None)
    rule = Rule(type='iconSet', iconSet=iconset)
    ws.conditional_formatting.add('F1:F100', rule)

    # highlights specific text in bold
    ft = Font(bold=True)
    for row_num in range(2, ws.max_row + 1):
        lis = ['This week', 'Next week', 'Last week', 'Two weeks',
               'Two weeks ago', 'This mth', 'Last mth', 'Next mth',
               '2 mths', '3 mths', '-2 mths', '-3 mths', '-2 weeks',
               'Today', 'Last BICC', 'Next BICC', 'This BICC',
               'Later this mth']
        if ws.cell(row=row_num, column=10).value in lis:
            ws.cell(row=row_num, column=10).font = ft
        if ws.cell(row=row_num, column=11).value in lis:
            ws.cell(row=row_num, column=11).font = ft
        if ws.cell(row=row_num, column=13).value in lis:
            ws.cell(row=row_num, column=13).font = ft
        if ws.cell(row=row_num, column=14).value in lis:
            ws.cell(row=row_num, column=14).font = ft

    return wb


def combine_narrtives(project_name, master_data, key_list):
    output = ''
    for key in key_list:
        output = output + str(master_data[project_name][key])

    return output

'''keys of used to get narratives'''
narrative_keys = ['Project Costs Narrative',
                 'Cost comparison with last quarters cost narrative',
                 'Cost comparison within this quarters cost narrative']


'''  INSTRUCTIONS FOR RUNNING THE PROGRAMME'''

'''ONE. Provide file path to empty dashboard document'''
wb = load_workbook('C:\\Users\\Standalone\\general\\masters folder\\portfolio_financial_profile\\'
                   'finance_dashboard master.xlsx')

'''TWO. place the right quarter information into programme'''
output = place_in_excel(q2_1920, q1_1920)

'''THREE. provide file path and specific name of output file.'''
output.save('C:\\Users\\Standalone\\general\\masters folder\\portfolio_financial_profile\\'
            'finance_dashboard_output_testing.xlsx')