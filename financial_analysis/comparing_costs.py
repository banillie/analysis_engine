'''
Programme to compare project financial data. It produces four wb outputs with data and calculations
only i.e. no graph. The outputs shows changes to:

1) overall wlc,
2) yearly cost profiles,

In the output files changes are highlighted in red if change is greater/less than £100m/-£100m or percentage change
greater/less than 5%/-5% of project value

It is from the data placed into the output documents that a charts can be built to show the most significant
changes.

See instructions below.
'''

from openpyxl import Workbook
from analysis.data import red_text, wlc_key, latest_cost_profiles, last_cost_profiles, \
    baseline_cost_profiles, latest_quarter_project_names, list_of_masters_all, fin_bc_index, root_path

def place_complex_comparision_excel(master_data_latest, master_data_last, master_data_baseline):
    '''
    Function that places all information structured via the get_wlc_costs and get_yearly_costs programmes into an
    excel spreadsheet. It does some calculations on the level of change that has taken place.
    This function places in data for a chart that shows changes in financial profile between latest, last and baseline
    :param master_data_latest: data representing latest quarter information
    :param master_data_last: data representing last quarter information.
    :param master_data_baseline: data representing baseline quarter information
    :return: excel workbook
    '''
    wb = Workbook()

    for i, key in enumerate(list(master_data_latest.keys())):
        ws = wb.create_sheet(key, i)  # creating worksheets
        ws.title = key  # title of worksheet

        data_latest = master_data_latest[key]
        data_last = master_data_last[key]
        data_baseline = master_data_baseline[key]

        for i, project_name in enumerate(data_latest):
            '''place project names into ws'''
            ws.cell(row=i+2, column=1).value = project_name

            '''loop for placing data into ws. highlight changes between quarters in red'''
            latest_value = data_latest[project_name]
            ws.cell(row=i + 2, column=2).value = latest_value

            '''comparision data against last quarter'''
            if project_name in data_last.keys():
                try:
                    last_value = data_last[project_name]
                    ws.cell(row=i + 2, column=3).value = last_value
                    change = latest_value - last_value
                    if last_value > 0:
                        percent_change = (latest_value - last_value)/last_value
                    else:
                        percent_change = (latest_value - last_value)/(last_value + 1)
                    ws.cell(row=i + 2, column=7).value = change
                    ws.cell(row=i + 2, column=8).value = percent_change
                    if change >= 100 or change <= -100:
                        ws.cell(row=i + 2, column=7).font = red_text
                    if percent_change >= 0.05 or percent_change <= -0.05:
                        ws.cell(row=i + 2, column=8).font = red_text
                except TypeError:
                    ws.cell(row=i + 2, column=3).value = 'check project data'
            else:
                ws.cell(row=i + 2, column=3).value = 'None'

            if project_name in data_baseline.keys():
                try:
                    last_value = data_last[project_name]
                    baseline_value = data_baseline[project_name]
                    ws.cell(row=i + 2, column=4).value = baseline_value
                    change = last_value - baseline_value
                    if baseline_value > 0:
                        percent_change = (last_value - baseline_value) / baseline_value
                    else:
                        percent_change = (last_value - baseline_value) / (baseline_value + 1)
                    ws.cell(row=i + 2, column=5).value = change
                    ws.cell(row=i + 2, column=6).value = percent_change
                    if change >= 100 or change <= -100:
                        ws.cell(row=i + 2, column=5).font = red_text
                    if percent_change >= 0.05 or percent_change <= -0.05:
                        ws.cell(row=i + 2, column=6).font = red_text
                except TypeError:
                    ws.cell(row=i + 2, column=4).value = 'check project data'
                except KeyError:
                    ws.cell(row=i + 2, column=4).value = 'not reporting'
            else:
                ws.cell(row=i + 2, column=4).value = 'None'


        # Note the ordering of data. Done in this manner so that data is displayed in graph in the correct way.
        ws.cell(row=1, column=1).value = 'Project Name'
        ws.cell(row=1, column=2).value = 'latest quarter (£m)'
        ws.cell(row=1, column=3).value = 'last quarter (£m)'
        ws.cell(row=1, column=4).value = 'baseline (£m)'
        ws.cell(row=1, column=7).value = '£m change between latest and last quarter'
        ws.cell(row=1, column=8).value = 'percentage change between latest and last quarter'
        ws.cell(row=1, column=5).value = '£m change between last and baseline quarter'
        ws.cell(row=1, column=6).value = 'percentage change between last and baseline quarter'

    return wb

def place_standard_comparision_excel(master_data_latest, master_data_baseline):
    '''
    Function that places all information structured via the get_wlc_costs and get_yearly_costs programmes into an
    excel spreadsheet. It does some calculations on the level of change that has taken place.
    This function places in data for a chart that shows changes in financial profile between latest and baseline.
    :param master_data_latest: data representing latest quarter information
    :param master_data_baseline: data representing baseline quarter information
    :return: excel workbook
    '''
    wb = Workbook()

    for i, key in enumerate(list(master_data_latest.keys())):
        ws = wb.create_sheet(key, i)  # creating worksheets
        ws.title = key  # title of worksheet

        data_latest = master_data_latest[key]
        data_baseline = master_data_baseline[key]

        for i, project_name in enumerate(data_latest):
            '''place project names into ws'''
            ws.cell(row=i+2, column=1).value = project_name

            '''loop for placing data into ws. highlight changes between quarters in red'''
            latest_value = data_latest[project_name]
            ws.cell(row=i + 2, column=2).value = latest_value

            '''comparision data against last quarter'''
            if project_name in data_baseline.keys():
                try:
                    baseline_value = data_baseline[project_name]
                    ws.cell(row=i + 2, column=3).value = baseline_value
                    change = latest_value - baseline_value
                    if baseline_value > 0:
                        percent_change = (latest_value - baseline_value)/baseline_value
                    else:
                        percent_change = (latest_value - baseline_value)/(baseline_value + 1)
                    ws.cell(row=i + 2, column=4).value = change
                    ws.cell(row=i + 2, column=5).value = percent_change
                    if change >= 100 or change <= -100:
                        ws.cell(row=i + 2, column=4).font = red_text
                    if percent_change >= 0.05 or percent_change <= -0.05:
                        ws.cell(row=i + 2, column=5).font = red_text
                except TypeError:
                    ws.cell(row=i + 2, column=3).value = 'check project data'
            else:
                ws.cell(row=i + 2, column=3).value = 'None'


        ws.cell(row=1, column=1).value = 'Project Name'
        ws.cell(row=1, column=2).value = 'latest quarter (£m)'
        ws.cell(row=1, column=3).value = 'baseline (£m)'
        ws.cell(row=1, column=4).value = '£m change between latest and baseline'
        ws.cell(row=1, column=5).value = 'percentage change between latest and baseline'

    return wb

def get_wlc(project_name_list, wlc_key, index):
    '''
    Function that gets projects wlc cost information and returns it in a python dictionary format.
    :param project_name_list: list of project names
    :param wlc_key: project whole life cost (wlc) key
    :param index: index value for which master to use from the q_master_data_list . 0 is for latest, 1 last and
    2 baseline. The actual index list q_master_list is set at a global level in this programme.
    :return: a dictionary structured 'wlc: 'project_name': total
    '''
    upper_dictionary = {}
    lower_dictionary = {}
    for project_name in project_name_list:
        try:
            project_data = list_of_masters_all[fin_bc_index[project_name][index]].data[project_name]
            total = project_data[wlc_key]
            lower_dictionary[project_name] = total
        except TypeError:
            lower_dictionary[project_name] = 0

    upper_dictionary['wlc'] = lower_dictionary

    return upper_dictionary


'''getting financial wlc cost breakdown'''
latest_wlc = get_wlc(latest_quarter_project_names, wlc_key, 0)
last_wlc = get_wlc(latest_quarter_project_names, wlc_key, 1)
baseline_wlc = get_wlc(latest_quarter_project_names, wlc_key, 2)

'''creating excel outputs'''
output_one = place_complex_comparision_excel(latest_wlc, last_wlc, baseline_wlc)
output_two = place_complex_comparision_excel(latest_cost_profiles, last_cost_profiles, baseline_cost_profiles)
output_three = place_standard_comparision_excel(latest_wlc, baseline_wlc)
output_four = place_standard_comparision_excel(latest_cost_profiles, baseline_cost_profiles)

'''INSTRUCTIONS FOR RUNNING PROGRAMME'''

'''Valid file paths for all the below need to be provided'''

'''ONE. Provide file path to where to save complex wlc breakdown'''
output_one.save(root_path/'output/comparing_wlc_complex_q3_1920.xlsx')

'''TWO. Provide file path to where to save complex yearly cost profile breakdown'''
output_two.save(root_path/'output/comparing_cost_profiles_complex_q3_1920.xlsx')

'''THREE. Provide file path to where to save standard wlc breakdown'''
output_three.save(root_path/'output/comparing_wlc_standard_q3_1920.xlsx')

'''FOUR. Provide file path to where to save standard yearly cost profile breakdown'''
output_four.save(root_path/'output/comparing_cost_profiles_standard_q3_1920.xlsx')