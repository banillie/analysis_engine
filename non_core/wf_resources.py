from openpyxl import load_workbook, Workbook
from datamaps.api import project_data_from_master


def create_keys():
    """
    Quick way to iterate the numbers in front of keys with the same names. To create the datamap.
    """
    wb = load_workbook("/home/will/Documents/workforce_review/dm_project_wf.xlsx")
    ws = wb.active
    int_list = []
    for cell in ws["A"]:
        if "1" in cell.value:
            int_list.append(cell.value)

    # print(int_list)  # if you need to get key names

    final_list = []
    for x in range(1, 200):
        for i in int_list:
            w = list(i)
            w[0] = str(x)
            final_list.append("".join(w))

    for i, x in enumerate(final_list):
        ws.cell(row=8 + i, column=1).value = x

    wb.save("/home/will/Documents/workforce_review/dm_project_wf_altered.xlsx")


key_list = [
    "1.1 Post (Job Title)",
    "1.2 Which function does this post belong to? (please select one from the drop-down list, if other, please specify in the comments column)",
    "1.3 Post Grade",
    "1.4 Post FTE on this project ",
    "1.5 If FTE in 2.3 is less than 1, please select the reason why from the drop-down list (e.g. working on another project, part-time etc.) ",
    "1.6 Is the post holder a project delivery professional?",
    "1.7 How is the post resourced? (e.g. substantive, on loan, seconded, FTA, FTC)",
    "1.8 Is this post currently vacant? If so, please provide comments on recruitment progress. Please specify whether 'recruitment is in hand' or whether there are any risks in filling the posts and how these are being mitigated?",
    "1.9 When did the current post holder join the project (date dd.mm.yyyy)?",
    "1.10 When is the current post holder expected to leave the project? (if applicable, dd.mm.yyyy)",
    "1.11 Comments",
]


def alter_key(key, number, ):
    """
    To handle issue around how keys are saved via project_data_to_master and the key list.
    """
    word = list(key)
    word[0] = str(number)
    word = [s.replace(',', '') for s in word]  # removes commas
    altered_key = "".join(word).rstrip()  # some keys have a trailing blank space.
    return altered_key


def data_for_pbi(master_path, save_path):
    """
    Transposes data from bcompiler format into the format for PBI.
    """
    wf_dict = project_data_from_master(master_path, 1, 2000)
    wb = Workbook()
    ws = wb.active
    chopped_dict = {}  # This dictionary chops off entries at the point they are none
    for project in wf_dict.projects:
        values = []
        stop = False
        for no in range(1, 200):
            none_check = []
            if stop is True:
                break
            for i, k in enumerate(key_list):
                key = alter_key(k, no)
                val = wf_dict.data[project][key]
                none_check.append(val)
                values.append((key, val))
                if i == 10:
                    if list(set(none_check)) == [None]:
                        del values[-11:-1]
                        stop = True
                        break

        chopped_dict[project] = dict(values)

    add = 0
    for project in wf_dict.projects:
        stop = False
        for no in range(1, 200):
            if stop is True:
                break
            if no == 1:
                ws.cell(row=1, column=1).value = "Project"
            ws.cell(row=add + no + 1, column=1).value = project
            for i, k in enumerate(key_list):
                if no == 1:
                    word = list(key)
                    ws.cell(row=1, column=2 + i).value = "".join(word[4:]).lstrip()
                key = alter_key(k, no)
                try:
                    val = chopped_dict[project][key]
                    ws.cell(row=add + no + 1, column=2 + i).value = val
                except KeyError:
                    ws.delete_rows(add + no + 1)
                    print(project, no)  # shows how many entries for each project.
                    add += (no - 1)
                    stop = True
                    break

    wb.save(save_path)


# create_keys()

m_path = "/home/will/Documents/workforce_review/work_force_master_bcompiler_version.xlsx"
s_path = "/home/will/Documents/workforce_review/work_force_master_pbi_version.xlsx"

data_for_pbi(m_path, s_path)
