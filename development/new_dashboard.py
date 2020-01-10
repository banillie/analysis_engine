'''

In development... function for creating different dashboard in line with PDIP work

Input documents:
1) Dashboard master document - this is an excel file.
2) Master data for two quarters - this will usually be latest and previous quarter

output document:
3) Dashboard with all project data placed into dashboard and formatted correctly.

Instructions:
1) provide path to dashboard master
2) provide path to master data sets
3) change bicc_date variable
4) provide path and specify file name for output document
'''


# comment

from openpyxl import load_workbook
import datetime
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, IconSet, FormatObject
from analysis.engine_functions import all_milestone_data_bulk, concatenate_dates, up_or_down, convert_rag_text, \
    convert_bc_stage_text
from analysis.data import q2_1920, q1_1920

def calculating_schedule_progression(proj_name, m_data):

    '''function that calculations the propotion of schedule that has been completed'''

    start_date = tuple(m_data[proj_name]['Start of Project'])[0]

    end_date = tuple(m_data[proj_name]['Project End Date'])[0]

    now = datetime.date.today()

    try:
        proj_length = (end_date - start_date).days
        remain_time = (end_date - now).days
        completed = round(100 - ((remain_time / proj_length) * 100), 0)
    except TypeError:
        completed = 'missing data'

    return completed

def calculating_cost_progression(master_data, proj_name):

    total = master_data.data[proj_name]['Total Forecast']


    pre_rdel = master_data.data[proj_name]['Pre 19-20 RDEL Forecast Total']
    if pre_rdel == None:
        pre_rdel = 0
    pre_cdel = master_data.data[proj_name]['Pre 19-20 CDEL Forecast Total']
    if pre_cdel == None:
        pre_cdel = 0
    pre_ngov = master_data.data[proj_name]['Pre 19-20 Forecast Non-Gov']
    if pre_ngov == None:
        pre_ngov = 0
    pre_spend = pre_rdel + pre_cdel + pre_ngov

    try:
        output = round((pre_spend / total) * 100, 0)
    except ZeroDivisionError:
        output = 0

    return output

'''function that places all information into the summary dashboard sheet'''
def placing_excel(project_name_list, master_data_one, master_data_two):

    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=3).value
        print(project_name)
        if project_name in project_name_list:
            ws.cell(row=row_num, column=4).value = convert_bc_stage_text(master_data_one.data[project_name]
                                                                         ['BICC approval point'])
            ws.cell(row=row_num, column=5).value = round(master_data_one.data[project_name]['Total Forecast'], 0)
            ws.cell(row=row_num, column=6).value = master_data_one.data[project_name]['Adjusted Benefits Cost Ratio (BCR)']

            p_m_data = all_milestone_data_bulk([project_name], master_data_one)
            try:
                ws.cell(row=row_num, column=7).value = (list(p_m_data[project_name]['Project End Date'])[0])
            except KeyError:
                ws.cell(row=row_num, column=7).value = 'No end date'

            ws.cell(row=row_num, column=10).value = convert_rag_text(master_data_one.data[project_name]['Departmental DCA'])
            try:
                dca_now = master_data_one.data[project_name]['Departmental DCA']
                dca_past = master_data_two.data[project_name]['Departmental DCA']
                ws.cell(row=row_num, column=9).value = up_or_down(dca_now, dca_past)
            except KeyError:
                ws.cell(row=row_num, column=9).value = 'NEW'
            ws.cell(row=row_num, column=11).value = calculating_schedule_progression(project_name, p_m_data)
            ws.cell(row=row_num, column=12).value = calculating_cost_progression(master_data_one, project_name)
            ws.cell(row=row_num, column=13).value = master_data_one.data[project_name]['Overall figure for Optimism Bias (£m)']
            ws.cell(row=row_num, column=14).value = master_data_one.data[project_name]['Overall contingency (£m)']

    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=3).value
        if project_name in master_data_two.projects:
            ws.cell(row=row_num, column=8).value = convert_rag_text(master_data_two.data[project_name]['Departmental DCA'])

    # Highlight cells that contain RAG text, with background and text the same colour. column G.

    # rag_txt_list = ["A/G", "A/R", "R", "G", "A"]
    #
    # for name in rag_txt_list:
    #     ag_text = Font(color="00a5b700")
    #     ag_fill = PatternFill(bgColor="00a5b700")
    #     dxf = DifferentialStyle(font=ag_text, fill=ag_fill)
    #     rule = Rule(type="containsText", operator="containsText", text=name, dxf=dxf)
    #     rule.formula = ['NOT(ISERROR(SEARCH('+name+',g1)))']
    #     ws.conditional_formatting.add('g1:g100', rule)
    #
    # for name in rag_txt_list:
    #     ag_text = Font(color="00a5b700")
    #     ag_fill = PatternFill(bgColor="00a5b700")
    #     dxf = DifferentialStyle(font=ag_text, fill=ag_fill)
    #     rule = Rule(type="containsText", operator="containsText", text=name, dxf=dxf)
    #     rule.formula = ['NOT(ISERROR(SEARCH('+name+',i1)))']
    #     ws.conditional_formatting.add('i1:i100', rule)

    # ar_text = Font(color="00f97b31")
    # ar_fill = PatternFill(bgColor="00f97b31")
    # dxf = DifferentialStyle(font=ar_text, fill=ar_fill)
    # rule = Rule(type="containsText", operator="containsText", text="A/R", dxf=dxf)
    # rule.formula = ['NOT(ISERROR(SEARCH("A/R",g1)))']
    # ws.conditional_formatting.add('g1:g100', rule)
    #
    # red_text = Font(color="00fc2525")
    # red_fill = PatternFill(bgColor="00fc2525")
    # dxf = DifferentialStyle(font=red_text, fill=red_fill)
    # rule = Rule(type="containsText", operator="containsText", text="R", dxf=dxf)
    # rule.formula = ['NOT(ISERROR(SEARCH("R",E1)))']
    # ws.conditional_formatting.add('E1:E100', rule)
    #
    # green_text = Font(color="0017960c")
    # green_fill = PatternFill(bgColor="0017960c")
    # dxf = DifferentialStyle(font=green_text, fill=green_fill)
    # rule = Rule(type="containsText", operator="containsText", text="G", dxf=dxf)
    # rule.formula = ['NOT(ISERROR(SEARCH("G",e1)))']
    # ws.conditional_formatting.add('E1:E100', rule)
    #
    # amber_text = Font(color="00fce553")
    # amber_fill = PatternFill(bgColor="00fce553")
    # dxf = DifferentialStyle(font=amber_text, fill=amber_fill)
    # rule = Rule(type="containsText", operator="containsText", text="A", dxf=dxf)
    # rule.formula = ['NOT(ISERROR(SEARCH("A",e1)))']
    # ws.conditional_formatting.add('e1:e100', rule)

    # highlight cells in column g

    # ag_text = Font(color="000000")
    # ag_fill = PatternFill(bgColor="00a5b700")
    # dxf = DifferentialStyle(font=ag_text, fill=ag_fill)
    # rule = Rule(type="containsText", operator="containsText", text="A/G", dxf=dxf)
    # rule.formula = ['NOT(ISERROR(SEARCH("A/G",g1)))']
    # ws.conditional_formatting.add('g1:g100', rule)
    #
    # ar_text = Font(color="000000")
    # ar_fill = PatternFill(bgColor="00f97b31")
    # dxf = DifferentialStyle(font=ar_text, fill=ar_fill)
    # rule = Rule(type="containsText", operator="containsText", text="A/R", dxf=dxf)
    # rule.formula = ['NOT(ISERROR(SEARCH("A/R",g1)))']
    # ws.conditional_formatting.add('g1:g100', rule)
    #
    # red_text = Font(color="000000")
    # red_fill = PatternFill(bgColor="00fc2525")
    # dxf = DifferentialStyle(font=red_text, fill=red_fill)
    # rule = Rule(type="containsText", operator="containsText", text="R", dxf=dxf)
    # rule.formula = ['NOT(ISERROR(SEARCH("R",g1)))']
    # ws.conditional_formatting.add('g1:g100', rule)
    #
    # green_text = Font(color="000000")
    # green_fill = PatternFill(bgColor="0017960c")
    # dxf = DifferentialStyle(font=green_text, fill=green_fill)
    # rule = Rule(type="containsText", operator="containsText", text="G", dxf=dxf)
    # rule.formula = ['NOT(ISERROR(SEARCH("G",g1)))']
    # ws.conditional_formatting.add('g1:g100', rule)
    #
    # amber_text = Font(color="000000")
    # amber_fill = PatternFill(bgColor="00fce553")
    # dxf = DifferentialStyle(font=amber_text, fill=amber_fill)
    # rule = Rule(type="containsText", operator="containsText", text="A", dxf=dxf)
    # rule.formula = ['NOT(ISERROR(SEARCH("A",g1)))']
    # ws.conditional_formatting.add('g1:g100', rule)
    #
    # # highlight cells in column H
    #
    # ag_text = Font(color="000000")
    # ag_fill = PatternFill(bgColor="00a5b700")
    # dxf = DifferentialStyle(font=ag_text, fill=ag_fill)
    # rule = Rule(type="containsText", operator="containsText", text="A/G", dxf=dxf)
    # rule.formula = ['NOT(ISERROR(SEARCH("A/G",h1)))']
    # ws.conditional_formatting.add('h1:h100', rule)
    #
    # ar_text = Font(color="000000")
    # ar_fill = PatternFill(bgColor="00f97b31")
    # dxf = DifferentialStyle(font=ar_text, fill=ar_fill)
    # rule = Rule(type="containsText", operator="containsText", text="A/R", dxf=dxf)
    # rule.formula = ['NOT(ISERROR(SEARCH("A/R",h1)))']
    # ws.conditional_formatting.add('h1:h100', rule)
    #
    # red_text = Font(color="000000")
    # red_fill = PatternFill(bgColor="00fc2525")
    # dxf = DifferentialStyle(font=red_text, fill=red_fill)
    # rule = Rule(type="containsText", operator="containsText", text="R", dxf=dxf)
    # rule.formula = ['NOT(ISERROR(SEARCH("R",h1)))']
    # ws.conditional_formatting.add('h1:h100', rule)
    #
    # green_text = Font(color="000000")
    # green_fill = PatternFill(bgColor="0017960c")
    # dxf = DifferentialStyle(font=green_text, fill=green_fill)
    # rule = Rule(type="containsText", operator="containsText", text="G", dxf=dxf)
    # rule.formula = ['NOT(ISERROR(SEARCH("G",h1)))']
    # ws.conditional_formatting.add('h1:h100', rule)
    #
    # amber_text = Font(color="000000")
    # amber_fill = PatternFill(bgColor="00fce553")
    # dxf = DifferentialStyle(font=amber_text, fill=amber_fill)
    # rule = Rule(type="containsText", operator="containsText", text="A", dxf=dxf)
    # rule.formula = ['NOT(ISERROR(SEARCH("A",h1)))']
    # ws.conditional_formatting.add('h1:h100', rule)
    #
    # # highlight cells in column H
    #
    # ag_text = Font(color="000000")
    # ag_fill = PatternFill(bgColor="00a5b700")
    # dxf = DifferentialStyle(font=ag_text, fill=ag_fill)
    # rule = Rule(type="containsText", operator="containsText", text="A/G", dxf=dxf)
    # rule.formula = ['NOT(ISERROR(SEARCH("A/G",l1)))']
    # ws.conditional_formatting.add('l1:l100', rule)
    #
    # ar_text = Font(color="000000")
    # ar_fill = PatternFill(bgColor="00f97b31")
    # dxf = DifferentialStyle(font=ar_text, fill=ar_fill)
    # rule = Rule(type="containsText", operator="containsText", text="A/R", dxf=dxf)
    # rule.formula = ['NOT(ISERROR(SEARCH("A/R",l1)))']
    # ws.conditional_formatting.add('l1:l100', rule)
    #
    # red_text = Font(color="000000")
    # red_fill = PatternFill(bgColor="00fc2525")
    # dxf = DifferentialStyle(font=red_text, fill=red_fill)
    # rule = Rule(type="containsText", operator="containsText", text="R", dxf=dxf)
    # rule.formula = ['NOT(ISERROR(SEARCH("R",l1)))']
    # ws.conditional_formatting.add('l1:l100', rule)
    #
    # green_text = Font(color="000000")
    # green_fill = PatternFill(bgColor="0017960c")
    # dxf = DifferentialStyle(font=green_text, fill=green_fill)
    # rule = Rule(type="containsText", operator="containsText", text="G", dxf=dxf)
    # rule.formula = ['NOT(ISERROR(SEARCH("G",l1)))']
    # ws.conditional_formatting.add('l1:l100', rule)
    #
    # amber_text = Font(color="000000")
    # amber_fill = PatternFill(bgColor="00fce553")
    # dxf = DifferentialStyle(font=amber_text, fill=amber_fill)
    # rule = Rule(type="containsText", operator="containsText", text="A", dxf=dxf)
    # rule.formula = ['NOT(ISERROR(SEARCH("A",l1)))']
    # ws.conditional_formatting.add('l1:l100', rule)


    # Highlight cells that contain RAG text, with background and black text columns G to L.
    # ag_text = Font(color="000000")
    # ag_fill = PatternFill(bgColor="00a5b700")
    # dxf = DifferentialStyle(font=ag_text, fill=ag_fill)
    # rule = Rule(type="uniqueValues", operator="equal", text="A/G", dxf=dxf)
    # rule.formula = ['NOT(ISERROR(SEARCH("A/G",G1)))']
    # ws.conditional_formatting.add('G1:L100', rule)
    #
    # ar_text = Font(color="000000")
    # ar_fill = PatternFill(bgColor="00f97b31")
    # dxf = DifferentialStyle(font=ar_text, fill=ar_fill)
    # rule = Rule(type="uniqueValues", operator="equal", text="A/R", dxf=dxf)
    # rule.formula = ['NOT(ISERROR(SEARCH("A/R",G1)))']
    # ws.conditional_formatting.add('G1:L100', rule)
    #
    # red_text = Font(color="000000")
    # red_fill = PatternFill(bgColor="00fc2525")
    # dxf = DifferentialStyle(font=red_text, fill=red_fill)
    # rule = Rule(type="uniqueValues", operator="equal", text="R", dxf=dxf)
    # rule.formula = ['NOT(ISERROR(SEARCH("R",G1)))']
    # ws.conditional_formatting.add('G1:L100', rule)
    #
    # green_text = Font(color="000000")
    # green_fill = PatternFill(bgColor="0017960c")
    # dxf = DifferentialStyle(font=green_text, fill=green_fill)
    # rule = Rule(type="uniqueValues", operator="equal", text="G", dxf=dxf)
    # rule.formula = ['NOT(ISERROR(SEARCH("Green",G1)))']
    # ws.conditional_formatting.add('G1:L100', rule)
    #
    # amber_text = Font(color="000000")
    # amber_fill = PatternFill(bgColor="00fce553")
    # dxf = DifferentialStyle(font=amber_text, fill=amber_fill)
    # rule = Rule(type="uniqueValues", operator="equal", text="A", dxf=dxf)
    # rule.formula = ['NOT(ISERROR(SEARCH("A",G1)))']
    # ws.conditional_formatting.add('G1:L100', rule)

    # highlighting new projects
    # red_text = Font(color="00fc2525")
    # white_fill = PatternFill(bgColor="000000")
    # dxf = DifferentialStyle(font=red_text, fill=white_fill)
    # rule = Rule(type="uniqueValues", operator="equal", text="NEW", dxf=dxf)
    # rule.formula = ['NOT(ISERROR(SEARCH("NEW",F1)))']
    # ws.conditional_formatting.add('F1:F100', rule)

    # assign the icon set to a rule
    first = FormatObject(type='num', val=-1)
    second = FormatObject(type='num', val=0)
    third = FormatObject(type='num', val=1)
    iconset = IconSet(iconSet='3Arrows', cfvo=[first, second, third], percent=None, reverse=None)
    rule = Rule(type='iconSet', iconSet=iconset)
    ws.conditional_formatting.add('H1:H100', rule)

    return wb



'''1) Provide file to empty dashboard document'''
wb = load_workbook(
    'C:\\Users\\Standalone\\general\\masters folder\\portfolio_dashboards\\pdip_dashboard_poc_master.xlsx')
ws = wb.active

p_names = q2_1920.projects

wb = placing_excel(p_names, q2_1920, q1_1920)

'''4) provide file path and specific name of output file.'''
wb.save('C:\\Users\\Standalone\\general\\masters folder\\portfolio_dashboards\\pdip_dashboard_poc_testing.xlsx')