'''
some new code to create costs v schedule graph
'''

from analysis.data import list_of_masters_all, bc_index, latest_quarter_project_names, financial_analysis_masters_list, \
    fin_bc_index
from analysis.engine_functions import all_milestone_data_bulk
from openpyxl import Workbook

def cost_v_schedule_chart():

    l_data = list_of_masters_all[0]

    wb = Workbook()
    ws = wb.active

    ws.cell(row=2, column=2).value = 'Project Name'
    ws.cell(row=2, column=3).value = 'Schedule change'
    ws.cell(row=2, column=4).value = 'WLC Change'
    ws.cell(row=2, column=5).value = 'WLC'
    ws.cell(row=2, column=6).value = 'DCA'

    for x, project_name in enumerate(l_data.projects):
        print(project_name)
        ws.cell(row=x+3, column=2).value = project_name
        ws.cell(row=x+3, column=3).value = calculate_schedule_change(project_name)
        ws.cell(row=x+3, column=4).value = calculate_wlc_change(project_name)
        ws.cell(row=x+3, column=5).value = l_data.data[project_name]['Total Forecast']
        ws.cell(row=x+3, column=6).value = l_data.data[project_name]['Departmental DCA']

    return wb

def calculate_wlc_change(project_name):

    '''Total WLC'''
    wlc_now = financial_analysis_masters_list[0].data[project_name]['Total Forecast']
    '''WLC variance against baseline quarter'''
    wlc_baseline = financial_analysis_masters_list[fin_bc_index[project_name][2]].data[project_name]['Total Forecast']

    percentage_change = int(((wlc_now - wlc_baseline) / wlc_now) * 100)

    return percentage_change

def calculate_schedule_change(project_name):

    '''full operation current date'''
    proj_milestones = all_milestone_data_bulk([project_name], list_of_masters_all[0])

    try:
        # foc_one = tuple(proj_milestones['Full Operating Capacity (FOC)'])[0]
        foc_one = tuple(proj_milestones[project_name]['Project End Date'])[0]
        if foc_one is None:
            try:
                foc_one = tuple(proj_milestones[project_name]['Full Operations'])[0]
            except (KeyError, TypeError):
                foc_one = None
    except KeyError:
        foc_one = None

    '''full operation baseline date'''
    proj_milestones_bl = all_milestone_data_bulk([project_name], list_of_masters_all[bc_index[project_name][2]])

    try:
        sop = tuple(proj_milestones_bl[project_name]['Start of Project'])[0]
    except KeyError:
        sop = None

    try:
        # foc_two = tuple(proj_milestones_bl['Full Operating Capacity (FOC)'])[0]
        foc_two = tuple(proj_milestones_bl[project_name]['Project End Date'])[0]

        if foc_two is None:
            try:
                foc_two = tuple(proj_milestones_bl[project_name]['Full Operations'])[0]
            except (KeyError, TypeError):
                foc_two = None
    except KeyError:
        foc_two = None

    try:
        proj_length = (foc_two - sop).days  # project length
    except TypeError:
        proj_length = None
    try:
        foc_change = (foc_one - foc_two).days
    except TypeError:
        foc_change = None

    try:
        percent_change = int((foc_change / proj_length) * 100)
    except TypeError:
        percent_change = 'couldn\'t calculate'

    return percent_change

def calculate_schedule_change_full_check(project_name, ws, x):
    '''this function isn't to be used but contains the workings for reaching the change figure so keeping in case
    helpful in future'''

    ws.cell(row=2, column=3).value = 'project Full Operation. NOW'
    ws.cell(row=2, column=4).value = 'project Start of Operation'
    ws.cell(row=2, column=5).value = 'project Full Operation BL'
    ws.cell(row=2, column=6).value = 'project length'
    ws.cell(row=2, column=7).value = 'length change'
    ws.cell(row=2, column=8).value = 'percentage change'

    '''full operation current date'''
    proj_milestones = all_milestone_data_bulk([project_name], list_of_masters_all[0])

    try:
        # foc_one = tuple(proj_milestones['Full Operating Capacity (FOC)'])[0]
        foc_one = tuple(proj_milestones[project_name]['Project End Date'])[0]

        if foc_one is None:
            try:
                foc_one = tuple(proj_milestones[project_name]['Full Operations'])[0]
                ws.cell(row=x + 3, column=3).value = foc_one
            except (KeyError, TypeError):
                foc_one = None
                ws.cell(row=x + 4, column=3).value = foc_one
        else:
            ws.cell(row=x + 3, column=3).value = foc_one

    except KeyError:
        foc_one = None
        ws.cell(row=x + 3, column=3).value = foc_one

    '''full operation baseline date'''
    proj_milestones_bl = all_milestone_data_bulk([project_name], list_of_masters_all[bc_index[project_name][2]])

    try:
        sop = tuple(proj_milestones_bl[project_name]['Start of Project'])[0]
        ws.cell(row=x + 3, column=4).value = sop
    except KeyError:
        sop = None
        ws.cell(row=x + 3, column=4).value = sop

    try:
        # foc_two = tuple(proj_milestones_bl['Full Operating Capacity (FOC)'])[0]
        foc_two = tuple(proj_milestones_bl[project_name]['Project End Date'])[0]

        if foc_two is None:
            try:
                foc_two = tuple(proj_milestones_bl[project_name]['Full Operations'])[0]
                ws.cell(row=x + 3, column=5).value = foc_two
            except (KeyError, TypeError):
                foc_two = None
                ws.cell(row=x + 3, column=5).value = foc_two
        else:
            ws.cell(row=x + 3, column=5).value = foc_two
    except KeyError:
        foc_two = None
        ws.cell(row=x + 3, column=5).value = foc_two

    try:
        proj_length = (foc_two - sop).days  # project length
        ws.cell(row=x + 3, column=6).value = proj_length
    except TypeError:
        proj_length = None
        ws.cell(row=x + 3, column=6).value = proj_length
    try:
        foc_change = (foc_one - foc_two).days
        ws.cell(row=x + 3, column=7).value = foc_change
    except TypeError:
        foc_change = None
        ws.cell(row=x + 3, column=7).value = foc_change

    try:
        percent_change = int((foc_change / proj_length) * 100)
        ws.cell(row=x + 3, column=8).value = percent_change
    except TypeError:
        ws.cell(row=x + 3, column=8).value = 'couldn\'t calculate'


current_milestones_all = all_milestone_data_bulk(latest_quarter_project_names, list_of_masters_all[0])

run = cost_v_schedule_chart()
run.save('/home/will/Documents/portfolio/cost_v_schedule.xlsx')