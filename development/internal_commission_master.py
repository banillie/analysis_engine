'''probably throw away code which converts data key names some that 'lst quarter' or some other text can be added to
the front of the keys. This will then be used for structuring overall commission master and migrating relevant
quarter data into the commission master

in development. not working, but keeping as could be useful'''

from openpyxl import load_workbook, Workbook
from analysis.data import list_of_masters_all, root_path

def get_key_names(workbook):
    key_list = []
    lst_q_key_list = []
    ws = workbook.active

    for row_num in range(2, ws.max_row + 1):
        key_list.append(ws.cell(row=row_num, column=1).value)

    for key in key_list:
        lst_q_key_list.append('lst qrt ' + str(key))

    wb = Workbook()
    ws = wb.active

    for i, key in enumerate(lst_q_key_list):
        ws.cell(row=1 + i, column=1).value = key

    return wb

run = get_key_names(list_of_masters_all[0])

run.save(root_path/'output/keys_with_lst_quarter.xlsx')