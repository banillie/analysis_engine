'''returns data required by no 10 commission'''

# TODO. conditional formatting, quarter stamp information.

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from analysis.data import list_of_masters_all, latest_quarter_project_names, bc_index, baseline_bc_stamp, salmon_fill, \
    root_path
from analysis.engine_functions import all_milestone_data_bulk, ap_p_milestone_data_bulk, assurance_milestone_data_bulk,\
    get_all_project_names, get_quarter_stamp, grey_conditional_formatting

def return_baseline_data_projects(project_name_list, data_key_list):

    wb = Workbook()

    for i, project_name in enumerate(project_name_list):
        '''worksheet is created for each project'''
        ws = wb.create_sheet(project_name, i)  # creating worksheets
        ws.title = project_name  # title of worksheet

        '''list project names, groups and stage in ws'''
        for y, key in enumerate(data_key_list):
            if key in list_of_masters_all[0].data[project_name].keys():
                #standard keys
                ws.cell(row=2+y, column=1, value=key) #returns key
                ws.cell(row=2+y, column=2).value = list_of_masters_all[0].data[project_name][
                    key] #returns latest value
                for x in range(0, len(baseline_bc_stamp[project_name])):
                    quarter_info = baseline_bc_stamp[project_name][x][1]
                    ws.cell(row=1, column=3+x, value=quarter_info) #returns quarter info
                    index = baseline_bc_stamp[project_name][x][2]
                    try:
                        ws.cell(row=2+y, column=3+x, value=list_of_masters_all[index].data[project_name][key]) #returns baselines
                    except KeyError:
                        ws.cell(row=2+y, column=3+x).value = 'missing data'

            else:
                #milestones keys
                ws.cell(row=2 + y, column=1, value=key)  # returns key
                milestones = all_milestone_data_bulk([project_name], list_of_masters_all[0])
                try:
                    ws.cell(row=2 + y, column=2).value = tuple(milestones[project_name][key])[0]  # returns latest value
                except KeyError:
                    ws.cell(row=2 + y, column=2).value = 'missing data'

                for x in range(0, len(baseline_bc_stamp[project_name])):
                    quarter_info = baseline_bc_stamp[project_name][x][1]
                    ws.cell(row=1, column=3 + x, value=quarter_info)  # returns quarter info
                    index = baseline_bc_stamp[project_name][x][2]
                    try:
                        milestones = all_milestone_data_bulk([project_name], list_of_masters_all[index])
                        ws.cell(row=2 + y, column=3 + x).value = tuple(milestones[project_name][key])[0]  # returns baselines
                    except KeyError:
                        ws.cell(row=2 + y, column=3 + x).value = 'project not reporting'

        ws.cell(row=1, column=1, value='Key')
        ws.cell(row=1, column=2, value='Latest')

        '''conditional formating'''
        for column in list_columns:
            for i, txt in enumerate(conditional_text):
                text = text_colours[i]
                fill = background_colours[i]
                dxf = DifferentialStyle(font=text, fill=fill)
                rule = Rule(type="containsText", operator="containsText", text=txt, dxf=dxf)
                for_rule_formula = 'NOT(ISERROR(SEARCH("' + txt + '",' + column + '1)))'
                rule.formula = [for_rule_formula]
                ws.conditional_formatting.add(column + '1:' + column + '60', rule)

    return wb

def return_data_projects(project_name_list, data_key_list):

    wb = Workbook()

    for i, project_name in enumerate(project_name_list):
        '''worksheet is created for each project'''
        ws = wb.create_sheet(project_name, i)  # creating worksheets
        ws.title = project_name  # title of worksheet

        '''list project names, groups and stage in ws'''
        for y, key in enumerate(data_key_list):
            ws.cell(row=2+y, column=1, value=key) #returns key
            for x, master in enumerate(list_of_masters_all):

                #standard keys

                try:
                    #standard keys
                    if key in master.data[project_name].keys():
                        value = master.data[project_name][key]
                        ws.cell(row=2+y, column=2+x, value=value) #retuns value

                        try:  #loop checks if value has changed since last quarter
                            lst_value = list_of_masters_all[x+1].data[project_name][key]
                            if value != lst_value:
                                ws.cell(row=2+y, column=2+x).fill = salmon_fill
                        except IndexError:
                            pass

                    # milestone keys
                    else:
                        milestones = all_milestone_data_bulk([project_name], list_of_masters_all[x])
                        value = tuple(milestones[project_name][key])[0]
                        ws.cell(row=2 + y, column=2 + x, value=value)
                        try:  # loop checks if value has changed since last quarter
                            old_milestones = all_milestone_data_bulk([project_name], list_of_masters_all[x + 1])
                            lst_value = tuple(old_milestones[project_name][key])[0]
                            if value != lst_value:
                                ws.cell(row=2 + y, column=2 + x).fill = salmon_fill
                        except IndexError:
                            pass

                except KeyError:
                    if project_name in master.projects: #loop calculates if project was not reporting or data missing
                        ws.cell(row=2+y, column=2+x, value='missing data')
                    else:
                        ws.cell(row=2+y, column=2+x, value='project not reporting')

        '''quarter tag information'''
        ws.cell(row=1, column=1, value='Key')
        quarter_labels = get_quarter_stamp(list_of_masters_all)
        for i, label in enumerate(quarter_labels):
            ws.cell(row=1, column=i + 2, value=label)

        '''conditional formating'''
        for column in list_columns:
            for i, txt in enumerate(conditional_text):
                text = text_colours[i]
                fill = background_colours[i]
                dxf = DifferentialStyle(font=text, fill=fill)
                rule = Rule(type="containsText", operator="containsText", text=txt, dxf=dxf)
                for_rule_formula = 'NOT(ISERROR(SEARCH("' + txt + '",' + column + '1)))'
                rule.formula = [for_rule_formula]
                ws.conditional_formatting.add(column + '1:' + column + '60', rule)

    return wb

def return_baseline_data_metrics(project_name_list, data_key_list):

    wb = Workbook()

    for i, key in enumerate(data_key_list):
        '''worksheet is created for each project'''
        ws = wb.create_sheet(key, i)  # creating worksheets
        ws.title = key  # title of worksheet

        '''list project names, groups and stage in ws'''
        for y, project_name in enumerate(project_name_list):
            ws.cell(row=y+2, column=2, value=project_name) # project name returned
            ws.cell(row=y+2, column=1).value = list_of_masters_all[0].data[project_name]['DfT Group']

            #TODO: group name across multiple masters

            if key in list_of_masters_all[0].data[project_name].keys():
                #standard keys
                ws.cell(row=y+2, column=3).value = list_of_masters_all[0].data[project_name][
                    key] #returns latest value
                for x in range(0, len(baseline_bc_stamp[project_name])):
                    index = baseline_bc_stamp[project_name][x][2]
                    try:
                        ws.cell(row=y+2, column=x+4, value=list_of_masters_all[index].data[project_name][key]) #returns baselines
                    except KeyError:
                        ws.cell(row=2+y, column=4+x).value = 'missing data'

            else:
                #milestones keys
                milestones = all_milestone_data_bulk([project_name], list_of_masters_all[0])
                try:
                    ws.cell(row=2 + y, column=3).value = tuple(milestones[project_name][key])[0]  # returns latest value
                except KeyError:
                    ws.cell(row=2 + y, column=3).value = 'missing data'

                for x in range(0, len(baseline_bc_stamp[project_name])):
                    index = baseline_bc_stamp[project_name][x][2]
                    try:
                        milestones = all_milestone_data_bulk([project_name], list_of_masters_all[index])
                        ws.cell(row=2 + y, column=3 + x).value = tuple(milestones[project_name][key])[0]  # returns baselines
                    except KeyError:
                        ws.cell(row=2 + y, column=3 + x).value = 'project not reporting'

        ws.cell(row=1, column=1, value='Group')
        ws.cell(row=1, column=2, value='Project')
        ws.cell(row=1, column=3, value='Latest')
        ws.cell(row=1, column=4, value='BL 1')
        ws.cell(row=1, column=5, value='BL 2')
        ws.cell(row=1, column=6, value='BL 3')
        ws.cell(row=1, column=7, value='BL 4')

        '''conditional formating'''
        for column in list_columns:
            for i, txt in enumerate(conditional_text):
                text = text_colours[i]
                fill = background_colours[i]
                dxf = DifferentialStyle(font=text, fill=fill)
                rule = Rule(type="containsText", operator="containsText", text=txt, dxf=dxf)
                for_rule_formula = 'NOT(ISERROR(SEARCH("' + txt + '",' + column + '1)))'
                rule.formula = [for_rule_formula]
                ws.conditional_formatting.add(column + '1:' + column + '60', rule)

    return wb

def return_data_metrics(project_name_list, data_key_list):

    wb = Workbook()

    for i, key in enumerate(data_key_list):
        '''worksheet is created for each project'''
        ws = wb.create_sheet(key, i)  # creating worksheets
        ws.title = key  # title of worksheet

        '''list project names, groups and stage in ws'''
        for y, project_name in enumerate(project_name_list):
            ws.cell(row=y + 2, column=2, value=project_name)  # project name returned
            ws.cell(row=y + 2, column=1).value = list_of_masters_all[0].data[project_name]['DfT Group']

            for x, master in enumerate(list_of_masters_all):
                try:
                    #standard keys
                    if key in master.data[project_name].keys():
                        value = master.data[project_name][key]
                        ws.cell(row=2+y, column=3+x, value=value) #retuns value

                        try:  #loop checks if value has changed since last quarter
                            lst_value = list_of_masters_all[x+1].data[project_name][key]
                            if value != lst_value:
                                ws.cell(row=2+y, column=3+x).fill = salmon_fill
                        except IndexError:
                            pass

                    # milestone keys
                    else:
                        milestones = all_milestone_data_bulk([project_name], list_of_masters_all[x])
                        value = tuple(milestones[project_name][key])[0]
                        ws.cell(row=2 + y, column=3 + x, value=value)
                        try:  # loop checks if value has changed since last quarter
                            old_milestones = all_milestone_data_bulk([project_name], list_of_masters_all[x + 1])
                            lst_value = tuple(old_milestones[project_name][key])[0]
                            if value != lst_value:
                                ws.cell(row=2 + y, column=3 + x).fill = salmon_fill
                        except IndexError:
                            pass

                except KeyError:
                    if project_name in master.projects: #loop calculates if project was not reporting or data missing
                        ws.cell(row=2+y, column=3+x, value='missing data')
                    else:
                        ws.cell(row=2+y, column=3+x, value='project not reporting')

        '''quarter tag information'''
        ws.cell(row=1, column=1, value='Group')
        ws.cell(row=1, column=2, value='Projects')
        quarter_labels = get_quarter_stamp(list_of_masters_all)
        for i, label in enumerate(quarter_labels):
            ws.cell(row=1, column=i + 3, value=label)

        '''conditional formating'''
        for column in list_columns:
            for i, txt in enumerate(conditional_text):
                text = text_colours[i]
                fill = background_colours[i]
                dxf = DifferentialStyle(font=text, fill=fill)
                rule = Rule(type="containsText", operator="containsText", text=txt, dxf=dxf)
                for_rule_formula = 'NOT(ISERROR(SEARCH("' + txt + '",' + column + '1)))'
                rule.formula = [for_rule_formula]
                ws.conditional_formatting.add(column + '1:' + column + '60', rule)

    return wb




#TODO: to put into analysis.data
list_columns = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'q', 's', 't', 'u', 'w']
conditional_text = ['missing data', 'project not reporting']
'''store of different colours'''
darkish_grey_text = Font(color="002e4053")
darkish_grey_fill = PatternFill(bgColor="002e4053")
light_grey_text = Font(color="0085929e")
light_grey_fill = PatternFill(bgColor="0085929e")
text_colours = [darkish_grey_text, light_grey_text]
background_colours = [darkish_grey_fill, light_grey_fill]

'''data keys of interest'''
key_list = ['BICC approval point',
            'Total Forecast',
           #'Start of Construction/build',
           'Start of Operation',
           'Full Operations',
           'Project End Date',
           'Departmental DCA',
           'Initial Benefits Cost Ratio (BCR)',
           'Adjusted Benefits Cost Ratio (BCR)',
           'VfM Category single entry']

'''running prog - step one'''
run_project_all = return_data_projects(latest_quarter_project_names, key_list)
run_project_bl = return_baseline_data_projects(latest_quarter_project_names, key_list)
run_metric_bl = return_baseline_data_metrics(latest_quarter_project_names, key_list)
run_metric_all = return_data_metrics(latest_quarter_project_names, key_list)


'''step two'''
run_project_all.save(root_path/'output/no_10_data_proj_all.xlsx')
run_project_bl.save(root_path/'output/no_10_data_proj_bl.xlsx')
run_metric_all.save(root_path/'output/no_10_data_metric_all.xlsx')
run_metric_bl.save(root_path/'output/no_10_data_metric_bl.xlsx')