from openpyxl.styles import Font, PatternFill

COLOUR_DICT = {
    "A": "#ffba00",
    "A/G": "#92a700",
    "A/R": "#e77200",
    "R": "#c00000",
    "G": "#007d00",
    "": "#FFFFFF",  # white if missing
    "W": "#ffffff",
    "Amber": "#ffba00",
    "Amber/Green": "#92a700",
    "Amber/Red": "#e77200",
    "Red": "#c00000",
    "Green": "#007d00",
    "None": "#FFFFFF",  # white if missing
    None: "#FFFFFF",  # white if missing
    "WHITE": "#ffffff",
    "Worsening": "#c00000",
    "No Change Expected": "#ffba00",
    "Improving": "#007d00",
    "GREY": "#808080",
    "BLACK": "#000000",
    "BLUE": "#1434A4",
}

FACE_COLOUR = "#a0c1d5"

SALMON_FILL = PatternFill(
    start_color="FFFF8080", end_color="FFFF8080", fill_type="solid"
)
AMBER_FILL = PatternFill(start_color="FFBF00", end_color="FFBF00", fill_type="solid")

black_text = Font(color="00000000")
ag_text = Font(color="00a5b700")  # text same colour as background
ag_fill = PatternFill(bgColor="00a5b700")
ar_text = Font(color="00f97b31")  # text same colour as background
ar_fill = PatternFill(bgColor="00f97b31")
red_text = Font(color="00fc2525")  # text same colour as background
red_fill = PatternFill(bgColor="00fc2525")
green_text = Font(color="0017960c")  # text same colour as background
green_fill = PatternFill(bgColor="0017960c")
amber_text = Font(color="00fce553")  # text same colour as background
amber_fill = PatternFill(bgColor="00fce553")

fill_colour_list = [ag_fill, ar_fill, red_fill, green_fill, amber_fill]

# for colouring group circles
# circle_colours = [
#     '#263552',
#     '#f1ad64',
#     '#983d3f',
#     '#6b4351'
# ]
# circle_colours = [
#     '#ae4553',
#     '#f28335',
#     '#2b7b62',
#     '#efc15f'
# ]
