"""
code for creating and maintaining test masters
"""

from datamaps.api import project_data_from_master
from openpyxl import load_workbook, Workbook
import random
from analysis_engine.data import root_path

# test_master_one = project_data_from_master("/home/will/"
#                                            "code/python/analysis_engine/"
#                                            "tests/resources/test_master_4_2016.xlsx", 4, 2016)
# test_master_two = project_data_from_master("/home/will/code/python/analysis_engine/"
#                                            "tests/resources"
#                                            "/test_master_4_2017.xlsx", 4, 2017)
# test_master_three = project_data_from_master("/home/will/code/python/analysis_engine/"
#                                              "tests/resources"
#                                              "/test_master_4_2018.xlsx", 4, 2018)
#
#
# test_master_four = project_data_from_master("/home/will/code/python/analysis_engine/"
#                                             "tests/resources"
#                                             "/test_master_4_2019.xlsx", 4, 2019)

test_master_five = project_data_from_master("/home/will/code/python/analysis_engine/"
                                            "tests/resources"
                                            "/test_master_1_2020.xlsx", 1, 2020)

test_master_data_list = [test_master_five]

# test_wb_one = load_workbook("/home/will/"
#                             "code/python/analysis_engine/"
#                             "tests/resources/test_master_4_2016.xlsx")
# test_wb_two = load_workbook("/home/will/code/python/analysis_engine/"
#                             "tests/resources"
#                             "/test_master_4_2017.xlsx")
# test_wb_three = load_workbook("/home/will/code/python/analysis_engine/"
#                               "tests/resources"
#                               "/test_master_4_2018.xlsx")
# test_wb_four = load_workbook("/home/will/code/python/analysis_engine/"
#                              "tests/resources"
#                              "/test_master_4_2019.xlsx")
test_wb_five = load_workbook("/home/will/code/python/analysis_engine/"
                             "tests/resources"
                             "/test_master_1_2020.xlsx")

def get_excel_data_in_list(wb):
    """
    Takes data from a single column in excel ws and puts into a list.
    """
    ws = wb.active
    output_list = []

    for r in range(1, ws.max_row + 1):
        output_list.append(ws.cell(row=r, column=1).value)

    return output_list


def altering_test_masters(wb, altered_keys, master_data, project_names_dict):
    """
    Places altered values for each project into excel wb. The source for altered data
    is a previous test wb.
    """
    ws = wb.active

    for project_name in project_names_dict:
        for i in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=1 + i).value == project_name:
                ws.cell(row=1, column=1 + i).value = project_names_dict[project_name]

                for row_num in range(2, ws.max_row + 1):
                    for key in altered_keys:
                        if ws.cell(row=row_num, column=1).value == key:
                            ws.cell(row=row_num, column=1 + i).value = \
                                master_data.data[project_names_dict[project_name]][key]

    return wb


def insert_random_text_milestone_notes(wb, random_key_list):
    """
    Places random text into wb as per keys in random keys list as well as for
    those keys containing the strings in code below
    """

    random_string_one = "Don't you know an apparition is just a cheap date. What have you been drinking these days"

    random_string_two = 'Memento mori and amor fati'

    random_string_three = "The sea gets deeper the further you go into it"

    random_string_four = "What you see if all there is"

    random_string_five = "Green lumber fallacy"

    random_string_six = "The mind is not a vessel to be filled but a fire to be lighted"

    random_string_list_1 = [random_string_one, random_string_two,
                            random_string_three, random_string_four,
                            random_string_five, random_string_six]
    random_string_list_2 = ['hammer', 'nail', 'screw', 'wood', 'plastic', 'screw driver']

    ws = wb.active

    for row_num in range(2, ws.max_row + 1):
        if ws.cell(row=row_num, column=1).value in random_key_list:
            for i in range(1, 7):
                text = random.choice(random_string_list_1 + random_string_list_2)
                ws.cell(row=row_num, column=1 + i).value = text
        if 'Notes' in ws.cell(row=row_num, column=1).value:
            for i in range(1, 6):
                text = random.choice(random_string_list_1)
                ws.cell(row=row_num, column=1 + i).value = text
        if 'Gov Type' in ws.cell(row=row_num, column=1).value:
            for i in range(1, 6):
                text = random.choice(random_string_list_2)
                ws.cell(row=row_num, column=1 + i).value = text
        if 'LoD' in ws.cell(row=row_num, column=1).value:
            for i in range(1, 6):
                text = random.choice(random_string_list_2)
                ws.cell(row=row_num, column=1 + i).value = text

    return wb


def milestones_list(project_name, master_data_list):
    """
    Function that places project milestone names into a list.
    """

    output = []

    for master_data in master_data_list:
        try:
            p_data = master_data.data[project_name]
            for i in range(1, 50):
                try:
                    try:
                        t = p_data['Approval MM' + str(i)]
                        output.append(str(t))
                    except KeyError:
                        t = p_data['Approval MM' + str(i)]
                        output.append(str(t))

                    t = p_data['Assurance MM' + str(i)]
                    output.append(str(t))

                except KeyError:
                    pass

            for i in range(18, 67):
                try:
                    t = p_data['Project MM' + str(i)]
                    output.append(str(t))
                except KeyError:
                    pass
        except KeyError:
            pass

    f_output = [x for x in output if x != 'None']  # remove None types
    list(set(f_output))

    return list(set(f_output))


def gen_random_milestone_names(milestone_name_list):
    """
    generates random code names for real milestone names and returns
    data in python dictionary
    """
    first = (
        "Apollo", "Space", "Moon", "Sputnik", "Team", "Soyuz", "Gemini", "Astronauts", "Mercury", "Man", "Tranquility",
        "Serenity", "Earth", "Orbital", "Gravity", "Meteorite", "Craters", "Spaceship", "Oxygen", "Lunar", "Columbia",
        "Challenger", "Inverted")
    second = (
        "Sea", "Ocean", "Magma", "Magnetic", "Cosmic", "Radiation", "Landing", "Module", "Walk", "Cosmonauts",
        "Shuttle", "Command", "Armstrong", "Aldridge", "Collins", "Hypatia", "Lade", "Kestrel", "Roving",
        "Eleven", "Checklist", "Liftoff")
    output_dict = {}
    code_name_list = []

    standard_milestones_dict = {'Start of Project': 'Standard A',
                                'SOBC - IPDC Approval': 'Standard B',
                                'OBC - IPDC Approval': 'Standard C',
                                'FBC - IPDC Approval': 'Standard D',
                                'Start of Construction/build': 'Standard E',
                                'Start of Operation': 'Standard F',
                                'Full Operations': 'Standard G',
                                'Project End Date': 'Standard H'}

    while len(code_name_list) < len(milestone_name_list):
        firrst = random.choice(first)
        seccond = random.choice(second)
        name = (firrst + " " + seccond)
        if name not in code_name_list:
            code_name_list.append(name)

    for i, name in enumerate(milestone_name_list):
        if name not in (standard_milestones_dict.keys()):
            output_dict[name] = code_name_list[i]
        else:
            output_dict[name] = standard_milestones_dict[name]

    return output_dict


def put_codenames_in_excel(code_name_dict):
    """
    places milestone code names in an excel file as persistent record.
    """
    wb = Workbook()

    ws = wb.create_sheet('codenames', 0)
    ws.title = 'codenames'

    for i, key in enumerate(code_name_dict.keys()):
        ws.cell(row=i + 1, column=1).value = key
        ws.cell(row=i + 1, column=2).value = code_name_dict[key]

    return wb


def put_codenames_in_dict(wb):
    """
    places codenames from wb into a python dictionary
    """
    ws = wb.active
    output_dict = {}

    for x in range(1, ws.max_row + 1):
        key = ws.cell(row=x, column=1).value
        codename = ws.cell(row=x, column=2).value
        output_dict[key] = codename

    return output_dict


def altering_test_masters_milestones(wb, project_name, codename_dict):
    """
    place milestone code names into excel wb writing over actual name. 
    """
    ws = wb.active

    for i in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=1 + i).value == project_name:
            for row_num in range(2, ws.max_row + 1):
                for key in codename_dict.keys():
                    if ws.cell(row=row_num, column=1 + i).value == key:
                        ws.cell(row=row_num, column=1 + i).value = \
                            codename_dict[key]

    return wb


"""Load list of random keys into a list"""
# random_key_wb = load_workbook("/home/will/Documents/analysis_engine/code_resources/tests/random_keys_for_test_masters"
#                               ".xlsx")
# random_key_list = get_excel_data_in_list(random_key_wb)
"""Place the random data into the wb"""
# insert_random = insert_random_text_milestone_notes(test_wb_five, random_key_wb)
# insert_random.save("/home/will/code/python/analysis_engine/tests/resources/test_master_1_2020.xlsx")
"""Load list of altered keys names into a list"""
# altered_key_wb = load_workbook("/home/will/Documents/analysis_engine/code_resources/tests/"
#                                "altered_keys_for_test_masters.xlsx")
# altered_key_list = get_excel_data_in_list(altered_key_wb)
"""Place the altered data into the wb"""
project_code_names = load_workbook("/home/will/Documents/analysis_engine/code_resources/tests"
                                   "/project_test_masters_codenames.xlsx")
project_code_names_dict = put_codenames_in_dict(project_code_names)
# run = altering_test_masters(test_wb_five, altered_key_list, test_master_four, project_code_names_dict)
# run.save("/home/will/code/python/analysis_engine/tests/resources/test_master_1_2020.xlsx")

"""Generate milestone code names and put them in excel file"""
#TODO adapt this part for next time. Only need to generate codenames for new milestones

# milestone_list = milestones_list('Mars', test_master_data_list)
# codename_milestones_dict = gen_random_milestone_names(milestone_list)
# """save milestones in excel file as persistent record"""
# save_milestone_code_names = put_codenames_in_excel(codename_milestones_dict)
# save_milestone_code_names.save("/home/will/Documents/analysis_engine/code_resources/tests/sarh2_codenames_2.xlsx")

"""Put milestone code names into a dictionary"""
# ltc_codename_wb = load_workbook(root_path / 'code_resources/tests/ltc_codenames.xlsx')
# ltc_codename_dict = put_codenames_in_dict(ltc_codename_wb)
# a14_codename_wb = load_workbook(root_path / 'code_resources/tests/a14_codenames.xlsx')
# a14_codename_dict = put_codenames_in_dict(a14_codename_wb)
# ewr_codename_wb = load_workbook(root_path / 'code_resources/tests/ewr_codenames.xlsx')
# ewr_codename_dict = put_codenames_in_dict(ewr_codename_wb)
# iep_codename_wb = load_workbook(root_path / 'code_resources/tests/iep_codenames.xlsx')
# iep_codename_dict = put_codenames_in_dict(iep_codename_wb)
# swrc_codename_wb = load_workbook(root_path / 'code_resources/tests/swrc_codenames.xlsx')
# swrc_codename_dict = put_codenames_in_dict(swrc_codename_wb)
sarh2_codename_wb = load_workbook(root_path / "code_resources/tests/sarh2_codenames_2.xlsx")
sarh2_codename_dict = put_codenames_in_dict(sarh2_codename_wb)

"""Put code names into masters"""
alter_milestones = altering_test_masters_milestones(test_wb_five, 'Mars', sarh2_codename_dict)
alter_milestones.save("/home/will/code/python/analysis_engine/"
                      "tests/resources"
                      "/test_master_1_2020.xlsx")




