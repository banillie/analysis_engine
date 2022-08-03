from openpyxl import Workbook
from openpyxl.formatting.rule import IconSetRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule

from analysis_engine.dictionaries import (
    DATA_KEY_DICT,
    BC_STAGE_DICT_FULL_TO_ABB,
    CONVERT_RAG,
    rag_txt_list,
    conf_list,
    risk_list,
    DASHBOARD_KEYS,
    DCA_KEYS,
    STANDARDISE_COST_KEYS,
    DANDELION_KEYS,
    FWD_LOOK_DICT,
    SCHEDULE_DASHBOARD_KEYS,
    DASHBOARD_RESOURCE_KEYS,
)
from analysis_engine.dandelion import dandelion_number_text
from analysis_engine.colouring import black_text, fill_colour_list
from analysis_engine.milestones import MilestoneData, get_milestone_date
from analysis_engine.error_msgs import InputError, logger
from analysis_engine.costs import CostData, convert_none_types
from analysis_engine.render_utils import plus_minus_days
from analysis_engine.settings import get_remove_income, get_board_date


def narrative_dashboard(master, wb: Workbook) -> None:
    ws = wb.active

    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=3).value
        if project_name in master["current_projects"]:
            # Group
            ws.cell(row=row_num, column=2).value = master["project_information"][
                project_name
            ]["Directorate"]
            # Abbreviation
            ws.cell(row=row_num, column=4).value = master["project_information"][
                project_name
            ]["Abbreviations"]
            # Stage
            bc_stage = master["master_data"][0]["data"][project_name][
                DATA_KEY_DICT["IPDC approval point"]
            ]
            ws.cell(row=row_num, column=5).value = BC_STAGE_DICT_FULL_TO_ABB[bc_stage]
            costs = master["master_data"][0]["data"][project_name][
                DATA_KEY_DICT["Total Forecast"]
            ]
            ws.cell(row=row_num, column=6).value = dandelion_number_text(costs)

            overall_dca = CONVERT_RAG[
                master["master_data"][0]["data"][project_name][
                    DATA_KEY_DICT["Departmental DCA"]
                ]
            ]
            ws.cell(row=row_num, column=7).value = overall_dca
            if overall_dca == "None":
                ws.cell(row=row_num, column=7).value = ""

            sro_n = master["master_data"][0]["data"][project_name]["SRO Narrative"]
            ws.cell(row=row_num, column=8).value = sro_n

        """list of columns with conditional formatting"""
        list_columns = ["g"]

        """same loop but the text is black. In addition these two loops go 
        through the list_columns list above"""
        for column in list_columns:
            for i, dca in enumerate(rag_txt_list):
                text = black_text
                fill = fill_colour_list[i]
                dxf = DifferentialStyle(font=text, fill=fill)
                rule = Rule(
                    type="containsText", operator="containsText", text=dca, dxf=dxf
                )
                for_rule_formula = 'NOT(ISERROR(SEARCH("' + dca + '",' + column + "5)))"
                rule.formula = [for_rule_formula]
                ws.conditional_formatting.add(column + "5:" + column + "60", rule)

        for row_num in range(2, ws.max_row + 1):
            for col_num in range(5, ws.max_column + 1):
                if ws.cell(row=row_num, column=col_num).value == 0:
                    ws.cell(row=row_num, column=col_num).value = "-"

    # return wb


def cdg_dashboard(master, wb: Workbook) -> None:
    ws = wb.active

    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=3).value
        if project_name in master["current_projects"]:
            ws.cell(row=row_num, column=2).value = master["project_information"][
                project_name
            ]["Directorate"]
            ws.cell(row=row_num, column=4).value = master["project_information"][
                project_name
            ]["Abbreviations"]
            bc_stage = master["master_data"][0]["data"][project_name][
                DATA_KEY_DICT["IPDC approval point"]
            ]
            ws.cell(row=row_num, column=5).value = BC_STAGE_DICT_FULL_TO_ABB[bc_stage]
            costs = master["master_data"][0]["data"][project_name][
                DATA_KEY_DICT["Total Forecast"]
            ]
            ws.cell(row=row_num, column=6).value = dandelion_number_text(
                costs, none_handle="none"
            )
            income = master["master_data"][0]["data"][project_name]["Total Income"]
            ws.cell(row=row_num, column=7).value = dandelion_number_text(
                income, none_handle="none"
            )
            benefits = master["master_data"][0]["data"][project_name]["Total Benefits"]
            ws.cell(row=row_num, column=8).value = dandelion_number_text(
                benefits, none_handle="none"
            )
            vfm = master["master_data"][0]["data"][project_name]["VfM Category"]
            ws.cell(row=row_num, column=9).value = vfm
            overall_dca = CONVERT_RAG[
                master["master_data"][0]["data"][project_name][
                    DATA_KEY_DICT["Departmental DCA"]
                ]
            ]
            ws.cell(row=row_num, column=10).value = overall_dca
            if overall_dca == "None":
                ws.cell(row=row_num, column=10).value = ""

            for i, key in enumerate(conf_list):
                dca = CONVERT_RAG[master["master_data"][0]["data"][project_name][key]]
                ws.cell(row=row_num, column=11 + i).value = dca

            for i, key in enumerate(risk_list):
                risk = master["master_data"][0]["data"][project_name][key]
                if risk == "YES":
                    ws.cell(row=row_num, column=14 + i).value = risk

        """list of columns with conditional formatting"""
        list_columns = ["j", "k", "l", "m"]

        """same loop but the text is black. In addition these two loops 
        go through the list_columns list above"""
        for column in list_columns:
            for i, dca in enumerate(rag_txt_list):
                text = black_text
                fill = fill_colour_list[i]
                dxf = DifferentialStyle(font=text, fill=fill)
                rule = Rule(
                    type="containsText", operator="containsText", text=dca, dxf=dxf
                )
                for_rule_formula = 'NOT(ISERROR(SEARCH("' + dca + '",' + column + "5)))"
                rule.formula = [for_rule_formula]
                ws.conditional_formatting.add(column + "5:" + column + "60", rule)

        for row_num in range(2, ws.max_row + 1):
            for col_num in range(5, ws.max_column + 1):
                if ws.cell(row=row_num, column=col_num).value == 0:
                    ws.cell(row=row_num, column=col_num).value = "-"



def ipdc_dashboard(md, wb: Workbook, **op_args) -> Workbook:
    IPDC_DATE = get_board_date(op_args)
    logger.info(
        f'The {op_args["report"].upper()} Board date has been taken from the GLOBALS date in the config file'
    )

    financial_dashboard(md, wb, **op_args)
    resource_dashboard(md, wb, **op_args)
    schedule_dashboard(md, wb, IPDC_DATE, **op_args)
    benefits_dashboard(md, wb, **op_args)
    overall_dashboard(md, wb, IPDC_DATE, **op_args)

    return wb


def resource_dashboard(md, wb: Workbook, **op_args) -> Workbook:

    ws = wb["Resource"]
    cmd = md["master_data"][0]["data"]  # cmd = current master data

    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=3).value
        if project_name not in md["current_projects"]:
            continue

        """BC Stage"""
        bc_stage = cmd[project_name][DASHBOARD_KEYS["BC_STAGE"]]
        ws.cell(row=row_num, column=4).value = BC_STAGE_DICT_FULL_TO_ABB[bc_stage]

        "Resourcing data"
        for i, key in enumerate(DASHBOARD_RESOURCE_KEYS):
            try:
                if key == "DfTc Resource Gap Criticality (RAG rating)":
                    ws.cell(row=row_num, column=5 + i).value = CONVERT_RAG[cmd[project_name][key]]
                else:
                    ws.cell(row=row_num, column=5 + i).value = cmd[project_name][key]
            except KeyError:
                raise InputError(
                    key + " key is not in quarter master. This key must"
                    " be present for dashboard compilation. Stopping. "
                    "Make sure all resource keys are in Master."
                )

        """DCA ratings"""
        for i, q in enumerate(md["quarter_list"]):
            try:
                ws.cell(row=row_num, column=12 + i).value = CONVERT_RAG[
                    md["master_data"][i]["data"][project_name][
                        DCA_KEYS["resource"]["resource"]
                    ]
                ]
            except KeyError:
                ws.cell(row=row_num, column=12 + i).value = ""

    """list of columns with conditional formatting"""
    list_columns = ["j", "l", "m", "n", "o"]

    """same loop but the text is black. In addition these two loops go through the list_columns list above"""
    for column in list_columns:
        for i, dca in enumerate(rag_txt_list):
            text = black_text
            fill = fill_colour_list[i]
            dxf = DifferentialStyle(font=text, fill=fill)
            rule = Rule(type="containsText", operator="containsText", text=dca, dxf=dxf)
            for_rule_formula = 'NOT(ISERROR(SEARCH("' + dca + '",' + column + "5)))"
            rule.formula = [for_rule_formula]
            ws.conditional_formatting.add("" + column + "5:" + column + "60", rule)

    # for row_num in range(2, ws.max_row + 1):
    #     for col_num in range(5, ws.max_column+1):
    #         if ws.cell(row=row_num, column=col_num).value == 0:
    #             ws.cell(row=row_num, column=col_num).value = '-'

    return wb


def financial_dashboard(
    md,
    wb: Workbook,
    **op_args,
) -> Workbook:

    ws = wb["Finance"]
    cmd = md["master_data"][0]["data"]  # cmd = current master data
    lmd = md["master_data"][1]["data"]  # lmd = last master data
    lymd = md["master_data"][3]["data"]  # lymd = last year master data

    rm = get_remove_income(op_args)

    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=3).value
        if project_name not in md["current_projects"]:
            continue

        """BC Stage"""
        bc_stage = cmd[project_name][DASHBOARD_KEYS["BC_STAGE"]]
        ws.cell(row=row_num, column=4).value = BC_STAGE_DICT_FULL_TO_ABB[bc_stage]
        """Total WLC"""
        wlc_now = convert_none_types(
            cmd[project_name][STANDARDISE_COST_KEYS[op_args["report"]]["total"]]
        )
        if project_name in rm:
            wlc_now = wlc_now - convert_none_types(
                cmd[project_name][
                    STANDARDISE_COST_KEYS[op_args["report"]]["income_total"]
                ]
            )
        ws.cell(row=row_num, column=6).value = wlc_now
        """WLC variance against lst quarter"""
        try:
            wlc_lst_qrt = convert_none_types(
                lmd[project_name][STANDARDISE_COST_KEYS[op_args["report"]]["total"]]
            )
            if project_name in rm:
                wlc_lst_qrt = wlc_lst_qrt - convert_none_types(
                    lmd[project_name][
                        STANDARDISE_COST_KEYS[op_args["report"]]["income_total"]
                    ]
                )
            diff = wlc_now - wlc_lst_qrt
            if float(diff) > 0.49 or float(diff) < -0.49:
                ws.cell(row=row_num, column=7).value = diff
            else:
                ws.cell(row=row_num, column=7).value = "-"
        except KeyError:
            ws.cell(row=row_num, column=7).value = "-"

        """WLC variance against last year"""
        try:
            wlc_lst_year = convert_none_types(
                lymd[project_name][STANDARDISE_COST_KEYS[op_args["report"]]["total"]]
            )
            if project_name in rm:
                wlc_lst_year = wlc_lst_year - convert_none_types(
                    lymd[project_name][
                        STANDARDISE_COST_KEYS[op_args["report"]]["income_total"]
                    ]
                )
            diff = wlc_now - wlc_lst_year
            if float(diff) > 0.49 or float(diff) < -0.49:
                ws.cell(row=row_num, column=8).value = diff
            else:
                ws.cell(row=row_num, column=8).value = "-"
        except KeyError:
            ws.cell(row=row_num, column=8).value = "-"

        """financial DCA ratings"""
        for i, q in enumerate(md["quarter_list"]):
            try:
                ws.cell(row=row_num, column=15 + i).value = CONVERT_RAG[
                    md["master_data"][i]["data"][project_name][
                        DCA_KEYS[op_args["report"]]["finance"]
                    ]
                ]
            except KeyError:
                ws.cell(row=row_num, column=15 + i).value = ""

    """list of columns with conditional formatting"""
    list_columns = ["o", "p", "q", "r", "s"]

    """same loop but the text is black. In addition these two loops go through the list_columns list above"""
    for column in list_columns:
        for i, dca in enumerate(rag_txt_list):
            text = black_text
            fill = fill_colour_list[i]
            dxf = DifferentialStyle(font=text, fill=fill)
            rule = Rule(type="containsText", operator="containsText", text=dca, dxf=dxf)
            for_rule_formula = 'NOT(ISERROR(SEARCH("' + dca + '",' + column + "5)))"
            rule.formula = [for_rule_formula]
            ws.conditional_formatting.add("" + column + "5:" + column + "60", rule)

    return wb


def schedule_dashboard(
    md,
    wb: Workbook,
    IPDC_DATE,
    **op_args,
) -> Workbook:
    ws = wb["Schedule"]
    # overall_ws = wb.worksheets[3]

    cmd = md["master_data"][0]["data"]  # cmd = current master data
    # lmd = md["master_data"][1]["data"]  # lmd = last master data

    ms = MilestoneData(md, **op_args)

    for row_num in range(2, ws.max_row + 1):
        p = ws.cell(row=row_num, column=3).value
        if p not in md["current_projects"]:
            continue
        abb = md["project_information"][p]["Abbreviations"]
        """IPDC approval point"""
        bc_stage = cmd[p][DASHBOARD_KEYS["BC_STAGE"]]
        ws.cell(row=row_num, column=4).value = BC_STAGE_DICT_FULL_TO_ABB[bc_stage]

        add_column = 0
        for m in SCHEDULE_DASHBOARD_KEYS:
            current = get_milestone_date(
                ms.milestone_dict, m, md["quarter_list"][0], abb
            )
            last_quarter = get_milestone_date(
                ms.milestone_dict, m, md["quarter_list"][0], abb
            )
            ws.cell(row=row_num, column=10 + add_column).value = current
            if current is not None and current <= IPDC_DATE:
                ws.cell(row=row_num, column=10 + add_column).value = "Completed"
            if current is None:
                ws.cell(row=row_num, column=10 + add_column).value = "-"
            try:
                last_change = (current - last_quarter).days
                ws.cell(row=row_num, column=11 + add_column).value = plus_minus_days(
                    last_change
                )
            except TypeError:
                pass
            add_column += 3

        """schedule DCA rating - this quarter"""
        for i, q in enumerate(md["quarter_list"]):
            try:
                ws.cell(row=row_num, column=22 + i).value = CONVERT_RAG[
                    md["master_data"][i]["data"][p][
                        DCA_KEYS[op_args["report"]]["schedule"]
                    ]
                ]
            except KeyError:
                ws.cell(row=row_num, column=22 + i).value = ""

    op_args["type"] = ["Approval"]
    ms.filter_chart_info(**op_args)
    ms.get_next_milestone(IPDC_DATE)
    for row_num in range(2, ws.max_row + 1):
        p = ws.cell(row=row_num, column=3).value
        if p not in md["current_projects"]:
            continue
        """Next milestone name and variance"""
        try:
            ws.cell(row=row_num, column=6).value = ms.next_milestone_dict[p][
                "milestone"
            ]
            ws.cell(row=row_num, column=7).value = ms.next_milestone_dict[p]["date"]
        except KeyError:
            ws.cell(row=row_num, column=6).value = "None Reported"
            ws.cell(row=row_num, column=7).value = None

    """list of columns with conditional formatting"""
    list_columns = ["v", "w", "x", "y", "z"]

    """same loop but the text is black. In addition these two loops go through the list_columns list above"""
    for column in list_columns:
        for i, dca in enumerate(rag_txt_list):
            text = black_text
            fill = fill_colour_list[i]
            dxf = DifferentialStyle(font=text, fill=fill)
            rule = Rule(type="containsText", operator="containsText", text=dca, dxf=dxf)
            for_rule_formula = 'NOT(ISERROR(SEARCH("' + dca + '",' + column + "5)))"
            rule.formula = [for_rule_formula]
            ws.conditional_formatting.add("" + column + "5:" + column + "60", rule)

    # for row_num in range(2, ws.max_row + 1):
    #     for col_num in range(5, ws.max_column + 1):
    #         if ws.cell(row=row_num, column=col_num).value == 0:
    #             ws.cell(row=row_num, column=col_num).value = "-"

    return wb


def benefits_dashboard(
    md,
    wb: Workbook,
    **op_args,
) -> Workbook:
    ws = wb["Benefits_VfM"]
    cmd = md["master_data"][0]["data"]  # cmd = current master data
    # lmd = md["master_data"][1]["data"]  # lmd = last master data

    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=3).value
        if project_name not in md["current_projects"]:
            continue

        """BICC approval point"""
        bc_stage = cmd[project_name][DASHBOARD_KEYS["BC_STAGE"]]
        ws.cell(row=row_num, column=4).value = BC_STAGE_DICT_FULL_TO_ABB[bc_stage]
        """initial bcr"""
        initial_bcr = cmd[project_name]["Initial Benefits Cost Ratio (BCR)"]
        ws.cell(row=row_num, column=6).value = initial_bcr
        """adjusted bcr"""
        adjusted_bcr = cmd[project_name]["Adjusted Benefits Cost Ratio (BCR)"]
        ws.cell(row=row_num, column=8).value = adjusted_bcr
        """vfm category now"""
        if cmd[project_name]["VfM Category single entry"] is None:
            vfm_cat = (
                str(cmd[project_name]["VfM Category lower range"])
                + " - "
                + str(cmd[project_name]["VfM Category upper range"])
            )
            if vfm_cat == "None - None":
                vfm_cat = "None"
            ws.cell(row=row_num, column=10).value = vfm_cat
        else:
            vfm_cat = cmd[project_name]["VfM Category single entry"]
            ws.cell(row=row_num, column=10).value = vfm_cat

        """DCA ratings"""
        for i, q in enumerate(md["quarter_list"]):
            try:
                ws.cell(row=row_num, column=16 + i).value = CONVERT_RAG[
                    md["master_data"][i]["data"][project_name][
                        DCA_KEYS[op_args["report"]]["benefits"]
                    ]
                ]
            except KeyError:
                ws.cell(row=row_num, column=16 + i).value = ""

    """list of columns with conditional formatting"""
    list_columns = ["p", "q", "r", "s", "t"]

    """loops below place conditional formatting (cf) rules into the wb. There are two as the dashboard currently has
    two distinct sections/headings, which do not require cf. Therefore, cf starts and ends at the stated rows. this
    is hard code that will need to be changed should the position of information in the dashboard change. It is an
    easy change however"""

    """same loop but the text is black. In addition these two loops go through the list_columns list above"""
    for column in list_columns:
        for i, dca in enumerate(rag_txt_list):
            text = black_text
            fill = fill_colour_list[i]
            dxf = DifferentialStyle(font=text, fill=fill)
            rule = Rule(type="containsText", operator="containsText", text=dca, dxf=dxf)
            for_rule_formula = 'NOT(ISERROR(SEARCH("' + dca + '",' + column + "5)))"
            rule.formula = [for_rule_formula]
            ws.conditional_formatting.add("" + column + "5:" + column + "60", rule)

    return wb


def overall_dashboard(
    md,
    wb: Workbook,
    IPDC_DATE,
    **op_args,
) -> Workbook:
    ws = wb["Overall"]

    cmd = md["master_data"][0]["data"]  # cmd = current master data
    lmd = md["master_data"][1]["data"]  # lmd = last master data
    lymd = md["master_data"][3]["data"]  # lymd = last year master data
    rm = get_remove_income(op_args)
    ms = MilestoneData(md, **op_args)

    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=3).value
        if project_name not in md["current_projects"]:
            continue

        """BC Stage"""
        bc_stage = cmd[project_name][DASHBOARD_KEYS["BC_STAGE"]]
        ws.cell(row=row_num, column=4).value = BC_STAGE_DICT_FULL_TO_ABB[bc_stage]

        """Total WLC"""
        wlc_now = convert_none_types(
            cmd[project_name][STANDARDISE_COST_KEYS[op_args["report"]]["total"]]
        )
        if project_name in rm:
            wlc_now = wlc_now - convert_none_types(
                cmd[project_name][
                    STANDARDISE_COST_KEYS[op_args["report"]]["income_total"]
                ]
            )
        ws.cell(row=row_num, column=6).value = wlc_now
        """WLC variance against lst quarter"""
        try:
            wlc_lst_qrt = convert_none_types(
                lmd[project_name][STANDARDISE_COST_KEYS[op_args["report"]]["total"]]
            )
            if project_name in rm:
                wlc_lst_qrt = wlc_lst_qrt - convert_none_types(
                    lmd[project_name][
                        STANDARDISE_COST_KEYS[op_args["report"]]["income_total"]
                    ]
                )
            diff = wlc_now - wlc_lst_qrt
            if float(diff) > 0.49 or float(diff) < -0.49:
                ws.cell(row=row_num, column=7).value = diff
            else:
                ws.cell(row=row_num, column=7).value = "-"
        except KeyError:
            ws.cell(row=row_num, column=7).value = "-"

        """WLC variance against last year"""
        try:
            wlc_lst_year = convert_none_types(
                lymd[project_name][STANDARDISE_COST_KEYS[op_args["report"]]["total"]]
            )
            if project_name in rm:
                wlc_lst_year = wlc_lst_year - convert_none_types(
                    lymd[project_name][
                        STANDARDISE_COST_KEYS[op_args["report"]]["income_total"]
                    ]
                )
            diff = wlc_now - wlc_lst_year
            if float(diff) > 0.49 or float(diff) < -0.49:
                ws.cell(row=row_num, column=8).value = diff
            else:
                ws.cell(row=row_num, column=8).value = "-"
        except KeyError:
            ws.cell(row=row_num, column=8).value = "-"

        """vfm category now"""
        if cmd[project_name]["VfM Category single entry"] is None:
            vfm_cat = (
                str(cmd[project_name]["VfM Category lower range"])
                + " - "
                + str(cmd[project_name]["VfM Category upper range"])
            )
            if vfm_cat == "None - None":
                vfm_cat = "None"
            ws.cell(row=row_num, column=9).value = vfm_cat

        else:
            vfm_cat = cmd[project_name]["VfM Category single entry"]
            ws.cell(row=row_num, column=9).value = vfm_cat

        abb = md["project_information"][project_name]["Abbreviations"]
        current = get_milestone_date(
            ms.milestone_dict, "Full Operations", md["quarter_list"][0], abb
        )
        # last_quarter = get_milestone_date(ms.milestone_dict, "Full Operations", md['quarter_list'][0], abb)
        ws.cell(row=row_num, column=10).value = current
        if current is not None and current < IPDC_DATE:
            ws.cell(row=row_num, column=10).value = "Completed"
        if current is None:
            ws.cell(row=row_num, column=10).value = "-"
        # try:
        #     last_change = (current - last_quarter).days
        #     ws.cell(
        #         row=row_num, column=11
        #     ).value = plus_minus_days(last_change)
        # except TypeError:
        #     pass

        """IPA DCA rating"""
        try:
            ipa_dca = CONVERT_RAG[cmd[project_name][DCA_KEYS["ipa"]["ipa"]]]
        except KeyError:
            raise InputError(
                f"No {DCA_KEYS['ipa']['ipa']} key in quarter master. This key must"
                " be present for dashboard compilation. Stopping."
            )
        ws.cell(row=row_num, column=16).value = ipa_dca
        if ipa_dca == "None":
            ws.cell(row=row_num, column=16).value = ""

        # SRO forward look
        try:
            ws.cell(row=row_num, column=19).value = FWD_LOOK_DICT[
                cmd[project_name][DANDELION_KEYS["forward_look"]]
            ]
        except KeyError:
            raise InputError(
                "No SRO Forward Look Assessment key in current quarter master. This key must"
                " be present for dashboard compilation. Stopping."
            )

        """DCA rating - this quarter"""
        for i, q in enumerate(md["quarter_list"]):
            try:
                ws.cell(row=row_num, column=20 + i).value = CONVERT_RAG[
                    md["master_data"][i]["data"][project_name][
                        DCA_KEYS[op_args["report"]]["sro"]
                    ]
                ]
            except KeyError:
                ws.cell(row=row_num, column=20 + i).value = ""

    # places arrow icons for sro forward look assessment
    icon_set_rule = IconSetRule("3Arrows", "num", ["1", "2", "3"], showValue=False)
    ws.conditional_formatting.add("S4:S40", icon_set_rule)

    """list of columns with conditional formatting"""
    list_columns = ["p", "t", "u", "v", "w"]

    """same loop but the text is black. In addition these two loops go 
    through the list_columns list above"""
    for column in list_columns:
        for i, dca in enumerate(rag_txt_list):
            text = black_text
            fill = fill_colour_list[i]
            dxf = DifferentialStyle(font=text, fill=fill)
            rule = Rule(type="containsText", operator="containsText", text=dca, dxf=dxf)
            for_rule_formula = 'NOT(ISERROR(SEARCH("' + dca + '",' + column + "5)))"
            rule.formula = [for_rule_formula]
            ws.conditional_formatting.add(column + "5:" + column + "60", rule)

    for row_num in range(2, ws.max_row + 1):
        for col_num in range(5, ws.max_column + 1):
            if ws.cell(row=row_num, column=col_num).value == 0:
                ws.cell(row=row_num, column=col_num).value = "-"

    return wb
