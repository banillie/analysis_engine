import math
import datetime
from typing import List

from openpyxl import Workbook
from openpyxl.workbook import workbook

from matplotlib import pyplot as plt

from analysis_engine.costs import CostData
from analysis_engine.benefits import BenefitsData
from analysis_engine.resourcing import ResourceData
from analysis_engine.milestones import MilestoneData, get_milestone_date
from analysis_engine.segmentation import (
    get_iter_list,
    get_group,
    get_quarter_index,
    get_correct_p_data,
)
from analysis_engine.error_msgs import InputError, logger
from analysis_engine.colouring import COLOUR_DICT, FACE_COLOUR
from analysis_engine.render_utils import make_file_friendly, handle_long_keys

from analysis_engine.dictionaries import (
    RAG_RANKING_DICT_NUMBER,
    RAG_RANKING_DICT_COLOUR,
    DCA_KEYS,
    FONT_TYPE,
    DANDELION_KEYS,
    BC_STAGE_DICT_FULL_TO_ABB,
    BC_STAGE_DICT_ABB_TO_FULL,
    DASHBOARD_KEYS,
    NEXT_STAGE_DICT,
)


def dandelion_project_text(number: int, project: str) -> str:
    total_len = len(str(int(number)))
    try:
        if total_len <= 3:
            round_total = int(round(number, -1))
            return "£" + str(round_total) + "m"
        if total_len == 4:
            round_total = int(round(number, -2))
            return "£" + str(round_total)[0] + "," + str(round_total)[1] + "bn"
        if total_len == 5:
            round_total = int(round(number, -2))
            return "£" + str(round_total)[:2] + "," + str(round_total)[2] + "bn"
        if total_len > 6:
            print(
                "Check total forecast and cost data reported by "
                + project
                + " total is £"
                + str(number)
                + "m"
            )
    except ValueError:
        print(
            "Check total forecast and cost data reported by "
            + project
            + " it is not reporting a number"
        )


def dandelion_number_text(number: int, **kwargs) -> str:
    total_len = len(str(int(number)))
    if "type" in kwargs:
        if kwargs["type"] in [
            "ps_resource",
            "contractor_resource",
            "total_resource",
            "funded_resource",
        ]:
            return str(round(number, 1))
    try:
        if number == 0:
            return "£0"
        if total_len <= 2:
            # round_total = round(number, -1)
            return "£" + str(int(round(number))) + "m"
        if total_len == 3:
            round_total = round(number, -1)
            return "£" + str(int(round_total)) + "m"
        if total_len == 4:
            round_total = round(number, -2)
            if str(round_total)[1] != "0":
                return "£" + str(round_total)[0] + "," + str(round_total)[1] + "bn"
            else:
                return "£" + str(round_total)[0] + "bn"
        if total_len == 5:
            round_total = int(round(number, -2))
            if str(round_total)[2] != "0":
                return "£" + str(round_total)[:2] + "," + str(round_total)[2] + "bn"
            else:
                return "£" + str(round_total)[:2] + "bn"
        if total_len == 6:
            round_total = int(round(number, -3))
            if str(round_total)[3] != "0":
                return "£" + str(round_total)[:3] + "," + str(round_total)[3] + "bn"
            else:
                return "£" + str(round_total)[:3] + "bn"

    except ValueError:
        print("not number")


#  there will be a better algorythm than this.
def cal_group_angle(dist_amount: int, group: List[str], inner_circles=False):
    """helper function for dandelion data class.
    Calculates distribution of first circle around center."""
    n_points = len(group)
    g_ang = dist_amount / n_points  # group_ang and distribution number
    output_list = []
    for i in range(n_points):
        if inner_circles:
            output_list.append(int((g_ang * i) + 270))
        else:
            output_list.append(int(g_ang * i))

    return output_list


#  switches the type of data displayed in dandelion graph output
#  change type to dandelion_type
def get_dandelion_type_total(master, kwargs) -> int or str:
    tp = kwargs["quarter"][0]  # only one quarter for dandelion
    if "type" in kwargs:
        if kwargs["type"] == "remaining_costs":
            cost = CostData(master, **kwargs)  # group costs data
            # cost.get_forecast_cost_profile()
            # return sum(cost.profiles[tp]["total"]) - sum(cost.profiles[tp]["std"])
            # return sum(cost.profiles[tp]["total"])
            return cost.totals[tp]["costs_remaining"]
        if kwargs["type"] == "spent":
            cost = CostData(master, **kwargs)
            # cost.get_forecast_cost_profile()
            # # return cost.c_totals[tp]['total'] - (sum(cost.profiles[tp]["total"]) - sum(cost.profiles[tp]["std"]))
            # return cost.totals[tp]["total"] - sum(cost.profiles[tp]["total"])
            return cost.totals[tp]["costs_spent"]
        if kwargs["type"] == "income":
            cost = CostData(master, **kwargs)
            # return cost.totals[tp]["income_total"]
            return cost.totals[tp]["income_total"]
        if kwargs["type"] == "benefits":
            benefits = BenefitsData(master, **kwargs)
            return benefits.b_totals[tp]["total"]
        if kwargs["type"] in (
            "ps_resource",
            "contractor_resource",
            "total_resource",
            "funded_resource",
        ):
            resource = ResourceData(master, **kwargs)
            return resource.totals[tp][kwargs["type"]]

    else:
        cost = CostData(master, **kwargs)  # group costs data
        return cost.totals[tp]["total"]


def calculate_circle_edge(sro_rag, fwd_look):
    if fwd_look == "No Change Expected":
        return sro_rag
    if fwd_look == "Improving":
        now = RAG_RANKING_DICT_COLOUR[sro_rag]
        return RAG_RANKING_DICT_NUMBER[now + 1]
    if fwd_look == "Worsening":
        now = RAG_RANKING_DICT_COLOUR[sro_rag]
        return RAG_RANKING_DICT_NUMBER[now - 1]


class DandelionData:
    """
    Data class for dandelion info graphic. Output dictionary to d_data()
    """

    def __init__(self, master, **kwargs):
        self.master = master
        self.kwargs = kwargs
        self.group = []
        self.group_stage_switch = ""
        self.handle_group()
        self.quarter = self.kwargs["quarter"][0]
        self.d_data = {}
        self.get_data()

    def handle_group(self):
        if "group" in self.kwargs:
            group_stage_switch = "group"
        if "stage" in self.kwargs:
            group_stage_switch = "stage"

        self.group_stage_switch = group_stage_switch
        self.group = self.kwargs[group_stage_switch]

    def get_data(self):
        g_d = {}  # group dictionary. first outer circle.
        l_g_d = {}  # lower group dictionary
        pf_wlc = get_dandelion_type_total(self.master, self.kwargs)  # portfolio wlc
        if "pc" in self.kwargs:  # pc portfolio colour
            pf_colour = COLOUR_DICT[self.kwargs["pc"]]
            pf_colour_edge = COLOUR_DICT[self.kwargs["pc"]]
        else:
            pf_colour = COLOUR_DICT["WHITE"]
            pf_colour_edge = COLOUR_DICT["GREY"]
        pf_text = "Portfolio\n" + dandelion_number_text(pf_wlc, **self.kwargs)

        ## center circle
        g_d["portfolio"] = {
            "axis": (0, 0),
            "tp": (0, 0),
            "r": math.sqrt(pf_wlc),
            "colour": pf_colour,
            "text": pf_text,
            "fill": "solid",
            "ec": pf_colour_edge,
            "alignment": ("center", "center"),
        }

        ## first outer circle
        g_radius_list = []
        for i, g in enumerate(self.group):
            self.kwargs[self.group_stage_switch] = [g]
            if g == "pipeline":
                g_wlc = self.master["pipeline_dict"]["pipeline"]["wlc"]
            else:
                g_wlc = get_dandelion_type_total(self.master, self.kwargs)

            if "same_size" in self.kwargs:
                if self.kwargs["same_size"] == "Yes":
                    g_wlc = 46000

            g_abb = g
            if self.group_stage_switch == "stage":
                try:
                    g_abb = BC_STAGE_DICT_FULL_TO_ABB[g]
                except KeyError:
                    if g in BC_STAGE_DICT_ABB_TO_FULL.keys():
                        g_abb = g

            g_text = (
                g_abb + "\n" + dandelion_number_text(g_wlc, **self.kwargs)
            )  # group text

            if g_wlc < pf_wlc / 10:  # adjusts any very small figures
                g_wlc_adjusted = pf_wlc / 10
            else:
                g_wlc_adjusted = g_wlc

            group_radius = math.sqrt(g_wlc_adjusted)

            if "values" in self.kwargs:
                if self.kwargs["values"] == "No":
                    g_text = g

            g_d[g] = {
                "r": group_radius,
                "wlc": g_wlc_adjusted,
                "text": g_text,
                "colour": "#FFFFFF",
                "fill": "dashed",
                "ec": COLOUR_DICT["GREY"],
            }

            g_radius_list.append(group_radius)

        g_ang_l = cal_group_angle(180, g_radius_list, inner_circles=True)
        if "angles" in self.kwargs:
            if len(self.kwargs["angles"]) == len(self.group):
                g_ang_l = self.kwargs["angles"]
            else:
                raise InputError(
                    "The number of groups and angles don't match. Stopping."
                )

        # multiplied here so a gap to central circle
        g_radius_dist = g_d["portfolio"]["r"] * 2
        if len(self.group) > 3:
            g_radius_dist = g_d["portfolio"]["r"] * 2.5

        for i, g in enumerate(self.group):
            if len(self.group) > 1:  # this needs testing
                y_axis = 0 + (
                    (math.sqrt(pf_wlc) + g_radius_dist)
                    * math.sin(math.radians(g_ang_l[i]))
                )
                x_axis = 0 + (math.sqrt(pf_wlc) + g_radius_dist) * math.cos(
                    math.radians(g_ang_l[i])
                )

                g_d[g]["axis"] = (y_axis, x_axis)
                g_d[g]["tp"] = (y_axis, x_axis)
                g_d[g]["angle"] = g_ang_l[i]

            else:
                g_d = {}  # delete the dictionary
                # to alter out circles with low values in line with group wlc
                pf_wlc = g_wlc
                g_d[g] = {
                    "axis": (0, 0),
                    "r": math.sqrt(g_wlc),
                    "wlc": g_wlc,
                    "colour": "#FFFFFF",
                    "text": g_text,
                    "fill": "dashed",
                    "ec": "grey",
                    "alignment": ("center", "center"),
                }

        logger.info("The angles for groups are " + str(g_ang_l))

        ## second outer circle
        for i, g in enumerate(self.group):
            self.kwargs[self.group_stage_switch] = [g]
            if g == "pipeline":
                project_group = self.master["pipeline_list"]
            else:
                project_group = get_group(
                    self.master, self.quarter, **self.kwargs
                )  # lower group

            p_list = []
            for p in project_group:
                self.kwargs[self.group_stage_switch] = [p]
                if g == "pipeline":
                    p_value = self.master["pipeline_dict"][p]["wlc"]
                    p_data = {}
                else:
                    p_value = get_dandelion_type_total(
                        self.master, self.kwargs
                    )  # project wlc
                    p_data = get_correct_p_data(self.master, p, self.quarter)

                if "same_size" in self.kwargs:
                    if self.kwargs["same_size"] == "Yes":
                        p_value = 6000

                if (
                    p_value < pf_wlc * 0.02
                ):  # achieve some consistency for zero / low values
                    p_value_adjusted = pf_wlc * 0.008
                else:
                    p_value_adjusted = p_value

                p_radius = math.sqrt(p_value_adjusted)

                p_schedule = None
                if "order_by" in self.kwargs:  # SOMETHING TO DO WITH SCHEDULE?
                    if g == "pipeline":
                        raise InputError(
                            "The argument order_by cannot be used with pipeline projects as milestone data is required"
                        )
                    else:
                        bc = BC_STAGE_DICT_FULL_TO_ABB[
                            self.master["master_data"][0]["data"][p][
                                DASHBOARD_KEYS["BC_STAGE"]
                            ]
                        ]
                        ms = MilestoneData(self.master, **self.kwargs)
                        next_stage = NEXT_STAGE_DICT[bc]
                        try:
                            d = get_milestone_date(
                                ms,
                                next_stage,
                                self.quarter,
                                p,
                            )
                            p_schedule = (d - datetime.date.today()).days
                        except TypeError:
                            p_schedule = 10000000000

                if "abbreviations" in self.kwargs:
                    abb = self.master["project_information"][p]["Abbreviations"]
                    project_text = handle_long_keys(
                        f"{abb}, {dandelion_number_text(p_value, **self.kwargs)}"
                    )
                else:
                    project_text = handle_long_keys(
                        f"{p}, {dandelion_number_text(p_value, **self.kwargs)}"
                    )

                if "type" in self.kwargs:
                    if self.kwargs["type"] in [
                        "ps resource",
                        "contract resource",
                        "total resource",
                        "funded resource",
                    ]:
                        project_text = (
                            self.master.abbreviations[p]["abb"]
                            + ", "
                            + dandelion_number_text(p_value, **self.kwargs)
                        )
                if "values" in self.kwargs:
                    if self.kwargs["values"] == "No":
                        project_text = self.master.abbreviations[p]["abb"]

                try:  # this is for pipeline projects
                    if "confidence" in self.kwargs:  # change confidence type here
                        rag = p_data[
                            DCA_KEYS[self.kwargs["report"]][self.kwargs["confidence"]]
                        ]
                    else:
                        rag = p_data[DCA_KEYS[self.kwargs["report"]]["sro"]]
                    colour = COLOUR_DICT[rag]  # bubble colour
                except KeyError:  # p_data is None for pipeline projects
                    colour = COLOUR_DICT["WHITE"]

                if "circle_colour" in self.kwargs:
                    if self.kwargs["circle_colour"] == "No":
                        colour = FACE_COLOUR

                if colour == COLOUR_DICT["WHITE"] or colour == FACE_COLOUR:
                    if p in self.master["meta_groupings"][self.quarter]["GMPP"]:
                        edge_colour = COLOUR_DICT["BLACK"]
                    else:
                        edge_colour = COLOUR_DICT["GREY"]
                else:
                    if p in self.master["meta_groupings"][self.quarter]["GMPP"]:
                        edge_colour = COLOUR_DICT["BLACK"]
                    else:
                        edge_colour = colour

                if g != "pipeline":
                    if "circle_edge" in self.kwargs:
                        if self.kwargs["circle_edge"] == "forward_look":
                            try:
                                fwd_look = p_data[
                                    DANDELION_KEYS[self.kwargs["circle_edge"]]
                                ]
                                edge_rag = calculate_circle_edge(rag, fwd_look)
                                edge_colour = COLOUR_DICT[edge_rag]
                            except KeyError:
                                raise InputError(
                                    "No SRO Forward Look Assessment key in quarter master. "
                                    "This key must be present for this dandelion command. Stopping."
                                )
                        if self.kwargs["circle_edge"] == "ipa":
                            edge_colour = COLOUR_DICT[p_data["GMPP - IPA DCA"]]

                g_d[p] = {
                    "r": p_radius,
                    "wlc": p_value,
                    "colour": colour,
                    "text": project_text,
                    "fill": "solid",
                    "ec": edge_colour,
                }

                p_list.append((p_value, p_radius, p_schedule, p))

            if "order_by" in self.kwargs:
                if self.kwargs["order_by"] == "schedule":
                    p_list.sort(key=lambda x: x[1])
                    if g == "pipeline":
                        l_g_d[g] = list(reversed(p_list))
                    else:
                        l_g_d[g] = p_list
            else:
                l_g_d[g] = list(reversed(sorted(p_list)))

        for g in self.group:
            g_radius = g_d[g]["r"]
            g_y_axis = g_d[g]["axis"][0]  # group y axis
            g_x_axis = g_d[g]["axis"][1]  # group x axis

            try:
                p_value, p_radius_list, p_schedule_list, p_list = zip(*l_g_d[g])
            except ValueError:  # handles no projects in l_g_d list
                continue

            if len(p_list) > 3:
                ang_l = cal_group_angle(360, p_list)
            else:
                if len(p_list) == 1:
                    ang_l = [g_d[g]["angle"]]
                if len(p_list) == 2:
                    ang_l = [g_d[g]["angle"], g_d[g]["angle"] + 70]
                if len(p_list) == 3:
                    ang_l = [
                        g_d[g]["angle"],
                        g_d[g]["angle"] + 70,
                        g_d[g]["angle"] + 140,
                    ]

            largest_p_radius = (
                max(p_radius_list) * 1.5
            )  # value used for distance from inner circle.
            for i, p in enumerate(p_list):
                angle = ang_l[i]
                p_y_axis = g_y_axis + (g_radius + largest_p_radius) * math.sin(
                    math.radians(ang_l[i])
                )
                p_x_axis = g_x_axis + (g_radius + largest_p_radius) * math.cos(
                    math.radians(ang_l[i])
                )
                yx_text_position = (
                    p_y_axis
                    + (g_d[p]["r"] + g_d[p]["r"] * 1 / 2)
                    * math.sin(math.radians(angle)),
                    p_x_axis
                    + (g_d[p]["r"] + g_d[p]["r"] * 1 / 2)
                    * math.cos(math.radians(angle)),
                )

                # This is an important part of how text is rendered into the
                # dandelion. To stop text overlapping with the circle.
                # the way angles are calculated above means they can be greater than 360
                ha = "center"  # default text position
                va = "center"

                if 0 <= angle <= 11 or 371 <= angle <= 380:
                    va = "bottom"
                elif 11 <= angle <= 174 or 381 <= angle <= 534:
                    ha = "left"
                elif 165 <= angle <= 195:
                    va = "top"
                elif 186 <= angle <= 339:
                    ha = "right"
                elif 340 <= angle <= 349:
                    va = "top"
                elif angle > 534:
                    print(f"{p} angle is {angle}")

                g_d[p]["axis"] = (p_y_axis, p_x_axis)
                g_d[p]["ha"] = ha
                g_d[p]["va"] = va
                g_d[p]["tp"] = yx_text_position
                g_d[p]["angle"] = angle

        self.d_data = g_d


def dandelion_data_into_wb(d_data: DandelionData) -> workbook:
    """
    Simple function that returns data required for the dandelion graph.
    """
    wb = Workbook()
    for tp in d_data.d_data.keys():
        ws = wb.create_sheet(
            make_file_friendly(tp)
        )  # creating worksheets. names restricted to 30 characters.
        ws.title = make_file_friendly(tp)  # title of worksheet
        for i, project in enumerate(d_data.d_data[tp]["projects"]):
            ws.cell(row=2 + i, column=1).value = d_data.d_data[tp]["group"][i]
            ws.cell(row=2 + i, column=2).value = d_data.d_data[tp]["abb"][i]
            ws.cell(row=2 + i, column=3).value = project
            ws.cell(row=2 + i, column=4).value = int(d_data.d_data[tp]["cost"][i])
            ws.cell(row=2 + i, column=5).value = d_data.d_data[tp]["rag"][i]

        ws.cell(row=1, column=1).value = "Group"
        ws.cell(row=1, column=2).value = "Project"
        ws.cell(row=1, column=3).value = "Graph details"
        ws.cell(row=1, column=4).value = "WLC (forecast)"
        ws.cell(row=1, column=5).value = "DCA"

    wb.remove(wb["Sheet"])
    return wb


def make_a_dandelion_auto(dl: DandelionData, **kwargs):
    """function used to compile dandelion graph. Data is taken from
    DandelionData class."""

    fig, ax = plt.subplots(figsize=(10, 8), facecolor=FACE_COLOUR)

    if "circle_edge" in kwargs:
        if kwargs["circle_edge"] == "forward_look" or "ipa":
            line_width = 2.0
    else:
        line_width = 1.0

    ts = []
    x_list = []
    y_list = []
    obj = []

    p_font_size = 10
    if kwargs["report"] == "ipdc":
        p_font_size = 6

    for c in dl.d_data.keys():
        circle = plt.Circle(
            dl.d_data[c]["axis"],  # x, y position
            radius=dl.d_data[c]["r"],
            fc=dl.d_data[c]["colour"],  # face colour
            ec=dl.d_data[c]["ec"],  # edge colour
            linewidth=line_width,
            zorder=2,
        )
        obj.append(circle)
        ax.add_patch(circle)
        text = dl.d_data[c]["text"].strip()  # what does strip do?
        if c in dl.group or c == "portfolio":
            ax.annotate(
                text,
                xy=dl.d_data[c]["axis"],
                xycoords="data",
                fontsize=10,
                fontname=FONT_TYPE,
                ha="center",
                va="center",
                weight="bold",
                zorder=3,
            )
        else:
            x = dl.d_data[c]["tp"][0]
            y = dl.d_data[c]["tp"][1]
            ts.append(
                ax.text(
                    x,
                    y,
                    text,
                    fontsize=p_font_size,
                    ha=dl.d_data[c]["ha"],
                    va=dl.d_data[c]["va"],
                )
            )
            x_list.append(x)
            y_list.append(y)

    # place lines
    line_clr = "#ececec"
    for i, g in enumerate(dl.group):
        dl.kwargs[dl.group_stage_switch] = [g]
        ax.arrow(
            0,
            0,
            dl.d_data[g]["axis"][0],
            dl.d_data[g]["axis"][1],
            fc=line_clr,
            ec=line_clr,
            zorder=1,
        )

        if g == "pipeline":
            lower_g = dl.master["pipeline_list"]
        else:
            lower_g = get_group(dl.master, dl.kwargs["quarter"][0], **dl.kwargs)
        for p in lower_g:
            ax.arrow(
                dl.d_data[g]["axis"][0],
                dl.d_data[g]["axis"][1],
                dl.d_data[p]["axis"][0] - dl.d_data[g]["axis"][0],
                dl.d_data[p]["axis"][1] - dl.d_data[g]["axis"][1],
                fc=line_clr,
                ec=line_clr,
                zorder=1,
            )

    plt.axis("scaled")
    plt.axis("off")
    if kwargs["chart"] != "save":
        plt.show()

    return fig
