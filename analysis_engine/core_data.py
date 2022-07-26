import configparser
import json
import sys
import datetime
from typing import List, Dict, Union, Optional, Tuple, TextIO, Callable

# from datetime import timedelta, date
from collections import OrderedDict

from openpyxl import load_workbook

from datamaps.process import Cleanser

from analysis_engine.segmentation import get_iter_list
from analysis_engine.settings import report_config
from analysis_engine.error_msgs import (
    ProjectNameError,
    ProjectGroupError,
    ProjectStageError,
    logger,
    historic_project_names_error,
    latest_project_names_error,
    historic_group_names_error,
    latest_group_names_error,
    abbreviation_error,
    latest_stage_names_error,
    historic_stage_names_error,
    config_issue,
)

## figure out way to set reporting type here
# CONFIG_DICT = report_config()


def get_master_data(
    config_dict,
) -> List[Dict[str, Union[str, int, datetime.date, float]]]:
    """Returns a list of dictionaries each containing quarter data"""
    config_path = config_dict["root_path"] + config_dict["config"]
    config = configparser.ConfigParser()
    config.read(config_path)
    master_data_list = []
    for key in config["MASTERS"]:
        text = config["MASTERS"][key].split(", ")
        year = text[2]
        quarter = text[1]
        m_path = config_dict["root_path"] + "/core_data/" + text[0]
        m = config_dict["callable"](m_path, int(quarter), int(year))
        master_data_list.append(m)

    return list(reversed(master_data_list))


def convert_none_types(x):
    if x is None:
        return 0
    else:
        return x


META_GROUP_DICT = {
    "cdg": "Directorate",
    "ipdc": "Group",
    "top_250": "Group",
}

META_STAGE_DICT = {
    "cdg": "Last Business Case (BC) achieved",
    "ipdc": "IPDC approval point",
}


class PythonMasterData:
    """
    Key part of the data building process. Takes master_data and project_info dicts and
    performs a number of high-level checks to stop any later bugs or crashes, as well as
    create some useful meta data information.
    """

    def __init__(
        self,
        master_data: List[Dict[str, Union[str, int, datetime.date, float]]],
        project_information: Dict[str, Union[str, int]],
        meta,
        **kwargs,
    ) -> None:
        self.master_data = master_data
        self.project_information = project_information
        self.all_groups = meta["all_groups"]
        self.portfolio_groups = meta["port_group"]
        self.stages = meta["stages"]
        self.all_projects = list(project_information.keys())
        self.kwargs = kwargs
        self.current_quarter = str(master_data[0].quarter)
        self.current_projects = master_data[0].projects
        self.abbreviations = {}
        self.meta_groupings = {}
        self.pipeline_dict = {}
        self.pipeline_list = []
        self.quarter_list = []
        self.check_project_abbreviations()
        self.check_project_names()
        self.get_and_check_groupings()
        self.get_quarter_list()
        self.pipeline_projects_information()
        self.get_current_tp()

    # why is this collecting full names also? Seems silly.
    def check_project_abbreviations(self) -> None:
        """gets the abbreviations for all current projects. held in the project info document"""
        abb_dict = {}
        abbreviation_errors = []
        for p in self.all_projects:
            abb = self.project_information[p]["Abbreviations"]
            abb_dict[abb] = p
            if abb is None:
                abbreviation_errors.append(p)

        abbreviation_error(abbreviation_errors)

        self.abbreviations = abb_dict
        # self.full_names = fn_dict

    def check_project_names(self) -> None:
        """Checks that project names in all master are present/the same as in project info.
        Stops the programme if not"""
        critical_error_cases = []
        info_error_cases = {}
        for i, master in enumerate(self.master_data):
            for p in self.current_projects:
                if p not in self.all_projects:
                    if i == 0:  # latest master
                        critical_error_cases.append(p)
                    else:
                        # try:
                        info_error_cases[p] = master.quarter
                        # except find the error

        latest_project_names_error(critical_error_cases)
        historic_project_names_error(info_error_cases)

    def get_and_check_groupings(self) -> None:
        """gets the groups that projects are part of"""

        critical_group_errors = []
        info_group_errors = []
        critical_stage_errors = []
        info_stage_errors = {}
        meta_groupings_meta = {}
        for i, master in enumerate(self.master_data):
            quarter_dict = {}
            for group in self.all_groups:
                group_list = []
                for p in master.projects:
                    #  This data comes from the project information document. Not master.
                    projects_group = self.project_information[p][
                        META_GROUP_DICT[self.kwargs["data_type"]]
                    ]
                    if projects_group is None or projects_group not in self.all_groups:
                        if i == 0:
                            if p not in critical_group_errors:
                                critical_group_errors.append(p)
                        else:
                            info_group_errors.append(p)
                    if projects_group == group:
                        group_list.append(p)

                quarter_dict[group] = group_list

            for stage in self.stages:
                stage_list = []
                for p in master.projects:
                    project_stage = master.data[p][
                        META_STAGE_DICT[self.kwargs["data_type"]]
                    ]
                    if project_stage is None or project_stage not in self.stages:
                        if i == 0:
                            if p not in critical_stage_errors:
                                critical_stage_errors.append(p)
                        else:
                            info_stage_errors[p] = master.quarter
                    if project_stage == stage:
                        stage_list.append(p)

                quarter_dict[stage] = stage_list

            gmpp_list = []
            for p in master.projects:
                if self.project_information[p]["GMPP"] == "YES":
                    gmpp_list.append(p)
                quarter_dict["GMPP"] = gmpp_list

            meta_groupings_meta[str(master.quarter)] = quarter_dict

        latest_group_names_error(critical_group_errors)
        historic_group_names_error(info_group_errors)
        latest_stage_names_error(critical_stage_errors)
        historic_stage_names_error(info_stage_errors)

        self.meta_groupings = meta_groupings_meta

    def get_quarter_list(self) -> None:
        output_list = []
        for master in self.master_data:
            try:
                output_list.append(str(master.month) + ", " + str(master.year))
            except KeyError:
                output_list.append(str(master.quarter))
        self.quarter_list = output_list

    def pipeline_projects_information(self) -> None:
        pipeline_dict = {}
        pipeline_list = []
        total_wlc = 0
        for p in self.all_projects:
            if self.project_information[p]["Pipeline"] is not None:
                wlc = convert_none_types(self.project_information[p]["WLC"])
                pipeline_dict[p] = {
                    "wlc": convert_none_types(self.project_information[p]["WLC"])
                }
                pipeline_list.append(p)
                total_wlc += wlc
        pipeline_dict["pipeline"] = {"wlc": total_wlc}

        self.pipeline_dict = pipeline_dict
        self.pipeline_list = pipeline_list

    def get_current_tp(self):
        try:
            self.current_quarter = (
                str(self.master_data[0].month) + ", " + str(self.master_data[0].year)
            )
        except KeyError:
            self.current_quarter = self.master_data[0].quarter


def json_date_converter(o):
    if isinstance(o, datetime.date):
        return o.__str__()


class JsonData:
    def __init__(
        self,
        master: List[Dict[str, Union[str, int, datetime.date, float]]],
        save_path: str,
    ):
        self.master = master
        self.path = save_path
        self.put_into_json()

    def put_into_json(self) -> None:
        master_list = []
        for m in self.master.master_data:
            data = m.data
            projects = m.projects
            qrt = str(str(m.quarter))
            d = {
                "data": data,
                "projects": projects,
                "quarter": qrt,
            }
            master_list.append(d)

        json_dict = {
            "abbreviations": self.master.abbreviations,
            "current_projects": self.master.current_projects,
            "current_quarter": str(self.master.current_quarter),
            "groups": self.master.portfolio_groups,
            "stages": self.master.stages,
            "meta_groupings": self.master.meta_groupings,  # change to meta_groupings
            "kwargs": self.master.kwargs,
            "json_master_data": master_list,
            "pipeline_dict": self.master.pipeline_dict,
            "pipeline_list": self.master.pipeline_list,
            "project_information": self.master.project_information,
            "quarter_list": self.master.quarter_list,
        }
        with open(self.path + ".json", "w") as write_file:
            json.dump(json_dict, write_file, default=json_date_converter)


def get_group_meta_data(settings_dict) -> Dict:
    """
    Gets group metadata types from config file. This is necessary as terminology is set in the config file
    and must correspond to terms used in project information document.
    """

    try:
        config_path = settings_dict["root_path"] + settings_dict["config"]
        config = configparser.ConfigParser()
        config.read(config_path)
        portfolio_group = json.loads(
            config.get("GROUPS", "portfolio_groups")
        )  # to return a list
        group_all = json.loads(config.get("GROUPS", "all_groups"))
    except:
        config_issue()

    group_meta_dict = {"port_group": portfolio_group, "all_groups": group_all}

    return group_meta_dict


def get_stage_meta_data(settings_dict) -> Dict:
    """
    Gets stage metadata types from config file. This is necessary as terminology is set in the config file
    and must correspond to terms used in project information document.
    """
    try:
        config_path = settings_dict["root_path"] + settings_dict["config"]
        config = configparser.ConfigParser()
        config.read(config_path)
        bc_stages = json.loads(config.get("GROUPS", "bc_stages"))
    except:
        config_issue()

    stage_meta_dict = {
        "stages": bc_stages,
    }

    return stage_meta_dict


def get_project_info_data(master_file: str) -> Dict:
    """
    Converts project_info document into a python dictionary. adapted from datamaps.api project_data_from_master
    """
    wb = load_workbook(master_file)
    ws = wb.active
    for cell in ws["A"]:
        if cell.value is None:
            continue
        c = Cleanser(cell.value)
        cell.value = c.clean()
    p_dict = {}
    for col in ws.iter_cols(min_col=2):
        project_name = ""
        o = OrderedDict()
        for cell in col:
            if cell.row == 1:
                project_name = cell.value
                p_dict[project_name] = o
            else:
                val = ws.cell(row=cell.row, column=1).value
                if type(cell.value) == datetime:
                    d_value = datetime.date(
                        cell.value.year, cell.value.month, cell.value.day
                    )
                    p_dict[project_name][val] = d_value
                else:
                    p_dict[project_name][val] = cell.value
    try:  # remove any 'None' projects that were pulled from the master
        del p_dict[None]
    except KeyError:
        pass
    return p_dict


def get_project_information(settings_dict) -> Dict[str, Union[str, int]]:
    """Returns dictionary containing all project meta data. confi_path is the config.ini file path.
    root_path is the core data root_path."""
    config_path = settings_dict["root_path"] + settings_dict["config"]
    config = configparser.ConfigParser()
    config.read(config_path)
    path = (
        str(settings_dict["root_path"])
        + "/core_data/"
        + config["PROJECT INFO"]["projects"]
    )
    return get_project_info_data(path)


def get_core(settings_dict) -> None:
    GROUP_META = get_group_meta_data(settings_dict)
    STAGE_META = get_stage_meta_data(settings_dict)
    META = {**GROUP_META, **STAGE_META}

    try:
        master = PythonMasterData(
            get_master_data(settings_dict),
            get_project_information(settings_dict),
            META,
            data_type=settings_dict["report"],
        )

    except (ProjectNameError, ProjectGroupError, ProjectStageError) as e:
        logger.critical(e)
        sys.exit(1)

    master_json_path = str(
        "{0}/core_data/json/master".format(settings_dict["root_path"])
    )
    JsonData(master, master_json_path)


def open_json_file(path: str, **op_args):
    with open(path, "r") as master_data:
        md = json.load(master_data)
        cli_qrts = get_iter_list(md, **op_args)
        cli_masters_list = []
        for x in cli_qrts:
            for y in md["json_master_data"]:
                if x == y["quarter"]:
                    cli_masters_list.append(y)
        md["master_data"] = cli_masters_list
        md["quarter_list"] = cli_qrts
        del md["json_master_data"]  # no longer required.
        return md
