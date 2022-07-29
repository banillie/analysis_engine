from openpyxl import load_workbook, Workbook
from datamaps.process import Cleanser
from datamaps.api import project_data_from_master


class Merge:
    def __init__(self, **op_args):
        self.op_args = op_args
        self.key_list = []
        self.p_names = []
        self.ignore_keys = [
            None,
            "file name",
            "Project/Programme Name",
            "Project Name (IPA Keys)",
            "Project Name (DfT Keys)",
        ]
        self.get_key_names()
        self.get_project_names()
        self.put_into_wb()

    def get_key_names(self):
        key_list = []
        for mst in self.op_args["masters_list"]:
            wb = load_workbook(self.op_args["root_path"] + f"/input/{mst}.xlsx")
            ws = wb.active
            for cell in ws["A"]:
                if cell.value in self.ignore_keys:
                    continue
                c = Cleanser(cell.value)
                cell.value = c.clean()
                if cell.value not in key_list:
                    key_list.append(cell.value)

        self.key_list = key_list

    def get_project_names(self):
        p_names = []
        for mst in self.op_args["masters_list"]:
            wb = load_workbook(self.op_args["root_path"] + f"/input/{mst}.xlsx")
            ws = wb.active
            for cell in ws["1"]:
                if cell.value in self.ignore_keys:
                    continue
                c = Cleanser(cell.value)
                cell.value = c.clean()
                if cell.value not in p_names:
                    p_names.append(cell.value)

        self.p_names = p_names

    def put_into_wb(self):
        wb = Workbook()
        ws = wb.active

        pd_msts = []  # pd_msts = python dictionary masters
        for mst in self.op_args["masters_list"]:
            pd_msts.append(
                project_data_from_master(
                    self.op_args["root_path"] + f"/input/{mst}.xlsx", 1, 2022
                )
            )

        for pn, project in enumerate(self.p_names):  # pn = project number
            ws.cell(row=1, column=2 + pn).value = project
            for m in pd_msts:
                for kn, key in enumerate(self.key_list):
                    try:
                        ws.cell(row=2 + kn, column=2 + pn).value = m.data[project][key]
                    except KeyError:
                        pass

        for kn, key in enumerate(self.key_list):
            ws.cell(row=2 + kn, column=1).value = key

        ws.cell(row=1, column=1).value = "Project/Programme Name"

        wb.save(self.op_args["root_path"] + "/output/merged_masters.xlsx")
