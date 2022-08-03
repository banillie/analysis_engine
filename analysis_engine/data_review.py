# code used to analysis outcomes from data review in march 2020.
# using https://realpython.com/python-counter/ along the way to better understand counter
# https://towardsdatascience.com/stacked-bar-charts-with-pythons-matplotlib-f4020e4eb4a7
# above was helpful for stacked bar charts, although difficult to follow as using pandas.

from collections import Counter, OrderedDict, defaultdict
from typing import Dict
import matplotlib.pyplot as plt
import numpy as np
from datamaps.process import Cleanser
from openpyxl import Workbook
from analysis_engine.data import (
    root_path,
    put_matplotlib_fig_into_word,
    get_input_doc,
    get_map,
)
from openpyxl import load_workbook


def get_review_info_data(master_file: str) -> Dict:

    """Taken from datamaps project_data_from_master api and adapted
    so dictionary nests keys via rows not columns"""

    wb = load_workbook(master_file)
    ws = wb.active
    for cell in ws["A"]:
        # we don't want to clean None...
        if cell.value is None:
            continue
        c = Cleanser(cell.value)
        cell.value = c.clean()
    review_dict = {}
    for row in ws.iter_rows(min_row=2):
        key_name = ""
        o = OrderedDict()
        for cell in row:
            if cell.column == 1:
                key_name = cell.value
                review_dict[key_name] = o
            else:
                val = ws.cell(row=1, column=cell.column).value
                review_dict[key_name][val] = cell.value
    # remove any "None" projects that were pulled from the master
    try:
        del review_dict[None]
    except KeyError:
        pass
    return review_dict


def generalise_score(mark):
    if mark == "EASY":
        return "HIGH"
    if mark == "HARD":
        return "LOW"
    if mark == "NONE":
        return "LOW"
    if mark == "MUST":
        return "HIGH"
    if mark == "SHOULD":
        return "MEDIUM"
    if mark == "COULD":
        return "LOW"

    else:
        return mark


def numbers(dictionary, review_category):
    counter = Counter()
    for key in dictionary.keys():
        score = dictionary[key][review_category]
        score = generalise_score(score)
        if score is not None:
            counter.update(Counter(score.split()))
            # counter[score] += 1  # with defaultdict which required counter = defaultdict(int)
            # counter[score] = counter.get(score, 0) + 1  # with counter = {}
            # if score not in counter:  # loop with counter = {}
            #     counter[score] = 0
            # counter[score] += 1

    return counter


def result(data):
    dicct = {}
    themes = [
        "Ease of Completion",
        "Use of Data",
        # 'Data Score',
        "MoSCoW",
        "Insightfulness",
    ]
    for t in themes:
        c = numbers(data, t)
        dicct[t] = c

    return dicct


def print_ascii_bar_chart(data, symbol="#"):
    counter = Counter(data).most_common()
    chart = {category: symbol * frequency for category, frequency in counter}
    max_len = max(len(category) for category in chart)
    for category, frequency in chart.items():
        padding = (max_len - len(category)) * " "
        print(f"{category}{padding} |{frequency}")


def place_in_matplotlib_order(data, type):
    l = []
    for x in data:
        l.append(data[x][type])
    return l


def bar_chart(data, title):
    fig, ax = plt.subplots(1, figsize=(12, 8))
    # fig, ax = plt.subplots()
    # labels = ['LOW', 'MEDIUM', 'HIGH']
    labels = list(data.keys())
    x = np.arange(0, len(labels))
    plt.bar(
        x - 0.3, place_in_matplotlib_order(data, "HIGH"), width=0.2
    )  # colour also an option
    plt.bar(x - 0.1, place_in_matplotlib_order(data, "MEDIUM"), width=0.2)
    plt.bar(x + 0.1, place_in_matplotlib_order(data, "LOW"), width=0.2)
    # plt.bar(x + 0.3, list(data['MoSCoW'].values()), width=0.2)
    # plt.bar(x + 0.6, list(data['Data Score'].values()), width=0.2)  # need to do data score also.

    ax.spines["right"].set_visible(False)
    ax.spines["top"].set_visible(False)

    plt.ylabel("Ratings count", fontweight="bold", fontsize=14)
    plt.xlabel("Themes", fontweight="bold", fontsize=14)
    plt.xticks(x, data.keys(), fontsize=14)
    plt.yticks(fontsize=14)
    # plt.xlim(-0.5, 31)

    # x, y = zip(*data)
    # plt.bar(x, y)
    plt.title(title, fontweight="bold", fontsize=16)
    plt.legend(["HIGH", "MEDIUM", "LOW"])
    # plt.show()

    return plt


def stacked_chart(data, title):
    fields = list(data.keys())
    # labels = ['LOW', 'MEDIUM', 'HIGH']
    labels = ["HIGH", "MEDIUM", "LOW"]

    fig, ax = plt.subplots(1, figsize=(12, 8))
    # fig, ax = plt.subplots()

    left = len(data) * [0]
    for i, name in enumerate(labels):
        plt.barh(list(data.keys()), place_in_matplotlib_order(data, name), left=left)
        left = [x + y for x, y in zip(left, place_in_matplotlib_order(data, name))]

    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_visible(False)
    ax.spines["top"].set_visible(False)
    ax.spines["bottom"].set_visible(False)

    # xticks = np.arange(0, 1.1, 0.1)
    # xlabels = ['{}%'.format(i) for i in np.arange(0, 101, 10)]

    ax.set_xticklabels(["0%", "20%", "40%", "60%", "80%", "100%"], fontsize=14)
    ax.xaxis.grid(color="grey", linestyle="dashed")

    # plt.ylabel('Ratings count', fontweight='bold')
    plt.xlabel("Percentage", fontweight="bold", fontsize=14)
    plt.yticks(fontsize=14)
    plt.title(title, fontweight="bold", fontsize=16)
    plt.legend(labels)
    plt.tight_layout()

    # plt.show()
    return plt


def sort_data_for_stacked_chart(data):
    percent_data = {}
    fields = ["LOW", "MEDIUM", "HIGH"]
    for d in data.keys():
        lower = {}
        for f in fields:
            lower[f] = data[d][f] / sum(data[d].values())
        percent_data[d] = lower

    return percent_data


def get_triage(
    data_slice: Dict,
    themes: list,
    # triage_method: list,
) -> Dict:

    # Helper funciton for data triage. Calculates triage for key ratings across all themes.
    # Retuns dict or nothing.

    scores_list = []
    for t in themes:
        if t == "GMPP Key":
            continue
        result = data_slice[t]
        scores_list.append(result)

    high = [x for x in scores_list if x not in ["HIGH", "EASY", "MUST"]]
    low = [x for x in scores_list if x not in ["NONE", "HARD", "LOW", "COULD"]]

    if scores_list == [None, None, None, None]:
        return "NONES"

    if high == []:
        return "KEEP"
    elif low == []:
        return "DELETE"
    else:
        return "DEBATE"


def put_triage_data_into_excel(
    workb: Workbook, result_dict: Dict, sheet_name: str, themes: list
) -> None:

    # Helper function to place each triage dictionary into excel across different tab.

    ws = workb.create_sheet(sheet_name)  # creating worksheets.
    # ws.title = sheet_name

    # dict for which keys are GMPP
    key_map = get_map(
        load_workbook(root_path / "input/GMPP_INTEGRATION_KEY_MAP.xlsx"),
        commas=True,
        gaps=True,
    )

    for x, k in enumerate(result_dict.keys()):
        ws.cell(row=x + 2, column=1).value = k
        for i, t in enumerate(themes):
            if t == "GMPP Key":
                try:
                    if k in key_map.keys():
                        ws.cell(row=x + 2, column=i + 2).value = "YES"
                except KeyError:
                    pass
            else:
                rating = result_dict[k][t]
                ws.cell(row=x + 2, column=i + 2).value = result_dict[k][t]

    for i, t in enumerate(themes):
        ws.cell(row=1, column=i + 2).value = t

    ws.cell(row=1, column=1).value = "Key Name"


def data_triage(data: Dict) -> None:
    """
    places raw data back into excel wb. Keys are triage into different tabs
    according to their score across different keys.
    """

    wb = Workbook()
    ws = wb.active

    themes = [
        "Ease of Completion",
        "Insightfulness",
        "Use of Data",
        # "Data Score",
        "MoSCoW",
        # "NEW",
        # "Notes",
        "GMPP Key",
    ]

    # high_triage = ["HIGH", "EASY", "MUST"]
    # medium_triage = ["HIGH", "EASY", "MUST", "MEDIUM", "SHOULD"]
    # low_triage = ["NONE", "HARD", "LOW", "COULD"]

    keep = {}
    debate = {}
    delete = {}

    x = 2
    for k in data.keys():
        rating = get_triage(data[k], themes)
        # high_triage)
        if rating == "KEEP":
            keep[k] = data[k]
        if rating == "DEBATE":
            debate[k] = data[k]
        if rating == "DELETE":
            delete[k] = data[k]

    put_triage_data_into_excel(wb, keep, "KEEP", themes)
    put_triage_data_into_excel(wb, debate, "DEBATE", themes)
    put_triage_data_into_excel(wb, delete, "DELETE", themes)

    wb.save(root_path / "output/data_review_triage.xlsx")

    pass


review_data = get_review_info_data(str(root_path) + "/data_review/DATA_REVIEW.xlsx")
all = result(review_data)
# print_ascii_bar_chart(uoa)
# bar = bar_chart(all, 'Ratings for each theme')
# p = sort_data_for_stacked_chart(all)
# stacked = stacked_chart(p, 'Ratings as a percentage of total')
#
# doc = get_input_doc(root_path / "input/summary_temp.docx")
# put_matplotlib_fig_into_word(doc, bar, size=5.5)
# put_matplotlib_fig_into_word(doc, stacked, size=5.5)
# doc.save(root_path / "data_review/review_charts.docx")
# data_triage(review_data)
