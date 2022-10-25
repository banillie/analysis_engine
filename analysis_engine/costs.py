from openpyxl.workbook import Workbook
from matplotlib import pyplot as plt
import numpy as np

from analysis_engine.segmentation import (
    get_group,
    get_correct_p_data,
)
from analysis_engine.cleaning import convert_none_types
from analysis_engine.dictionaries import (
    STANDARDISE_COST_KEYS,
    YEAR_LIST,
    RDEL_FORECAST_COST_KEYS,
    CDEL_FORECAST_COST_KEYS,
    RDEL_BL_COST_KEYS,
    CDEL_BL_COST_KEYS,
)
from analysis_engine.error_msgs import ProjectNameError, logger
from analysis_engine.settings import get_remove_income
from analysis_engine.render_utils import make_file_friendly
from analysis_engine.render_utils import set_fig_size, get_chart_title


class CostData:
    def __init__(self, master, **kwargs):
        self.master = master
        self.kwargs = kwargs
        self.report = kwargs["report"]
        self.quarters = self.master["quarter_list"]
        self.totals = {}
        self.baseline = {}
        self.profiles = {}
        self.get_totals()

    def get_totals(self) -> None:
        lower_dict = {}
        rm = get_remove_income(self.kwargs)
        for tp in self.quarters:
            spent = 0
            remaining = 0
            total = 0
            income_achieved = 0
            income_remaining = 0
            income_total = 0
            group = get_group(self.master, tp, **self.kwargs)
            for project_name in group:
                p_data = get_correct_p_data(self.master, project_name, tp)
                #  All try statements temporary until consistency between rpting datasets
                total += convert_none_types(
                    p_data[STANDARDISE_COST_KEYS[self.report]["total"]]
                )
                try:
                    spent += convert_none_types(
                        p_data[STANDARDISE_COST_KEYS[self.report]["spent"]]
                    )
                except KeyError:
                    spent += 0
                try:
                    income_achieved += convert_none_types(
                        p_data[STANDARDISE_COST_KEYS[self.report]["income_achieved"]]
                    )
                except KeyError:
                    income_achieved += 0
                try:
                    income_remaining += convert_none_types(
                        p_data[STANDARDISE_COST_KEYS[self.report]["income_remaining"]]
                    )
                except KeyError:
                    income_remaining += 0

                income_total += convert_none_types(
                    p_data[STANDARDISE_COST_KEYS[self.report]["income_total"]]
                )

                # option here handled via config file
                if "env_funds" not in self.kwargs:
                    if project_name in rm:
                        try:
                            total = total - convert_none_types(
                                p_data[
                                    STANDARDISE_COST_KEYS[self.report]["income_total"]
                                ]
                            )
                            if [project_name] == group:
                                logger.info(
                                    f"income has been removed from the total of {project_name}"
                                )
                        except KeyError:  # some older masters don't have key.
                            raise ProjectNameError()

            try:
                remaining += total - spent
            except KeyError:
                remaining += 0

            lower_dict[tp] = {
                "costs_spent": spent,
                "costs_remaining": remaining,
                "total": total,
                "income_achieved": income_achieved,
                "income_remaining": income_remaining,
                "income_total": income_total,
            }

        self.totals = lower_dict

    # def get_cost_profile(self) -> None:
    #     """Returns several lists which contain the sum of different cost profiles for the group of project
    #     contained with the master"""
    #     self.iter_list = get_iter_list(self.kwargs, self.master)
    #     lower_dict = {}
    #     for tp in self.iter_list:
    #         yearly_profile = []
    #         rdel_yearly_profile = []
    #         cdel_yearly_profile = []
    #         ngov_yearly_profile = []
    #         self.group = get_group(self.master, tp, self.kwargs)
    #         for year in YEAR_LIST:
    #             cost_total = 0
    #             rdel_total = 0
    #             cdel_total = 0
    #             ngov_total = 0
    #             for cost_type in COST_KEY_LIST:
    #                 for p in self.group:
    #                     p_data = get_correct_p_data(
    #                         self.kwargs, self.master, self.baseline_type, p, tp
    #                     )
    #                     if p_data is None:
    #                         continue
    #                     try:
    #                         cost = p_data[year + cost_type]
    #                         if cost is None:
    #                             cost = 0
    #                         cost_total += cost
    #                     except KeyError:  # handles data across different financial years via proj_info
    #                         try:
    #                             cost = self.master.project_information[p][
    #                                 year + cost_type
    #                                 ]
    #                         except KeyError:
    #                             cost = 0
    #                         if cost is None:
    #                             cost = 0
    #                         cost_total += cost
    #
    #                     if cost_type == COST_KEY_LIST[0]:  # rdel
    #                         rdel_total += cost
    #                     if cost_type == COST_KEY_LIST[1]:  # cdel
    #                         cdel_total += cost
    #                     if cost_type == COST_KEY_LIST[2]:  # ngov
    #                         ngov_total += cost
    #
    #             yearly_profile.append(cost_total)
    #             rdel_yearly_profile.append(rdel_total)
    #             cdel_yearly_profile.append(cdel_total)
    #             ngov_yearly_profile.append(ngov_total)
    #         lower_dict[tp] = {
    #             "prof": yearly_profile,
    #             "prof_ra": moving_average(yearly_profile, 2),
    #             "rdel": rdel_yearly_profile,
    #             "cdel": cdel_yearly_profile,
    #             "ngov": ngov_yearly_profile,
    #         }
    #     self.c_profiles = lower_dict
    #
    def get_forecast_cost_profile(self) -> None:
        COST_CAT = [" RDEL ", " CDEL "]
        profile_dict = {}
        group = get_group(self.master, self.quarters[0], **self.kwargs)
        lst_qrt = None
        for tp in self.quarters:
            f_rdel_list = []
            f_cdel_list = []
            f_total_list = []
            b_rdel_list = []
            b_cdel_list = []
            b_total_list = []
            for y in YEAR_LIST:
                f_year_rdel = 0
                f_year_cdel = 0
                b_year_rdel = 0
                b_year_cdel = 0
                for p in group:
                    try:
                        p_data = get_correct_p_data(self.master, p, tp)
                    except KeyError:
                        try:
                            p_data = get_correct_p_data(self.master, p, lst_qrt)
                        except KeyError:
                            p_data = None

                    if p_data is None:
                        continue

                    for cat in COST_CAT:
                        if cat == " RDEL ":
                            for k in RDEL_FORECAST_COST_KEYS.keys():
                                # if y in ["16-17", "17-18", "18-19"]:
                                #     try:
                                #         rdel = convert_none_types(self.master.project_information[p][y + cat + k])
                                #     except KeyError:
                                #         rdel = 0
                                #         print(y + cat + k + " not found.")
                                # else:
                                f_local_rdel = convert_none_types(p_data[y + cat + k])
                                f_year_rdel += f_local_rdel
                            for k in RDEL_BL_COST_KEYS.keys():
                                b_local_rdel = convert_none_types(p_data[y + cat + k])
                                b_year_rdel += b_local_rdel
                        if cat == " CDEL ":
                            for k in CDEL_FORECAST_COST_KEYS.keys():
                                try:
                                    f_local_cdel = convert_none_types(
                                        p_data[y + cat + k]
                                    )
                                except KeyError:
                                    # try:
                                    f_local_cdel = convert_none_types(p_data[y + k])
                                    # except KeyError:
                                    #     local_cdel = 0
                                f_year_cdel += f_local_cdel
                            for k in CDEL_BL_COST_KEYS.keys():
                                try:
                                    b_local_cdel = convert_none_types(
                                        p_data[y + cat + k]
                                    )
                                except KeyError:
                                    # try:
                                    b_local_cdel = convert_none_types(p_data[y + k])
                                    # except KeyError:
                                    #     # user messaging if necessary
                                    #     # print(tp + " " + y + k + ' could not be found. Check')
                                    #     cdel = 0
                                b_year_cdel += b_local_cdel

                f_rdel_list.append(f_year_rdel)
                f_cdel_list.append(f_year_cdel)
                f_total_list.append(f_year_rdel + f_year_cdel)
                b_rdel_list.append(b_year_rdel)
                b_cdel_list.append(b_year_cdel)
                b_total_list.append(b_year_rdel + b_year_cdel)

            lst_qrt = tp

            profile_dict[tp] = {
                "Total Forecast": f_total_list,
                "Total Baseline": b_total_list,
                "RDEL Forcast": f_rdel_list,
                "CDEL Forecast": f_cdel_list,
                "RDEL Baseline": b_rdel_list,
                "CDEL Baseline": b_cdel_list,
            }

        self.profiles = profile_dict

    # def get_baseline_cost_profile(self) -> None:
    #     COST_CAT = [" RDEL ", " CDEL "]
    #     # self.iter_list = get_iter_list(self.kwargs, self.master)
    #     self.iter_list = [self.master.current_quarter]
    #     tp_dict = {}
    #     for tp in self.iter_list:
    #         self.group = get_group(self.master, tp, self.kwargs)
    #         project_dict = {}
    #         list_total_total = []
    #         list_rdel_total = []
    #         list_cdel_total = []
    #         list_ngov_total = []
    #         for p in self.group:
    #             RDEL_BL_COST_KEYS = {
    #                 "BL one off new costs": [],
    #                 "BL recurring new costs": [],
    #                 "BL recurring old costs": [],
    #                 "BL Non Gov costs": [],
    #                 "BL Total": [],
    #                 "BL Income": [],
    #             }
    #             CDEL_BL_COST_KEYS = {
    #                 "BL one off new costs": [],
    #                 "BL recurring new costs": [],
    #                 "BL recurring old costs": [],
    #                 " BL Non-Gov": [],
    #                 "BL WLC": [],
    #                 " BL Income both Revenue and Capital": [],
    #             }
    #             # at moment bl only coming from current quarters data
    #             p_data = self.master.master_data[0]["data"][p]
    #             # p_data = get_correct_p_data(
    #             #     self.kwargs, self.master, self.baseline_type, p, tp
    #             # )
    #             if p_data is None:
    #                 continue
    #             for y in YEAR_LIST:
    #                 for cat in COST_CAT:
    #                     if cat == ' RDEL ':
    #                         for k in RDEL_BL_COST_KEYS.keys():
    #                             if y in ["16-17", "17-18", "18-19"]:
    #                                 rdel = 0
    #                             else:
    #                                 rdel = convert_none_types(p_data[y + cat + k])
    #                             RDEL_BL_COST_KEYS[k].append(rdel)
    #                     if cat == ' CDEL ':
    #                         for k in CDEL_BL_COST_KEYS.keys():
    #                             if y in ["16-17", "17-18", "18-19"]:
    #                                 cdel = 0
    #                             else:
    #                                 try:
    #                                     cdel = convert_none_types(p_data[y + cat + k])
    #                                 except KeyError:
    #                                     try:
    #                                         cdel = convert_none_types(p_data[y + k])
    #                                     except KeyError:
    #                                         # user messaging if necessary
    #                                         # print(tp + " " + y + k + ' could not be found. Check')
    #                                         cdel = 0
    #                             CDEL_BL_COST_KEYS[k].append(cdel)
    #                 total_adding = [RDEL_BL_COST_KEYS["BL Total"], CDEL_BL_COST_KEYS["BL WLC"]]
    #                 year_total = [sum(x) for x in zip(*total_adding)]
    #                 ngov_adding = [RDEL_BL_COST_KEYS["BL Non Gov costs"], CDEL_BL_COST_KEYS[" BL Non-Gov"]]
    #                 ngov_total = [sum(x) for x in zip(*ngov_adding)]
    #             # project_dict[p] = {
    #             #     "rdel": RDEL_BL_COST_KEYS["BL Total"],
    #             #     "cdel": CDEL_BL_COST_KEYS["BL WLC"],
    #             #     "total": year_total,
    #             # }
    #             list_total_total.append(year_total)
    #             list_ngov_total.append(ngov_total)
    #             list_rdel_total.append(RDEL_BL_COST_KEYS["BL Total"])
    #             list_cdel_total.append(CDEL_BL_COST_KEYS["BL WLC"])
    #         project_dict["total"] = [sum(x) for x in zip(*list_total_total)]
    #         project_dict["rdel_total"] = [sum(x) for x in zip(*list_rdel_total)]
    #         project_dict["cdel_total"] = [sum(x) for x in zip(*list_cdel_total)]
    #         project_dict["ngov_total"] = [sum(x) for x in zip(*list_ngov_total)]
    #         # tp_dict["baseline"] = project_dict
    #         self.profiles["baseline"] = project_dict
    #
    # def get_wlc_data(self) -> None:
    #     """
    #     calculates the quarters total wlc change
    #     """
    #     self.iter_list = get_iter_list(self.kwargs, self.master)
    #     wlc_dict = {}
    #     for tp in self.iter_list:
    #         self.group = get_group(self.master, tp, self.kwargs)
    #         wlc_dict = {}
    #         p_total = 0  # portfolio total
    #
    #         for i, g in enumerate(self.group):
    #             l_group = get_group(self.master, tp, self.kwargs, i)  # lower group
    #             g_total = 0
    #             l_g_l = []  # lower group list
    #             for p in l_group:
    #                 p_data = get_correct_p_data(
    #                     self.kwargs, self.master, self.baseline_type, p, tp
    #                 )
    #                 wlc = p_data["Total Forecast"]
    #                 if isinstance(wlc, (float, int)) and wlc is not None and wlc != 0:
    #                     if wlc > 50000:
    #                         logger.info(
    #                             tp
    #                             + ", "
    #                             + str(p)
    #                             + " is £"
    #                             + str(round(wlc))
    #                             + " please check this is correct. For now analysis_engine has recorded it as £0"
    #                         )
    #                     # wlc_dict[p] = wlc
    #                 if wlc == 0:
    #                     logger.info(
    #                         tp
    #                         + ", "
    #                         + str(p)
    #                         + " wlc is currently £"
    #                         + str(wlc)
    #                         + " note this is key information that should be provided by the project"
    #                     )
    #                     # wlc_dict[p] = wlc
    #                 if wlc is None:
    #                     logger.info(
    #                         tp
    #                         + ", "
    #                         + str(p)
    #                         + " wlc is currently None note this is key information that should be provided by the project"
    #                     )
    #                     wlc = 0
    #
    #                 l_g_l.append((wlc, p))
    #                 g_total += wlc
    #
    #             wlc_dict[g] = list(reversed(sorted(l_g_l)))
    #             p_total += g_total
    #
    #         wlc_dict["total"] = p_total
    #         wlc_dict[tp] = wlc_dict
    #
    #     self.wlc_dict = wlc_dict
    #
    # def calculate_wlc_change(self) -> None:
    #     wlc_change_dict = {}
    #     for i, tp in enumerate(self.wlc_dict.keys()):
    #         p_wlc_change_dict = {}
    #         for p in self.wlc_dict[tp].keys():
    #             wlc_one = self.wlc_dict[tp][p]
    #             try:
    #                 wlc_two = self.wlc_dict[self.iter_list[i + 1]][p]
    #                 try:
    #                     percentage_change = int(((wlc_one - wlc_two) / wlc_one) * 100)
    #                     p_wlc_change_dict[p] = percentage_change
    #                 except ZeroDivisionError:
    #                     logger.info(
    #                         "As "
    #                         + str(p)
    #                         + " has no wlc total figure for "
    #                         + tp
    #                         + " change has been calculated as zero"
    #                     )
    #             except IndexError:  # handles NoneTypes.
    #                 pass
    #
    #         wlc_change_dict[tp] = p_wlc_change_dict
    #
    #     self.wlc_change = wlc_change_dict


def cost_profile_into_wb_new(costs: CostData) -> Workbook:
    wb = Workbook()
    # ws = wb.active

    # type_list = [
    #     "rdel_total", "cdel_total", "total", "baseline_total"]

    for tp in list(costs.profiles.keys()):
        ws = wb.create_sheet(
            make_file_friendly(tp)
        )  # creating worksheets. names restricted to 30 characters.
        ws.title = make_file_friendly(tp)  # title of worksheet
        ws.cell(row=1, column=1).value = "F/Y"
        ws.cell(row=1, column=2).value = tp
        for x, key in enumerate(costs.profiles[tp].keys()):
            ws.cell(row=1, column=2 + x).value = key
            for i, cv in enumerate(costs.profiles[tp][key]):
                ws.cell(row=2 + i, column=1).value = YEAR_LIST[i]
                ws.cell(row=2 + i, column=2 + x).value = cv

    wb.remove(wb["Sheet"])

    return wb


def cost_profile_graph_new(costs: CostData, **kwargs) -> plt.figure:
    """Compiles a matplotlib line chart for costs of GROUP of projects contained within cost_master class"""

    fig, (ax1) = plt.subplots(1)  # two subplots for this chart
    fig = set_fig_size(kwargs, fig)

    title = get_chart_title(**kwargs)
    plt.suptitle(title, fontweight="bold", fontsize=20)

    # Overall cost profile chart
    if "baseline" not in kwargs:
        for i, tp in enumerate(list(costs.profiles.keys())):
            # try:
            #     label = tidy_label[iter]
            # except KeyError:
            label = tp
            ax1.plot(
                YEAR_LIST,
                np.array(costs.profiles[tp]["Total Forecast"]),
                label=label,
                linewidth=5.0,
                marker="o",
                zorder=10 - i,
            )
    else:
        profile_list = ["Total Forecast", "Total Baseline"]
        for i, p_type in enumerate(profile_list):
            label = p_type
            ax1.plot(
                YEAR_LIST,
                np.array(costs.profiles[costs.quarters[0]][p_type]),
                label=label,
                linewidth=5.0,
                marker="o",
                zorder=10 - i,
            )

    # Chart styling
    plt.xticks(rotation=45, size=16)
    plt.yticks(size=16)
    # ax1.tick_params(axis="series_one", which="major")  # matplotlib version issue
    ax1.set_ylabel("Cost (£m)")
    ax1.set_xlabel("Financial Year")
    xlab1 = ax1.xaxis.get_label()
    xlab1.set_style("italic")
    xlab1.set_size(20)
    ylab1 = ax1.yaxis.get_label()
    ylab1.set_style("italic")
    ylab1.set_size(20)
    ax1.grid(color="grey", linestyle="-", linewidth=0.2)
    ax1.legend(prop={"size": 20})

    fig.tight_layout(rect=[0, 0.03, 1, 0.95])  # size/fit of chart

    if "chart" in kwargs:
        if kwargs["chart"]:
            plt.show()

    return fig
