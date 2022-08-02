import datetime
from openpyxl import load_workbook, Workbook

from analysis_engine.error_msgs import config_issue, ProjectNameError


def get_map(wb, commas=False, gaps=False, flip=False):
    ws = wb.active
    output_dict = {}

    for x in range(2, ws.max_row + 1):
        ipa_key = ws.cell(row=x, column=2).value
        if ipa_key in output_dict.keys():
            pass
        if ipa_key is None:
            pass
        else:
            dft_key = ws.cell(row=x, column=1).value
            ipa_key = ws.cell(row=x, column=2).value
            if not commas:
                ipa_key = ipa_key.replace(",", "")
            if not gaps:
                ipa_key = ipa_key.replace("  ", " ")
            if flip:
                output_dict[ipa_key] = dft_key
            else:
                output_dict[dft_key] = ipa_key

    return output_dict


class GmppOnlineCosts:
    def __init__(self, **op_args):
        self.ws = None
        self.gmpp_dict = None
        self.op_args = op_args
        self.income = {}
        self.cost_total = {}
        self.income_ach = {}
        self.spent = {}
        self.get_workbook()
        self.get_data()
        self.get_cost_totals()

    def get_workbook(self):
        file_name = self.op_args["gmpp_data_path"]
        try:
            wb = load_workbook(
                self.op_args["root_path"]
                + "/input/{}.xlsm".format(file_name, data_only=True)
            )
        except FileNotFoundError:
            try:
                wb = load_workbook(
                    self.op_args["root_path"]
                    + "/input/{}.xlsm".format(file_name, data_only=True)
                )
            except FileNotFoundError:
                config_issue()

        self.ws = wb.active

    def get_data(self):
        from datetime import datetime
        import xlrd

        key_map = get_map(
            load_workbook(
                self.op_args["root_path"] + "/input/GMPP_INTEGRATION_KEY_MAP.xlsx"
            ),
            commas=True,
            gaps=True,
        )

        gmpp_dict = {}
        missing_keys = []
        for x in range(24, self.ws.max_row + 1):
            project_name = self.ws.cell(row=x, column=2).value
            key = self.ws.cell(row=x, column=6).value
            if key not in key_map.values():
                if key not in missing_keys:
                    missing_keys.append(key)
            s_value = self.ws.cell(row=x, column=7).value
            n_value = self.ws.cell(row=x, column=8).value
            if n_value != 0:
                s_value = n_value
            if "Date" in key or "date" in key or "6.03c: To" in key:
                if n_value > 20000:
                    s_value = datetime(*xlrd.xldate_as_tuple(n_value, 0))
            if "Grade" in key:  # to make grade 6 consistent with dft data
                try:
                    s_value = int(s_value)
                except ValueError:
                    pass

            if project_name in list(gmpp_dict.keys()):
                gmpp_dict[project_name][key] = s_value
            else:
                gmpp_dict[project_name] = {key: s_value}

        # print("These keys are missing")
        # for x in list(set(missing_keys)):
        #     print(x)

        self.gmpp_dict = gmpp_dict

    def get_cost_totals(self):
        income = {}
        total = {}
        income_ach = {}
        spent = {}
        for p in self.gmpp_dict.keys():
            # all forecast (not baseline).
            # There is no recurring old cost keys or non gov keys in here at mo
            rdel_income = 0  # unlikely to be any rdel income
            rdel_oonc = 0  # oonc = one off new costs!
            rdel_rnc = 0  # rnc = recurring new costs!
            rdel_ng = 0  # ng = non gov
            rdel_income_ach = 0
            rdel_spent = 0
            cdel_income = 0
            cdel_oonc = 0
            cdel_rnc = 0
            cdel_ng = 0
            cdel_income_ach = 0
            cdel_spent = 0
            for i in range(2, 51, 2):
                try:
                    rdel_income += self.gmpp_dict[p][f"9.01.{i}: RDEL Income"]
                except KeyError:
                    rdel_income += 0
                try:
                    rdel_oonc += self.gmpp_dict[p][f"9.01.{i}: One off new costs"]
                except KeyError:
                    rdel_oonc += 0
                try:
                    rdel_rnc += self.gmpp_dict[p][f"9.01.{i}: Recurring new costs"]
                except KeyError:
                    rdel_rnc += 0
                try:
                    rdel_ng += self.gmpp_dict[p][f"9.01.{i}: Non Gov"]
                except KeyError:
                    rdel_ng += 0

                try:
                    cdel_income += self.gmpp_dict[p][f"9.02.{i}: CDEL Income"]
                except KeyError:
                    cdel_income += 0
                try:
                    cdel_oonc += self.gmpp_dict[p][
                        f"9.02.{i}: One off new costs"
                    ]  # oonc = one off new costs!
                except KeyError:
                    cdel_oonc += 0
                try:
                    cdel_rnc += self.gmpp_dict[p][f"9.02.{i}: Recurring new costs"]
                except KeyError:
                    cdel_rnc += 0
                try:
                    cdel_ng += self.gmpp_dict[p][f"9.02.{i}: Non Gov"]
                except KeyError:
                    cdel_ng += 0

                if i == 8:
                    rdel_income_ach += rdel_income
                    rdel_spent += rdel_oonc + rdel_rnc + rdel_ng
                    cdel_income_ach += cdel_income
                    cdel_spent += cdel_oonc + cdel_rnc + cdel_ng

            income[p] = rdel_income + cdel_income
            total[p] = rdel_oonc + rdel_rnc + rdel_ng + cdel_oonc + cdel_rnc + cdel_ng
            income_ach[p] = rdel_income_ach + cdel_income_ach
            spent[p] = rdel_spent + cdel_spent

        self.income = income
        self.cost_total = total
        self.income_ach = income_ach
        self.spent = spent

    def put_cost_totals_into_wb(self):
        wb = Workbook()
        ws = wb.active

        project_map = get_map(
            load_workbook(
                self.op_args["root_path"] + "/input/GMPP_INTEGRATION_PROJECT_MAP.xlsx"
            )
        )

        for x, project in enumerate(list(self.cost_total.keys())):
            ws.cell(row=2 + x, column=1).value = project_map[project]
            ws.cell(row=2 + x, column=2).value = self.cost_total[project]
            ws.cell(row=2 + x, column=3).value = self.spent[project]
            ws.cell(row=2 + x, column=4).value = self.income[project]
            ws.cell(row=2 + x, column=5).value = self.income_ach[project]

        ws.cell(row=1, column=1).value = "Project Name"
        ws.cell(row=1, column=2).value = "Total Forecast Cost"
        ws.cell(row=1, column=3).value = "Spent"
        ws.cell(row=1, column=4).value = "Income"
        ws.cell(row=1, column=5).value = "Income Realised"

        wb.save(self.op_args["root_path"] + "/output/gmpp_online_total_costs.xlsx")

    def place_into_dft_master_format(self):
        wb = Workbook()
        ws = wb.active

        key_map = get_map(
            load_workbook(
                self.op_args["root_path"]
                + "/input/{}.xlsx".format(self.op_args["key_map_path"])
            ),
            commas=True,
            gaps=True,
            # flip=True,
        )
        project_map = get_map(
            load_workbook(
                self.op_args["root_path"] + "/input/GMPP_INTEGRATION_PROJECT_MAP.xlsx"
            )
        )

        for x, project in enumerate(list(self.gmpp_dict.keys())):
            try:
                ws.cell(row=1, column=3 + x).value = project_map[project]
            except KeyError:
                raise ProjectNameError(
                    f"{project} not in the key map. Please up date and re-run"
                )
            for i, k in enumerate(key_map.keys()):
                if x == 0:
                    try:
                        ws.cell(row=2 + i, column=1).value = k
                        ws.cell(row=2 + i, column=2).value = key_map[k]
                    except KeyError:
                        pass
                try:
                    if k == "Total Forecast":
                        v = self.cost_total[project]
                        ws.cell(row=2 + i, column=3 + x).value = v
                    if k == "Spent Costs":
                        v = self.spent[project]
                        ws.cell(row=2 + i, column=3 + x).value = v
                    if k == "Total Forecast - Income both Revenue and Capital":
                        v = self.income[project]
                        ws.cell(row=2 + i, column=3 + x).value = v
                    else:
                        v = self.gmpp_dict[project][key_map[k]]
                        ws.cell(row=2 + i, column=3 + x).value = v
                        if isinstance(v, datetime.datetime):
                            ws.cell(row=2 + i, column=3 + x).number_format = "dd/mm/yy"
                except KeyError:
                    pass

        ws.cell(row=1, column=1).value = "Project Name (IPA Keys)"
        ws.cell(row=1, column=2).value = "Project Name (DfT Keys)"

        wb.save(
            self.op_args["root_path"]
            + "/output/gmpp_online_data_dft_master_format.xlsx"
        )

    # # Code not currently in use.
    # def data_check_print_out(self):
    #     gmpp_data = project_data_from_master(root_path / "input/gmpp_online_data_temp.xlsx", 2, 2021)
    #     os.remove(root_path / "input/gmpp_online_data_temp.xlsx")
    #     ipdc_data = project_data_from_master(root_path / "core_data/{}.xlsx".format(ipdc_d_file_path), 2, 2021)
    #     key_map = get_map(load_workbook
    #                       (root_path / "input/{}.xlsx".format(km_file_name)), flip=True)
    #     project_map = get_map(load_workbook
    #                           (root_path / "input/{}.xlsx".format(pn_file_name)))
    #
    #     wb = Workbook()
    #     ws = wb.active
    #
    #     def remove_keys(key):
    #         output = key
    #         for rk in RK_LIST:  # remove key
    #             if rk in key:
    #                 output = "remove"
    #         return output
    #
    #     start_row = 2
    #     project_check_list = []
    #     for x, project in enumerate(list(project_map.keys())):  # could be project_map.keys()
    #         project_check_list.append(project)
    #         try:  # exception so only projects in ipdc data compared.
    #             for i, k in enumerate(gmpp_data.data[project]):
    #                 if k is None:
    #                     continue
    #                 check_key = remove_keys(k)
    #                 if check_key == "remove":
    #                     continue
    #                 ws.cell(row=start_row, column=1).value = project
    #                 try:
    #                     dft_project_name = project_map[project]
    #                     project_check = "PASS"
    #                 except KeyError:
    #                     dft_project_name = ""
    #                     project_check = "FAILED"
    #                 ws.cell(row=start_row, column=2).value = dft_project_name
    #                 ws.cell(row=start_row, column=3).value = project_check
    #                 ws.cell(row=start_row, column=4).value = k
    #                 try:
    #                     dft_key_name = key_map[k]
    #                     if dft_key_name == "None":
    #                         continue
    #                     key_check = "PASS"
    #                 except KeyError:
    #                     # print(k)
    #                     dft_key_name = ""
    #                     key_check = "FAILED"
    #                 ws.cell(row=start_row, column=5).value = dft_key_name
    #                 ws.cell(row=start_row, column=6).value = key_check
    #
    #                 gmpp_val = gmpp_data[project][k]
    #
    #                 try:
    #                     dft_val = ipdc_data[dft_project_name][dft_key_name]
    #                     if "Ver No" in dft_key_name or "Version No" in dft_key_name:
    #                         if dft_val is not None:
    #                             dft_val = str(dft_val)
    #                         if gmpp_val is not None:
    #                             gmpp_val = str(gmpp_val)
    #                     # if 'Phone No' in dft_key_name:  # started to think about tele nos but leaving for now.
    #                     #     print(gmpp_val)
    #                     #     print(dft_val)
    #                 except KeyError:
    #                     dft_val = ""
    #
    #                 ws.cell(row=start_row, column=7).value = gmpp_val
    #                 if isinstance(gmpp_val, datetime.datetime):
    #                     gmpp_val = gmpp_val.date()
    #                     ws.cell(row=start_row, column=7, value=gmpp_val).number_format = "dd/mm/yy"
    #
    #                 ws.cell(row=start_row, column=8).value = dft_val
    #                 if isinstance(dft_val, datetime.datetime):
    #                     dft_val = dft_val.date()
    #                     ws.cell(row=start_row, column=8, value=dft_val).number_format = "dd/mm/yy"
    #
    #                 if gmpp_val in list(GMPP_M_DICT.keys()):
    #                     if GMPP_M_DICT[gmpp_val] == dft_val:
    #                         ws.cell(row=start_row, column=9).value = "MATCH"
    #                         start_row += 1
    #                         continue
    #
    #                 if isinstance(gmpp_val, str) and isinstance(dft_val, str):
    #                     if "Ver No" in dft_key_name or "Version No" in k:
    #                         try:
    #                             gmpp_val = int(float(gmpp_val))
    #                             dft_val = int(float(dft_val))
    #                         except ValueError:
    #                             pass
    #                     else:
    #                         gmpp_val = gmpp_val.split()
    #                         dft_val = dft_val.split()
    #
    #                 # get floats of different lengths to match
    #                 if isinstance(dft_val, float) and isinstance(gmpp_val, float):
    #                     dft_val = float("{:.2f}".format(dft_val))
    #                     gmpp_val = float("{:.2f}".format(gmpp_val))
    #
    #                 if isinstance(dft_val, float) and isinstance(gmpp_val, int):
    #                     dft_val = round(dft_val)
    #
    #                 if isinstance(dft_val, int) and isinstance(gmpp_val, float):
    #                     gmpp_val = round(gmpp_val)
    #
    #                 if gmpp_val == dft_val:
    #                     ws.cell(row=start_row, column=9).value = "MATCH"
    #                 elif gmpp_val is None and dft_val == "":
    #                     ws.cell(row=start_row, column=9).value = "MATCH"
    #                 elif gmpp_val == "" and dft_val is None:
    #                     ws.cell(row=start_row, column=9).value = "MATCH"
    #                 elif gmpp_val is None and dft_val == 0:
    #                     ws.cell(row=start_row, column=9).value = "MATCH"
    #                 elif dft_key_name in IGNORE_LIST:
    #                     # print(dft_key_name)
    #                     ws.cell(row=start_row, column=9).value = "IGNORE"
    #                 else:
    #                     ws.cell(row=start_row, column=9).value = "DIFFERENT"
    #
    #                 start_row += 1
    #         except KeyError:
    #             pass
    #     ws.cell(row=1, column=1).value = "GMPP PROJECT NAME"
    #     ws.cell(row=1, column=2).value = "DFT PROJECT NAME"
    #     ws.cell(row=1, column=3).value = "NAME CHECK"
    #     ws.cell(row=1, column=4).value = "GMPP KEY"
    #     ws.cell(row=1, column=5).value = "DFT KEY"
    #     ws.cell(row=1, column=6).value = "KEY CHECK"
    #     ws.cell(row=1, column=7).value = "GMPP VALUE"
    #     ws.cell(row=1, column=8).value = "DFT VALUE"
    #     ws.cell(row=1, column=9).value = "VALUE CHECK"
    #
    #     p_check = [x for x in list(project_map.keys()) if x not in project_check_list]
    #     if not p_check:
    #         pass
    #     else:
    #         print("note following projects missing:")
    #         for x in p_check:
    #             print(x)
    #
    #     wb.save(root_path / f"output/GMPP_IPDC_DATA_CHECK_USING_{ipdc_d_file_path}.xlsx")
