import datetime

from openpyxl import Workbook

from analysis_engine.dictionaries import DCA_KEYS, PROJECT_INFO_KEYS
from analysis_engine.milestones import MilestoneData, get_milestone_date, convert_date

from analysis_engine.segmentation import get_iter_list, get_group, get_correct_p_data
from analysis_engine.render_utils import make_file_friendly
from analysis_engine.colouring import AMBER_FILL, SALMON_FILL


def data_query_into_wb(md, **kwargs) -> Workbook:
    """
    Returns data values for keys of interest. Keys placed on one page.
    Quarter data placed across different wbs.
    """

    wb = Workbook()
    md["quarter_list"].reverse()
    ms = MilestoneData(md, **kwargs).milestone_dict
    lst_qrt = None
    for z, tp in enumerate(md["quarter_list"]):
        ws = wb.create_sheet(
            make_file_friendly(tp), 0
        )  # creating worksheets. names restricted to 30 characters.
        ws.title = make_file_friendly(tp)  # title of worksheet
        qrt_group = get_group(md, tp, **kwargs)

        """list project names, groups and stage in ws"""
        for y, p in enumerate(qrt_group):  # p is project name
            p_data = get_correct_p_data(md, p, tp)
            abb = md["project_information"][p]["Abbreviations"]
            ws.cell(row=2 + y, column=1).value = md["project_information"][p][
                PROJECT_INFO_KEYS[kwargs["report"]]["group"]
            ]
            ws.cell(row=2 + y, column=2).value = p
            ws.cell(row=2 + y, column=3).value = abb
            ws.cell(row=2 + y, column=4).value = md["project_information"][p]["GMPP"]

            for x, key in enumerate(kwargs["key"]):
                ws.cell(row=1, column=5 + x, value=key)
                try:  # standard keys
                    value = p_data[key]
                    if isinstance(value, datetime.date):  # in case value direct date.
                        value = convert_date(value)
                except KeyError:  # milestone keys
                    value = get_milestone_date(ms, key, tp, abb)

                if value is None:
                    ws.cell(row=2 + y, column=5 + x).value = "md"
                    ws.cell(row=2 + y, column=5 + x).fill = AMBER_FILL
                else:
                    ws.cell(row=2 + y, column=5 + x, value=value)
                    if isinstance(value, datetime.date):
                        ws.cell(
                            row=2 + y, column=5 + x, value=value
                        ).number_format = "dd/mm/yy"

                if lst_qrt:
                    try:
                        last_q_data = get_correct_p_data(md, p, lst_qrt)
                        lst_value = last_q_data[key]
                    except KeyError:
                        lst_value = get_milestone_date(ms, key, lst_qrt, abb)

                    if value != lst_value:
                        ws.cell(row=2 + y, column=5 + x).fill = SALMON_FILL

        lst_qrt = tp

        ws.cell(row=1, column=1).value = "Group"
        ws.cell(row=1, column=2).value = "Project Name"
        ws.cell(row=1, column=3).value = "Project Acronym"
        ws.cell(row=1, column=4).value = "GMPP"

    wb.remove(wb["Sheet"])
    return wb


# def convert_date(date_str: str):
#     """
#     When date converted into json file the dates take the standard python format
#     year-month-day. This function converts format to year-day-month. This function is
#     used when the MilestoneData class is created. Seems to be the best place to deploy.
#     """
#     try:
#         return parser.parse(date_str)  # returns datetime
#     except TypeError:  # for a different data value e.g integer.
#         return date_str
#     except ValueError:  # for string data that is not a date.
#         return date_str
#     # is a ParserError necessary here also?
