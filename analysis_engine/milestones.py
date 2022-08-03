from typing import List, Dict, Union
from collections import Counter
from datetime import datetime
from matplotlib import pyplot as plt
import matplotlib.dates as mdates
from openpyxl import Workbook
from dateutil import parser

from analysis_engine.colouring import COLOUR_DICT
from analysis_engine.render_utils import set_fig_size, get_chart_title, handle_long_keys
from analysis_engine.segmentation import get_group, get_correct_p_data
from analysis_engine.error_msgs import logger
from analysis_engine.settings import get_board_date, convert_date


def milestone_info_handling(output_list: list, t_list: list, **kwargs) -> list:
    """helper function for handling and cleaning up milestone date generated
    via MilestoneDate class. Removes none type milestone names and non date
    string values"""

    if t_list[1][1] is None or t_list[1][1] == "Project - Business Case End Date":
        pass
    else:
        try:
            t_list[3] = ("Date", t_list[3][1].date())
            return output_list.append(t_list)
        except AttributeError:  # Non-datetime values
            if "type" in kwargs:
                if kwargs["type"] == "central support":
                    if t_list[3][1] is None:
                        return output_list.append(t_list)
                    else:
                        logger.info(
                            t_list[0][1]
                            + ": incorrect date format for entry '"
                            + t_list[1][1]
                            + ""
                            "', requires amending or will not be included. "
                            + str(kwargs["tp"])
                            + " data."
                        )
            if t_list[3][1] is None:
                pass
            else:
                logger.info(
                    t_list[0][1]
                    + ": incorrect date format for entry '"
                    + t_list[1][1]
                    + ""
                    "', requires amending or will not be included. "
                    + str(kwargs["tp"])
                    + " data."
                )


def get_milestone_date(
    ms_dict: Dict[str, Union[datetime.date, str]],
    milestone_name: str,
    tp: str,
    p_name: str,
) -> datetime:
    qrt_ms_dict = ms_dict[tp]
    for k in qrt_ms_dict:
        if qrt_ms_dict[k]["Project"] == p_name:
            if qrt_ms_dict[k]["Milestone"] == milestone_name:
                return qrt_ms_dict[k]["Date"]


class MilestoneData:
    def __init__(self, master, **kwargs):
        self.master = master
        self.kwargs = kwargs
        self.report = kwargs["report"]
        self.quarters = self.master["quarter_list"]
        self.milestone_dict = {}
        self.sorted_milestone_dict = {}
        self.next_milestone_dict = {}
        self.max_date = None
        self.min_date = None
        # self.schedule_change = {}
        # self.schedule_key_last = None
        # self.schedule_key_baseline = None
        self.get_milestones()
        self.get_chart_info()
        # self.get_next_milestone()
        # # self.calculate_schedule_changes()

    def get_milestones(self) -> None:
        """
        Creates project milestone dictionaries
        """
        m_dict = {}
        for tp in self.quarters:  # tp time period
            # self.kwargs["tp"] = tp
            lower_dict = {}
            raw_list = []
            group = get_group(self.master, tp, **self.kwargs)
            for project_name in group:
                project_milestones = []
                p_data = get_correct_p_data(self.master, project_name, tp)
                if p_data is None:
                    continue
                # i loops below removes None Milestone names and rejects non-datetime date values.
                p = self.master["project_information"][project_name]["Abbreviations"]
                category = "Milestone"
                if self.kwargs["report"] == "cdg":
                    # report = "CDG"
                    for i in range(1, 15):
                        entry = [
                            ("Project", p),
                            ("Milestone", p_data["MM" + str(i)]),
                            ("Notes", p_data["MM" + str(i) + " NOTES"]),
                            ("Date", convert_date(p_data["MM" + str(i) + " DATE"])),
                            ("Status", p_data["MM" + str(i) + " STATUS"]),
                            ("Report", self.kwargs["report"]),
                            ("Cat", category),
                        ]
                        milestone_info_handling(
                            project_milestones, entry, **self.kwargs
                        )

                if self.kwargs["report"] == "ipdc":
                    for i in range(1, 50):
                        try:
                            entry = [
                                ("Project", p),
                                ("Milestone", p_data["Approval MM" + str(i)]),
                                ("Type", "Approval"),
                                (
                                    "Date",
                                    convert_date(
                                        p_data[
                                            "Approval MM"
                                            + str(i)
                                            + " Forecast / Actual"
                                        ]
                                    ),
                                ),
                                ("Notes", p_data["Approval MM" + str(i) + " Notes"]),
                                ("Report", self.kwargs["report"]),
                                ("Cat", category),
                            ]
                            milestone_info_handling(
                                project_milestones, entry, **self.kwargs
                            )
                            entry = [
                                ("Project", p),
                                ("Milestone", p_data["Assurance MM" + str(i)]),
                                ("Type", "Assurance"),
                                (
                                    "Date",
                                    convert_date(
                                        p_data[
                                            "Assurance MM"
                                            + str(i)
                                            + " Forecast - Actual"
                                        ]
                                    ),
                                ),
                                ("Notes", p_data["Assurance MM" + str(i) + " Notes"]),
                                ("Report", self.kwargs["report"]),
                                ("Cat", category),
                            ]
                            milestone_info_handling(
                                project_milestones, entry, **self.kwargs
                            )
                        except KeyError:  # handles inconsistent keys naming for approval milestones.
                            try:
                                entry = [
                                    ("Project", p),
                                    ("Milestone", p_data["Approval MM" + str(i)]),
                                    ("Type", "Approval"),
                                    (
                                        "Date",
                                        convert_date(
                                            p_data[
                                                "Approval MM"
                                                + str(i)
                                                + " Forecast - Actual"
                                            ]
                                        ),
                                    ),
                                    (
                                        "Notes",
                                        p_data["Approval MM" + str(i) + " Notes"],
                                    ),
                                    ("Report", self.kwargs["report"]),
                                    ("Cat", category),
                                ]
                                milestone_info_handling(
                                    project_milestones, entry, **self.kwargs
                                )
                            except KeyError:
                                pass

                    # handles inconsistent number of Milestone. could be incorporated above.
                    for i in range(18, 67):
                        try:
                            entry = [
                                ("Project", p),
                                ("Milestone", p_data["Project MM" + str(i)]),
                                ("Type", "Delivery"),
                                (
                                    "Date",
                                    convert_date(
                                        p_data[
                                            "Project MM" + str(i) + " Forecast - Actual"
                                        ]
                                    ),
                                ),
                                ("Notes", p_data["Project MM" + str(i) + " Notes"]),
                                ("Report", self.kwargs["report"]),
                                ("Cat", category),
                            ]
                            milestone_info_handling(
                                project_milestones, entry, **self.kwargs
                            )
                        except KeyError:
                            pass

                    # change in Q3. Some milestones collected via HMT approval section.
                    # this loop picks them up
                    for i in range(1, 4):
                        try:
                            entry = [
                                ("Project", p),
                                ("Milestone", p_data["HMT Approval " + str(i)]),
                                ("Type", "Approval"),
                                (
                                    "Date",
                                    convert_date(
                                        p_data[
                                            "HMT Approval "
                                            + str(i)
                                            + " Forecast / Actual"
                                        ]
                                    ),
                                ),
                                ("Notes", p_data["HMT Approval " + str(i) + " Notes"]),
                                ("Report", self.kwargs["report"]),
                                ("Cat", category),
                            ]
                            milestone_info_handling(
                                project_milestones, entry, **self.kwargs
                            )
                        except KeyError:
                            pass

                # loop to stop keys names being the same. Done at project level.
                # not particularly concise code.
                upper_counter_list = []
                for entry in project_milestones:
                    upper_counter_list.append(entry[1][1])
                upper_count = Counter(upper_counter_list)
                lower_counter_list = []
                for entry in project_milestones:
                    if upper_count[entry[1][1]] > 1:
                        lower_counter_list.append(entry[1][1])
                        lower_count = Counter(lower_counter_list)
                        new_milestone_key = (
                            entry[1][1] + " (" + str(lower_count[entry[1][1]]) + ")"
                        )
                        entry[1] = ("Milestone", new_milestone_key)
                        raw_list.append(entry)
                    else:
                        raw_list.append(entry)

            # puts the list in chronological order
            sorted_list = sorted(raw_list, key=lambda k: (k[3][1] is None, k[3][1]))

            for r in range(len(sorted_list)):
                lower_dict["Milestone " + str(r)] = dict(sorted_list[r])

            m_dict[tp] = lower_dict
            # HERE POTENTIAL TO MERGE DICTS
        self.milestone_dict = m_dict

    def get_chart_info(self) -> None:
        """returns data lists for matplotlib chart"""
        # Note this code could refactored so that it collects all milestones
        # reported across current, last and baseline. At the moment it only
        # uses milestones that are present in the current quarter.

        output_dict = {}
        for i in self.milestone_dict:
            report = []
            category = []
            p_names = []
            key_names = []
            g_dates = []  # graph dates
            r_dates = []  # raw dates
            notes = []
            status = []
            for v in self.milestone_dict[self.quarters[0]].values():
                p = None  # project
                mn = None  # milestone name
                d = None  # date
                for x in self.milestone_dict[i].values():
                    if (
                        x["Project"] == v["Project"]
                        and x["Milestone"] == v["Milestone"]
                    ):
                        p = x["Project"]
                        mn = x["Milestone"]
                        p_names.append(p)
                        key_names.append(mn)
                        d = x["Date"]
                        g_dates.append(d)
                        r_dates.append(d)
                        notes.append(x["Notes"])
                        report.append(x["Report"])
                        category.append(x["Cat"])
                        try:
                            status.append(x["Status"])
                        except KeyError:
                            pass
                        break
                if p is None and mn is None and d is None:
                    p = v["Project"]
                    mn = v["Milestone"]
                    p_names.append(p)
                    key_names.append(mn)
                    g_dates.append(v["Date"])
                    r_dates.append(None)
                    notes.append(None)
                    status.append(None)
                    report.append(x["Report"])
                    category.append(x["Cat"])

            output_dict[i] = {
                "project": p_names,
                "names": key_names,
                "g_dates": g_dates,
                "r_dates": r_dates,
                "notes": notes,
                "status": status,  # only present for top35
                "report": report,
                "cat": category,
            }

        self.sorted_milestone_dict = output_dict

    # def get_chart_info_old(self) -> None:
    #     """returns data lists for matplotlib chart"""
    #     # Note this code could refactored so that it collects all milestones
    #     # reported across current, last and baseline. At the moment it only
    #     # uses milestones that are present in the current quarter.
    #     key_names = []
    #     key_names_last = []
    #     keys_names_baseline = []
    #     md_current = []
    #     md_last = []
    #     md_last_po = []  # po is for printout
    #     md_baseline = []
    #     md_baseline_po = []
    #     md_baseline_two_po = []
    #     md_baseline_two = []
    #     type_list = []
    #
    #     for m in self.milestone_dict[self.iter_list[0]].values():
    #         m_project = m["Project"]
    #         m_name = m["Milestone"]
    #         m_date = m["Date"]
    #         m_type = m["Type"]
    #         key_names.append(m_project + ", " + m_name)
    #         md_current.append(m_date)
    #         type_list.append(m_type)
    #
    #         # In two loops below NoneType has to be replaced with a datetime object
    #         # due to matplotlib being unable to handle NoneTypes when milestone_chart
    #         # is created. Haven't been able to find a solution to this.
    #         try:
    #             m_last_date = None
    #             for m_last in self.milestone_dict[self.iter_list[1]].values():
    #                 if m_last["Project"] == m_project:
    #                     if m_last["Milestone"] == m_name:
    #                         key_names_last.append(m_project + ", " + m_name)
    #                         m_last_date = m_last["Date"]
    #                         md_last.append(m_last_date)
    #                         md_last_po.append(m_last_date)
    #                         continue
    #             if m_last_date is None:
    #                 md_last.append(m_date)
    #                 md_last_po.append(None)
    #
    #             m_bl_date = None
    #             for m_bl in self.milestone_dict[self.iter_list[2]].values():
    #                 if m_bl["Project"] == m_project:
    #                     if m_bl["Milestone"] == m_name:
    #                         keys_names_baseline.append(m_project + ", " + m_name)
    #                         m_bl_date = m_bl["Date"]
    #                         md_baseline.append(m_bl_date)
    #                         md_baseline_po.append(m_bl_date)
    #                         continue
    #             if m_bl_date is None:
    #                 md_baseline.append(m_date)
    #                 md_baseline_po.append(None)
    #
    #             m_bl_two_date = None
    #             for m_bl_two in self.milestone_dict[self.iter_list[3]].values():
    #                 if m_bl_two["Project"] == m_project:
    #                     if m_bl_two["Milestone"] == m_name:
    #                         m_bl_two_date = m_bl_two["Date"]
    #                         md_baseline_two.append(m_bl_two_date)
    #                         md_baseline_two_po.append(m_bl_two_date)
    #                         continue
    #             if m_bl_two_date is None:
    #                 md_baseline_two.append(m_date)
    #                 md_baseline_two_po.append(None)
    #
    #         except IndexError:
    #             pass
    #
    #     if len(self.group) == 1:
    #         key_names = remove_project_name_from_milestone_key(
    #             self.master.abbreviations[self.group[0]]["abb"], key_names
    #         )
    #     else:
    #         pass
    #
    #     self.key_names = key_names
    #     self.key_names_last = key_names_last
    #     self.key_names_baseline = keys_names_baseline
    #     self.md_current = md_current
    #     self.md_last = md_last
    #     self.md_last_po = md_last_po
    #     self.md_baseline = md_baseline
    #     self.md_baseline_po = md_baseline_po
    #     self.md_baseline_two = md_baseline_two
    #     self.md_baseline_two_po = md_baseline_two_po
    #     self.type_list = type_list
    #     self.max_date = max(
    #         remove_none_types(self.md_current)
    #         + remove_none_types(self.md_last)
    #         + remove_none_types(self.md_baseline)
    #     )
    #     self.min_date = min(
    #         remove_none_types(self.md_current)
    #         + remove_none_types(self.md_last)
    #         + remove_none_types(self.md_baseline)
    #     )

    def filter_chart_info(self, **filter_kwargs):
        # bug handling required in the event that there are no milestones with the filter.
        # i.e. the filter returns no milestones.
        filtered_dict = {}
        if (
            "type" in filter_kwargs
            and "key" in filter_kwargs
            and "dates" in filter_kwargs
        ):
            start_date, end_date = zip(*filter_kwargs["dates"])
            start = parser.parse(start_date, dayfirst=True)
            end = parser.parse(end_date, dayfirst=True)
            for i, v in enumerate(self.milestone_dict[self.iter_list[0]].values()):
                if v["Type"] in filter_kwargs["type"]:
                    if v["Milestone"] in filter_kwargs["keys"]:
                        if start.date() <= filter_kwargs["dates"] <= end.date():
                            filtered_dict["Milestone " + str(i)] = v
                            continue

        elif "type" in filter_kwargs and "key" in filter_kwargs:
            for i, v in enumerate(self.milestone_dict[self.quarters[0]].values()):
                if v["Type"] in filter_kwargs["type"]:
                    if v["Milestone"] in filter_kwargs["keys"]:
                        filtered_dict["Milestone " + str(i)] = v
                        continue

        elif "type" in filter_kwargs and "dates" in filter_kwargs:
            start_date, end_date = zip(filter_kwargs["dates"])
            start = parser.parse(start_date[0], dayfirst=True)
            end = parser.parse(end_date[0], dayfirst=True)
            for i, v in enumerate(self.milestone_dict[self.quarters[0]].values()):
                if v["Type"] in filter_kwargs["type"]:
                    if start.date() <= v["Date"] <= end.date():
                        filtered_dict["Milestone " + str(i)] = v
                        continue

        elif "key" in filter_kwargs and "dates" in filter_kwargs:
            start_date, end_date = zip(filter_kwargs["dates"])
            start = parser.parse(start_date[0], dayfirst=True)
            end = parser.parse(end_date[0], dayfirst=True)
            for i, v in enumerate(self.milestone_dict[self.quarters[0]].values()):
                if v["Milestone"] in filter_kwargs["key"]:
                    if start.date() <= v["Date"] <= end.date():
                        filtered_dict["Milestone " + str(i)] = v
                        continue

        elif "type" in filter_kwargs:
            for i, v in enumerate(self.milestone_dict[self.quarters[0]].values()):
                if v["Type"] in filter_kwargs["type"]:
                    filtered_dict["Milestone " + str(i)] = v
                    continue

        elif "key" in filter_kwargs:
            for i, v in enumerate(self.milestone_dict[self.quarters[0]].values()):
                if v["Milestone"] in filter_kwargs["key"]:
                    filtered_dict["Milestone " + str(i)] = v
                    continue

        elif "dates" in filter_kwargs:
            start_date, end_date = zip(filter_kwargs["dates"])
            start = parser.parse(start_date[0], dayfirst=True)
            end = parser.parse(end_date[0], dayfirst=True)
            for i, v in enumerate(self.milestone_dict[self.quarters[0]].values()):
                if start.date() <= v["Date"] <= end.date():
                    filtered_dict["Milestone " + str(i)] = v
                    continue

        output_dict = {}
        for dict in self.milestone_dict.keys():
            if dict == self.quarters[0]:
                output_dict[dict] = filtered_dict
            else:
                output_dict[dict] = self.milestone_dict[dict]

        self.milestone_dict = output_dict
        self.get_chart_info()

    def calculate_schedule_changes(self) -> None:
        """calculates the changes in project schedules. If standard key for calculation
        not available it using the best next one available"""

        self.filter_chart_info(milestone_type=["Delivery", "Approval"])
        m_dict_keys = list(self.milestone_dict.keys())

        def schedule_info(
            project_name: str,
            other_key_list: List[str],
            c_key_list: List[str],
            miles_dict: dict,
            dict_l_current: str,
            dict_l_other: str,
        ):
            output_dict = {}
            schedule_info = []
            for key in reversed(other_key_list):
                if key in c_key_list:
                    sop = get_milestone_date(
                        project_name, miles_dict, dict_l_other, " Start of Project"
                    )
                    if sop is None:
                        sop = get_milestone_date(
                            project_name, miles_dict, dict_l_current, other_key_list[0]
                        )
                        schedule_info.append(("start key", other_key_list[0]))
                    else:
                        schedule_info.append(("start key", " Start of Project"))
                    schedule_info.append(("start", sop))
                    schedule_info.append(("end key", key))
                    date = get_milestone_date(
                        project_name, miles_dict, dict_l_current, key
                    )
                    schedule_info.append(("end current date", date))
                    other_date = get_milestone_date(
                        project_name, miles_dict, dict_l_other, key
                    )
                    schedule_info.append(("end other date", other_date))
                    project_length = (other_date - sop).days
                    schedule_info.append(("project length", project_length))
                    change = (date - other_date).days
                    schedule_info.append(("change", change))
                    p_change = int((change / project_length) * 100)
                    schedule_info.append(("percent change", p_change))
                    output_dict[dict_l_other] = dict(schedule_info)
                    break

            return output_dict

        output_dict = {}
        for project_name in self.group:
            project_name = self.master.abbreviations[project_name]
            current_key_list = []
            last_key_list = []
            baseline_key_list = []
            for key in self.key_names:
                try:
                    p = key.split(",")[0]
                    milestone_key = key.split(",")[1]
                    if project_name == p:
                        if milestone_key != " Project - Business Case End Date":
                            current_key_list.append(milestone_key)
                except IndexError:
                    # patch of single project group. In this instance the project name
                    # is removed from the key_name via remove_project_name function as
                    # part of get chart info.
                    if len(self.group) == 1:
                        current_key_list.append(" " + key)
            for last_key in self.key_names_last:
                p = last_key.split(",")[0]
                milestone_key_last = last_key.split(",")[1]
                if project_name == p:
                    if milestone_key_last != " Project - Business Case End Date":
                        last_key_list.append(milestone_key_last)
            for baseline_key in self.key_names_baseline:
                p = baseline_key.split(",")[0]
                milestone_key_baseline = baseline_key.split(",")[1]
                if project_name == p:
                    if (
                        milestone_key_baseline
                        != " Project - Business Case End Date"
                        # and milestone_key_baseline != " Project End Date"
                    ):
                        baseline_key_list.append(milestone_key_baseline)

            b_dict = schedule_info(
                project_name,
                baseline_key_list,
                current_key_list,
                self.milestone_dict,
                m_dict_keys[0],
                m_dict_keys[2],
            )
            l_dict = schedule_info(
                project_name,
                last_key_list,
                current_key_list,
                self.milestone_dict,
                m_dict_keys[0],
                m_dict_keys[1],
            )
            lower_dict = {**b_dict, **l_dict}

            output_dict[project_name] = lower_dict

        self.schedule_change = output_dict

    def get_next_milestone(
        self,
        date: datetime.date,
    ) -> list:
        next_ms_dict = {}
        group = get_group(self.master, self.quarters[0], **self.kwargs)
        for p in group:
            for x in self.milestone_dict[self.quarters[0]].values():
                if (
                    x["Project"]
                    == self.master["project_information"][p]["Abbreviations"]
                ):
                    d = x["Date"]
                    ms = x["Milestone"]
                    if d >= date:
                        if p not in list(next_ms_dict.keys()):
                            next_ms_dict[p] = {
                                "milestone": ms,
                                "date": d,
                            }

        self.next_milestone_dict = next_ms_dict


def merge_project_milestone_name(project, ms_names, no):
    return project[no] + ", " + ms_names[no]


def calculate_max_min_date(milestones: MilestoneData, **kwargs) -> int:
    m_list = []
    for i in milestones.sorted_milestone_dict.keys():
        m_list += milestones.sorted_milestone_dict[i]["g_dates"]
    # This step required for central support milestone dates which can be None.
    final_m_list = [x for x in m_list if x is not None]
    if kwargs["value"] == "max":
        return max(final_m_list)
    if kwargs["value"] == "min":
        return min(final_m_list)


# LEGEND_DICT = {
#     "Q1 22/23": "THIS QUARTER",
#     "Q4 21/22": "LAST QUARTER",
#     "Q1 21/22": "ONE YEAR AG0",
# }


def milestone_chart(
    milestones: MilestoneData,
    **kwargs,
) -> plt.figure:
    fig, ax1 = plt.subplots()
    fig = set_fig_size(kwargs, fig)
    title = get_chart_title(**kwargs)
    plt.suptitle(title, fontweight="bold", fontsize=20)
    project = milestones.sorted_milestone_dict[milestones.quarters[0]]["project"]
    ms_names = milestones.sorted_milestone_dict[milestones.quarters[0]]["names"]
    combined = [
        merge_project_milestone_name(project, ms_names, i)
        for i, p in enumerate(project)
    ]
    ms_names = handle_long_keys(combined, output_type="milestones")

    colour_start = 1
    for i, tp in enumerate(milestones.quarters):
        m = [
            x for x in milestones.sorted_milestone_dict[tp]["g_dates"] if x is not None
        ]
        ax1.scatter(
            m,
            ms_names[0 : len(m)],
            label=tp,  # add functionality so user can add via config
            s=200,
            zorder=20 - i,
            edgecolor=COLOUR_DICT["BLUE"],
            fc=COLOUR_DICT["BLUE"],
            alpha=colour_start,
        )
        colour_start = colour_start - (1 / len(milestones.quarters))

    ax1.legend(prop={"size": 14})  # insert legend
    plt.yticks(size=10)
    # reverse series_two axis so order is earliest to oldest
    ax1 = plt.gca()
    ax1.set_ylim(ax1.get_ylim()[::-1])
    ax1.yaxis.grid()  # horizontal lines
    ax1.set_axisbelow(True)

    years = mdates.YearLocator()  # every year
    months = mdates.MonthLocator()  # every month
    weeks = mdates.WeekdayLocator()
    years_fmt = mdates.DateFormatter("%Y")
    months_fmt = mdates.DateFormatter("%b")
    weeks_fmt = mdates.DateFormatter("%d")

    max_date = calculate_max_min_date(milestones, value="max")
    min_date = calculate_max_min_date(milestones, value="min")
    td = (max_date - min_date).days
    if td >= 365 * 3:
        ax1.xaxis.set_major_locator(years)
        ax1.xaxis.set_minor_locator(months)
        ax1.xaxis.set_major_formatter(years_fmt)
        plt.setp(ax1.xaxis.get_minorticklabels(), rotation=45, size=12)
        plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, weight="bold", size=14)
    elif 365 * 3 >= td >= 90:
        ax1.xaxis.set_major_locator(years)
        ax1.xaxis.set_minor_locator(months)
        ax1.xaxis.set_major_formatter(years_fmt)
        ax1.xaxis.set_minor_formatter(months_fmt)
        plt.setp(ax1.xaxis.get_minorticklabels(), rotation=45, size=12)
        plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, weight="bold", size=14)
    else:
        ax1.xaxis.set_major_locator(months)
        ax1.xaxis.set_minor_locator(weeks)
        ax1.xaxis.set_major_formatter(months_fmt)
        ax1.xaxis.set_minor_formatter(weeks_fmt)
        plt.setp(ax1.xaxis.get_minorticklabels(), rotation=45, size=12)
        plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, weight="bold", size=14)

    if "show_keys" in kwargs:
        if kwargs["show_keys"] == "no":
            ax1.get_yaxis().set_visible(False)

    # Add line of analysis_engine date, but only if in the time period
    if "blue_line" in kwargs:
        blue_line = kwargs["blue_line"]
        if blue_line == "today":
            if min_date <= datetime.today().date() <= max_date:
                plt.axvline(datetime.today().date())
        if blue_line == "config_date":
            board_date = get_board_date(kwargs)
            if min_date <= board_date <= max_date:
                plt.axvline(board_date)
    # size of chart and fit
    fig.tight_layout(rect=[0, 0.03, 1, 0.95])  # for title

    if kwargs["chart"] != "save":
        plt.show()

    return fig


def put_milestones_into_wb(milestones: MilestoneData) -> Workbook:
    wb = Workbook()
    ws = wb.active

    row_num = 2
    ms_names = milestones.sorted_milestone_dict[milestones.quarters[0]]["names"]

    for i, m in enumerate(ms_names):
        for x, tp in enumerate(milestones.quarters):
            ms_date = milestones.sorted_milestone_dict[tp]["r_dates"][i]
            ws.cell(row=row_num + i, column=5 + x).value = ms_date
            ws.cell(row=row_num + i, column=5 + x).number_format = "dd/mm/yy"

    for i, m in enumerate(ms_names):  # want the latest notes.
        ws.cell(row=row_num + i, column=1).value = milestones.sorted_milestone_dict[
            milestones.quarters[0]
        ]["project"][
            i
        ]  # project name
        ws.cell(row=row_num + i, column=2).value = milestones.sorted_milestone_dict[
            milestones.quarters[0]
        ]["report"][
            i
        ]  # milestone
        ws.cell(row=row_num + i, column=3).value = milestones.sorted_milestone_dict[
            milestones.quarters[0]
        ]["cat"][
            i
        ]  # milestone
        ws.cell(row=row_num + i, column=4).value = milestones.sorted_milestone_dict[
            milestones.quarters[0]
        ]["names"][
            i
        ]  # milestone
        try:
            notes = milestones.sorted_milestone_dict[milestones.quarters[0]]["notes"][i]
            ws.cell(row=row_num + i, column=len(milestones.quarters) + 9).value = notes
        except KeyError:  # "notes" not in central support dict
            pass
        try:  # only present in top250 milestones
            status = milestones.sorted_milestone_dict[milestones.quarters[0]]["status"][
                i
            ]
            ws.cell(row=row_num + i, column=len(milestones.quarters) + 5).value = status
        except (IndexError, KeyError):
            # IndexError for ipdc rpting. "status" is in dict, but list is empty.
            # KeyError for top250 central support, which does not have "status" in dict.
            pass
        try:  # only present in top250 central support
            escalated = milestones.sorted_milestone_dict[milestones.quarters[0]][
                "escalated"
            ][i]
            cs_type = milestones.sorted_milestone_dict[milestones.quarters[0]]["type"][
                i
            ]
            cs_response = milestones.sorted_milestone_dict[milestones.quarters[0]][
                "cs_response"
            ][i]
            secured = milestones.sorted_milestone_dict[milestones.quarters[0]][
                "secured"
            ][i]
            ws.cell(
                row=row_num + i, column=len(milestones.quarters) + 6
            ).value = escalated
            ws.cell(
                row=row_num + i, column=len(milestones.quarters) + 7
            ).value = cs_type
            ws.cell(
                row=row_num + i, column=len(milestones.quarters) + 8
            ).value = secured
        except KeyError:  # all above not in ipdc or top250 milestones
            pass

    ws.cell(row=1, column=1).value = "Project"
    ws.cell(row=1, column=2).value = "Report"
    ws.cell(row=1, column=3).value = "Type"
    ws.cell(row=1, column=4).value = "Milestone"
    for x, tp in enumerate(milestones.quarters):
        ws.cell(row=1, column=5 + x).value = tp
    ws.cell(row=1, column=len(milestones.quarters) + 5).value = "Status (top 250 ms)"
    ws.cell(row=1, column=len(milestones.quarters) + 6).value = "Escalated (top 250 cs)"
    ws.cell(row=1, column=len(milestones.quarters) + 7).value = "Type (top 250 cs)"
    ws.cell(row=1, column=len(milestones.quarters) + 8).value = "Secured (top 250 cs)"
    ws.cell(
        row=1, column=len(milestones.quarters) + 9
    ).value = "Notes / Central Response (top 250 cs)"

    return wb
