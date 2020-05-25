from openpyxl import load_workbook
from analysis.data import list_of_masters_all, root_path, red_text, a14

def get_data(wb):
    '''Function that takes data from a wb across all ws in use and places them into
     a python dictionary
     wb: workbook containing the data
     returns and python dictionary'''

    bl_dict = {}

    for name in (list_of_masters_all[0].projects):
        '''worksheet is created for each project'''
        try:
            ws = wb[name[0:29]]  # opening project ws

            other_dict = {}

            for x in range(2, ws.max_column+1):
                lower_dictionary = {}
                quarter = ws.cell(row=1, column=x).value
                for i in range(2, ws.max_row):
                    value = ws.cell(row=i, column=x).value
                    key = ws.cell(row=i, column=1).value
                    lower_dictionary[key] = value

                other_dict[quarter] = lower_dictionary

            bl_dict[name] = other_dict

        except KeyError:
            pass

    return bl_dict

def place_in_masters(project_list, master_wb_title_list, baseline_data):
    '''
    Places data into masters and removes unnecessary data by putting in blanks.
    :param project_list: list of project names
    :param master_wb_title_list: list of quarter master data workbooks. This is different from the master list of
    python dictionaries usually passed into functions.
    :param baseline_data: list of keys that need to be changed.
    :return: the wbs in the master_wb_title_list with the data amended.
    '''

    wb_key_list = ['IPDC approval point',
                   'Re-baseline this quarter',
                   'Re-baseline ALB/Programme milestones',
                   'Re-baseline ALB/Programme cost',
                   'Re-baseline ALB/Programme benefits',
                   'Re-baseline IPDC milestones',
                   'Re-baseline IPDC cost',
                   'Re-baseline IPDC benefits',
                   'Re-baseline HMT milestones',
                   'Re-baseline HMT cost',
                   'Re-baseline HMT benefits']

    for name in project_list:
        for master in master_wb_title_list:
            wb = load_workbook(master)
            ws = wb.active
            for col_num in range(2, ws.max_column + 1):
                project_name = ws.cell(row=1, column=col_num).value
                if project_name == name:
                    print(name)
                    project_bl_dict = baseline_data[name]
                    for row_num in range(2, ws.max_row + 1):
                        if ws.cell(row=row_num, column=1).value == 'Reporting period (GMPP - Snapshot Date)':
                            quarter = ws.cell(row=row_num, column=col_num).value
                            project_bl_dict_keys = list(project_bl_dict[quarter].keys())

                    for i, dict_key in enumerate(project_bl_dict_keys):
                        for row_num in range(2, ws.max_row + 1):
                            wb_key = ws.cell(row=row_num, column=1).value
                            if wb_key == wb_key_list[i]:
                                ws.cell(row=row_num, column=col_num).value = project_bl_dict[quarter][dict_key]
                                ws.cell(row=row_num, column=col_num).font = red_text

                else:
                    pass

            wb.save(master)

def check_keys():
    '''small function to check if masters have the baseline keys. to delete once masters updated'''
    for master in list_of_masters_all:
        q_data = master.data[a14]
        print(q_data['Reporting period (GMPP - Snapshot Date)'])
        q_data_keys = list(q_data.keys())
        if 'Re-baseline ALB/Programme milestones' in q_data_keys:
            print('Yes')

def change_quarter_string():

    for master in bl_data_list:
        wb = load_workbook(master)
        for ws in wb.worksheets:
            ws.cell(row=1, column=6).value = 'Q4 1819'
            ws.cell(row=1, column=6).font = red_text
            ws.cell(row=1, column=7).value = 'Q3 Oct - Dec 2018'
            ws.cell(row=1, column=7).font = red_text
            ws.cell(row=1, column=13).value = 'Q1 Apr - Jun 2017'
            ws.cell(row=1, column=13).font = red_text
            ws.cell(row=1, column=14).value = 'Q4 Jan - Mar 2017'
            ws.cell(row=1, column=14).font = red_text
        wb.save(master)


master_list = [root_path/'core_data/master_4_2019.xlsx',
               root_path/'core_data/master_3_2019.xlsx',
               root_path/'core_data/master_2_2019.xlsx',
               root_path/'core_data/master_1_2019.xlsx',
               root_path/'core_data/master_4_2018.xlsx',
               root_path/'core_data/master_3_2018.xlsx',
               root_path/'core_data/master_2_2018.xlsx',
               root_path/'core_data/master_1_2018.xlsx',
               root_path/'core_data/master_4_2017.xlsx',
               root_path/'core_data/master_3_2017.xlsx',
               root_path/'core_data/master_2_2017.xlsx',
               root_path/'core_data/master_1_2017.xlsx',
               root_path/'core_data/master_4_2016.xlsx',
               root_path/'core_data/master_3_2016.xlsx']

bl_data_list = [root_path / 'input/baseline_info_HSMRPG_projects.xlsx',
                    root_path / 'input/baseline_info_amis_projects.xlsx',
                    root_path / 'input/baseline_info_FTTS.XLSX',
                    root_path / 'input/baseline_info_HE_projects.xlsx',
                    root_path / 'input/baseline_info_rail.xlsx']

baseline_data_wb = load_workbook(root_path / 'input/baseline_info_HSMRPG_projects.xlsx')

bl_data = get_data(baseline_data_wb)

place_in_masters(bl_data.keys(), master_list, bl_data)

#check_keys()

#change_quarter_string()