'''code that creates wb with all project bl meta data point'''


from openpyxl import Workbook
from analysis.data import list_of_masters_all, root_path, baseline_bc_stamp
from analysis.engine_functions import get_quarter_stamp, baseline_information

def get_baseline(project_name_list, baseline_data, comp_baseline_list):
    '''functions places all bl data into a wb'''

    wb = Workbook()

    for i, name in enumerate(project_name_list):
        '''worksheet is created for each project'''
        ws = wb.create_sheet(name[0:29], i)  # creating worksheets
        ws.title = name[0:29]  # title of worksheet

        for data in (baseline_data[name]):
            column_index = data[2]
            bc_stage = data[0]
            ws.cell(row=2, column=column_index+2, value=bc_stage)

        for x, baseline_type in enumerate(comp_baseline_list):
            for albm in (baseline_type[name]):
                column_index = albm[2]
                bc_stage = albm[0]
                ws.cell(row=x+3, column=column_index+2, value=bc_stage)

        quarter_labels = get_quarter_stamp(list_of_masters_all)
        for l, label in enumerate(quarter_labels):
            ws.cell(row=1, column=l+2, value=label)

        for n, key in enumerate(baseline_name_list[1:]):
            ws.cell(row=n+4, column=1, value=key)

        ws.cell(row=1, column=1, value='Quarter')
        ws.cell(row=2, column=1, value='IPDC BC approval')
        ws.cell(row=3, column=1, value='Re-baselined in quarter')
        ws.cell(row=13, column=1, value='Notes')

    return wb

baseline_name_list = ['this quarter',
                      'ALB/Programme milestones',
                      'ALB/Programme cost',
                      'ALB/Programme benefits',
                      'IPDC milestones',
                      'IPDC cost',
                      'IPDC benefits',
                      'HMT milestones',
                      'HMT cost',
                      'HMT benefits']

all_baselines_list = []
for x in baseline_name_list:
    baselines = baseline_information(list_of_masters_all[0].projects, list_of_masters_all, x)
    all_baselines_list.append(baselines)

'''Running the programme'''
'''output one - all data'''
run = get_baseline(list_of_masters_all[0].projects, baseline_bc_stamp, all_baselines_list)

'''Specify name of the output document here'''
run.save(root_path/'output/baseline_info.xlsx')
