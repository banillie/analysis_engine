"""
Returns list of project values from a single master for specified keys of interest. Return is contained within one wb.
Code can handle both standard and project milestone keys, as well as project name lists across
multiple quarters.

There is one outputs.
1) wb containing all values.

Conditional formatting is placed in the files as follows:
rag_rating colours
missing data (md) = black grey
project not reporting (pnr) = light grey
keys not collected (knc) = light blue grey
"""


from openpyxl import Workbook
from analysis.data import list_of_masters_all, root_path, gen_txt_list, \
    gen_txt_colours, gen_fill_colours, list_column_ltrs, list_of_rag_keys, rag_txt_list_full, \
    rag_fill_colours, rag_txt_colours, salmon_fill, stakeholders
from analysis.engine_functions import all_milestone_data_bulk, conditional_formatting

def return_data(data_key_list, quarter_master):
    """
    Returns all data of interest into a excel wb.
    master: excel wb master from which data should be taken.
    data_key_list: list of data keys for which values should be returned.
    """

    wb = Workbook()
    ws = wb.active

    for i in range(len(list_of_masters_all)):
        if quarter_master == str(list_of_masters_all[i].quarter):
            master = list_of_masters_all[i]

            '''list project names, groups and stage in ws'''
            for y, project_name in enumerate(master.projects):

                group = master.data[project_name]['DfT Group']

                ws.cell(row=2 + y, column=1, value=group) # group info
                ws.cell(row=2 + y, column=2, value=project_name)  # project name returned

                for x, key in enumerate(data_key_list):
                    ws.cell(row=1, column=3 + x, value=key)
                    try: # standard keys
                        value = master.data[project_name][key]
                        if value is None:
                            ws.cell(row=2 + y, column=3 + x).value = 'md'
                        else:
                            ws.cell(row=2 + y, column=3 + x, value=value)
                        try:  # checks for change against last quarter
                            lst_value = list_of_masters_all[i+1].data[project_name][key]
                            if value != lst_value:
                                ws.cell(row=2 + y, column=3 + x).fill = salmon_fill
                        except (KeyError, IndexError):
                            pass
                    except KeyError:
                        try: # milestone keys
                            milestones = all_milestone_data_bulk([project_name], master)
                            value = tuple(milestones[project_name][key])[0]
                            if value is None:
                                ws.cell(row=2 + y, column=3 + x).value = 'md'
                            else:
                                ws.cell(row=2 + y, column=3 + x).value = value
                                ws.cell(row=2 + y, column=3 + x).number_format = 'dd/mm/yy'
                            try:  # loop checks if value has changed since last quarter
                                old_milestones = all_milestone_data_bulk([project_name], list_of_masters_all[i+1])
                                lst_value = tuple(old_milestones[project_name][key])[0]
                                if value != lst_value:
                                    ws.cell(row=2 + y, column=3 + x).fill = salmon_fill
                            except (KeyError, IndexError):
                                pass
                        except KeyError: # exception catches both standard and milestone keys
                            ws.cell(row=2 + y, column=3 + x).value = 'knc'
                        except TypeError:
                            ws.cell(row=2 + y, column=3 + x).value = 'pnr'

            for z, key in enumerate(data_key_list):
                if key in list_of_rag_keys:
                    conditional_formatting(ws, [list_column_ltrs[z+2]], rag_txt_list_full, rag_txt_colours, rag_fill_colours,
                                           '1', '60') # plus 2 in column ltrs as values start being placed in at col 2.
            '''quarter tag information'''
            ws.cell(row=1, column=1, value='Group')
            ws.cell(row=1, column=2, value='Projects')

            conditional_formatting(ws, list_column_ltrs, gen_txt_list, gen_txt_colours, gen_fill_colours, '1', '60')

    return wb

# Instructions for running the programme

# Place all keys of interest as stings in to a list
# or use one of the imported lists from the data file
data_interest = ['IPDC approval point',
                 'Total Forecast',
                 'SOBC - IPDC Approval',
                 'OBC - IPDC Approval',
                 'FBC - IPDC Approval',
                 'Start of Construction/build',
                 'Start of Operation',
                 'Full Operations',
                 'Project End Date']

# Specify which quarter to take the data from
# format 'quarter fy' e.g. Q1 20/21
quarter = 'Q2 20/21'

# Default name for output file is s_data_query_output.
# See general guidance re saving output files names
# if other name to be used.
run = return_data(stakeholders, quarter)
run.save(root_path/'output/major_project_stakeholders.xlsx')






