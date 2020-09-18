#  Places vfm data into an excel wb. Three different methods can be used.
#  db options are the fastest.

from openpyxl import Workbook
from data_mgmt.data import get_master_data, get_current_project_names, root_path
from vfm.database import convert_db_python_dict, get_project_names
import sqlite3


#  Places data in excel wb. Using python dictionary structure.
def compile_data(masters, project_name_list):
    wb = Workbook()
    ws = wb.active

    for row, project in enumerate(project_name_list):
        i = row + 2  # has to start with row 1 in excel
        ws.cell(row=i, column=1).value = masters[0].data[project]['DfT Group']
        ws.cell(row=i, column=2).value = project
        ws.cell(row=i, column=3).value = masters[0].data[project]['NPV for all projects ' \
                                                                  'and NPV for programmes if available']
        try:
            ws.cell(row=i, column=4).value = masters[1].data[project]['NPV for all projects ' \
                                                                      'and NPV for programmes if available']
        except KeyError:
            ws.cell(row=i, column=4).value = 'Not reporting'
        ws.cell(row=i, column=5).value = masters[0].data[project]['Adjusted Benefits Cost Ratio (BCR)']
        try:
            ws.cell(row=i, column=6).value = masters[1].data[project]['Adjusted Benefits Cost Ratio (BCR)']
        except KeyError:
            ws.cell(row=i, column=6).value = 'Not reporting'
        ws.cell(row=i, column=7).value = masters[0].data[project]['Initial Benefits Cost Ratio (BCR)']
        try:
            ws.cell(row=i, column=8).value = masters[1].data[project]['Initial Benefits Cost Ratio (BCR)']
        except KeyError:
            ws.cell(row=i, column=8).value = 'Not reporting'
        ws.cell(row=i, column=9).value = masters[0].data[project]['VfM Category single entry']
        try:
            ws.cell(row=i, column=10).value = masters[1].data[project]['VfM Category single entry']
        except KeyError:
            ws.cell(row=i, column=10).value = 'Not reporting'
        ws.cell(row=i, column=11).value = masters[0].data[project]['Present Value Cost (PVC)']
        try:
            ws.cell(row=i, column=12).value = masters[1].data[project]['Present Value Cost (PVC)']
        except KeyError:
            ws.cell(row=i, column=12).value = 'Not reporting'
        ws.cell(row=i, column=13).value = masters[0].data[project]['Present Value Benefit (PVB)']
        try:
            ws.cell(row=i, column=14).value = masters[1].data[project]['Present Value Benefit (PVB)']
        except KeyError:
            ws.cell(row=i, column=14).value = 'Not reporting'

    ws.cell(row=1, column=1).value = 'DfT Group'
    ws.cell(row=1, column=2).value = 'Project'
    ws.cell(row=1, column=3).value = 'NPV'
    ws.cell(row=1, column=4).value = 'NPV lst qrt'
    ws.cell(row=1, column=5).value = 'Adjusted BCR'
    ws.cell(row=1, column=6).value = 'Adjusted BCR lst qrt'
    ws.cell(row=1, column=7).value = 'Initial BCR'
    try:
        ws.cell(row=i, column=8).value = masters[1].data[project]['Initial Benefits Cost Ratio (BCR)']
    except KeyError:
        ws.cell(row=i, column=8).value = 'Not reporting'
    ws.cell(row=i, column=9).value = masters[0].data[project]['VfM Category single entry']
    try:
        ws.cell(row=i, column=10).value = masters[1].data[project]['VfM Category single entry']
    except KeyError:
        ws.cell(row=i, column=10).value = 'Not reporting'
    ws.cell(row=i, column=11).value = masters[0].data[project]['Present Value Cost (PVC)']
    try:
        ws.cell(row=i, column=12).value = masters[1].data[project]['Present Value Cost (PVC)']
    except KeyError:
        ws.cell(row=i, column=12).value = 'Not reporting'
    ws.cell(row=i, column=13).value = masters[0].data[project]['Present Value Benefit (PVB)']
    try:
        ws.cell(row=i, column=14).value = masters[1].data[project]['Present Value Benefit (PVB)']
    except KeyError:
        ws.cell(row=i, column=14).value = 'Not reporting'

    return wb


#  Places data in excel wb. Using python dictionary structure via sqlite db
def compile_data_db(masters, project_name_list):
    wb = Workbook()
    ws = wb.active

    for row, project in enumerate(project_name_list):
        i = row + 2  # has to start with row 1 in excel. data entered into second row
        # print(masters)
        # m = masters['q1_2021']
        # print(m)

        ws.cell(row=i, column=1).value = masters['q1_2021'][project]['project_group']
        ws.cell(row=i, column=2).value = project
        ws.cell(row=i, column=3).value = masters['q1_2021'][project]['npv']
        try:
            ws.cell(row=i, column=4).value = masters['q4_1920'][project]['npv']
        except KeyError:
            ws.cell(row=i, column=4).value = 'Not reporting'
        ws.cell(row=i, column=5).value = masters['q1_2021'][project]['adjusted_bcr']
        try:
            ws.cell(row=i, column=6).value = masters['q4_1920'][project]['adjusted_bcr']
        except KeyError:
            ws.cell(row=i, column=6).value = 'Not reporting'
        ws.cell(row=i, column=7).value = masters['q1_2021'][project]['initial_bcr']
        try:
            ws.cell(row=i, column=8).value = masters['q4_1920'][project]['initial_bcr']
        except KeyError:
            ws.cell(row=i, column=8).value = 'Not reporting'
        ws.cell(row=i, column=9).value = masters['q1_2021'][project]['vfm_cat_single']
        try:
            ws.cell(row=i, column=10).value = masters['q4_1920'][project]['vfm_cat_single']
        except KeyError:
            ws.cell(row=i, column=10).value = 'Not reporting'
        ws.cell(row=i, column=11).value = masters['q1_2021'][project]['pvc']
        try:
            ws.cell(row=i, column=12).value = masters['q4_1920'][project]['pvc']
        except KeyError:
            ws.cell(row=i, column=12).value = 'Not reporting'
        ws.cell(row=i, column=13).value = masters['q1_2021'][project]['pvb']
        try:
            ws.cell(row=i, column=14).value = masters['q4_1920'][project]['pvb']
        except KeyError:
            ws.cell(row=i, column=14).value = 'Not reporting'

    ws.cell(row=1, column=1).value = 'DfT Group'
    ws.cell(row=1, column=2).value = 'Project'
    ws.cell(row=1, column=3).value = 'NPV'
    ws.cell(row=1, column=4).value = 'NPV lst qrt'
    ws.cell(row=1, column=5).value = 'Adjusted BCR'
    ws.cell(row=1, column=6).value = 'Adjusted BCR lst qrt'
    ws.cell(row=1, column=7).value = 'Initial BCR'
    ws.cell(row=1, column=8).value = 'Initial BCR lst qrt'
    ws.cell(row=1, column=9).value = 'VfM Category'
    ws.cell(row=1, column=10).value = 'VfM Category lst qrt'
    ws.cell(row=1, column=11).value = 'PVC'
    ws.cell(row=1, column=12).value = 'PVC lst qrt'
    ws.cell(row=1, column=13).value = 'PVB'
    ws.cell(row=1, column=14).value = 'PVB lst qrt'

    return wb


# Places data in excel wb straight from sqlite db.
def compile_data_pure_db(db_name, project_names, key_list, column_index):
    conn = sqlite3.connect(db_name + '.db')
    c = conn.cursor()

    wb = Workbook()
    ws = wb.active

    for i, p in enumerate(project_names):
        row = i + 2
        project_name = p
        ws.cell(row=row, column=2).value = project_name
        for x, key in enumerate(key_list[2:]):
            c.execute("SELECT {key} FROM q1_2021 WHERE "
                      "project_name = '{pn}'".format(key=key, pn=str(project_name)))
            vfm_db_q1_2021 = c.fetchone()
            # print(vfm_db_q1_2021[0])
            try:
                ws.cell(row=row, column=column_index[x][0]).value = vfm_db_q1_2021[0]
            except TypeError:
                pass
            c.execute("SELECT {key} FROM q4_1920 WHERE "
                      "project_name = '{pn}'".format(key=key, pn=str(project_name)))
            vfm_db_q4_1920 = c.fetchone()
            try:
                ws.cell(row=row, column=column_index[x][1]).value = vfm_db_q4_1920[0]
            except TypeError:
                pass

    conn.commit()
    conn.close()

    return wb


#  places vfm category figures into excel wb
def compile_vfm_cat_data_db(masters, cat_list):
    wb = Workbook()
    ws = wb.active

    for row, cat in enumerate(cat_list):
        i = row + 2  # has to start with row 1 in excel. data entered into second row
        for col, m in enumerate(masters):
            counter = 0
            total_projects = 0
            project_name = list(masters[m].keys())
            for p in project_name:
                if masters[m][p]['vfm_cat_single'] == cat:
                    counter += 1
                total_projects += 1

            ws.cell(row=1, column=2 + col).value = m  # values entered from col 2 onwards
            ws.cell(row=i, column=2 + col).value = counter  # values entered from col 2 onwards
            ws.cell(row=len(cat_list) + 3, column=2 + col).value = total_projects

        if cat is not None:
            ws.cell(row=i, column=1).value = cat
        else:
            ws.cell(row=i, column=1).value = 'None'

        ws.cell(row=len(cat_list) + 3, column=1).value = 'Total'

    return wb


def calculate_pvc(masters, cat_list):
    wb = Workbook()
    ws = wb.active

    row = 2
    for x, cat in enumerate(cat_list):
        for i, m in enumerate(masters):
            total = 0
            projects = (masters[m].keys())
            for p in projects:
                if masters[m][p]['vfm_cat_single'] == cat:
                    value = masters[m][p]['pvc']
                    if value is not None:
                        total += value
                    else:
                        pass
                else:
                    pass

            ws.cell(row=row + x, column=i + 2).value = total
            ws.cell(row=1, column=i + 2).value = m
        ws.cell(row=row + x, column=1).value = cat

    row = 10
    for i, m in enumerate(masters):
        hs2_total = 0
        total = 0
        other_total = 0
        projects = list(masters[m].keys())

        for p in projects:
            value = masters[m][p]['pvc']
            if value is not None:
                total += value
                if 'HS2 P' in p:
                    hs2_total += value
                else:
                    other_total += value
            else:
                pass

        ws.cell(row=row + 2, column=i + 2).value = hs2_total
        ws.cell(row=row + 3, column=i + 2).value = other_total
        ws.cell(row=row + 4, column=i + 2).value = total
        ws.cell(row=row + 1, column=i + 2).value = m

    ws.cell(row=row + 2, column=1).value = 'HS2'
    ws.cell(row=row + 3, column=1).value = 'Other'
    ws.cell(row=row + 4, column=1).value = 'Total'

    row = 16

    for i, m in enumerate(masters):
        high_total = 0
        poor_total = 0
        high_total_no_hs2 = 0
        poor_total_no_hs2 = 0
        projects = list(masters[m].keys())

        try:
            projects.remove('High Speed Rail Programme (HS2)') # stop double counting
        except ValueError:
            pass

        for cat in cat_list:
            for p in projects:
                if masters[m][p]['vfm_cat_single'] == cat:
                    value = masters[m][p]['pvc']
                    if value is not None:
                        if cat in ['Poor', 'Low', 'Medium']:
                            poor_total += value
                            if 'HS2 P' not in p:
                                poor_total_no_hs2 += value
                        if cat in ['High', 'Very High']:
                            high_total += value
                            if 'HS2 P' not in p:
                                high_total_no_hs2 += value
                    else:
                        pass

        ws.cell(row=row + 2, column=i + 2).value = poor_total
        ws.cell(row=row + 3, column=i + 2).value = poor_total_no_hs2
        ws.cell(row=row + 4, column=i + 2).value = high_total
        ws.cell(row=row + 5, column=i + 2).value = high_total_no_hs2
        ws.cell(row=row + 1, column=i + 2).value = m

    ws.cell(row=row + 2, column=1).value = 'total poor-medium'
    ws.cell(row=row + 3, column=1).value = 'poor-medium excluding hs2'
    ws.cell(row=row + 4, column=1).value = 'total high-very high'
    ws.cell(row=row + 5, column=1).value = 'high-very high excluding hs2'

    return wb


#  METHOD USES ALL PYTHON DICTIONARIES
# master_data = get_master_data()
# current_project_name_list = get_current_project_names()
#
# run = compile_data(master_data, current_project_name_list)
# run.save(root_path / "output/vfm_data_output_dict_way.xlsx")

#  METHOD USES SQLITE DB AND PYTHON DICTIONARIES
# q_list = ['q1_2021', 'q4_1920']
# master_dict = convert_db_python_dict('vfm', q_list)
# project_names = get_project_names('vfm', 'q1_2021')
#
# run = compile_data_db(master_dict, project_names)
# run.save(root_path / "output/vfm_data_output_db_dict_way.xlsx")


#  METHOD USES SQLTE DB ONLY
# vfm_key_list = ['project_name text',
#                 'project_group text',
#                 'npv real',
#                 'adjusted_bcr real',
#                 'initial_bcr real',
#                 'vfm_cat_single text',
#                 'pvc real',
#                 'pvb real']
# p_names = get_project_names('vfm', 'q1_2021')
# c_index = [(3, 4), (5, 6), (7, 8), (9, 10), (11, 12), (13, 14), (15, 16)] # column index
#
# run = compile_data_pure_db('vfm', p_names, vfm_key_list, c_index)
# run.save(root_path / "output/vfm_data_output_db_way.xlsx")


#  COMPILE VFM CAT DATA
ordered_cat_list = ['Poor', 'Low', 'Medium', 'High', 'Very High',
                    'Very High and Financially Positive', 'Economically Positive',
                    None]
q_list = ['q1_2021', 'q4_1920']
master_data = convert_db_python_dict('vfm', q_list)
project_names = get_project_names('vfm', 'q1_2021')
# run = compile_vfm_cat_data_db(master_data, ordered_cat_list)
# run.save(root_path / "output/vfm_cat_count.xlsx")
run = calculate_pvc(master_data, ordered_cat_list)
run.save(root_path / "output/pvc_count.xlsx")
