#import sqlite3
from vfm.database import create_connect_db, get_vfm_values, get_quarter_values, \
    create_vfm_table
from data_mgmt.data import get_master_data, root_path
from openpyxl import Workbook

conn = create_connect_db('vfm')
c = conn.cursor()

# m = get_master_data()
# vfm_db_list = get_vfm_values(m)
# vfm_q1_2021 = get_quarter_values(vfm_db_list, "Q1 20/21")
# vfm_q4_1920 = get_quarter_values(vfm_db_list, "Q4 19/20")


vfm_key_list = ['project_name text',
            'project_group text',
            'npv real',
            'adjusted_bcr real',
            'initial_bcr real',
            'vfm_cat_single text',
            'pvc real',
            'pvb real']

def get_project_names(conn, quarter):
    c = conn.cursor()
    c.execute("SELECT project_name FROM '{table}'".format(table=quarter))
    names = c.fetchall()
    return names


p_names = get_project_names(conn, 'q1_2021')


def place_vfm_excel(project_names, key_list):
    wb = Workbook()
    ws = wb.active

    for i, p in enumerate(project_names):
        row = i+2
        project_name = p[0]
        ws.cell(row=row, column=2).value = project_name
        for x, key in enumerate(key_list[1:]):
            col = x+2
            c.execute("SELECT {key} FROM q1_2021 WHERE "
                      "project_name = '{pn}'".format(key=key, pn=str(project_name)))
            vfm_db_q1_2021 = c.fetchone()
            ws.cell(row=row, column=col+1).value = vfm_db_q1_2021[0]
            c.execute("SELECT {key} FROM q4_1920 WHERE "
                      "project_name = '{pn}'".format(key=key, pn=str(project_name)))
            vfm_db_q4_1920 = c.fetchone()
            try:
                ws.cell(row=row, column=col*2).value = vfm_db_q4_1920[0]
            except TypeError:
                pass

    return wb

run = place_vfm_excel(p_names, vfm_key_list)
run.save(root_path / "output/vfm_data_output_other_way.xlsx")

conn.commit()
conn.close()
