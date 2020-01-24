'''

This programme creates a master spreadsheet to share with IPA for gmpp reporting. The 'master' print out is then
shared with the IPA which runs an excel macro to populate individual gmpp reporting templates.

Documents required to run the programme are set out below.

Documents required are:
1) the gmpp datamap (make sure you have the latest/final version).
2) The latest quarter DfT_master spreadsheet (i.e. the quarter that is being reported).

IMPORTANT to note:
- Handling of Hs2 data currently not done by this programme.

'''

from openpyxl import load_workbook
from analysis.data import q3_1920, root_path, hs2_programme, tru
from analysis.engine_functions import filter_gmpp


def create_master(gmpp_wb, master_data):
    ws = gmpp_wb.active

    type_list = ['RDEL', 'CDEL', 'Non-Gov', 'Income'] # list of cost types. used to amend Hs2 data
    zero_list = ['RDEL', 'CDEL', 'Non-Gov', 'Income', 'BEN'] # list of cost/ben types. used to remove none value entries

    # this section filters out only gmpp project names. Subsequent list is then used to populate ws
    gmpp_project_names = filter_gmpp(master_data)

    for i, project_name in enumerate([hs2_programme, tru]):
        print(project_name)
        ws.cell(row=1, column=6+i).value = project_name  # place project names in file

        # for loop for placing data into the worksheet
        keys_not_found = []
        for row_num in range(2, ws.max_row+1):
            key = ws.cell(row=row_num, column=1).value
            # this loop places all latest raw data into the worksheet
            if key in master_data.data[project_name].keys():
                ws.cell(row=row_num, column=6+i).value = master_data.data[project_name][key]
            elif key not in master_data.data[project_name].keys():
                keys_not_found.append(key)

                # this section of the code ensures that all financial costs / benefit forecasts have a zero
                for cost_type in zero_list:
                    if cost_type in key:
                        try:
                            if master_data.data[project_name][key] is None:
                                ws.cell(row=row_num, column=6 + i).value = 0
                        except KeyError:
                            keys_not_found.append(key)

        #This is where loop to amend Hs2 data could be placed

    print(keys_not_found)

    return gmpp_wb

latest_dm = load_workbook(root_path/'input/gmpp_datamap_master_q3_1920.xlsx')

run = create_master(latest_dm, q3_1920)

run.save(root_path/'output/hs2_tru_gmpp_dataset_q3_1920.xlsx')
