'''

Programme for compiling the IPA AR data dashboard.

Follow instructions below.

'''

from openpyxl import load_workbook
from datamaps.api import project_data_from_master
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, IconSet, FormatObject
from analysis.engine_functions import up_or_down

def cal_date_difference(milestone_date, old_milestone_date):
    '''
    Small helper function for this programme only. calculates the difference between two dates.
    :param milestone_date: latest date
    :param old_milestone_date: last year date
    :return: time delta difference between two dates
    '''
    try:
        time_delta = (milestone_date - old_milestone_date).days
    except TypeError:
        time_delta = 0
    return time_delta

def placing_excel(master_data_one, master_data_two):
    '''
    function that places all information into the summary dashboard sheet
    :param master_data_one: python dictionary of latest ar data.
    :param master_data_two: python dictionary of last ar data.
    :return: populated Excel dashboard.
    '''

    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=2).value
        print(project_name)
        if project_name in master_data_one.projects:
            dca_one = master_data_one.data[project_name]['DCA']
            try:
                dca_two = master_data_two.data[project_name]['DCA']
                change = up_or_down(dca_one, dca_two)
                ws.cell(row=row_num, column=4).value = change
            except KeyError:
                ws.cell(row=row_num, column=4).value = 'NEW'
            ws.cell(row=row_num, column=5).value = master_data_one.data[project_name]['DCA']

            start_date_one = master_data_one.data[project_name]['Start Date']
            ws.cell(row=row_num, column=6).value = start_date_one
            try:
                start_date_two = master_data_two.data[project_name]['Start Date']
                s_date_diff = cal_date_difference(start_date_one, start_date_two)
                ws.cell(row=row_num, column=7).value = s_date_diff
            except KeyError:
                ws.cell(row=row_num, column=7).value = 0

            end_date_one = master_data_one.data[project_name]['End Date']
            ws.cell(row=row_num, column=8).value = end_date_one
            try:
                end_date_two = master_data_two.data[project_name]['End Date']
                e_date_diff = cal_date_difference(end_date_one, end_date_two)
                ws.cell(row=row_num, column=9).value = e_date_diff
            except KeyError:
                ws.cell(row=row_num, column=9).value = 0

            ws.cell(row=row_num, column=10).value = master_data_one.data[project_name]['in year baseline']
            ws.cell(row=row_num, column=11).value = master_data_one.data[project_name]['in year forecast']
            ws.cell(row=row_num, column=12).value = master_data_one.data[project_name]['in year variance']
            wlc_one = master_data_one.data[project_name]['WLC baseline']
            ws.cell(row=row_num, column=13).value = wlc_one
            try:
                wlc_two = master_data_two.data[project_name]['WLC baseline']
                wlc_diff = wlc_one - wlc_two
                ws.cell(row=row_num, column=14).value = wlc_diff
            except KeyError:
                ws.cell(row=row_num, column=14).value = 0
            except TypeError:
                ws.cell(row=row_num, column=14).value = 'Check wlc value/data'


    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=2).value
        if project_name in master_data_two.data:
            ws.cell(row=row_num, column=3).value = master_data_two[project_name]['DCA']

    # Highlight cells that contain RAG text, with background and text the same colour. column E.
    ag_text = Font(color="00a5b700")
    ag_fill = PatternFill(bgColor="00a5b700")
    dxf = DifferentialStyle(font=ag_text, fill=ag_fill)
    rule = Rule(type="containsText", operator="containsText", text="Amber/Green", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Amber/Green",e1)))']
    ws.conditional_formatting.add('e1:e100', rule)

    ar_text = Font(color="00f97b31")
    ar_fill = PatternFill(bgColor="00f97b31")
    dxf = DifferentialStyle(font=ar_text, fill=ar_fill)
    rule = Rule(type="containsText", operator="containsText", text="Amber/Red", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Amber/Red",e1)))']
    ws.conditional_formatting.add('e1:e100', rule)

    red_text = Font(color="00fc2525")
    red_fill = PatternFill(bgColor="00fc2525")
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="containsText", operator="containsText", text="Red", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Red",E1)))']
    ws.conditional_formatting.add('E1:E100', rule)

    green_text = Font(color="0017960c")
    green_fill = PatternFill(bgColor="0017960c")
    dxf = DifferentialStyle(font=green_text, fill=green_fill)
    rule = Rule(type="containsText", operator="containsText", text="Green", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Green",e1)))']
    ws.conditional_formatting.add('E1:E100', rule)

    amber_text = Font(color="00fce553")
    amber_fill = PatternFill(bgColor="00fce553")
    dxf = DifferentialStyle(font=amber_text, fill=amber_fill)
    rule = Rule(type="containsText", operator="containsText", text="Amber", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Amber",e1)))']
    ws.conditional_formatting.add('e1:e100', rule)

    # Highlight cells that contain RAG text, with background and black text columns G to L.
    ag_text = Font(color="000000")
    ag_fill = PatternFill(bgColor="00a5b700")
    dxf = DifferentialStyle(font=ag_text, fill=ag_fill)
    rule = Rule(type="containsText", operator="containsText", text="Amber/Green", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Amber/Green",G1)))']
    ws.conditional_formatting.add('G1:L100', rule)

    ar_text = Font(color="000000")
    ar_fill = PatternFill(bgColor="00f97b31")
    dxf = DifferentialStyle(font=ar_text, fill=ar_fill)
    rule = Rule(type="containsText", operator="containsText", text="Amber/Red", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Amber/Red",G1)))']
    ws.conditional_formatting.add('G1:L100', rule)

    red_text = Font(color="000000")
    red_fill = PatternFill(bgColor="00fc2525")
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="containsText", operator="containsText", text="Red", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Red",G1)))']
    ws.conditional_formatting.add('G1:L100', rule)

    green_text = Font(color="000000")
    green_fill = PatternFill(bgColor="0017960c")
    dxf = DifferentialStyle(font=green_text, fill=green_fill)
    rule = Rule(type="containsText", operator="containsText", text="Green", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Green",G1)))']
    ws.conditional_formatting.add('G1:L100', rule)

    amber_text = Font(color="000000")
    amber_fill = PatternFill(bgColor="00fce553")
    dxf = DifferentialStyle(font=amber_text, fill=amber_fill)
    rule = Rule(type="containsText", operator="containsText", text="Amber", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Amber",G1)))']
    ws.conditional_formatting.add('G1:L100', rule)

    # highlighting new projects
    red_text = Font(color="00fc2525")
    white_fill = PatternFill(bgColor="000000")
    dxf = DifferentialStyle(font=red_text, fill=white_fill)
    rule = Rule(type="containsText", operator="containsText", text="NEW", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("NEW",F1)))']
    ws.conditional_formatting.add('D1:D100', rule)

    # assign the icon set to a rule
    first = FormatObject(type='num', val=-1)
    second = FormatObject(type='num', val=0)
    third = FormatObject(type='num', val=1)
    iconset = IconSet(iconSet='3Arrows', cfvo=[first, second, third], percent=None, reverse=None)
    rule = Rule(type='iconSet', iconSet=iconset)
    ws.conditional_formatting.add('D1:D100', rule)

    return wb

'''INSTRUCTIONS'''

'''1) Provide file path to empty master dashboard document. Make sure the dashboard is set-up/structured correctly'''
wb = load_workbook(
    'C:\\Users\\Standalone\\general\\masters folder\\ipa_annual_report\\ipa_annual_report_dashboard_master.xlsx')
ws = wb.active

'''2) Provide file paths to the master IAP AR data sets. Make sure the keys names are the same as last year'''
latest_ar_data = project_data_from_master(
    'C:\\Users\\Standalone\\general\\masters folder\\ipa_annual_report\\DfT AR 2019 Data.xlsx', 2, 2019)
last_ar_data = project_data_from_master(
    'C:\\Users\\Standalone\\general\\masters folder\\ipa_annual_report\\ipa_annual_report_2018.xlsx', 2, 2018)

wb = placing_excel(latest_ar_data, last_ar_data)

'''4) provide file path and specific name of output file.'''
wb.save('C:\\Users\\Standalone\\general\\ar_dashboard_test.xlsx')