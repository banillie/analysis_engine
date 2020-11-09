"""
This code changes key names in master wb documents
"""

import typing
from typing import Dict
from openpyxl import load_workbook
from openpyxl.workbook import workbook
from data_mgmt.data import YEAR_LIST, get_master_data_file_paths, root_path


def put_key_change_master_into_dict(key_change_file: typing.TextIO) -> Dict[str, str]:
    """
    places key information i.e. keys old and new names from wb into a python dictionary
    """
    wb = load_workbook(key_change_file)
    ws = wb.active

    output_dict = {}
    for x in range(1, ws.max_row + 1):
        key = ws.cell(row=x, column=1).value
        codename = ws.cell(row=x, column=2).value
        output_dict[key] = codename

    return output_dict


def alter_wb_master_file_key_names(master_file: typing.TextIO, key_change_dict: Dict[str, str]) -> workbook:
    """
    places altered key names, from the key change master dictionary, into master wb(s).
    """
    wb = load_workbook(master_file)
    ws = wb.active

    for row_num in range(2, ws.max_row + 1):
        for key in key_change_dict.keys():  # changes stored in the altered key change log wb
            if ws.cell(row=row_num, column=1).value == key:
                ws.cell(row=row_num, column=1).value = \
                    key_change_dict[key]
        for year in YEAR_LIST:  # changes to yearly profile keys
            if ws.cell(row=row_num, column=1).value == year + " CDEL Forecast Total":
                ws.cell(row=row_num, column=1).value = year + " CDEL Forecast one off new costs"

    return wb.save(master_file)


def run_change_keys(master_files_list: list, key_dict: Dict[str, str]) -> None:
    """
    runs code which replaces old key names with new names in master excel workbooks.
    """
    for f in master_files_list:
        alter_wb_master_file_key_names(f, key_dict)


keys_dict = put_key_change_master_into_dict(root_path / "core_data/data_mgmt/key_change_log_2.xlsx")
run_change_keys(get_master_data_file_paths(), keys_dict)
