"""
code in development which changes cost key names in masters.
"""

import typing
from typing import Dict

from openpyxl import load_workbook
from openpyxl.workbook import workbook

from data_mgmt.data import YEAR_LIST, get_key_change_log_file_path, get_master_data_file_paths, get_project_information

from datamaps.api import project_data_from_master


def put_key_change_master_into_dict(key_change_file: typing.TextIO) -> Dict[str, str]:
    """
    places key information i.e. keys old and new names from wb into a python dictionary
    """
    wb = load_workbook(key_change_file)
    ws = wb.active

    output_dict = {}
    for x in range(1, ws.max_row + 1):
        key = ws.cell(row=x, column=1).value
        codename = ws.cell(row=x, column=2).value
        output_dict[key] = codename

    return output_dict


def alter_wb_master_file_key_names(master_file: typing.TextIO, key_change_dict: Dict[str, str]) -> workbook:
    """
    places altered key names, from the key change master dictionary, into master wb(s).
    """
    wb = load_workbook(master_file)
    ws = wb.active

    for row_num in range(2, ws.max_row + 1):
        for key in key_change_dict.keys():  # changes stored in the altered key change log wb
            if ws.cell(row=row_num, column=1).value == key:
                ws.cell(row=row_num, column=1).value = \
                    key_change_dict[key]
        for year in YEAR_LIST:  # changes to yearly profile keys
            if ws.cell(row=row_num, column=1).value == year + " CDEL Forecast Total":
                ws.cell(row=row_num, column=1).value = year + " CDEL Forecast one off new costs"

    return wb.save(master_file)


def run_change_keys(master_files_list: list, key_dict: Dict[str, str]) -> None:
    """
    runs code which replaces old key names with new names in master excel workbooks.
    """
    for f in master_files_list:
        alter_wb_master_file_key_names(f, key_dict)


def get_old_fy_cost_data(master_file: typing.TextIO, project_id_wb: typing.TextIO) -> None:
    """
    Gets all old financial data across quarters and places into project id document.
    """
    master = project_data_from_master(master_file, 1, 2010)  # random year specified as not in use
    wb = load_workbook(project_id_wb)
    ws = wb.active

    for i in range(1, ws.max_column + 1):
        project_name = ws.cell(row=1, column=1 + i).value
        for row_num in range(2, ws.max_row + 1):
            key = ws.cell(row=row_num, column=1).value
            try:
                if key in master.data[project_name].keys():
                    ws.cell(row=row_num, column=1 + i).value = master[project_name][key]
            except KeyError:  # project might not be present in quarter
                pass

    wb.save(project_id_wb)


def run_get_old_fy_data(master_files_list: list, project_id_wb: typing.TextIO) -> None:
    for f in master_files_list:
        get_old_fy_cost_data(f, project_id_wb)


def place_old_fy_data_into_master_wb(master_file: typing.TextIO, project_id_wb: typing.TextIO) -> None:
    """
    places all old financial year data into master files.
    """
    id_master = project_data_from_master(project_id_wb, 2, 2020)  # random year specify as not used
    wb = load_workbook(master_file)
    ws = wb.active

    for i in range(1, ws.max_column + 1):
        project_name = ws.cell(row=1, column=1 + i).value
        for row_num in range(2, ws.max_row + 1):
            key = ws.cell(row=row_num, column=1).value
            try:
                if key in id_master.data[project_name].keys():
                    ws.cell(row=row_num, column=1 + i).value = id_master[project_name][key]
            except KeyError:  # project might not be present in quarter
                pass

    wb.save(master_file)


def run_place_old_fy_data_into_masters(master_files_list: list, project_id_wb: typing.TextIO) -> None:
    for f in master_files_list:
        place_old_fy_data_into_master_wb(f, project_id_wb)


# keys_dict = put_key_change_master_into_dict(get_key_change_log_file_path())
# run_change_keys(get_master_data_file_paths(), keys_dict)
# run_get_old_fy_data(get_master_data_file_paths(), get_project_information())
# run_place_old_fy_data_into_masters(get_master_data_file_paths(), get_project_information())

