'''

This programme creates a master spreadsheet for commissioning templates.

Documents required to run the programme are set out below. Make sure you are using the latest version of these
files.

Documents required are:
1) The internal commission datamap in excel file format. This is the dm with all keys necessary for master to
templates Datamap command.
2) Master data commission file. See note below
3) last quarter DfT_master spreadsheet. See note below.

Note that 2 and 3 need to be separate because they may have different data, if a project has changed the date it
reported after the DL for providing quarter data, it should go in the master commission file.

Note In this programme masters uploaded using the datamaps api rather than via the analysis.data file. This allows more
flexibility around which masters are being used. Which is useful for handling the master data commission file.

'''

from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from analysis.data import root_path
from datamaps.api import project_data_from_master


def create_master(workbook, project_name_list, commission_q_data, latest_q_data):
    ws = workbook.active
    type_list_2 = ['RDEL', 'CDEL', 'Non-Gov', 'Income', 'BEN'] # list of cost/ben types. used to remove none value entries

    red_text = Font(color="00fc2525")

    for i, name in enumerate(project_name_list):
        print(name)
        ws.cell(row=1, column=2+i).value = name  # place project names in file

        # for loop for placing data into the worksheet
        for row_num in range(2, ws.max_row+1):
            key = ws.cell(row=row_num, column=1).value

            if key in commission_q_data.data[name].keys():
                ws.cell(row=row_num, column=2+i).value = commission_q_data[name][key]
            if "lst qrt" in key:
                altered_lst_quarter = key.replace("lst qrt ", "")
                try:
                    ws.cell(row=row_num, column=2 + i).value = latest_q_data.data[name][altered_lst_quarter]
                except KeyError:
                    ws.cell(row=row_num, column=40).value = 'Couldnt match'


                '''handling of DN types'''
                if key == "lst qrt DN Type 1":
                    ws.cell(row=row_num, column=2 + i).value = dn_combine_text(latest_q_data.data[name],['DN Type 1', 'DN Description 1'])
                if key == "lst qrt DN Type 2":
                    ws.cell(row=row_num, column=2 + i).value = dn_combine_text(latest_q_data.data[name],['DN Type 2', 'DN Description 2'])
                if key == "lst qrt DN Type 3":
                    ws.cell(row=row_num, column=2 + i).value = dn_combine_text(latest_q_data.data[name],['DN Type 3', 'DN Description 3'])
                if key == "lst qrt DN Type 4":
                    ws.cell(row=row_num, column=2 + i).value = dn_combine_text(latest_q_data.data[name],['DN Type 4', 'DN Description 4'])
                if key == "lst qrt DN Type 5":
                    ws.cell(row=row_num, column=2 + i).value = dn_combine_text(latest_q_data.data[name],['DN Type 5', 'DN Description 5'])

                '''handling of strategic outcomes'''
                if key == "lst qrt List Strategic Outcomes (GMPP - Intended Outcome 1)":
                    ws.cell(row=row_num, column=2 + i).value = \
                        strategic_combine_text(latest_q_data.data[name], ['List Strategic Outcomes (GMPP - Intended Outcome 1)',
                                                                    'IO1', 'IO1 - Monetised?', 'IO1 PESTLE'])
                if key == "lst qrt List Strategic Outcomes (GMPP - Intended Outcome 2)":
                    ws.cell(row=row_num, column=2 + i).value = \
                        strategic_combine_text(latest_q_data.data[name], ['List Strategic Outcomes (GMPP - Intended Outcome 2)',
                                                                    'IO2', 'IO2 - Monetised?', 'IO2 PESTLE'])
                if key == "lst qrt List Strategic Outcomes (GMPP - Intended Outcome 3)":
                    ws.cell(row=row_num, column=2 + i).value = \
                        strategic_combine_text(latest_q_data.data[name], ['List Strategic Outcomes (GMPP - Intended Outcome 3)',
                                                                    'IO3', 'IO3 - Monetised?', 'IO3 PESTLE'])
                if key == "lst qrt List Strategic Outcomes (GMPP - Intended Outcome 4)":
                    ws.cell(row=row_num, column=2 + i).value = \
                        strategic_combine_text(latest_q_data.data[name], ['List Strategic Outcomes (GMPP - Intended Outcome 4)',
                                                                    'IO4', 'IO4 - Monetised?', 'IO4 PESTLE'])
                if key == "lst qrt List Strategic Outcomes (GMPP - Intended Outcome 5)":
                    ws.cell(row=row_num, column=2 + i).value = \
                        strategic_combine_text(latest_q_data.data[name], ['List Strategic Outcomes (GMPP - Intended Outcome 5)',
                                                                    'IO5', 'IO5 - Monetised?', 'IO5 PESTLE'])
                if key == "lst qrt List Strategic Outcomes (GMPP - Intended Outcome 6)":
                    ws.cell(row=row_num, column=2 + i).value = \
                        strategic_combine_text(latest_q_data.data[name], ['List Strategic Outcomes (GMPP - Intended Outcome 6)',
                                                                    'IO6', 'IO6 - Monetised?', 'IO6 PESTLE'])
                if key == "lst qrt List Strategic Outcomes (GMPP - Intended Outcome 7)":
                    ws.cell(row=row_num, column=2 + i).value = \
                        strategic_combine_text(latest_q_data.data[name], ['List Strategic Outcomes (GMPP - Intended Outcome 7)',
                                                                    'IO7', 'IO7 - Monetised?', 'IO7 PESTLE'])
                if key == "lst qrt List Strategic Outcomes (GMPP - Intended Outcome 8)":
                    ws.cell(row=row_num, column=2 + i).value = \
                        strategic_combine_text(latest_q_data.data[name], ['List Strategic Outcomes (GMPP - Intended Outcome 8)',
                                                                    'IO8', 'IO8 - Monetised?', 'IO8 PESTLE'])
                if key == "lst qrt List Strategic Outcomes (GMPP - Intended Outcome 8)":
                    ws.cell(row=row_num, column=2 + i).value = \
                        strategic_combine_text(latest_q_data.data[name], ['List Strategic Outcomes (GMPP - Intended Outcome 8)',
                                                                    'IO8', 'IO8 - Monetised?', 'IO8 PESTLE'])
                if key == "lst qrt List Strategic Outcomes (GMPP - Intended Outcome 9)":
                    ws.cell(row=row_num, column=2 + i).value = \
                        strategic_combine_text(latest_q_data.data[name], ['List Strategic Outcomes (GMPP - Intended Outcome 9)',
                                                                    'IO9', 'IO9 - Monetised?', 'IO9 PESTLE'])
                if key == "lst qrt List Strategic Outcomes (GMPP - Intended Outcome 10)":
                    ws.cell(row=row_num, column=2 + i).value = \
                        strategic_combine_text(latest_q_data.data[name], ['List Strategic Outcomes (GMPP - Intended Outcome 10)',
                                                                    'IO10', 'IO10 - Monetised?', 'IO10 PESTLE'])

                '''handling of investment objectives'''
                if key == "lst qrt Primary investment Objective":
                    ws.cell(row=row_num, column=2 + i).value = dn_combine_text(latest_q_data.data[name],
                                               ['Primary investment Objective', 'IO11 Monetised'])
                if key == "lst qrt Secondary investment Objective":
                    ws.cell(row=row_num, column=2 + i).value = dn_combine_text(latest_q_data.data[name],
                                               ['Secondary investment Objective', 'IO12 Monetised'])

                '''handling of risk descriptions'''
                if key == "lst qrt Brief Risk Decription 1":
                    ws.cell(row=row_num, column=2 + i).value = \
                        risk_combine_text(latest_q_data.data[name], [
                            'Brief Risk Decription 1', 'BRD 1Risk Category', 'BRD 1 Primary Risk to',
                            'BRD 1 Internal Control', 'BRD 1 Residual Impact', 'BRD 1 Residual Likelihood'])
                if key == "lst qrt Brief Risk Decription 2":
                    ws.cell(row=row_num, column=2 + i).value = \
                        risk_combine_text(latest_q_data.data[name], [
                            'Brief Risk Decription 2', 'BRD 2 Risk Category', 'BRD 2 Primary Risk to',
                            'BRD 2 Internal Control', 'BRD 2 Residual Impact', 'BRD 2 Residual Likelihood'])
                if key == "lst qrt Brief Risk Decription 3":
                    ws.cell(row=row_num, column=2 + i).value = \
                        risk_combine_text(latest_q_data.data[name], [
                            'Brief Risk Decription 3', 'BRD 3 Risk Category', 'BRD 3 Primary Risk to',
                            'BRD 3 Internal Control', 'BRD 3 Residual Impact', 'BRD 3 Residual Likelihood'])
                if key == "lst qrt Brief Risk Decription 4":
                    ws.cell(row=row_num, column=2 + i).value = \
                        risk_combine_text(latest_q_data.data[name], [
                            'Brief Risk Decription 4', 'BRD 4 Risk Category', 'BRD 4 Primary Risk to',
                            'BRD 4 Internal Control', 'BRD 4 Residual Impact', 'BRD 4 Residual Likelihood'])
                if key == "lst qrt Brief Risk Decription 5":
                    ws.cell(row=row_num, column=2 + i).value = \
                        risk_combine_text(latest_q_data.data[name], [
                            'Brief Risk Decription 5', 'BRD 5 Risk Category', 'BRD 5 Primary Risk to',
                            'BRD 5 Internal Control', 'BRD 5 Residual Impact', 'BRD 5 Residual Likelihood'])

                '''handling of leaders info'''
                if key == "lst qrt Job Title / Grade":
                    ws.cell(row=row_num, column=2 + i).value = \
                        dn_combine_text(latest_q_data.data[name], ['Job Title / Grade', 'SRO Grade'])
                if key == "lst qrt Job Title":
                    ws.cell(row=row_num, column=2 + i).value = \
                        dn_combine_text(latest_q_data.data[name], ['Job Title', 'PD Grade'])

                '''handling of resourcing info'''
                if key == "lst qrt Digital - Now":
                    ws.cell(row=row_num, column=2 + i).value = \
                        resource_combine_text(latest_q_data.data[name], ['Digital - Now', 'Digital - Future',
                                                                    'Digital Short Narrative (Amber or Red)'])
                if key == "lst qrt Information Technology - Now":
                    ws.cell(row=row_num, column=2 + i).value = \
                        resource_combine_text(latest_q_data.data[name], ['Information Technology - Now', 'Information Technology - Future',
                                                                    'Information Technology Short Narrative (Amber or Red)'])
                if key == "lst qrt Legal Commercial Contract Management - Now":
                    ws.cell(row=row_num, column=2 + i).value = \
                        resource_combine_text(latest_q_data.data[name], ['Legal Commercial Contract Management - Now', 'Legal Commercial Contract Management - Future',
                                                                    'Legal Commercial & Contract Management Short Narrative (Amber or Red)'])
                if key == "lst qrt Project Delivery - Now":
                    ws.cell(row=row_num, column=2 + i).value = \
                        resource_combine_text(latest_q_data.data[name], ['Project Delivery - Now', 'Project Delivery - Future',
                                                                    'Project Delivery Short Narrative (Amber or Red)'])
                if key == "lst qrt Change Implementation - Now":
                    ws.cell(row=row_num, column=2 + i).value = \
                        resource_combine_text(latest_q_data.data[name], ['Change Implementation - Now', 'Change Implementation - Future',
                                                                    'Change Implementation Short Narrative (Amber or Red)'])
                if key == "lst qrt Technical - Now":
                    ws.cell(row=row_num, column=2 + i).value = \
                        resource_combine_text(latest_q_data.data[name], ['Technical - Now', 'Technical - Future',
                                                                    'Technical Short Narrative (Amber or Red)'])
                if key == "lst qrt Industry Knowledge - Now":
                    ws.cell(row=row_num, column=2 + i).value = \
                        resource_combine_text(latest_q_data.data[name], ['Industry Knowledge - Now', 'Industry Knowledge - Future',
                                                                    'Industry Knowledge Short Narrative (Amber or Red)'])
                if key == "lst qrt Finance - Now":
                    ws.cell(row=row_num, column=2 + i).value = \
                        resource_combine_text(latest_q_data.data[name], ['Finance - Now', 'Finance - Future',
                                                                    'Finance Short Narrative (Amber or Red)'])
                # if key == "lst qrt Analysis Now":
                #     ws.cell(row=row_num, column=2 + i).value = \
                #         resource_combine_text(latest_q_data.data[name], ['Analysis - Now', 'Analysis - Future',
                #                                                     'Analysis Short Narrative (Amber or Red)'])
                if key == "lst qrt Communications & Stakeholder Engagement - Now":
                    ws.cell(row=row_num, column=2 + i).value = \
                        resource_combine_text(latest_q_data.data[name], ['Communications & Stakeholder Engagement - Now', 'Communications & Stakeholder Engagement - Future',
                                                                    'Communications & Stakeholder Engagement Short Narrative (Amber or Red)'])

                '''handling of extra resource'''
                if key == "lst qrt Additional Capability 1 Descriptor":
                    ws.cell(row=row_num, column=2 + i).value = \
                        extra_resource_combine_text(latest_q_data.data[name], ['Additional Capability 1 Descriptor', 'Additional Capability 1 Now',
                                                                    'Additional Capability 1 Future', 'Additional Capability 1 Short Narrative (Amber or Red)'])
                if key == "lst qrt Additional Capability 2 Descriptor":
                    ws.cell(row=row_num, column=2 + i).value = \
                        extra_resource_combine_text(latest_q_data.data[name], ['Additional Capability 2 Descriptor', 'Additional Capability 2 Now',
                                                                    'Additional Capability 2 Future', 'Additional Capability 2 Short Narrative (Amber or Red)'])
                if key == "lst qrt Additional Capability 3 Descriptor":
                    ws.cell(row=row_num, column=2 + i).value = \
                        extra_resource_combine_text(latest_q_data.data[name], ['Additional Capability 3 Descriptor', 'Other Capability 3 - Now',
                                                                    'Other Capability 3 - Future', 'Additional Capability 3 Short Narrative (Amber or Red)'])
                if key == "lst qrt Additional Capability 4 Descriptor":
                    ws.cell(row=row_num, column=2 + i).value = \
                        extra_resource_combine_text(latest_q_data.data[name], ['Additional Capability 4 Descriptor', 'Other Capability 4 - Now',
                                                                    'Other Capability 4 - Future', 'Additional Capability 4 Short Narrative (Amber or Red)'])
                if key == "lst qrt Additional Capability 5 Descriptor":
                    ws.cell(row=row_num, column=2 + i).value = \
                        extra_resource_combine_text(latest_q_data.data[name], ['Additional Capability 5 Descriptor', 'Additional Capability 5 Now',
                                                                    'Additional Capability 5 Future', 'Additional Capability 5 Short Narrative (Amber or Red)'])


                '''handling of project costs'''
                if key == "lst qrt RDEL one off costs":
                    ws.cell(row=row_num, column=2 + i).value = \
                        combine_figures(latest_q_data.data[name], ['Total RDEL BL one off new costs', 'Total RDEL Forecast one off new costs'])
                if key == "lst qrt RDEL recurring new costs":
                    ws.cell(row=row_num, column=2 + i).value = \
                        combine_figures(latest_q_data.data[name], ['Total RDEL BL recurring new costs', 'Total RDEL Forecast recurring new costs'])
                if key == "lst qrt RDEL recurring old costs":
                    ws.cell(row=row_num, column=2 + i).value = \
                        combine_figures(latest_q_data.data[name], ['Total RDEL BL recurring old costs', 'Total RDEL Forecast recurring old costs'])
                if key == "lst qrt RDEL non gov":
                    ws.cell(row=row_num, column=2 + i).value = \
                        combine_figures(latest_q_data.data[name], ['Total RDEL BL Non Gov costs', 'Total RDEL Forecast Non Gov costs'])
                if key == "lst qrt CDEL one off costs":
                    ws.cell(row=row_num, column=2 + i).value = \
                        combine_figures(latest_q_data.data[name], ['Total CDEL BL one off new costs', 'Total CDEL Forecast Total'])
                if key == "lst qrt CDEL recurring new costs":
                    ws.cell(row=row_num, column=2 + i).value = \
                        combine_figures(latest_q_data.data[name], ['Total CDEL BL recurring new costs', 'Total CDEL Forecast recurring new costs'])
                if key == "lst qrt CDEL recurring old costs":
                    ws.cell(row=row_num, column=2 + i).value = \
                        combine_figures(latest_q_data.data[name], ['Total CDEL BL recurring old costs', 'Total CDEL Forecast recurring old costs'])
                if key == "lst qrt CDEL non gov":
                    ws.cell(row=row_num, column=2 + i).value = \
                        combine_figures(latest_q_data.data[name], ['Non-Gov Total Budget/BL', 'Non-Gov Total Forecast'])
                if key == "lst qrt Total Budget/BL":
                    ws.cell(row=row_num, column=2 + i).value = combine_figures(latest_q_data.data[name],
                                                                               ['Total Budget/BL', 'Total Forecast'])
                if key == "lst qrt In-Year Spend Total":
                    ws.cell(row=row_num, column=2 + i).value = add_combine_figures(latest_q_data.data[name],
                                                                               ['19-20 RDEL BL Total', '19-20 CDEL BL WLC',
                                                                                '19-20 RDEL Forecast Total', '19-20 CDEL Forecast Total WLC'])
                if key == "lst qrt Income total":
                    ws.cell(row=row_num, column=2 + i).value = combine_figures(latest_q_data.data[name],
                                                                               ['Total Baseline - Income both Revenue and Capital',
                                                                                'Total Forecast - Income both Revenue and Capital'])
                if key == "lst qrt In-Year Income Total":
                    ws.cell(row=row_num, column=2 + i).value = add_combine_figures(latest_q_data.data[name],
                                                                               ['19-20 RDEL BL Income', '19-20 BL Income both Revenue and Capital',
                                                                                '19-20 RDEL Forecast Income',
                                                                                '19-20 Forecast - Income both Revenue and Capital'])
                #if key == "lst qrt 19-20 RDEL Forecast one off new costs":
                #     ws.cell(row=row_num, column=2 + i).value = combine_figures(latest_q_data.data[name],
                #                                                                ['19-20 RDEL Forecast one off new costs', '19-20 RDEL Forecast Income'])
                # if key == "lst qrt 20-21 RDEL BL one off new costs":
                #     ws.cell(row=row_num, column=2 + i).value = combine_figures(latest_q_data.data[name],
                #                                                                ['20-21 RDEL BL one off new costs', '20-21 RDEL BL Income'])
                # if key == "lst qrt 20-21 RDEL Forecast one off new costs":
                #     ws.cell(row=row_num, column=2 + i).value = combine_figures(latest_q_data.data[name],
                #                                                                ['20-21 RDEL Forecast one off new costs', '20-21 RDEL Forecast Income'])

                ws.cell(row=row_num, column=2 + i).font = red_text

                # this section of the code ensures that all financial costs / benefit forecasts have a zero
                for cost_type in type_list_2:
                    if cost_type in key:
                        try:
                            if latest_q_data.data[name][key] is None:
                                ws.cell(row=row_num, column=2 + i).value = 0
                        except KeyError:
                            pass

    return workbook

def dn_combine_text(q_data, string_list):

    '''essentaly used to combine two text strings'''

    combined_string = str(q_data[string_list[0]]) + ' - ' + str(q_data[string_list[1]])

    if 'None' in combined_string:
        combined_string = ''

    return combined_string

def strategic_combine_text(q_data, string_list):

    '''essentially used to combined four text strings'''

    combined_string = str(q_data[string_list[0]]) + ' - ' + str(q_data[string_list[1]]) + ' - ' + \
                      str(q_data[string_list[2]]) + ' - ' + str(q_data[string_list[3]])

    if 'None' in combined_string:
        combined_string = ''

    return combined_string

def resource_combine_text(q_data, string_list):

    '''essentially used to combined four text strings. different out to strategy above'''

    text = string_list[0]
    first_bit = ''
    for i in text:
        if i is not '-':
            first_bit += i
        else:
            break

    '''handling for none types'''
    other_list = []
    for x in string_list:
        if q_data[x] is not None:
            other_list.append(q_data[x])
        else:
            other_list.append('')

    combined_string = first_bit[0:-1] + ' - ' + str(other_list[0]) + ' : ' + str(other_list[1]) + ' ; ' + \
                      str(other_list[2])

    return combined_string

def extra_resource_combine_text(q_data, string_list):

    '''essentially used to combined four text strings. different out to strategy above'''


    combined_string = str(q_data[string_list[0]]) + ' - ' + str(q_data[string_list[1]]) + ' : ' + \
                      str(q_data[string_list[2]]) + ' ; ' + str(q_data[string_list[3]])

    return combined_string

def risk_combine_text(q_data, string_list):

    '''essentially used to combined six text strings'''

    combined_string = str(q_data[string_list[0]]) + ' - ' + str(q_data[string_list[1]]) + \
                      ' - ' + str(q_data[string_list[2]]) + ' - ' + str(q_data[string_list[3]]) + \
                      ' - ' + str(q_data[string_list[4]]) + ' - ' + str(q_data[string_list[5]])

    if 'None' in combined_string:
        combined_string = ''

    return combined_string

def combine_figures(q_data, string_list):

    combined_string = '(B) £' + str(q_data[string_list[0]]) + 'm / (F) £' + str(q_data[string_list[1]]) + 'm'

    return combined_string

def add_combine_figures(q_data, string_list):

    try:
        one = q_data[string_list[0]] + q_data[string_list[1]]
    except TypeError:
        one = ''

    try:
        two = q_data[string_list[2]] + q_data[string_list[3]]
    except TypeError:
        two = ''

    combined_string = '(B) £' + str(one) + 'm / (F) £' + str(two) + 'm'

    return combined_string


master_dm = load_workbook(root_path/'input/commission_master_dm.xlsx')

commission_master = project_data_from_master(root_path/'core_data/master_4_2019_commission_initial.xlsx', 4, 2019)

latest_quarter_master = project_data_from_master(root_path/'core_data/master_3_2019.xlsx', 3, 2019)


'''compile list of project names. This step is necessary to remove projects that finished reporting last quarter'''
project_name_list_commission = [x for x in commission_master.projects if x in latest_quarter_master.projects]

run = create_master(master_dm, project_name_list_commission, commission_master, latest_quarter_master)

run.save(root_path/'output/Q4_1920_commission_data_final.xlsx')
