'''Returns latest project values for specified keys of interest into single sheet workbook'''

from openpyxl import Workbook
from analysis.data import list_of_masters_all, latest_quarter_project_names, root_path, conditional_text, \
    text_colours, background_colours
from analysis.engine_functions import all_milestone_data_bulk, conditional_formatting

def return_data(project_name_list, data_key_list):

    wb = Workbook()
    ws = wb.active

    for i, project_name in enumerate(project_name_list):
        '''list project names, groups and stage in ws'''
        ws.cell(row=2 + i, column=1).value = project_name
        for y, key in enumerate(data_key_list):
            ws.cell(row=1, column=2+y, value=key) #returns key

            try:
                #standard keys
                if key in list_of_masters_all[0].data[project_name].keys():
                    value = list_of_masters_all[0].data[project_name][key]
                    ws.cell(row=2+i, column=2+y, value=value) #retuns value
                    if value is None:
                        ws.cell(row=2 + i, column=2 + y, value='missing data')

                # milestone keys
                else:
                    milestones = all_milestone_data_bulk([project_name], list_of_masters_all[0])
                    value = tuple(milestones[project_name][key])[0]
                    ws.cell(row=2 + i, column=2 + y, value=value)
                    ws.cell(row=2 + i, column=2 + y).number_format = 'dd/mm/yy'
                    if value is None:
                        ws.cell(row=2 + i, column=2 + y, value='missing data')


            except KeyError:
                if project_name in list_of_masters_all[0].projects: #loop calculates if project was not reporting or data missing
                    ws.cell(row=2+ i, column=2+y, value='missing data')
                else:
                    ws.cell(row=2+i, column=2+y, value='project not reporting')

    conditional_formatting(ws, list_columns, conditional_text, text_colours, background_colours, '1', '60')

    '''quarter tag information'''
    ws.cell(row=1, column=1, value='Project')

    return wb

'''data keys of interest'''
key_list = ['SRO Full Name',
            'SRO Email',
            'SRO Phone No.',
            'PD Full Name',
            'PD Email',
            'PD Phone No.',
            'Working Contact Name',
            'Working Contact Email']

list_columns = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'q', 's', 't', 'u', 'w']

'''running prog - step one'''
run_project_all = return_data(latest_quarter_project_names, key_list)

'''step two'''
run_project_all.save(root_path/'output/major_projects_contacts.xlsx')

old_key_search_list = ['BICC approval point',
            'SRO Full Name',
            'SRO Email',
            'SRO Phone No.',
            'Brief project description (GMPP - brief descripton)',
            'Project End Date']