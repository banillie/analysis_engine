'''

Programme for querying and returning multiple milestone data from master data set.

returns multiple key/values of interest for each project.

Some formatting is placed into the output file:
1) When projects are not reporting data -  grey out cell,

Follow instruction as set out below are provided

'''

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from analysis.data import list_of_masters_all, latest_quarter_project_names, bc_index, baseline_bc_stamp, salmon_fill, \
    root_path
from analysis.engine_functions import all_milestone_data_bulk, ap_p_milestone_data_bulk, assurance_milestone_data_bulk,\
    get_all_project_names, get_quarter_stamp, grey_conditional_formatting

def return_multiple_milestone_data(project_name_list, data_key_list):
    ''' places all milestone data of interest into excel file output

    project_name_list: list of project to return data for
    data_key: the data key of interest
    '''

    master = list_of_masters_all[0]

    salmon_fill = PatternFill(start_color='ff8080', end_color='ff8080', fill_type='solid')

    wb = Workbook()
    ws = wb.active

    '''list project names, groups and stage in ws'''
    for x in range(0, len(project_name_list)):
        ws.cell(row=x + 2, column=2, value=project_name_list[x])
        ws.cell(row=x+2, column=4, value=master.data[project_name_list[x]]['Departmental DCA'])
        try:
            ws.cell(row=x + 2, column=1).value = master.data[project_name_list[x]]['DfT Group']
        except KeyError:
            pass
        try:
            ws.cell(row=x + 2, column=3).value = master.data[project_name_list[x]]['BICC approval point']
        except KeyError:
            pass


    'place data into workbook'
    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=2).value
        milestone_data = all_milestone_data_bulk([project_name], master)

        if project_name in master.projects:

            for i, data_key in enumerate(data_key_list):
                ws.cell(row=1, column=5+i).value = data_key

                try:
                    ws.cell(row=row_num, column=5+i).value = tuple(milestone_data[project_name][data_key])[0]

                    if tuple(milestone_data[project_name][data_key])[0] is None:
                        ws.cell(row=row_num, column=5+i).value = 'None'

                except KeyError:
                    ws.cell(row=row_num, column=5+i).value = 'Data not collected'


        '''quarter tag / meta data into ws'''
        ws.cell(row=1, column=1, value='Group')
        ws.cell(row=1, column=2, value='Project')
        ws.cell(row=1, column=3, value='BC stage')
        ws.cell(row=1, column=4, value='DCA')
        # for i, label in enumerate(quarter_labels):
        #     ws.cell(row=1, column=i + 3, value=label)

        grey_conditional_formatting(ws)  # apply conditional formatting

    return wb

''' RUNNING PROGRAMME '''

'''Note that the all master data is taken from the data file. Make sure that this is up to date and that all relevant
  data is being imported'''

'''TWO. Set data of interest. there are two options here. hash out whichever option you are not using'''

'''option two - milestone data. NOTE. this must be in a list [] even if just one data key'''
milestone_data_interest = ['Start of Project',
                           'Start of Construction/build', 'Start of Operation', 'Project End Date']

'''THREE. Run the programme'''
'''option one - run the return_milestone_data for all milestone data'''
run_standard = return_multiple_milestone_data(latest_quarter_project_names, milestone_data_interest)

'''FOUR. specify the file path and name of the output document'''
run_standard.save(root_path/'output/heathrow_analysis_data.xlsx')

'''old lists stored here for use in future'''
old_entries = ['Project End Date', 'Start of Project', 'Start of Project', 'SOBC - BICC Approval', 'OBC - BICC Approval', 'FBC - BICC Approval', \
               'Start of Construction/build', 'Start of Operation', 'Project End Date']

