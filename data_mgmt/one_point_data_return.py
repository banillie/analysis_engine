'''

Programme for querying and returning data from master data set.

There are several options that need to be specified below:
1) returning single data of interest (in to one tab)
2) returning several data of interest (across multiple tabs)... in development 
3) return data across all masters 
4) return data all pertaining to latest, last and baseline data. 

some formatting is placed into the output file:
2) changes in reported data - highlighted by salmon pink background,
3) when projects were not reporting data -  grey out cell,
4) if a rag status is returned the colour of the rag status

Follow instruction as set out below are provided

TODO: Needs some further testing in relation to returning values that aren't reported in quarters.

'''

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from analysis.data import q2_1920, q1_1920, one_quarter_master_list, bespoke_group_masters_list, list_of_masters_all
from analysis.engine_functions import all_milestone_data_bulk, ap_p_milestone_data_bulk, assurance_milestone_data_bulk,\
    get_all_project_names, get_quarter_stamp, bc_ref_stages, master_baseline_index

def return_data(masters_list, project_name_list, data_key_list):
    '''
    places all (non-milestone) data of interest into excel file output

    masters_list: list of masters containing quarter information
    project_name_list: list of project to return data for
    data_key_list: the data key of interest

    '''

    salmon_fill = PatternFill(start_color='ff8080', end_color='ff8080', fill_type='solid')

    wb = Workbook()

    '''project data into ws'''
    for i, data_key in enumerate(data_key_list):
        ws = wb.create_sheet(data_key, i)  # creating worksheets
        ws.title = data_key  # title of worksheet

        '''lists project names in ws'''
        for x in range(0, len(project_name_list)):
            try:
                ws.cell(row=x + 2, column=1).value = masters_list[0].data[project_name_list[x]]['DfT Group']
            except KeyError:
                pass
            ws.cell(row=x + 2, column=2, value=project_name_list[x])


        for row_num in range(2, ws.max_row + 1):
            project_name = ws.cell(row=row_num, column=2).value
            print(project_name)
            col_start = 3
            for i, master in enumerate(masters_list):
                if project_name in master.projects:
                    try:
                        ws.cell(row=row_num, column=col_start).value = master.data[project_name][data_key]
                        if master.data[project_name][data_key] == None:
                            ws.cell(row=row_num, column=col_start).value = 'None'
                        try:
                            if masters_list[i+1].data[project_name][data_key] != master.data[project_name][data_key]:
                                ws.cell(row=row_num, column=col_start).fill = salmon_fill
                        except (IndexError, KeyError):
                            pass
                        col_start += 1
                    except KeyError:
                        ws.cell(row=row_num, column=col_start).value = 'data key not collected'
                else:
                    ws.cell(row=row_num, column=col_start).value = 'Not reporting'
                    col_start += 1

        '''quarter tag / meta data into ws'''
        quarter_labels = get_quarter_stamp(masters_list)
        ws.cell(row=1, column=1, value='Group')
        ws.cell(row=1, column=2, value='Project')
        for i, label in enumerate(quarter_labels):
            ws.cell(row=1, column=i + 3, value=label)

        grey_conditional_formatting(ws)  # apply grey formatting

    return wb

def return_baseline_data(masters_list, baseline_list, baseline_ref, project_name_list, data_key_list):
    '''
    places all non-milestone data into output document with latest, last and baseline data. Also states which quarter
    is being used as baseline
    :param masters_list: list of master quarter data
    :param baseline_list: list indexing where latest, last and baseline master data for each project
    :param project_name_list: list of project names
    :param data_key_list: data of interest/to return
    :return: excel spreadsheet
    '''

    salmon_fill = PatternFill(start_color='ff8080', end_color='ff8080', fill_type='solid')

    wb = Workbook()

    '''project data into ws'''
    for i, data_key in enumerate(data_key_list):
        ws = wb.create_sheet(data_key, i)  # creating worksheets
        ws.title = data_key  # title of worksheet

        '''lists project names in ws'''
        for x in range(0, len(project_name_list)):
            try:
                ws.cell(row=x + 2, column=1).value = masters_list[0].data[project_name_list[x]]['DfT Group']
            except KeyError:
                pass
            ws.cell(row=x + 2, column=2, value=project_name_list[x])

        '''project data into ws'''
        for row_num in range(2, ws.max_row + 1):
            project_name = ws.cell(row=row_num, column=2).value
            ws.cell(row=row_num, column=8).value = baseline_ref[project_name][2][0] # ref to baseline quarter
            print(project_name)
            col_start = 3
            for i in baseline_list[project_name]:
                try:
                    ws.cell(row=row_num, column=col_start).value = masters_list[i].data[project_name][data_key]
                    if masters_list[i].data[project_name][data_key] == None:
                        ws.cell(row=row_num, column=col_start).value = 'None'
                    try:
                        if masters_list[i+1].data[project_name][data_key] != masters_list[i].data[project_name][data_key]:
                            ws.cell(row=row_num, column=col_start).fill = salmon_fill
                    except (IndexError, KeyError):
                        pass
                    col_start += 1
                except KeyError:
                    ws.cell(row=row_num, column=col_start).value = 'Data not collected'
                    col_start += 1

        '''quarter tag / meta data into ws'''
        baseline_labels = ['This quarter', 'Last quarter', 'Baseline quarter']
        ws.cell(row=1, column=1, value='Group')
        ws.cell(row=1, column=2, value='Project')
        for i, label in enumerate(baseline_labels):
            ws.cell(row=1, column=i + 3, value=label)
        ws.cell(row=1, column=8, value='Quarter from which baseline data taken')

    return wb

def return_baseline_milestone_data(masters_list, baseline_list, baseline_ref, project_name_list, data_key_list):
    '''
    places all milestone data into output document with latest, last and baseline data. Also states which quarter
    is being used as baseline
    :param masters_list: list of master quarter data
    :param baseline_list: list indexing where latest, last and baseline master data for each project
    :param project_name_list: list of project names
    :param data_key_list: data of interest/to return
    :return: excel spreadsheet
    '''

    salmon_fill = PatternFill(start_color='ff8080', end_color='ff8080', fill_type='solid')

    wb = Workbook()

    '''project data into ws'''
    for i, data_key in enumerate(data_key_list):
        ws = wb.create_sheet(data_key, i)  # creating worksheets
        ws.title = data_key  # title of worksheet

        '''lists project names in ws'''
        for x in range(0, len(project_name_list)):
            try:
                ws.cell(row=x + 2, column=1).value = masters_list[0].data[project_name_list[x]]['DfT Group']
            except KeyError:
                pass
            ws.cell(row=x + 2, column=2, value=project_name_list[x])

        '''project data into ws'''
        for row_num in range(2, ws.max_row + 1):
            project_name = ws.cell(row=row_num, column=2).value
            ws.cell(row=row_num, column=8).value = baseline_ref[project_name][2][0]  # ref to baseline quarter
            print(project_name)
            col_start = 3
            for i in baseline_list[project_name]:
                milestone_data = all_milestone_data_bulk([project_name], masters_list[i])

                try:
                    ws.cell(row=row_num, column=col_start).value = tuple(milestone_data[project_name][data_key])[0]

                    if tuple(milestone_data[project_name][data_key])[0] is None:
                        ws.cell(row=row_num, column=col_start).value = 'None'
                except KeyError:
                    ws.cell(row=row_num, column=col_start).value = 'None'

                try:

                    last_milestone_data = all_milestone_data_bulk([project_name], masters_list[i + 1])

                    if tuple(last_milestone_data[project_name][data_key])[0] != \
                            tuple(milestone_data[project_name][data_key])[0]:
                        ws.cell(row=row_num, column=col_start).fill = salmon_fill
                except (IndexError, KeyError):
                    pass
                col_start += 1

        '''quarter tag / meta data into ws'''
        baseline_labels = ['This quarter', 'Last quarter', 'Baseline quarter']
        ws.cell(row=1, column=1, value='Group')
        ws.cell(row=1, column=2, value='Project')
        for i, label in enumerate(baseline_labels):
            ws.cell(row=1, column=i + 3, value=label)
        ws.cell(row=1, column=8, value='Quarter from which baseline data taken')

    return wb

def return_milestone_data(masters_list, project_name_list, data_key_list):
    ''' places all milestone data of interest into excel file output

    master_list: list of masters containing quarter information
    project_name_list: list of project to return data for
    data_key: the data key of interest
    '''

    salmon_fill = PatternFill(start_color='ff8080', end_color='ff8080', fill_type='solid')
    # red_text = Font(color="FF0000") #currently not in use

    wb = Workbook()

    '''project data into ws'''
    for i, data_key in enumerate(data_key_list):
        ws = wb.create_sheet(data_key, i)  # creating worksheets
        ws.title = data_key  # title of worksheet

        '''lists project names in ws'''
        for x in range(0, len(project_name_list)):
            try:
                ws.cell(row=x + 2, column=1).value = masters_list[0].data[project_name_list[x]]['DfT Group']
            except KeyError:
                pass
            ws.cell(row=x + 2, column=2, value=project_name_list[x])

        '''project data into ws'''
        for row_num in range(2, ws.max_row + 1):
            project_name = ws.cell(row=row_num, column=2).value
            print(project_name)
            col_start = 3
            for i, master in enumerate(masters_list):
                if project_name in master.projects:

                    milestone_data = all_milestone_data_bulk([project_name], master)

                    try:
                        ws.cell(row=row_num, column=col_start).value = tuple(milestone_data[project_name][data_key])[0]

                        if tuple(milestone_data[project_name][data_key])[0] is None:
                            ws.cell(row=row_num, column=col_start).value = 'None'
                    except KeyError:
                        ws.cell(row=row_num, column=col_start).value = 'None'

                    try:

                        last_milestone_data = all_milestone_data_bulk([project_name], masters_list[i + 1])

                        if tuple(last_milestone_data[project_name][data_key])[0] != \
                                tuple(milestone_data[project_name][data_key])[0]:
                            ws.cell(row=row_num, column=col_start).fill = salmon_fill
                    except (IndexError, KeyError):
                        pass
                    col_start += 1
                else:
                    ws.cell(row=row_num, column=col_start).value = 'Not reporting'
                    col_start += 1

        '''quarter tag / meta data into ws'''
        quarter_labels = get_quarter_stamp(masters_list)
        ws.cell(row=1, column=1, value='Group')
        ws.cell(row=1, column=2, value='Project')
        for i, label in enumerate(quarter_labels):
            ws.cell(row=1, column=i + 3, value=label)

        grey_conditional_formatting(ws)  # apply conditional formatting

    return wb

def conditional_formatting(worksheet):
    '''
    function applie conditional formating for RAG colors... in development.
    :param worksheet:
    :return:
    '''

    ag_text = Font(color="000000")
    ag_fill = PatternFill(bgColor="00a5b700")
    dxf = DifferentialStyle(font=ag_text, fill=ag_fill)
    rule = Rule(type="containsText", operator="containsText", text="Amber/Green", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Amber/Green",A1)))']
    worksheet.conditional_formatting.add('A1:X80', rule)

    ar_text = Font(color="000000")
    ar_fill = PatternFill(bgColor="00f97b31")
    dxf = DifferentialStyle(font=ar_text, fill=ar_fill)
    rule = Rule(type="containsText", operator="containsText", text="Amber/Red", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Amber/Red",A1)))']
    worksheet.conditional_formatting.add('A1:X80', rule)

    red_text = Font(color="000000")
    red_fill = PatternFill(bgColor="00fc2525")
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="containsText", operator="containsText", text="Red", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Red",A1)))']
    worksheet.conditional_formatting.add('A1:X80', rule)

    green_text = Font(color="000000")
    green_fill = PatternFill(bgColor="0017960c")
    dxf = DifferentialStyle(font=green_text, fill=green_fill)
    rule = Rule(type="containsText", operator="containsText", text="Green", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Green",A1)))']
    worksheet.conditional_formatting.add('A1:X80', rule)

    amber_text = Font(color="000000")
    amber_fill = PatternFill(bgColor="00fce553")
    dxf = DifferentialStyle(font=amber_text, fill=amber_fill)
    rule = Rule(type="containsText", operator="containsText", text="Amber", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Amber",A1)))']
    worksheet.conditional_formatting.add('A1:X80', rule)

    return worksheet

def grey_conditional_formatting(worksheet):
    '''
    function applies grey conditional formatting for 'Not Reporting'.
    :param worksheet: ws
    :return: cf of sheet
    '''

    grey_text = Font(color="f0f0f0")
    grey_fill = PatternFill(bgColor="f0f0f0")
    dxf = DifferentialStyle(font=grey_text, fill=grey_fill)
    rule = Rule(type="containsText", operator="containsText", text="Not reporting", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Not reporting",A1)))']
    worksheet.conditional_formatting.add('A1:X80', rule)

    return worksheet

''' RUNNING PROGRAMME '''

'''Note that the all master data is taken from the data file. Make sure that this is up to date and that all relevant
  data is being imported'''

''' ONE. Set relevant list of projects. This needs to be done in accordance with the data you are working with via the
 data.py file '''
one_quarter_list = q2_1920.projects
combined_quarters_list = get_all_project_names(list_of_masters_all)
specific_project_list = ['Oxford-Cambridge Expressway'] # opportunity to provide manual list of projects

'''TWO. calculate baseline meta data'''
baseline_bc = bc_ref_stages(one_quarter_list, list_of_masters_all)
baseline_list = master_baseline_index(one_quarter_list, list_of_masters_all, baseline_bc)

'''THREE. Set data of interest. there are two options here. hash out whichever option you are not using'''

'''option one - non-milestone data. NOTE. this must be in a list [] even if just one data key'''
data_interest = ['Working Contact Name', 'Working Contact Email', 'Brief project description (GMPP - brief descripton)',
                 'Business Case & Version No.', 'NPV for all projects and NPV for programmes if available',
                 'Initial Benefits Cost Ratio (BCR)', 'Adjusted Benefits Cost Ratio (BCR)',
                 'VfM Category single entry', 'VfM Category', 'Present Value Cost (PVC)', 'Present Value Benefit (PVB)']

'''option two - milestone data. NOTE. this must be in a list [] even if just one data key'''
#milestone_data_interest = ['Project End Date', 'Start of Project']

'''THREE. Run the programme'''

'''option one - run the return_data function for all non-milestone data'''
#run = return_data(list_of_masters_all, combined_quarters_list, data_interest)

'''option two - run the return_baseline_data function for all non-milestone data'''
run = return_baseline_data(list_of_masters_all, baseline_list, baseline_bc, specific_project_list, data_interest)

'''option three - run the return_milestone_data for all milestone data'''
#run = return_milestone_data(list_of_masters_all, one_quarter_list, milestone_data_interest)

'''option four - run the return_baseline_milestone_data function for all milestone data'''
#run = return_baseline_milestone_data(list_of_masters_all, baseline_list, baseline_bc, one_quarter_list,
#                                     milestone_data_interest)

'''FOUR. specify the file path and name of the output document'''
run.save('C:\\Users\\Standalone\\general\\vfm_data_ox_cam_express_baseline.xlsx')



'''old lists stored here for use in future'''

vfm_analysis_list = ['Working Contact Name', 'Working Contact Email', 'Brief project description (GMPP - brief descripton)',
                 'Business Case & Version No.', 'NPV for all projects and NPV for programmes if available',
                 'Initial Benefits Cost Ratio (BCR)', 'Adjusted Benefits Cost Ratio (BCR)',
                 'VfM Category single entry', 'VfM Category', 'Present Value Cost (PVC)', 'Present Value Benefit (PVB)']
