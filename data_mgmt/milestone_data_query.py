'''

Programme for querying and returning milestone data from master data set.

There are several options that need to be specified below:
1) returning single data of interest (in to one tab)
2) returning several data of interest (across multiple tabs)... in development
3) return data across all masters
4) return data all pertaining to latest, last and baseline data.

some formatting is placed into the output file:
2) changes in reported data - highlighted by salmon pink background,
3) when projects were not reporting data -  grey out cell,
4) if a rag status is returned the colour of the rag status

Follow instruction as set out below are provided

'''

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from analysis.data import list_of_masters_all, latest_quarter_project_names, bc_index, baseline_bc_stamp, salmon_fill, \
    root_path
from analysis.engine_functions import all_milestone_data_bulk, ap_p_milestone_data_bulk, assurance_milestone_data_bulk,\
    get_all_project_names, get_quarter_stamp, grey_conditional_formatting


def return_baseline_milestone_data(project_name_list, data_key_list):
    '''
    places all milestone data into output document with latest, last and baseline data. Also states which quarter
    is being used as baseline
    :param project_name_list: list of project names
    :param data_key_list: data of interest/to return
    :return: excel spreadsheet
    '''

    wb = Workbook()

    '''project data into ws'''
    for i, data_key in enumerate(data_key_list):
        ws = wb.create_sheet(data_key, i)  # creating worksheets
        ws.title = data_key  # title of worksheet

        '''lists project names in ws'''
        for x in range(0, len(project_name_list)):
            ws.cell(row=x + 2, column=2, value=project_name_list[x])
            try:
                ws.cell(row=x + 2, column=1).value = list_of_masters_all[0].data[project_name_list[x]]['DfT Group']
            except KeyError:
                pass

        '''project data into ws'''
        for row_num in range(2, ws.max_row + 1):
            project_name = ws.cell(row=row_num, column=2).value
            ws.cell(row=row_num, column=8).value = baseline_bc_stamp[project_name][0][1]  # ref to baseline quarter
            print(project_name)
            col_start = 3
            for i in bc_index[project_name]:
                try:
                    milestone_data = all_milestone_data_bulk([project_name], list_of_masters_all[i])

                    try:
                        ws.cell(row=row_num, column=col_start).value = tuple(milestone_data[project_name][data_key])[0]

                        if tuple(milestone_data[project_name][data_key])[0] is None:
                            ws.cell(row=row_num, column=col_start).value = 'None'
                    except KeyError:
                        ws.cell(row=row_num, column=col_start).value = 'None'

                    try:

                        last_milestone_data = all_milestone_data_bulk([project_name], list_of_masters_all[i + 1])

                        if tuple(last_milestone_data[project_name][data_key])[0] != \
                                tuple(milestone_data[project_name][data_key])[0]:
                            ws.cell(row=row_num, column=col_start).fill = salmon_fill
                    except (IndexError, KeyError):
                        pass
                except TypeError:
                    ws.cell(row=row_num, column=col_start).value = 'None'

                col_start += 1

        '''quarter tag / meta data into ws'''
        baseline_labels = ['This quarter', 'Last quarter', 'Baseline quarter']
        ws.cell(row=1, column=1, value='Group')
        ws.cell(row=1, column=2, value='Project')
        for i, label in enumerate(baseline_labels):
            ws.cell(row=1, column=i + 3, value=label)
        ws.cell(row=1, column=8, value='Quarter from which baseline data taken')

    return wb

def return_milestone_data(project_name_list, data_key_list):
    ''' places all milestone data of interest into excel file output

    master_list: list of masters containing quarter information
    project_name_list: list of project to return data for
    data_key: the data key of interest
    '''

    salmon_fill = PatternFill(start_color='ff8080', end_color='ff8080', fill_type='solid')
    # red_text = Font(color="FF0000") #currently not in use

    wb = Workbook()

    '''project data into ws'''
    for i, data_key in enumerate(data_key_list):
        ws = wb.create_sheet(data_key, i)  # creating worksheets
        ws.title = data_key  # title of worksheet

        '''lists project names in ws'''
        for x in range(0, len(project_name_list)):
            try:
                ws.cell(row=x + 2, column=1).value = list_of_masters_all[0].data[project_name_list[x]]['DfT Group']
            except KeyError:
                pass
            ws.cell(row=x + 2, column=2, value=project_name_list[x])

        '''project data into ws'''
        for row_num in range(2, ws.max_row + 1):
            project_name = ws.cell(row=row_num, column=2).value
            print(project_name)
            col_start = 3
            for i, master in enumerate(list_of_masters_all):
                if project_name in master.projects:

                    milestone_data = all_milestone_data_bulk([project_name], master)

                    try:
                        ws.cell(row=row_num, column=col_start).value = tuple(milestone_data[project_name][data_key])[0]

                        if tuple(milestone_data[project_name][data_key])[0] is None:
                            ws.cell(row=row_num, column=col_start).value = 'None'
                    except KeyError:
                        ws.cell(row=row_num, column=col_start).value = 'Data not collected'

                    try:

                        last_milestone_data = all_milestone_data_bulk([project_name], list_of_masters_all[i + 1])

                        if tuple(last_milestone_data[project_name][data_key])[0] != \
                                tuple(milestone_data[project_name][data_key])[0]:
                            ws.cell(row=row_num, column=col_start).fill = salmon_fill
                    except (IndexError, KeyError):
                        pass
                    col_start += 1
                else:
                    ws.cell(row=row_num, column=col_start).value = 'Not reporting'
                    col_start += 1

        '''quarter tag / meta data into ws'''
        quarter_labels = get_quarter_stamp(list_of_masters_all)
        ws.cell(row=1, column=1, value='Group')
        ws.cell(row=1, column=2, value='Project')
        for i, label in enumerate(quarter_labels):
            ws.cell(row=1, column=i + 3, value=label)

        grey_conditional_formatting(ws)  # apply conditional formatting

    return wb

''' RUNNING PROGRAMME '''

'''Note that the all master data is taken from the data file. Make sure that this is up to date and that all relevant
  data is being imported'''

''' ONE. Set relevant list of projects. This needs to be done in accordance with the data you are working with via the
 data.py file '''
one_quarter_list = latest_quarter_project_names
combined_quarters_list = get_all_project_names(list_of_masters_all)
specific_project_list = [] # opportunity to provide manual list of projects. get project name in import statement

'''TWO. Set data of interest. there are two options here. hash out whichever option you are not using'''

'''option two - milestone data. NOTE. this must be in a list [] even if just one data key'''
milestone_data_interest = ['Full Operating Capacity (FOC)', 'Full Operations', 'Project End Date']

'''THREE. Run the programme'''
'''option one - run the return_milestone_data for all milestone data'''
run_standard = return_milestone_data(one_quarter_list, milestone_data_interest)

'''option two - run the return_baseline_milestone_data function for all milestone data'''
run_baseline = return_baseline_milestone_data(one_quarter_list, milestone_data_interest)

'''FOUR. specify the file path and name of the output document'''
run_standard.save(root_path/'output/foc_data.xlsx')

run_baseline.save(root_path/'output/foc_baseline_data.xlsx')


'''old lists stored here for use in future'''
old_entries = ['Project End Date', 'Start of Project']

