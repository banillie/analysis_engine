'''probably throw away code which converts data key names some that 'lst quarter' or some other text can be added to
the front of the keys. This will then be used for structuring overall commission master and migrating relevant
quarter data into the commission master

in development'''

from openpyxl import load_workbook, Workbook

def get_key_names(workbook):
    key_list = []
    lst_q_key_list = []
    ws = workbook.active

    for row_num in range(2, ws.max_row + 1):
        key_list.append(ws.cell(row=row_num, column=1).value)

    for key in key_list:
        lst_q_key_list.append('lst qrt ' + str(key))

    wb = Workbook()
    ws = wb.active

    for i, key in enumerate(lst_q_key_list):
        ws.cell(row=1 + i, column=1).value = key

    return wb


master = load_workbook(
        'C:\\Users\\Standalone\\general\\masters folder\\core_data\\master_1_2019_wip_(25_7_19).xlsx')

test = get_key_names(master)

test.save(
    'C:\\Users\\Standalone\\general\\lst_quarter_keys.xlsx')