'''Returns project values across multiple masters for specified keys of interest. Return for each key is provided
on a separate wb. Code can handle both standard and project milestone keys, as well as project name lists across
multiple quarters.

There are two outputs.
1) wb containing all values
2) wb containing bl values only'''


from openpyxl import Workbook
from analysis.data import list_of_masters_all, latest_quarter_project_names, root_path, gen_txt_list, \
    gen_txt_colours, gen_fill_colours, list_column_ltrs, list_of_rag_keys, rag_txt_list_full, \
    rag_fill_colours, rag_txt_colours, all_project_names, salmon_fill, bc_index
from analysis.engine_functions import all_milestone_data_bulk, conditional_formatting, get_quarter_stamp

def return_data(project_name_list, data_key_list):
    """Returns project values across multiple masters for specified keys of interest:
    project_names_list: list of project names
    data_key_list: list of data keys
    """
    wb = Workbook()

    for i, key in enumerate(data_key_list):
        '''worksheet is created for each project'''
        ws = wb.create_sheet(key, i)  # creating worksheets
        ws.title = key  # title of worksheet

        '''list project names, groups and stage in ws'''
        for y, project_name in enumerate(project_name_list):

            # get project group info
            try:
                group = list_of_masters_all[0].data[project_name]['DfT Group']
            except KeyError:
                for m, master in enumerate(list_of_masters_all):
                    if project_name in master.projects:
                        group = list_of_masters_all[m].data[project_name]['DfT Group']

            ws.cell(row=2 + y, column=1, value=group) # group info return
            ws.cell(row=2 + y, column=2, value=project_name)  # project name returned

            for x, master in enumerate(list_of_masters_all):
                if project_name in master.projects:
                    try:
                        #standard keys
                        if key in list_of_masters_all[x].data[project_name].keys():
                            value = list_of_masters_all[x].data[project_name][key]
                            ws.cell(row=2 + y, column=3 + x, value=value) #returns value
                            if value is None:
                                ws.cell(row=2 + y, column=3 + x, value='md')

                            try: # checks for change against last quarter
                                lst_value = list_of_masters_all[x + 1].data[project_name][key]
                                if value != lst_value:
                                    ws.cell(row=2 + y, column=3 + x).fill = salmon_fill
                            except (KeyError, IndexError):
                                pass

                        # milestone keys
                        else:
                            milestones = all_milestone_data_bulk([project_name], list_of_masters_all[x])
                            value = tuple(milestones[project_name][key])[0]
                            ws.cell(row=2 + y, column=3 + x, value=value)
                            ws.cell(row=2 + y, column=3 + x).number_format = 'dd/mm/yy'
                            if value is None:
                                ws.cell(row=2 + y, column=3 + x, value='md')

                            try:  # loop checks if value has changed since last quarter
                                old_milestones = all_milestone_data_bulk([project_name], list_of_masters_all[x + 1])
                                lst_value = tuple(old_milestones[project_name][key])[0]
                                if value != lst_value:
                                    ws.cell(row=2 + y, column=3 + x).fill = salmon_fill
                            except (KeyError, IndexError):
                                pass

                    except KeyError:
                        if project_name in master.projects:
                            #loop calculates if project was not reporting or data missing
                            ws.cell(row=2 + y, column=3 + x, value='knc')
                        else:
                            ws.cell(row=2 + y, column=3 + x, value='pnr')

                else:
                    ws.cell(row=2 + y, column=3 + x, value='pnr')

        '''quarter tag information'''
        ws.cell(row=1, column=1, value='Group')
        ws.cell(row=1, column=2, value='Projects')
        quarter_labels = get_quarter_stamp(list_of_masters_all)
        for l, label in enumerate(quarter_labels):
            ws.cell(row=1, column=l + 3, value=label)

        list_columns = list_column_ltrs[0:len(list_of_masters_all)+3]

        if key in list_of_rag_keys:
            conditional_formatting(ws, list_columns, rag_txt_list_full, rag_txt_colours, rag_fill_colours, '1', '60')

        conditional_formatting(ws, list_columns, gen_txt_list, gen_txt_colours, gen_fill_colours, '1', '60')

    return wb

def return_baseline_data(project_name_list, data_key_list):
    '''
    returns values of interest across multiple ws for baseline values only.
    project_name_list: list of project names
    data_key_list: list of data keys containing values of interest.
    '''
    wb = Workbook()

    for i, key in enumerate(data_key_list):
        '''worksheet is created for each project'''
        ws = wb.create_sheet(key, i)  # creating worksheets
        ws.title = key  # title of worksheet

        '''list project names, groups and stage in ws'''
        for y, project_name in enumerate(project_name_list):

            # get project group info
            try:
                group = list_of_masters_all[0].data[project_name]['DfT Group']
            except KeyError:
                for m, master in enumerate(list_of_masters_all):
                    if project_name in master.projects:
                        group = list_of_masters_all[m].data[project_name]['DfT Group']

            ws.cell(row=2 + y, column=1, value=group) # group info
            ws.cell(row=2 + y, column=2, value=project_name)  # project name returned

            for x in range(0, len(bc_index[project_name])):
                index = bc_index[project_name][x]
                try: # standard keys
                    value = list_of_masters_all[index].data[project_name][key]
                    if value is None:
                        ws.cell(row=2 + y, column=3 + x).value = 'md'
                    else:
                        ws.cell(row=2 + y, column=3 + x, value=value)
                except KeyError:
                    try: # milestone keys
                        milestones = all_milestone_data_bulk([project_name], list_of_masters_all[index])
                        value = tuple(milestones[project_name][key])[0]
                        if value is None:
                            ws.cell(row=2 + y, column=3 + x).value = 'md'
                        else:
                            ws.cell(row=2 + y, column=3 + x).value = value
                            ws.cell(row=2 + y, column=3 + x).number_format = 'dd/mm/yy'
                    except KeyError: # exception catches both standard and milestone keys
                        ws.cell(row=2 + y, column=3 + x).value = 'knc'
                except TypeError:
                    ws.cell(row=2 + y, column=3 + x).value = 'pnr'

        ws.cell(row=1, column=1, value='Group')
        ws.cell(row=1, column=2, value='Project')
        ws.cell(row=1, column=3, value='Latest')
        ws.cell(row=1, column=4, value='Last quarter')
        ws.cell(row=1, column=5, value='BL 1')
        ws.cell(row=1, column=6, value='BL 2')
        ws.cell(row=1, column=7, value='BL 3')
        ws.cell(row=1, column=8, value='BL 4')
        ws.cell(row=1, column=9, value='BL 5')

        list_columns = list_column_ltrs[2:9] # hard coded so not ideal

        if key in list_of_rag_keys:
            conditional_formatting(ws, list_columns, rag_txt_list_full, rag_txt_colours, rag_fill_colours, '1', '60')

        conditional_formatting(ws, list_columns, gen_txt_list, gen_txt_colours, gen_fill_colours, '1', '60')

    return wb

'''data keys of interest'''
data_interest = ['VfM Category single entry', 'VfM Category lower range', 'VfM Category upper range',
                 'SRO Benefits RAG', 'Start of Operation', 'Full Operations', 'Project End Date']

'''Running the programme'''
'''output one - all data'''
run_standard = return_data(latest_quarter_project_names, data_interest)
'''output two - bl data'''
run_baseline = return_baseline_data(all_project_names, data_interest)

'''Specify name of the output document here'''
run_standard.save(root_path/'output/data_query_testing.xlsx')
run_baseline.save(root_path/'output/data_query_testing_bl.xlsx')

'''old lists stored here for use in future'''
old_entries = ['GMPP - IPA DCA', 'BICC approval point',
               'Brief project description (GMPP - brief descripton)',
               'Delivery Narrative']

vfm_analysis_list = ['Departmental DCA', 'Working Contact Name', 'Working Contact Email',
                 'Brief project description (GMPP - brief descripton)',
                 'Business Case & Version No.', 'BICC approval point',
                 'NPV for all projects and NPV for programmes if available',
                 'Initial Benefits Cost Ratio (BCR)', 'Adjusted Benefits Cost Ratio (BCR)',
                 'VfM Category single entry', 'VfM Category lower range', 'VfM Category upper range',
                 'Present Value Cost (PVC)', 'Present Value Benefit (PVB)', 'SRO Benefits RAG',
                 'Benefits Narrative', 'Ben comparison with last quarters cost - narrative']

ipa_ar_fields_1920 =  ['Department', '19-20 RDEL BL Total', '19-20 CDEL BL WLC',
                 '19-20 RDEL Forecast Total', '19-20 CDEL Forecast Total WLC', 'Total BL',
                 'GMPP - IPA ID Number']
