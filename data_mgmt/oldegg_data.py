#  start of what was library/data

from datamaps.api import project_data_from_master
from analysis.engine_functions import baseline_information_bc, baseline_index, get_project_income_profile, \
    get_project_cost_profile, get_all_project_names, baseline_information, project_all_milestones_dict, \
    project_time_difference
from openpyxl.styles import Font, PatternFill
import platform, datetime
from pathlib import Path

'''file path'''


def _platform_docs_dir() -> Path:
    if platform.system() == "Linux":
        return Path.home() / "Documents" / "analysis_engine"
    if platform.system() == "Darwin":
        return Path.home() / "Documents" / "analysis_engine"
    else:
        return Path.home() / "Documents" / "analysis_engine"


root_path = _platform_docs_dir()

'''master data'''
q2_2021 = project_data_from_master(root_path / 'core_data/master_2_2020.xlsx', 2, 2020)
q1_2021 = project_data_from_master(root_path / 'core_data/master_1_2020.xlsx', 1, 2020)
q4_1920 = project_data_from_master(root_path / 'core_data/master_4_2019.xlsx', 4, 2019)
q3_1920 = project_data_from_master(root_path / 'core_data/master_3_2019.xlsx', 3, 2019)
q2_1920 = project_data_from_master(root_path / 'core_data/master_2_2019.xlsx', 2, 2019)
q1_1920 = project_data_from_master(root_path / 'core_data/master_1_2019.xlsx', 1, 2019)
q4_1819 = project_data_from_master(root_path / 'core_data/master_4_2018.xlsx', 4, 2018)
q3_1819 = project_data_from_master(root_path / 'core_data/master_3_2018.xlsx', 3, 2018)
q2_1819 = project_data_from_master(root_path / 'core_data/master_2_2018.xlsx', 2, 2018)
q1_1819 = project_data_from_master(root_path / 'core_data/master_1_2018.xlsx', 1, 2018)
q4_1718 = project_data_from_master(root_path / 'core_data/master_4_2017.xlsx', 4, 2017)
q3_1718 = project_data_from_master(root_path / 'core_data/master_3_2017.xlsx', 3, 2017)
q2_1718 = project_data_from_master(root_path / 'core_data/master_2_2017.xlsx', 2, 2017)
q1_1718 = project_data_from_master(root_path / 'core_data/master_1_2017.xlsx', 1, 2017)
q4_1617 = project_data_from_master(root_path / 'core_data/master_4_2016.xlsx', 4, 2016)
q3_1617 = project_data_from_master(root_path / 'core_data/master_3_2016.xlsx', 3, 2016)

"""list of dictionaries"""
# one_quarter_master_list = []
# bespoke_group_masters_list = []

list_of_masters_all = [q2_2021,
                       q1_2021,
                       q4_1920,
                       q3_1920,
                       q2_1920,
                       q1_1920,
                       q4_1819,
                       q3_1819,
                       q2_1819,
                       q1_1819,
                       q4_1718,
                       q3_1718,
                       q2_1718,
                       q1_1718,
                       q4_1617,
                       q3_1617]

# financial bls only go back to q1_1819
# financial_analysis_masters_list = list_of_masters_all[0:9]

'''list of project names. useful to have here and import into programme'''
all_project_names = get_all_project_names(list_of_masters_all)

'''dates for functions. python date format is Year, Month, day'''
ipdc_date = datetime.date(2020, 11, 2)

abbreviations = {'2nd Generation UK Search and Rescue Aviation': 'SARH2',
                 'A12 Chelmsford to A120 widening': 'A12',
                 'A14 Cambridge to Huntingdon Improvement Scheme': 'A14',
                 'A303 Amesbury to Berwick Down': 'A303',
                 'A358 Taunton to Southfields Dualling': 'A358',
                 'A417 Air Balloon': 'A417',
                 'A428 Black Cat to Caxton Gibbet': 'A428',
                 'A66 Northern Trans-Pennine': 'A66',
                 'Brighton Mainline Upgrade Programme': 'Brighton Mainline',
                 'Crossrail Programme': 'Crossrail',
                 'East Coast Digital Programme': 'ECDP',
                 'East Coast Mainline Programme': 'ECMP',
                 'East West Rail Programme (Central Section)': 'EWR (Central)',
                 'East West Rail Programme (Western Section)': 'EWR (Western)',
                 "East West Rail Configuration State 1": "EWR Config 1",
                 "East West Rail Configuration State 2": "EWR Config 2",
                 "East West Rail Configuration State 3": "EWR Config 3",
                 'Future Theory Test Service (FTTS)': 'FTTS',
                 'Great Western Route Modernisation (GWRM) including electrification': 'GWRM',
                 'Heathrow Expansion': 'HEP',
                 'Hexagon': 'Hexagon',
                 'High Speed Rail Programme (HS2)': 'HS2 Prog',
                 'HS2 Phase 2b': 'HS2 2b',
                 'HS2 Phase1': 'HS2 1',
                 'HS2 Phase2a': 'HS2 2a',
                 'Integrated and Smart Ticketing - creating an account based back office': 'IST',
                 'Intercity Express Programme': 'IEP',
                 'Lower Thames Crossing': 'LTC',
                 'M4 Junctions 3 to 12 Smart Motorway': 'M4',
                 'Manchester North West Quadrant': 'MNWQ',
                 'Midland Main Line Programme': 'MML Prog',
                 'Midlands Rail Hub': 'Mid Rail Hub',
                 'North Western Electrification': 'NWE',
                 'Northern Powerhouse Rail': 'NPR',
                 'Oxford-Cambridge Expressway': 'Ox-Cam Expressway',
                 'Rail Franchising Programme': 'Rail Franchising',
                 'South West Route Capacity': 'SWRC',
                 'Thameslink Programme': 'Thameslink',
                 'Transpennine Route Upgrade': 'TRU',
                 'Western Rail Link to Heathrow': 'WRLtH'}

'''specific project names. Useful to have them captured here so don't have to keep cutting and pasting string 
name from excel master'''
a12 = 'A12 Chelmsford to A120 widening'
a14 = 'A14 Cambridge to Huntingdon Improvement Scheme'
a303 = 'A303 Amesbury to Berwick Down'
a358 = 'A358 Taunton to Southfields Dualling'
a417 = 'A417 Air Balloon'
a428 = 'A428 Black Cat to Caxton Gibbet'
a66 = 'A66 Northern Trans-Pennine'
bright_ml = 'Brighton Mainline Upgrade Programme'
cvs = 'Commercial Vehicle Services (CVS)'
east_coast_digital = 'East Coast Digital Programme'
east_coast_mainline = 'East Coast Mainline Programme'
em_franchise = 'East Midlands Franchise'
ewr_central = 'East West Rail Programme (Central Section)'
ewr_western = 'East West Rail Programme (Western Section)'
ewr_config1 = "East West Rail Configuration State 1"
ewr_config2 = "East West Rail Configuration State 2"
ewr_config3 = "East West Rail Configuration State 3"
ftts = 'Future Theory Test Service (FTTS)'
heathrow_expansion = 'Heathrow Expansion'
hexagon = 'Hexagon'
hs2_programme = 'High Speed Rail Programme (HS2)'
hs2_2b = 'HS2 Phase 2b'
hs2_1 = 'HS2 Phase1'
hs2_2a = 'HS2 Phase2a'
ist = 'Integrated and Smart Ticketing - creating an account based back office'
lower_thames_crossing = 'Lower Thames Crossing'
m4 = 'M4 Junctions 3 to 12 Smart Motorway'
manchester_north_west_quad = 'Manchester North West Quadrant'
midland_mainline = 'Midland Main Line Programme'
midlands_rail_hub = 'Midlands Rail Hub'
north_of_england = 'North of England Programme'
northern_powerhouse = 'Northern Powerhouse Rail'
nwe = 'North Western Electrification'
ox_cam_expressway = 'Oxford-Cambridge Expressway'
rail_franchising = 'Rail Franchising Programme'
west_coast_partnership = 'West Coast Partnership Franchise'
crossrail = 'Crossrail Programme'
gwrm = 'Great Western Route Modernisation (GWRM) including electrification'
iep = 'Intercity Express Programme'
sarh2 = '2nd Generation UK Search and Rescue Aviation'
south_west_route_capacity = 'South West Route Capacity'
thameslink = 'Thameslink Programme'
tru = 'Transpennine Route Upgrade'
wrlth = 'Western Rail Link to Heathrow'

'''project groups'''
hsmrpg = [hs2_1, hs2_2a, hs2_2b, hexagon, northern_powerhouse, ewr_western, ewr_central]

obc_fbc = []

'''baselining information'''
# business case approval baseline. Used for overall DCA rating
baseline_bc_stamp = baseline_information_bc(all_project_names, list_of_masters_all)
bc_index = baseline_index(baseline_bc_stamp, list_of_masters_all)
# # finance baseline information
# fin_baseline_bcs = baseline_information_bc(all_project_names, financial_analysis_masters_list)
# fin_bc_index = baseline_index(fin_baseline_bcs, list_of_masters_all)

milestone_bl_stamp = baseline_information(all_project_names, list_of_masters_all,
                                          'Re-baseline IPDC milestones')
milestone_bl_index = baseline_index(milestone_bl_stamp, list_of_masters_all)

costs_bl_stamp = baseline_information(all_project_names, list_of_masters_all,
                                      'Re-baseline IPDC cost')
costs_bl_index = baseline_index(costs_bl_stamp, list_of_masters_all)

benefits_bl_stamp = baseline_information(all_project_names, list_of_masters_all,
                                         'Re-baseline IPDC benefits')
benefits_bl_index = baseline_index(benefits_bl_stamp, list_of_masters_all)

'''lists and keys for running programmes'''
income_list = [' Forecast - Income both Revenue and Capital']
cost_list = [' RDEL Forecast Total', ' CDEL Forecast Total', ' Forecast Non-Gov']
year_list = ['19-20', '20-21', '21-22', '22-23', '23-24', '24-25', '25-26', '26-27', '27-28', '28-29', 'Unprofiled']
wlc_key = 'Total Forecast'

'''Financial information'''
latest_income_profiles = get_project_income_profile(list_of_masters_all[0].projects, list_of_masters_all,
                                                    income_list, year_list, costs_bl_index, 0)
last_income_profiles = get_project_income_profile(list_of_masters_all[0].projects, list_of_masters_all,
                                                  income_list, year_list, costs_bl_index, 1)
baseline_1_income_profiles = get_project_income_profile(list_of_masters_all[0].projects, list_of_masters_all,
                                                        income_list, year_list, costs_bl_index, 2)
# baseline_2_income_profiles = get_project_income_profile(list_of_masters_all[0].projects, financial_analysis_masters_list,
#                                                       income_list, year_list, fin_bc_index, 3)


latest_cost_profiles = get_project_cost_profile(list_of_masters_all[0].projects, list_of_masters_all,
                                                cost_list, year_list, costs_bl_index, 0)
last_cost_profiles = get_project_cost_profile(list_of_masters_all[0].projects, list_of_masters_all,
                                              cost_list, year_list, costs_bl_index, 1)
baseline_1_cost_profiles = get_project_cost_profile(list_of_masters_all[0].projects, list_of_masters_all,
                                                    cost_list, year_list, costs_bl_index, 2)
# baseline_2_cost_profiles = get_project_cost_profile(list_of_masters_all[0].projects, list_of_masters_all,
#                                                         cost_list, year_list, costs_bl_index, 3)

# milestone information
'''get all milestone data'''
p_current_milestones = project_all_milestones_dict(list_of_masters_all[0].projects,
                                                   list_of_masters_all,
                                                   milestone_bl_index, 0)
p_last_milestones = project_all_milestones_dict(list_of_masters_all[0].projects,
                                                list_of_masters_all,
                                                milestone_bl_index, 1)
p_baseline_milestones = project_all_milestones_dict(list_of_masters_all[0].projects,
                                                    list_of_masters_all,
                                                    milestone_bl_index, 2)
# p_baseline_milestones_two = project_all_milestones_dict(list_of_masters_all[0].projects,
#                                         list_of_masters_all,
#                                         milestone_bl_index, 3)

'''calculate time current and last quarter'''
first_diff_data = project_time_difference(p_current_milestones, p_last_milestones)

# for financial dca dashboard
financial_narrative_keys = ['Project Costs Narrative',
                            'Cost comparison with last quarters cost narrative',
                            'Cost comparison within this quarters cost narrative']

# for project summary pages
SRO_conf_table_list = ['SRO DCA', 'Finance DCA', 'Benefits DCA', 'Resourcing DCA', 'Schedule DCA']
SRO_conf_key_list = ['Departmental DCA', 'SRO Finance confidence', 'SRO Benefits RAG', 'Overall Resource DCA - Now',
                     'SRO Schedule Confidence']

'''list of projects to exclude from counting of totals in portfolio financial profile'''
dont_double_count = [hs2_programme, northern_powerhouse, east_coast_digital, heathrow_expansion]

'''Store of different colours'''
ag_text = Font(color="00a5b700")  # text same colour as background
ag_fill = PatternFill(bgColor="00a5b700")
ar_text = Font(color="00f97b31")  # text same colour as background
ar_fill = PatternFill(bgColor="00f97b31")
red_text = Font(color="00fc2525")  # text same colour as background
red_fill = PatternFill(bgColor="00fc2525")
green_text = Font(color="0017960c")  # text same colour as background
green_fill = PatternFill(bgColor="0017960c")
amber_text = Font(color="00fce553")  # text same colour as background
amber_fill = PatternFill(bgColor="00fce553")

black_text = Font(color="00000000")
red_text = Font(color="FF0000")

darkish_grey_text = Font(color="002e4053")
darkish_grey_fill = PatternFill(bgColor="002e4053")
light_grey_text = Font(color="0085929e")
light_grey_fill = PatternFill(bgColor="0085929e")
greyblue_text = Font(color="85c1e9")
greyblue_fill = PatternFill(bgColor="85c1e9")

salmon_fill = PatternFill(start_color='FFFF8080',
                          end_color='FFFF8080',
                          fill_type='solid')

'''Conditional formatting, cell colouring and text colouring'''
# reference for column names when applying conditional fomatting
list_column_ltrs = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l',
                    'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z']
# list of keys that have rag values for conditional formatting.
list_of_rag_keys = ['SRO Schedule Confidence', 'Departmental DCA',
                    'SRO Finance confidence', 'SRO Benefits RAG',
                    'GMPP - IPA DCA']

# lists of text and backfround colours and list of values for conditional formating rules.
rag_txt_colours = [ag_text, ar_text, red_text, green_text, amber_text]
rag_fill_colours = [ag_fill, ar_fill, red_fill, green_fill, amber_fill]
rag_txt_list_acroynms = ["A/G", "A/R", "R", "G", "A"]
rag_txt_list_full = ["Amber/Green", "Amber/Red", "Red", "Green", "Amber"]
gen_txt_colours = [darkish_grey_text, light_grey_text, greyblue_text]
gen_fill_colours = [darkish_grey_fill, light_grey_fill, greyblue_fill]
gen_txt_list = ['md', 'pnr', 'knc']
'''
keeping as colour coding is useful
s1 = chart.series[0]
s1.graphicalProperties.line.solidFill = "cfcfea" #light blue
s2 = chart.series[1]
s2.graphicalProperties.line.solidFill = "e2f1bb" #light green
s3 = chart.series[2]
s3.graphicalProperties.line.solidFill = "eaba9d" #light red
s4 = chart.series[3]
s4.graphicalProperties.line.solidFil = "5097a4" #medium blue
s5 = chart.series[4]
s5.graphicalProperties.line.solidFill = "a0db8e" #medium green
s6 = chart.series[5]
s6.graphicalProperties.line.solidFill = "b77575" #medium red
s7 = chart.series[6]
s7.graphicalProperties.line.solidFil = "0e2f44" #dark blue
s8 = chart.series[7]
s8.graphicalProperties.line.solidFill = "29ab87" #dark green
s9 = chart.series[8]
s9.graphicalProperties.line.solidFill = "691c1c" #dark red
'''

'''lists with groups of useful keys for data querying'''

stakeholders = ['SRO Full Name',
                'SRO Email',
                'SRO Phone No.',
                'PD Full Name',
                'PD Email',
                'PD Phone No.',
                'Working Contact Name',
                'Working Contact Email',
                'Working Contact Telephone']

vfm = ['IPDC approval point',
       'Total Forecast',
       'VfM Category single entry',
       'VfM Category lower range',
       'VfM Category upper range',
       'VfM Category single entry',
       'Present Value Cost (PVC)',
       'Present Value Benefit (PVB)',
       'Initial Benefits Cost Ratio (BCR)',
       'Adjusted Benefits Cost Ratio (BCR)',
       'Start of Construction/build',
       'Start of Operation',
       'Full Operations',
       'Project End Date']

vfm_list_two = ['Departmental DCA',
                'Working Contact Name',
                'Working Contact Email',
                'Brief project description (GMPP - brief descripton)',
                'Business Case & Version No.',
                'BICC approval point',
                'NPV for all projects and NPV for programmes if available',
                'Initial Benefits Cost Ratio (BCR)',
                'Adjusted Benefits Cost Ratio (BCR)',
                'VfM Category single entry',
                'VfM Category lower range',
                'VfM Category upper range',
                'Present Value Cost (PVC)',
                'Present Value Benefit (PVB)',
                'SRO Benefits RAG',
                'Benefits Narrative',
                'Ben comparison with last quarters cost - narrative']

ipa_ar_fields_1920 = ['Department',
                      '19-20 RDEL BL Total',
                      '19-20 CDEL BL WLC',
                      '19-20 RDEL Forecast Total',
                      '19-20 CDEL Forecast Total WLC',
                      'Total BL',
                      'GMPP - IPA ID Number']

project_basics = ['Brief project description (GMPP - brief descripton)',
                  'Delivery Narrative']

milestones = ['Start of Project',
              'SOBC - IPDC Approval',
              'OBC - IPDC Approval',
              'FBC - IPDC Approval',
              'Start of Construction/build',
              'Start of Operation',
              'Full Operations',
              'Project End Date']

rags = ['SRO Benefits RAG',
        'GMPP - IPA DCA']

baselines = ['IPDC approval point']

ben_change_key_list = ['Pre-profile BEN Total',
                       'Total BEN Forecast - Total Monetised Benefits',
                       'Unprofiled Remainder BEN Forecast - Total Monetised Benefits']

ben_type_key_list = ['Total BEN Forecast - Total Monetised Benefits',
                     'Pre-profile BEN Forecast Gov Cashable',
                     'Pre-profile BEN Forecast Gov Non-Cashable',
                     'Pre-profile BEN Forecast - Economic (inc Private Partner)',
                     'Pre-profile BEN Forecast - Disbenefit UK Economic',
                     'Unprofiled Remainder BEN Forecast - Gov. Cashable',
                     'Unprofiled Remainder BEN Forecast - Gov. Non-Cashable',
                     'Unprofiled Remainder BEN Forecast - Economic (inc Private Partner)',
                     'Unprofiled Remainder BEN Forecast - Disbenefit UK Economic',
                     'Total BEN Forecast - Gov. Cashable',
                     'Total BEN Forecast - Gov. Non-Cashable',
                     'Total BEN Forecast - Economic (inc Private Partner)',
                     'Total BEN Forecast - Disbenefit UK Economic']

cost_key_list = ['Total Forecast',
                 'Pre-profile RDEL',
                 'Pre-profile CDEL',
                 'Total RDEL Forecast Total',
                 'Total CDEL Forecast Total WLC',
                 'Unprofiled RDEL Forecast Total',
                 'Unprofiled CDEL Forecast Total WLC']

financal_key_list = ['Source of Finance',
                     'Overall contingency (£m)',
                     'Is this Continency amount included within the WLC?',
                     'Overall figure for Optimism Bias (£m)',
                     'Is this Optimism Bias included within the WLC?']

baselining_keys = ['IPDC approval point',
                   'Re-baseline this quarter',
                   'Re-baseline ALB/Programme milestones',
                   'Re-baseline ALB/Programme cost',
                   'Re-baseline ALB/Programme benefits',
                   'Re-baseline IPDC milestones',
                   'Re-baseline IPDC cost',
                   'Re-baseline IPDC benefits',
                   'Re-baseline HMT milestones',
                   'Re-baseline HMT cost',
                   'Re-baseline HMT benefits']
