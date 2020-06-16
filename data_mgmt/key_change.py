'''
Changes key names contained in master data workbooks. Works across all masters. Changes are highlighted with red
text.

Useful code of tidying up master data key names.
'''

from openpyxl import load_workbook
from datamaps.api import project_data_from_master
from analysis.data import root_path, red_text

#TODO include master datamap in this list also
list_master_names = [root_path/'core_data/master_4_2019.xlsx',
                     root_path/'core_data/master_3_2019.xlsx',
               root_path/'core_data/master_2_2019.xlsx',
               root_path/'core_data/master_1_2019.xlsx',
               root_path/'core_data/master_4_2018.xlsx',
               root_path/'core_data/master_3_2018.xlsx',
               root_path/'core_data/master_2_2018.xlsx',
               root_path/'core_data/master_1_2018.xlsx',
               root_path/'core_data/master_4_2017.xlsx',
               root_path/'core_data/master_3_2017.xlsx',
               root_path/'core_data/master_2_2017.xlsx',
               root_path/'core_data/master_1_2017.xlsx',
               root_path/'core_data/master_4_2016.xlsx',
               root_path/'core_data/master_3_2016.xlsx']

key_data = project_data_from_master(root_path/'core_data/key_change_log.xlsx', 1, 2020)

def change_keys(masters_list, key_change_data):
    key_dict = key_change_data.data['Change Key']

    for master in masters_list:
        wb = load_workbook(master)
        ws = wb.active

        for row_num in range(2, ws.max_row + 1):
            keye = ws.cell(row=row_num, column=1).value
            if keye in key_dict.keys():
                print(keye)
                ws.cell(row=row_num, column=1).value = key_dict[keye]
                ws.cell(row=row_num, column=1).font = red_text

        wb.save(master)


'''Runs programme'''
change_keys(list_master_names, key_data)