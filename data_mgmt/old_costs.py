"""
Code used for getting old financial year data that should be present in current masters.
"""


import typing

from datamaps.api import project_data_from_master
from openpyxl import load_workbook

from data_mgmt.data import root_path, get_master_data_file_paths_fy_17_18, get_master_data_file_paths_fy_18_19, \
    get_master_data_file_paths_fy_19_20, get_master_data_file_paths_fy_20_21


def get_old_fy_cost_data(master_file: typing.TextIO, project_id_wb: typing.TextIO) -> None:
    """
    Gets all old financial data from a specified master and places into project id document.
    """
    master = project_data_from_master(master_file, 1, 2010)  # random year specified as not in use
    wb = load_workbook(project_id_wb)
    ws = wb.active

    for i in range(1, ws.max_column + 1):
        project_name = ws.cell(row=1, column=1 + i).value
        for row_num in range(2, ws.max_row + 1):
            key = ws.cell(row=row_num, column=1).value
            try:
                if key in master.data[project_name].keys():
                    ws.cell(row=row_num, column=1 + i).value = master.data[project_name][key]
            except KeyError:  # project might not be present in quarter
                pass

    wb.save(project_id_wb)


def run_get_old_fy_data(master_files_list: list, project_id_wb: typing.TextIO) -> None:
    for f in reversed(master_files_list):  # reversed so it gets the latest data in masters
        get_old_fy_cost_data(f, project_id_wb)


def place_old_fy_data_into_master_wb(master_file: typing.TextIO, project_id_wb: typing.TextIO) -> None:
    """
    places all old financial year data into master files.
    """
    id_master = project_data_from_master(project_id_wb, 2, 2020)  # random year specify as not used
    wb = load_workbook(master_file)
    ws = wb.active

    for i in range(1, ws.max_column + 1):
        project_name = ws.cell(row=1, column=1 + i).value
        for row_num in range(2, ws.max_row + 1):
            key = ws.cell(row=row_num, column=1).value
            try:
                if key in id_master.data[project_name].keys():
                    ws.cell(row=row_num, column=1 + i).value = id_master.data[project_name][key]
            except KeyError:  # project might not be present in quarter
                pass

    wb.save(master_file)


def run_place_old_fy_data_into_masters(master_files_list: list, project_id_wb: typing.TextIO) -> None:
    for f in master_files_list:
        place_old_fy_data_into_master_wb(f, project_id_wb)


# get_old_fy_cost_data(root_path / "core_data/master_4_2018.xlsx",
#                      root_path / "core_data/other/project_info_fy_18_19_cost_info.xlsx")

run_place_old_fy_data_into_masters(get_master_data_file_paths_fy_18_19(),
                                   root_path / "core_data/other/project_info_fy_17_18_cost_info.xlsx")
