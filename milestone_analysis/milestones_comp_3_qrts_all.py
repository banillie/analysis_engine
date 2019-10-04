'''

This programme to calculate time difference between reported milestones

input documents:
There quarters master information, typically:
1) latest quarter data
2) last quarter data
3) year ago quarter data

output document:
1) excel workbook with all project milestone information

See instructions on how to operate programme below.

'''

#TODO solve problem re filtering in excle when values have + sign in front of the them

import datetime
from openpyxl import Workbook
from engine_functions import all_milestone_data_bulk, ap_p_milestone_data_bulk, assurance_milestone_data_bulk, \
    project_time_difference, bc_ref_stages, get_master_baseline_dict
from data import q1_1920, list_of_dicts_all

'''function for putting all data into excel for this programme'''
def put_into_wb_all(project_list, t_dict, td_dict, td_dict2, wb):
    ws = wb.active

    row_num = 2
    for name in project_list:
        for i, milestone in enumerate(td_dict[name].keys()):
            ws.cell(row=row_num + i, column=1).value = name
            ws.cell(row=row_num + i, column=2).value = milestone
            try:
                milestone_date = tuple(t_dict[name][milestone])[0]
                ws.cell(row=row_num + i, column=3).value = milestone_date
            except KeyError:
                ws.cell(row=row_num + i, column=3).value = 0

            try:
                value = td_dict[name][milestone]
                # try:
                    # if int(value) > 0:
                    #     ws.cell(row=row_num + i, column=4).value = '+' + str(value) + ' (days)'
                    # elif int(value) < 0:
                    #     ws.cell(row=row_num + i, column=4).value = str(value) + ' (days)'
                    # elif int(value) == 0:
                ws.cell(row=row_num + i, column=4).value = value
                # except ValueError:
                #     ws.cell(row=row_num + i, column=4).value = value
            except KeyError:
                ws.cell(row=row_num + i, column=4).value = 0

            try:
                value = td_dict2[name][milestone]
                # try:
                    # if int(value) > 0:
                    #     ws.cell(row=row_num + i, column=5).value = '+' + str(value) + ' (days)'
                    # elif int(value) < 0:
                    #     ws.cell(row=row_num + i, column=5).value = str(value) + ' (days)'
                    # elif int(value) == 0:
                ws.cell(row=row_num + i, column=5).value = value
                # except ValueError:
                #     ws.cell(row=row_num + i, column=5).value = value
            except KeyError:
                ws.cell(row=row_num + i, column=5).value = 0

            try:
                milestone_date = tuple(t_dict[name][milestone])[0]
                ws.cell(row=row_num + i, column=6).value = t_dict[name][milestone][milestone_date]  # provides notes
            except IndexError:
                ws.cell(row=row_num + i, column=6).value = 0

        row_num = row_num + len(td_dict[name])

    ws.cell(row=1, column=1).value = 'Project'
    ws.cell(row=1, column=2).value = 'Milestone'
    ws.cell(row=1, column=3).value = 'Date'
    ws.cell(row=1, column=4).value = '3/m change (days)'
    ws.cell(row=1, column=5).value = 'Baseline change (days)'
    ws.cell(row=1, column=6).value = 'Notes'

    return wb

'''
Function that runs this programme...
Notes: 1)It does not check to see whether milestones have been removed, 
'''
def run_milestone_comparator(function, proj_list, q_masters_dict_list, date_of_interest):
    wb = Workbook()

    '''firstly business cases of interest are filtered out by bc_ref_stage function'''
    baseline_bc = bc_ref_stages(proj_list, q_masters_dict_list)
    print(baseline_bc)
    q_masters_list = get_master_baseline_dict(proj_list, q_masters_dict_list, baseline_bc)
    print(q_masters_list)

    '''gather mini-dictionaries for each quarter'''

    current_milestones_dict = {}
    last_milestones_dict = {}
    oldest_milestones_dict = {}
    for proj_name in proj_list:
        p_current_milestones_dict = function([proj_name], q_masters_dict_list[q_masters_list[proj_name][0]])
        current_milestones_dict.update(p_current_milestones_dict)
        p_last_milestones_dict = function([proj_name], q_masters_dict_list[q_masters_list[proj_name][1]])
        last_milestones_dict.update(p_last_milestones_dict)
        p_oldest_milestones_dict = function([proj_name], q_masters_dict_list[q_masters_list[proj_name][2]])
        oldest_milestones_dict.update(p_oldest_milestones_dict)

    '''calculate time current and last quarter'''
    first_diff_dict = project_time_difference(current_milestones_dict, last_milestones_dict, date_of_interest)
    second_diff_dict = project_time_difference(current_milestones_dict, oldest_milestones_dict, date_of_interest)

    run = put_into_wb_all(proj_list, current_milestones_dict, first_diff_dict, second_diff_dict, wb)

    return run

'''INSTRUCTIONS FOR RUNNING THE PROGRAMME'''

''' 3) set list of projects to be included in output. Still in development'''
'''option one - all projects'''
latest_q_list = q1_1920.projects

'''option two - group of projects... in development'''
group_projects_list = ['Rail Group', 'HSMRPG', 'International Security and Environment', 'Roads Devolution & Motoring']

'''option three - single project'''
one_proj_list = ['Thameslink Programme']

'''4) Specify date after which project milestones should be returned. NOTE: Python date format is (YYYY,MM,DD)'''
start_date = datetime.date(2019, 6, 1)

'''5) choose the type of variables that you would like to place in run_milestone_comparator function, below. 
The type of milestone you wish to analysis can be specified through choosing
all_milestone_data_bulk, ap_p_milestone_data_bulk, or assurance_milestone_data_bulk functions. This choice should be the 
first to be inserted into the below function. After this select the list of the projects on which to perform analysis 
and then the three quarters data that you have put into variables above, in order of newest to oldest.'''
print_miles = \
    run_milestone_comparator(assurance_milestone_data_bulk, latest_q_list, list_of_dicts_all, start_date)

'''5) specify file path to output document'''
print_miles.save('C:\\Users\\Standalone\\general\\assurance.xlsx')