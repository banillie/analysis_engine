"""
To transfer a CombinedData object into excel wb.
"""

from openpyxl import Workbook, load_workbook
# from analysis.engine_functions import project_time_difference
# from analysis.data import list_of_masters_all, bc_index, \
#     p_current_milestones, p_last_milestones, p_baseline_milestones, \
#     p_baseline_milestones_two
from analysis_engine.data import MilestoneData, MilestoneChartData, \
    Master, Projects, master_data_list, root_path, blue_line_date, \
    abbreviations, CombinedData


mst = Master(master_data_list[1:], Projects.hsmrpg)  # get master data and specify projects
mst.get_baseline_data('Re-baseline IPDC milestones')  # get baseline information of interest
milestone_data = MilestoneData(mst, abbreviations)  # get milestone data

hsmrpg_milestone_wb = load_workbook("/home/will/Documents/analysis_engine/input/exported_milestones_HSMRPG.xlsx")
hsmrpg_combined_milestone_data = CombinedData(hsmrpg_milestone_wb, milestone_data)

def put_into_wb_all(combined_data):
    """
    places combined_data object into excel wb. Data in wb
    is milestone name, current data, movement from baseline
    data and milestone notes.
    """

    wb = Workbook()
    ws = wb.active

    row_num = 2

    for i, milestone in enumerate(combined_data.group_current.keys()):
        ws.cell(row=row_num + i, column=2).value = milestone
        try:
            milestone_date = tuple(combined_data.group_current[milestone])[0]
            ws.cell(row=row_num + i, column=3).value = milestone_date
            ws.cell(row=row_num + i, column=3).number_format = 'dd/mm/yy'
        except KeyError:
            ws.cell(row=row_num + i, column=3).value = ''

        try:
            baseline_milestone_date = tuple(combined_data.group_baseline[milestone])[0]
            time_delta = (milestone_date - baseline_milestone_date).days
            ws.cell(row=row_num + i, column=4).value = time_delta
        except (KeyError, TypeError):
            ws.cell(row=row_num + i, column=4).value = ''

        try:
            ws.cell(row=row_num + i, column=5).value = combined_data.group_current[milestone][
                milestone_date]  # provides notes
        except (IndexError, KeyError):
            ws.cell(row=row_num + i, column=5).value = ''


    #ws.cell(row=1, column=1).value = 'Project'
    ws.cell(row=1, column=2).value = 'Milestone'
    ws.cell(row=1, column=3).value = 'Date'
    #ws.cell(row=1, column=4).value = '3/m change'
    ws.cell(row=1, column=4).value = 'Movement from baseline'
    # ws.cell(row=1, column=6).value = 'Baseline change (last)'
    ws.cell(row=1, column=5).value = 'Notes'

    return wb


# def run_milestone_comparator(project_name_list):
#     '''
#     Function that runs this programme.
#
#     function: The type of milestone you wish to analysis can be specified through choosing all_milestone_data_bulk,
#     ap_p_milestone_data_bulk, or assurance_milestone_data_bulk functions, all available from engine_function import
#     statement above.
#     project_name_list: list of project to return data for
#     masters_list: list of masters containing quarter information
#     date_of_interest: the date after which project milestones should be returned.
#
#     '''
#
#     wb = Workbook()
#
#     '''gather mini-dictionaries for each quarter'''
#
#     '''calculate time current and last quarter'''
#     first_diff_data = project_time_difference(p_current_milestones, p_last_milestones)
#     second_diff_data = project_time_difference(p_current_milestones, p_baseline_milestones)
#     third_diff_data = project_time_difference(p_current_milestones, p_baseline_milestones_two)
#
#     run = put_into_wb_all(project_name_list,
#                           p_current_milestones,
#                           first_diff_data,
#                           second_diff_data,
#                           third_diff_data,
#                           wb)
#
#     return run
#


run = put_into_wb_all(hsmrpg_combined_milestone_data)
run.save(root_path / 'output/hsmrpg_milestones_printout.xlsx')