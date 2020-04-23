'''
programme for altering across all masters changes in milestone keys.

Issue: not currently marking new text as red. not clear why this line of code not working

code doesn't/can't handle keys that have been removed.
'''


from datamaps.api import project_data_from_master
from openpyxl import load_workbook
from openpyxl.styles import Font
from analysis.data import bc_index, root_path

def change_key(project_list, master_wb_title_list, change_key):
    '''
    :param project_list: list of project names
    :param master_wb_title_list: list of quarter master data workbooks. This is different from the master list of
    python dictionaries usually passed into functions.
    :param baseline_list: list which indexes where baseline information sits.
    :param change_key: list of keys that need to be changed.
    :return: the wbs in the master_wb_title_list with the data amended.
    '''

    red_text = Font(color="00fc2525")

    for name in project_list:
        #print(name)
        for master in master_wb_title_list:
            wb = load_workbook(master)
            ws = wb.active
            for col_num in range(2, ws.max_column + 1):
                project_name = ws.cell(row=1, column=col_num).value
                if project_name == name:
                    print(name)
                    for row_num in range(2, ws.max_row + 1):
                        for i in range(1, 4): # TODO: non-hard code fix.
                            if ws.cell(row=row_num, column=col_num).value == change_key[project_name]['Key '+ str(i)]:
                                    print(change_key[project_name]['Key '+ str(i)])
                                    ws.cell(row=row_num, column=col_num).value = change_key.data[project_name]['Key '+ str(i)+' change']
                                    print(change_key[project_name]['Key '+ str(i)+' change'])
                                    ws.cell(row=row_num, column=col_num).font = red_text
                            else:
                                pass

            wb.save(master)


'''INSTRUCTIONS FOR RUNNING THE PROGRAMME'''

'''ONE. List of file paths to masters'''
master_list = (root_path/'core_data/master_4_2019.xlsx',
               root_path/'core_data/master_3_2019.xlsx',
               root_path/'core_data/master_2_2019.xlsx',
               root_path/'core_data/master_1_2019.xlsx',
               root_path/'core_data/master_4_2018.xlsx',
               root_path/'core_data/master_3_2018.xlsx',
               root_path/'core_data/master_2_2018.xlsx',
               root_path/'core_data/master_1_2018.xlsx',
               root_path/'core_data/master_4_2017.xlsx',
               root_path/'core_data/master_3_2017.xlsx',
               root_path/'core_data/master_2_2017.xlsx',
               root_path/'core_data/master_1_2017.xlsx',
               root_path/'core_data/master_4_2016.xlsx',
               root_path/'core_data/master_3_2016.xlsx')

'''TWO. Provide file path to document which contains information on the data that needs to be changed'''
key_change = project_data_from_master(root_path/'input/change_milestone_keys_q4_1920.xlsx', 2, 2019)

'''THREE. List of projects. taken from the key change document - as this contains the only projects that need 
information changed'''
project_name_list = key_change.projects
#project_name_list = ['Commercial Vehicle Services (CVS)']

'''FOUR. enter relevant variables in the change_key function'''
change_key(project_name_list, master_list, key_change)