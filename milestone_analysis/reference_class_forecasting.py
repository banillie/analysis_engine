from openpyxl import load_workbook, Workbook
from bcompiler.utils import project_data_from_master
import datetime
from openpyxl.chart import ScatterChart, Reference, Series


def ref_class_fore(master_dict, project_title, start_row, output_wb):
    # output_wb = Workbook()
    data = project_data_from_master(master_dict)
    project_data = data[project_title]

    cells_we_want_to_capture = ['Reporting period (GMPP - Snapshot Date)',
                                'Approval MM1',
                                'Approval MM1 Forecast / Actual',
                                'Approval MM3',
                                'Approval MM3 Forecast / Actual',
                                'Approval MM10',
                                'Approval MM10 Forecast / Actual',
                                'Project MM18',
                                'Project MM18 Forecast - Actual',
                                'Project MM19',
                                'Project MM19 Forecast - Actual',
                                'Project MM20',
                                'Project MM20 Forecast - Actual',
                                'Project MM21',
                                'Project MM21 Forecast - Actual']
    output_list = []
    for item in project_data.items():
        if item[0] in cells_we_want_to_capture:
            output_list.append(item)

    output_list = list(enumerate(output_list, start=1))
    print(output_list)

    output_list2 = [output_list[2][1][1],
                    output_list[4][1][1],
                    output_list[6][1][1],
                    output_list[8][1][1],
                    output_list[10][1][1],
                    output_list[12][1][1],
                    output_list[14][1][1]]

    SOBC = output_list2[0]
    print('SOBC', SOBC)
    OBC = output_list2[1]
    print('OBC', OBC)
    FBC = output_list2[2]
    print('FBC', FBC)
    start_project = output_list2[3]
    print('Start of Project', start_project)
    start_construction = output_list2[4]
    print('Start of construction', start_construction)
    start_ops = output_list2[5]
    print('Start of Ops', start_ops)
    end_project = output_list2[6]
    print('End of project', end_project)

    try:
        time_delta1 = (SOBC - start_project).days
    except TypeError:
        time_delta1 = None
    print(time_delta1)
    try:
        time_delta2 = (OBC - SOBC).days
    except TypeError:
        time_delta2 = None
    print(time_delta2)
    try:
        time_delta3 = (FBC - OBC).days
    except TypeError:
        time_delta3 = None
    print(time_delta3)
    try:
        time_delta4 = (start_construction - FBC).days
    except TypeError:
        time_delta4 = None
    print(time_delta4)
    try:
        time_delta5 = (start_ops - start_construction).days
    except TypeError:
        time_delta5 = None
    print(time_delta5)
    try:
        time_delta6 = (end_project - start_ops).days
    except TypeError:
        time_delta6 = None
    print(time_delta6)

    ws = output_wb.active

    for x in output_list[:3]:
        ws.cell(row=2, column=x[0] + 1, value=x[1][0])
        ws.cell(row=start_row + 1, column=x[0] + 1, value=x[1][1])
        ws.cell(row=start_row + 1, column=x[0] + 2, value=time_delta1)

    for x in output_list[3:5]:
        ws.cell(row=2, column=x[0] + 2, value=x[1][0])
        ws.cell(row=start_row + 1, column=x[0] + 2, value=x[1][1])
        ws.cell(row=start_row + 1, column=x[0] + 3, value=time_delta2)

    for x in output_list[5:7]:
        ws.cell(row=2, column=x[0] + 3, value=x[1][0])
        ws.cell(row=start_row + 1, column=x[0] + 3, value=x[1][1])
        ws.cell(row=start_row + 1, column=x[0] + 4, value=time_delta3)

    for x in output_list[7:9]:
        ws.cell(row=2, column=x[0] + 4, value=x[1][0])
        ws.cell(row=start_row + 1, column=x[0] + 4, value=x[1][1])
        # ws.cell(row=start_row+1, column=series_one[0]+5, value=time_delta3)

    for x in output_list[9:11]:
        ws.cell(row=2, column=x[0] + 5, value=x[1][0])
        ws.cell(row=start_row + 1, column=x[0] + 5, value=x[1][1])
        ws.cell(row=start_row + 1, column=x[0] + 6, value=time_delta4)

    for x in output_list[11:13]:
        ws.cell(row=2, column=x[0] + 6, value=x[1][0])
        ws.cell(row=start_row + 1, column=x[0] + 6, value=x[1][1])
        ws.cell(row=start_row + 1, column=x[0] + 7, value=time_delta5)

    for x in output_list[13:15]:
        ws.cell(row=2, column=x[0] + 7, value=x[1][0])
        ws.cell(row=start_row + 1, column=x[0] + 7, value=x[1][1])
        ws.cell(row=start_row + 1, column=x[0] + 8, value=time_delta6)

    for x in output_list[1:3]:
        ws.cell(row=start_row + 13, column=x[0] + 1, value=x[1][1])
        # ws.cell(row=start_row+10, column=series_one[0]+2, value=series_one[1][1])
        ws.cell(row=start_row + 13, column=5, value=time_delta1)
        ws.cell(row=start_row + 13, column=6, value=1)

    for x in output_list[3:5]:
        ws.cell(row=start_row + 13 + len(master_list), column=x[0] - 1, value=x[1][1])
        ws.cell(row=start_row + 13 + len(master_list), column=5, value=time_delta2)
        ws.cell(row=start_row + 13 + len(master_list), column=6, value=2)

    for x in output_list[5:7]:
        ws.cell(row=start_row + 13 + (len(master_list) * 2), column=x[0] - 3, value=x[1][1])
        ws.cell(row=start_row + 13 + (len(master_list) * 2), column=5, value=time_delta3)
        ws.cell(row=start_row + 13 + (len(master_list) * 2), column=6, value=3)

    for x in output_list[9:11]:
        ws.cell(row=start_row + 13 + (len(master_list) * 3), column=x[0] - 7, value=x[1][1])
        ws.cell(row=start_row + 13 + (len(master_list) * 3), column=5, value=time_delta4)
        ws.cell(row=start_row + 13 + (len(master_list) * 3), column=6, value=4)

    for x in output_list[11:13]:
        ws.cell(row=start_row + 13 + (len(master_list) * 4), column=x[0] - 9, value=x[1][1])
        ws.cell(row=start_row + 13 + (len(master_list) * 4), column=5, value=time_delta5)
        ws.cell(row=start_row + 13 + (len(master_list) * 4), column=6, value=5)

    for x in output_list[13:15]:
        ws.cell(row=start_row + 13 + (len(master_list) * 5), column=x[0] - 11, value=x[1][1])
        ws.cell(row=start_row + 13 + (len(master_list) * 5), column=5, value=time_delta6)
        ws.cell(row=start_row + 13 + (len(master_list) * 5), column=6, value=6)

    return output_wb


def make_chart(data, p, output_wb):
    # wb = load_workbook(workbook)
    ws = output_wb.active
    # approval_point = data[p]['BICC approval point']
    chart = ScatterChart()
    # chart.title = 'Time Delta Schedule \n Last BC agreed by BICC ' + str(approval_point)
    chart.style = 18
    chart.x_axis.title = 'Days'
    chart.y_axis.title = 'Time Delta'
    chart.height = 11  # default is 7.5
    chart.width = 22  # default is 15

    xvalues = Reference(ws, min_col=5, min_row=1 + (len(master_list) * 3), max_row=(len(master_list) * 4) - 2)
    yvalues = Reference(ws, min_col=6, min_row=1 + (len(master_list) * 3), max_row=(len(master_list) * 4) - 2)
    series = Series(values=yvalues, xvalues=xvalues, title=None)
    chart.series.append(series)
    s2 = chart.series[0]
    s2.marker.symbol = "diamond"
    s2.marker.size = 10
    s2.marker.graphicalProperties.solidFill = "dcc7aa"  # Marker filling grey
    s2.marker.graphicalProperties.line.solidFill = "dcc7aa"  # Marker outline grey
    s2.graphicalProperties.line.noFill = True

    xvalues = Reference(ws, min_col=5, min_row=len(master_list) * 3, max_row=len(master_list) * 3)
    yvalues = Reference(ws, min_col=6, min_row=len(master_list) * 3, max_row=len(master_list) * 3)
    series = Series(values=yvalues, xvalues=xvalues, title=None)
    chart.series.append(series)
    s1 = chart.series[1]
    s1.marker.symbol = "diamond"
    s1.marker.size = 10
    s1.marker.graphicalProperties.solidFill = "f7c331"  # Marker filling yellow
    s1.marker.graphicalProperties.line.solidFill = "f7c331"  # Marker outline yellow
    s1.graphicalProperties.line.noFill = True

    xvalues = Reference(ws, min_col=5, min_row=(len(master_list) * 4) - 1, max_row=(len(master_list) * 4) - 1)
    yvalues = Reference(ws, min_col=6, min_row=(len(master_list) * 4) - 1, max_row=(len(master_list) * 4) - 1)
    series = Series(values=yvalues, xvalues=xvalues, title=None)
    chart.series.append(series)
    s3 = chart.series[2]
    s3.marker.symbol = "diamond"
    s3.marker.size = 10
    s3.marker.graphicalProperties.solidFill = "f7882f"  # Marker filling orange
    s3.marker.graphicalProperties.line.solidFill = "f7882f"  # Marker outline orange
    s3.graphicalProperties.line.noFill = True

    xvalues = Reference(ws, min_col=5, min_row=1 + (len(master_list) * 4), max_row=(len(master_list) * 5) - 2)
    yvalues = Reference(ws, min_col=6, min_row=1 + (len(master_list) * 4), max_row=(len(master_list) * 5) - 2)
    series = Series(values=yvalues, xvalues=xvalues, title=None)
    chart.series.append(series)
    s4 = chart.series[3]
    s4.marker.symbol = "diamond"
    s4.marker.size = 10
    s4.marker.graphicalProperties.solidFill = "dcc7aa"  # Marker filling grey
    s4.marker.graphicalProperties.line.solidFill = "dcc7aa"  # Marker outline grey
    s4.graphicalProperties.line.noFill = True

    xvalues = Reference(ws, min_col=5, min_row=len(master_list) * 4, max_row=len(master_list) * 4)
    yvalues = Reference(ws, min_col=6, min_row=len(master_list) * 4, max_row=len(master_list) * 4)
    series = Series(values=yvalues, xvalues=xvalues, title=None)
    chart.series.append(series)
    s5 = chart.series[4]
    s5.marker.symbol = "diamond"
    s5.marker.size = 10
    s5.marker.graphicalProperties.solidFill = "f7c331"  # Marker filling yellow
    s5.marker.graphicalProperties.line.solidFill = "f7c331"  # Marker outline yellow
    s5.graphicalProperties.line.noFill = True

    xvalues = Reference(ws, min_col=5, min_row=(len(master_list) * 5) - 1, max_row=(len(master_list) * 5) - 1)
    yvalues = Reference(ws, min_col=6, min_row=(len(master_list) * 5) - 1, max_row=(len(master_list) * 5) - 1)
    series = Series(values=yvalues, xvalues=xvalues, title=None)
    chart.series.append(series)
    s6 = chart.series[5]
    s6.marker.symbol = "diamond"
    s6.marker.size = 10
    s6.marker.graphicalProperties.solidFill = "f7882f"  # Marker filling orange
    s6.marker.graphicalProperties.line.solidFill = "f7882f"  # Marker outline orange
    s6.graphicalProperties.line.noFill = True

    xvalues = Reference(ws, min_col=5, min_row=1 + (len(master_list) * 5), max_row=(len(master_list) * 6) - 2)
    yvalues = Reference(ws, min_col=6, min_row=1 + (len(master_list) * 5), max_row=(len(master_list) * 6) - 2)
    series = Series(values=yvalues, xvalues=xvalues, title=None)
    chart.series.append(series)
    s7 = chart.series[6]
    s7.marker.symbol = "diamond"
    s7.marker.size = 10
    s7.marker.graphicalProperties.solidFill = "dcc7aa"  # Marker filling grey
    s7.marker.graphicalProperties.line.solidFill = "dcc7aa"  # Marker outline grey
    s7.graphicalProperties.line.noFill = True

    xvalues = Reference(ws, min_col=5, min_row=len(master_list) * 5, max_row=len(master_list) * 5)
    yvalues = Reference(ws, min_col=6, min_row=len(master_list) * 5, max_row=len(master_list) * 5)
    series = Series(values=yvalues, xvalues=xvalues, title=None)
    chart.series.append(series)
    s8 = chart.series[7]
    s8.marker.symbol = "diamond"
    s8.marker.size = 10
    s8.marker.graphicalProperties.solidFill = "f7c331"  # Marker filling yellow
    s8.marker.graphicalProperties.line.solidFill = "f7c331"  # Marker outline yellow
    s8.graphicalProperties.line.noFill = True

    xvalues = Reference(ws, min_col=5, min_row=(len(master_list) * 6) - 1, max_row=(len(master_list) * 6) - 1)
    yvalues = Reference(ws, min_col=6, min_row=(len(master_list) * 6) - 1, max_row=(len(master_list) * 6) - 1)
    series = Series(values=yvalues, xvalues=xvalues, title=None)
    chart.series.append(series)
    s9 = chart.series[8]
    s9.marker.symbol = "diamond"
    s9.marker.size = 10
    s9.marker.graphicalProperties.solidFill = "f7882f"  # Marker filling orange
    s9.marker.graphicalProperties.line.solidFill = "f7882f"  # Marker outline orange
    s9.graphicalProperties.line.noFill = True

    xvalues = Reference(ws, min_col=5, min_row=1 + (len(master_list) * 6), max_row=(len(master_list) * 7) - 2)
    yvalues = Reference(ws, min_col=6, min_row=1 + (len(master_list) * 6), max_row=(len(master_list) * 7) - 2)
    series = Series(values=yvalues, xvalues=xvalues, title=None)
    chart.series.append(series)
    s10 = chart.series[9]
    s10.marker.symbol = "diamond"
    s10.marker.size = 10
    s10.marker.graphicalProperties.solidFill = "dcc7aa"  # Marker filling grey
    s10.marker.graphicalProperties.line.solidFill = "dcc7aa"  # Marker outline grey
    s10.graphicalProperties.line.noFill = True

    xvalues = Reference(ws, min_col=5, min_row=len(master_list) * 6, max_row=len(master_list) * 6)
    yvalues = Reference(ws, min_col=6, min_row=len(master_list) * 6, max_row=len(master_list) * 6)
    series = Series(values=yvalues, xvalues=xvalues, title=None)
    chart.series.append(series)
    s11 = chart.series[10]
    s11.marker.symbol = "diamond"
    s11.marker.size = 10
    s11.marker.graphicalProperties.solidFill = "f7c331"  # Marker filling yellow
    s11.marker.graphicalProperties.line.solidFill = "f7c331"  # Marker outline yellow
    s11.graphicalProperties.line.noFill = True

    xvalues = Reference(ws, min_col=5, min_row=(len(master_list) * 7) - 1, max_row=(len(master_list) * 7) - 1)
    yvalues = Reference(ws, min_col=6, min_row=(len(master_list) * 7) - 1, max_row=(len(master_list) * 7) - 1)
    series = Series(values=yvalues, xvalues=xvalues, title=None)
    chart.series.append(series)
    s12 = chart.series[11]
    s12.marker.symbol = "diamond"
    s12.marker.size = 10
    s12.marker.graphicalProperties.solidFill = "f7882f"  # Marker filling orange
    s12.marker.graphicalProperties.line.solidFill = "f7882f"  # Marker outline orange
    s12.graphicalProperties.line.noFill = True

    xvalues = Reference(ws, min_col=5, min_row=1 + (len(master_list) * 7), max_row=(len(master_list) * 8) - 2)
    yvalues = Reference(ws, min_col=6, min_row=1 + (len(master_list) * 7), max_row=(len(master_list) * 8) - 2)
    series = Series(values=yvalues, xvalues=xvalues, title=None)
    chart.series.append(series)
    s13 = chart.series[12]
    s13.marker.symbol = "diamond"
    s13.marker.size = 10
    s13.marker.graphicalProperties.solidFill = "dcc7aa"  # Marker filling grey
    s13.marker.graphicalProperties.line.solidFill = "dcc7aa"  # Marker outline grey
    s13.graphicalProperties.line.noFill = True

    xvalues = Reference(ws, min_col=5, min_row=len(master_list) * 7, max_row=len(master_list) * 7)
    yvalues = Reference(ws, min_col=6, min_row=len(master_list) * 7, max_row=len(master_list) * 7)
    series = Series(values=yvalues, xvalues=xvalues, title=None)
    chart.series.append(series)
    s14 = chart.series[13]
    s14.marker.symbol = "diamond"
    s14.marker.size = 10
    s14.marker.graphicalProperties.solidFill = "f7c331"  # Marker filling yellow
    s14.marker.graphicalProperties.line.solidFill = "f7c331"  # Marker outline yellow
    s14.graphicalProperties.line.noFill = True

    xvalues = Reference(ws, min_col=5, min_row=(len(master_list) * 8) - 1, max_row=(len(master_list) * 8) - 1)
    yvalues = Reference(ws, min_col=6, min_row=(len(master_list) * 8) - 1, max_row=(len(master_list) * 8) - 1)
    series = Series(values=yvalues, xvalues=xvalues, title=None)
    chart.series.append(series)
    s15 = chart.series[14]
    s15.marker.symbol = "diamond"
    s15.marker.size = 10
    s15.marker.graphicalProperties.solidFill = "f7882f"  # Marker filling orange
    s15.marker.graphicalProperties.line.solidFill = "f7882f"  # Marker outline orange
    s15.graphicalProperties.line.noFill = True

    xvalues = Reference(ws, min_col=5, min_row=1 + (len(master_list) * 8), max_row=(len(master_list) * 9) - 2)
    yvalues = Reference(ws, min_col=6, min_row=1 + (len(master_list) * 8), max_row=(len(master_list) * 9) - 2)
    series = Series(values=yvalues, xvalues=xvalues, title=None)
    chart.series.append(series)
    s16 = chart.series[15]
    s16.marker.symbol = "diamond"
    s16.marker.size = 10
    s16.marker.graphicalProperties.solidFill = "dcc7aa"  # Marker filling grey
    s16.marker.graphicalProperties.line.solidFill = "dcc7aa"  # Marker outline grey
    s16.graphicalProperties.line.noFill = True

    xvalues = Reference(ws, min_col=5, min_row=len(master_list) * 8, max_row=len(master_list) * 8)
    yvalues = Reference(ws, min_col=6, min_row=len(master_list) * 8, max_row=len(master_list) * 8)
    series = Series(values=yvalues, xvalues=xvalues, title=None)
    chart.series.append(series)
    s17 = chart.series[16]
    s17.marker.symbol = "diamond"
    s17.marker.size = 10
    s17.marker.graphicalProperties.solidFill = "f7c331"  # Marker filling yellow
    s17.marker.graphicalProperties.line.solidFill = "f7c331"  # Marker outline yellow
    s17.graphicalProperties.line.noFill = True

    xvalues = Reference(ws, min_col=5, min_row=(len(master_list) * 9) - 1, max_row=(len(master_list) * 9) - 1)
    yvalues = Reference(ws, min_col=6, min_row=(len(master_list) * 9) - 1, max_row=(len(master_list) * 9) - 1)
    series = Series(values=yvalues, xvalues=xvalues, title=None)
    chart.series.append(series)
    s18 = chart.series[17]
    s18.marker.symbol = "diamond"
    s18.marker.size = 10
    s18.marker.graphicalProperties.solidFill = "f7882f"  # Marker filling orange
    s18.marker.graphicalProperties.line.solidFill = "f7882f"  # Marker outline orange
    s18.graphicalProperties.line.noFill = True

    ws.add_chart(chart, "I12")

    return output_wb


def get_project_names(data):
    output_list = []
    for x in data:
        output_list.append(x)
    # output_list.remove('East Coast Mainline Programme')
    # output_list.remove('Hexagon')
    return output_list


master_list = ['C:\\Users\\Standalone\\Will\\masters folder\\master_3_2016.xlsx',
               'C:\\Users\\Standalone\\Will\\masters folder\\master_4_2016.xlsx',
               'C:\\Users\\Standalone\\Will\\masters folder\\master_1_2017.xlsx',
               'C:\\Users\\Standalone\\Will\\masters folder\\master_2_2017.xlsx',
               'C:\\Users\\Standalone\\Will\\masters folder\\master_3_2017_full_report.xlsx']
# 'C:\\Users\\Standalone\\Will\\masters folder\\master_4_2017.xlsx']

master_list.reverse()

m1 = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_4_2017.xlsx')
m5 = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_3_2016.xlsx')
m1_list = get_project_names(m1)
m5_list = get_project_names(m5)
joint_list = m1_list + m5_list
project_list = list(set(joint_list))
project_list = ['M20 Lorry Area',
                'M20 Op Stack Interim Solution (Project BROCK)']
# print(project_list)

# wb = Workbook()

for p in project_list:
    wb = Workbook()
    try:
        for start_row, m in list(enumerate(master_list, start=2)):
            output_wb = ref_class_fore(m, p, start_row, wb)
    except KeyError:
        pass
    output_wb = make_chart(m1, p, output_wb)
    output_wb.save('C:\\Users\\Standalone\\Will\\{}_RCF.xlsx'.format(p))

