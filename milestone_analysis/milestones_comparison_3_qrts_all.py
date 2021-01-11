"""
Transfers MilestoneData object into an excel wb. Wb includes calculation
of time differences between milestone dates at current, last and
baseline quarter.
"""

from openpyxl import Workbook
from data_mgmt.data import MilestoneData, Master, Projects, get_master_data, root_path, get_project_information


def put_into_wb_all(milestones: MilestoneData) -> Workbook:
    wb = Workbook()
    ws = wb.active

    row_num = 2
    for i, m in enumerate(milestones.key_names):
        ws.cell(row=row_num + i, column=1).value = m.split(",")[0]  # project name
        ws.cell(row=row_num + i, column=2).value = m.split(",")[1][1:]  # milestone
        ws.cell(row=row_num + i, column=3).value = milestones.md_current[i].strftime("%d/%m/%Y")
        try:
            ws.cell(row=row_num + i, column=4).value = milestones.md_last_po[i].strftime("%d/%m/%Y")
        except AttributeError:
            pass
        try:
            ws.cell(row=row_num + i, column=5).value = milestones.md_baseline_po[i].strftime("%d/%m/%Y")
        except AttributeError:
            pass
        try:
            ws.cell(row=row_num + i, column=6).value = milestones.md_baseline_two_po[i].strftime("%d/%m/%Y")
        except AttributeError:
            pass
        # note =
        # ws.cell(row=row_num + i, column=7).value = milestones.md_baseline_two_po[i].strftime("%d/%m/%Y")


    #     try:
    #         milestone_date = tuple(milestones.current[abb][m])[0]
    #         ws.cell(row=row_num + i, column=3).value = milestone_date
    #         ws.cell(row=row_num + i, column=3).number_format = 'dd/mm/yy'm.split(",")[0]
    #     except KeyError:
    #         ws.cell(row=row_num + i, column=3).value = ''
    #
    #     try:
    #         last_date = tuple(milestones.last_quarter[abb][m])[0]
    #         ws.cell(row=row_num + i, column=4).value = (milestone_date - last_date).days
    #     except (KeyError, TypeError):
    #         ws.cell(row=row_num + i, column=4).value = ''
    #
    #     try:
    #         baseline_date = tuple(milestones.baseline_type[abb][m])[0]
    #         ws.cell(row=row_num + i, column=5).value = (milestone_date - baseline_date).days
    #     except (KeyError, TypeError):
    #         ws.cell(row=row_num + i, column=5).value = ''
    #
    #     try:
    #         notes = milestones.current[abb][m][milestone_date]
    #         ws.cell(row=row_num + i, column=7).value = notes
    #     except (IndexError, KeyError):
    #         ws.cell(row=row_num + i, column=7).value = ''
    #
    # row_num = row_num + len(milestones.current[abb].keys())

    ws.cell(row=1, column=1).value = 'Project'
    ws.cell(row=1, column=2).value = 'Milestone'
    ws.cell(row=1, column=3).value = 'Current date'
    ws.cell(row=1, column=4).value = 'Last quarter'
    ws.cell(row=1, column=5).value = 'Baseline one'
    ws.cell(row=1, column=6).value = 'Baseline two'
    # ws.cell(row=1, column=7).value = 'Notes'

    return wb


mst = Master(get_master_data(), get_project_information())  # get Master object and specify projects of interest
projects = mst.project_stage["Q2 20/21"]["Full Business Case"] + mst.project_stage["Q2 20/21"]["Outline Business Case"]
milestone_data = MilestoneData(mst, projects)  # create MilestoneData object
milestone_data.filter_chart_info(milestone_type=["Approval", "Delivery"])
run = put_into_wb_all(milestone_data)
run.save(root_path/"output/milestone_data_output.xlsx")