'''

This programme checks if there has been changes in reported milestone keys.

Output document:
1) excel workbook with all project milestone information for each project

See instructions below.

Note: all master data is taken from the data file. Make sure this is up to date and that all relevant data is in
the import statement.

'''

import datetime
from openpyxl import Workbook
from analysis.engine_functions import all_milestone_data_bulk
from analysis.data import list_of_masters_all, red_text, bc_index, root_path

def check_m_keys_in_excel_single():
    '''
    function for placing all information into an excel wb

    :param t_data_one: dictionary containing latest milestone data for projects.
    dictionary structure is {'project name': {'milestone name': datetime.date: 'notes'}}
    :param t_data_two: dictionary containing last milestone date for projects. same structure as above.
    :param t_data_three: dictionary containing baseline milestone date for projects. same structure as above.

    '''

    wb = Workbook()

    for i, project_name in enumerate(list_of_masters_all[0].projects):
        '''worksheet is created for each project'''
        ws = wb.create_sheet(project_name[:29], i)  # creating worksheets
        ws.title = project_name[:29]  # title of worksheet

        row_num = 2

        one = list(p_current_milestones_data[project_name].keys())
        [x for x in one if x is not None].sort()
        two = list(p_last_milestones_data[project_name].keys())
        [x for x in two if x is not None].sort()

        p_baseline_milestone_data = all_milestone_data_bulk([project_name],
                                                            list_of_masters_all[bc_index[project_name][2]])

        three = list(p_baseline_milestone_data[project_name].keys())
        [x for x in three if x is not None].sort()

        long = longest_list(one, two, three)
        for i in range(0, len(long)):
            ws.cell(row=row_num + i, column=1).value = project_name
            try:
                ws.cell(row=row_num + i, column=2).value = one[i]
            except IndexError:
                pass
            try:
                ws.cell(row=row_num + i, column=3).value = two[i]
                if two[i] not in one:
                    ws.cell(row=row_num + i, column=3).font = red_text
            except IndexError:
                pass
            try:
                ws.cell(row=row_num + i, column=4).value = three[i]
                if three[i] not in one:
                    ws.cell(row=row_num + i, column=4).font = red_text
            except IndexError:
                pass


        row_heading_list = ['Project', 'This quarter', 'Last quarter', 'Baseline quarter', 'KEY MATCH']
        for i, heading in enumerate(row_heading_list):
            ws.cell(row=1, column=i+1).value = heading

        column_ltr_list = ['A', 'B', 'C', 'D', 'E']
        for ltr in (column_ltr_list):
            ws.column_dimensions[ltr].width = 40

    return wb

def longest_list(one, two, three):
    '''
    Function. Helper for the above. Calculates the longest list and therefore the one to use for iteration.

    :param one: list_one
    :param two: list_two
    :param three: list_three

    Returns the longest list.
    '''

    list_list = [one, two, three]
    a = len(one)
    b = len(two)
    c = len(three)

    out = [a, b, c]
    out.sort()
    for x in list_list:
        if out[-1] == len(x):
            return x

p_current_milestones_data = all_milestone_data_bulk(list_of_masters_all[0].projects, list_of_masters_all[0])
p_last_milestones_data = all_milestone_data_bulk(list_of_masters_all[0].projects, list_of_masters_all[1])

''' RUNNING THE PROGRAMME'''

'''Note: all master data is taken from the data file. Make sure this is up to date and that all relevant data is in 
the import statement.'''

'''TWO. Specify date after which project milestones should be returned. NOTE: Python date format is (YYYY,MM,DD)'''
start_date = datetime.date(2019, 11, 1)

'''THREE. the following for statement prompts the programme to run'''

output = check_m_keys_in_excel_single()
output.save(root_path/'output/checking_milestones_q4_1920.xlsx')