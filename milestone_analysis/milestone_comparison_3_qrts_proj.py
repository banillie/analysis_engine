'''

This programme calculates the time difference between reported milestones

Output document:
1) excel workbook with all project milestone information for each project

See instructions below.

Note: all master data is taken from the data file. Make sure this is up to date and that all relevant data is in
the import statement.

'''

#TODO solve problem re filtering in excel when values have + sign in front of the them

from openpyxl import Workbook
from analysis.engine_functions import all_milestone_data_bulk, ap_p_milestone_data_bulk, assurance_milestone_data_bulk, \
    project_time_difference, filter_project_group
from analysis.data import list_of_masters_all, bc_index, baseline_bc_stamp, root_path


def put_into_wb_all_single(function):
    '''

    Function that places all data into excel wb for this programme

    project_name_list: list of project to return data for
    t_data: dictionary containing milestone data for projects.
    dictionary structure is {'project name': {'milestone name': datetime.date: 'notes'}}
    td_data: dictionary containing time_delta milestone data for projects.
    dictionary structure is {'project name': {'milestone name': 'time delta info'}}
    td_data_two: dictionary containing second time_delta data for projects.
    same structure as for td_data.
    wb: blank excel wb

    '''

    '''get all milestone data'''
    p_current_milestones = function(list_of_masters_all[0].projects, list_of_masters_all[0])
    p_last_milestones = function(list_of_masters_all[1].projects, list_of_masters_all[1])

    '''calculate time current and last quarter'''
    first_diff_data = project_time_difference(p_current_milestones, p_last_milestones)

    wb = Workbook()

    for x, project_name in enumerate(list_of_masters_all[0].projects):
        '''worksheet is created for each project'''
        ws = wb.create_sheet(project_name, x)  # creating worksheets
        ws.title = project_name  # title of worksheet

        row_num = 2
        for i, milestone in enumerate(first_diff_data[project_name].keys()):
            ws.cell(row=row_num + i, column=1).value = project_name
            ws.cell(row=row_num + i, column=2).value = milestone
            try:
                milestone_date = tuple(p_current_milestones[project_name][milestone])[0]
                ws.cell(row=row_num + i, column=3).value = milestone_date
                ws.cell(row=row_num + i, column=3).number_format = 'dd/mm/yy'
            except KeyError:
                ws.cell(row=row_num + i, column=3).value = 0

            try:
                value = first_diff_data[project_name][milestone]
                try:
                    if int(value) > 0:
                        ws.cell(row=row_num + i, column=4).value = '+' + str(value) + ' (days)'
                    elif int(value) < 0:
                        ws.cell(row=row_num + i, column=4).value = str(value) + ' (days)'
                    elif int(value) == 0:
                        ws.cell(row=row_num + i, column=4).value = value
                except ValueError:
                    ws.cell(row=row_num + i, column=4).value = value
            except KeyError:
                ws.cell(row=row_num + i, column=4).value = 0

            p_oldest_milestones = function([project_name], list_of_masters_all[bc_index[project_name][2]])
            second_diff_data = project_time_difference(p_current_milestones, p_oldest_milestones)

            try:
                value = second_diff_data[project_name][milestone]
                try:
                    if int(value) > 0:
                        ws.cell(row=row_num + i, column=5).value = '+' + str(value) + ' (days)'
                    elif int(value) < 0:
                        ws.cell(row=row_num + i, column=5).value = str(value) + ' (days)'
                    elif int(value) == 0:
                        ws.cell(row=row_num + i, column=5).value = value
                except ValueError:
                    ws.cell(row=row_num + i, column=5).value = value
            except KeyError:
                ws.cell(row=row_num + i, column=5).value = 0

            try:
                milestone_date = tuple(p_current_milestones[project_name][milestone])[0]
                ws.cell(row=row_num + i, column=6).value = p_current_milestones[project_name][milestone][milestone_date] # provided notes
            except IndexError:
                ws.cell(row=row_num + i, column=6).value = 0

        ws.cell(row=1, column=1).value = 'Project'
        ws.cell(row=1, column=2).value = 'Milestone'
        ws.cell(row=1, column=3).value = 'Date'
        ws.cell(row=1, column=4).value = '3/m change (days)'
        ws.cell(row=1, column=5).value = 'baseline change (days)'
        ws.cell(row=1, column=6).value = 'Notes'

        ws.cell(row=1, column=8).value = 'data baseline quarter'
        ws.cell(row=2, column=8).value = baseline_bc_stamp[project_name][0][1]

    return wb


''' RUNNING THE PROGRAMME'''

'''
Only one part of programme is to be amended each quarter. place which ever quarter information being produced at 
end of output file name e.g. q4_1920. Note make sure file ends with .xlsx format

Note when code completes it may state:
UserWarning: Title is more than 31 characters. Some applications may not be able to read the file 
warnings.warn("Title is more than 31 characters. Some applications may not be able to read the file"). 
However, you can ignore and open the file as usual.
'''

output = put_into_wb_all_single(all_milestone_data_bulk)
output.save(root_path/'output/ind_project_milestone_analysis_q4_1920.xlsx')